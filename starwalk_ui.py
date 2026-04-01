"""
SharkNinja Review Analyst + Symptomizer — Combined App
=======================================================
Primary shell : SNeviews AI (review analytics)
5th tab       : Symptomizer (row-level AI tagging)
Shared        : OpenAI key/model, CSS style, review dataframe, export
Style         : Symptomizer gradient + chips + badge rows
"""

from __future__ import annotations

# ── stdlib ──────────────────────────────────────────────────────────────────
import difflib
import gc
import hashlib
import html
import io
import json
import math
import os
import random
import re
import sqlite3
import tempfile
import textwrap
import time
from collections import Counter
from dataclasses import dataclass
from datetime import date
from typing import (
    Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple
)
from urllib.parse import urlparse

# ── third-party ──────────────────────────────────────────────────────────────
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import (
    column_index_from_string, get_column_letter
)
from openpyxl.worksheet.worksheet import Worksheet
from plotly.subplots import make_subplots

try:
    from openai import OpenAI
    _HAS_OPENAI = True
except ImportError:
    OpenAI = None  # type: ignore
    _HAS_OPENAI = False

try:
    import tiktoken
    _HAS_TIKTOKEN = True
except ImportError:
    tiktoken = None  # type: ignore
    _HAS_TIKTOKEN = False

# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG + SHARED CSS
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="SharkNinja Review Analyst",
    layout="wide",
)

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
  :root {
    --bg1: rgba(124,58,237,.12);
    --bg2: rgba(6,182,212,.10);
    --line: #e5e7eb;
    --muted: #667085;
    --text: #101828;
    --brand: #7c3aed;
    --brand-2: #06b6d4;
    --good: #16a34a;
    --bad: #dc2626;
    --warn: #d97706;
  }
  html, body, .stApp {
    font-family: 'Inter', system-ui, -apple-system, Segoe UI, Roboto, sans-serif;
    color: var(--text);
  }
  .stApp {
    background:
      radial-gradient(1000px 420px at 0% -10%, var(--bg1), transparent 60%),
      radial-gradient(900px 420px at 100% 0%,  var(--bg2), transparent 60%);
  }
  /* ── cards ── */
  .hero-card, .metric-card, .info-card, .report-card {
    background: linear-gradient(180deg,rgba(255,255,255,.97),rgba(255,255,255,.90));
    border: 1px solid rgba(226,232,240,.95);
    border-radius: 18px;
    padding: 14px 16px;
    box-shadow: 0 6px 20px rgba(16,24,40,.06);
    margin-bottom: .6rem;
  }
  .hero-card { border-radius: 22px; padding: 18px 20px 16px; }
  .hero-title { font-size:22px; font-weight:800; letter-spacing:-.02em; }
  .hero-sub   { color:var(--muted); font-size:13px; margin-top:4px; }
  .metric-card { min-height:148px; display:flex; flex-direction:column; justify-content:space-between; }
  .metric-label { color:#64748b; font-size:11px; text-transform:uppercase; letter-spacing:.08em; }
  .metric-value { font-size:clamp(1.55rem,2vw,2rem); font-weight:800; color:#16213e; line-height:1.05; }
  .metric-sub   { color:#4b5563; font-size:.83rem; line-height:1.3; }
  .section-title { font-size:18px; font-weight:800; margin:6px 0 10px; }
  .section-sub   { color:var(--muted); font-size:13px; margin:-2px 0 12px; }
  /* ── chips / badges ── */
  .badge-row, .chip-wrap { display:flex; gap:8px; flex-wrap:wrap; }
  .badge, .chip {
    padding:6px 10px; border-radius:999px; font-size:12.5px; line-height:1;
    border:1px solid #e5e7eb; background:#fff;
    box-shadow:0 1px 2px rgba(16,24,40,.04);
  }
  .chip.blue   { background:#eff6ff; border-color:#bfdbfe; }
  .chip.green  { background:#ecfdf3; border-color:#bbf7d0; }
  .chip.red    { background:#fff1f2; border-color:#fecdd3; }
  .chip.yellow { background:#fffbeb; border-color:#fde68a; }
  .chip.purple { background:#f5f3ff; border-color:#ddd6fe; }
  .chip.gray   { background:#f8fafc; border-color:#e2e8f0; }
  /* ── hero stat grid ── */
  .hero-grid {
    display:grid; grid-template-columns:repeat(5,minmax(0,1fr));
    gap:12px; margin-top:14px;
  }
  .hero-stat {
    background:#fff; border:1px solid #e5e7eb; border-radius:18px;
    padding:14px 16px; box-shadow:0 2px 8px rgba(16,24,40,.05);
  }
  .hero-stat.accent { border-color:rgba(124,58,237,.35); box-shadow:0 6px 18px rgba(124,58,237,.14); }
  .hero-stat .label { color:#64748b; font-size:11px; text-transform:uppercase; letter-spacing:.08em; }
  .hero-stat .value { font-size:28px; font-weight:800; margin-top:2px; }
  /* ── run plan ── */
  .run-plan { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:10px; }
  .kv { background:#fff; border:1px solid #e5e7eb; border-radius:14px; padding:12px 14px; }
  .kv .k { color:#667085; font-size:11px; text-transform:uppercase; letter-spacing:.06em; }
  .kv .v { font-size:20px; font-weight:800; margin-top:2px; }
  /* ── misc ── */
  .danger-box {
    background:linear-gradient(180deg,rgba(254,242,242,.92),rgba(254,226,226,.72));
    border:1px solid #fecaca; border-radius:16px; padding:12px 14px;
  }
  .tiny { font-size:12px; color:#667085; }
  .mono { font-family:ui-monospace,SFMono-Regular,Menlo,Monaco,Consolas,monospace; }
  mark.hl { background:#fde68a; padding:0 .16em; border-radius:.25em; }
  /* ── thinking overlay ── */
  .thinking-overlay {
    position:fixed; inset:0; background:rgba(15,23,42,.30);
    display:flex; align-items:center; justify-content:center; z-index:99999;
  }
  .thinking-card {
    width:min(430px,92vw); background:rgba(255,255,255,.98);
    border:1px solid rgba(49,51,63,.12); border-radius:20px;
    box-shadow:0 24px 60px rgba(15,23,42,.18); padding:1.2rem 1.3rem; text-align:center;
  }
  .thinking-spinner {
    width:40px; height:40px; border:4px solid rgba(17,24,39,.14);
    border-top-color:#111827; border-radius:50%;
    margin:0 auto .8rem; animation:thinking-spin .9s linear infinite;
  }
  .thinking-title { color:#16213e; font-weight:700; font-size:1.08rem; margin-bottom:.3rem; }
  .thinking-sub   { color:#4b5563; font-size:.95rem; line-height:1.35; }
  @keyframes thinking-spin { to { transform:rotate(360deg); } }
  /* ── buttons ── */
  .stButton > button {
    height:42px; border-radius:12px; font-weight:700;
    border:1px solid #d0d5dd;
    background:linear-gradient(180deg,#ffffff,#f8fafc);
    box-shadow:0 1px 2px rgba(16,24,40,.04);
  }
  .stButton > button:hover {
    border-color:rgba(124,58,237,.35);
    box-shadow:0 4px 12px rgba(124,58,237,.14);
  }
  /* ── AI response ── */
  .ai-response-html { color:#16213e; font-size:.86rem; line-height:1.5; }
  .ai-response-html h2,.ai-response-html h3,.ai-response-html h4 {
    font-size:.97rem; line-height:1.25; margin:.4rem 0 .32rem; color:#16213e;
  }
  .ai-response-html p,.ai-response-html li { font-size:.86rem; line-height:1.48; margin-bottom:.38rem; }
  .ai-response-html ul,.ai-response-html ol { padding-left:1.08rem; margin:.1rem 0 .35rem; }
  /* ── inline evidence chips ── */
  .inline-evidence-chip {
    position:relative; display:inline-flex; align-items:center;
    border:1px solid rgba(49,51,63,.14); border-radius:999px;
    padding:.12rem .44rem; background:rgba(245,247,250,.98);
    color:#16213e; font-size:.73rem; line-height:1.18;
    cursor:help; white-space:nowrap; margin-right:.1rem;
  }
  .inline-evidence-chip:hover::after {
    content:attr(data-tooltip);
    position:absolute; left:50%; top:calc(100% + 9px);
    transform:translateX(-50%);
    width:min(340px,72vw); background:rgba(17,24,39,.96);
    color:#f9fafb; border-radius:12px; padding:.65rem .75rem;
    font-size:.75rem; line-height:1.32;
    box-shadow:0 16px 34px rgba(15,23,42,.26);
    white-space:normal; z-index:1000; text-align:left;
  }
  @media(max-width:1100px) {
    .hero-grid { grid-template-columns:repeat(2,minmax(0,1fr)); }
    .run-plan  { grid-template-columns:1fr; }
  }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

APP_TITLE              = "SharkNinja Review Analyst"
DEFAULT_PASSKEY        = "caC6wVBHos09eVeBkLIniLUTzrNMMH2XMADEhpHe1ewUw"
DEFAULT_DISPLAYCODE    = "15973_3_0-en_us"
DEFAULT_API_VERSION    = "5.5"
DEFAULT_PAGE_SIZE      = 100
DEFAULT_SORT           = "SubmissionTime:desc"
DEFAULT_CONTENT_LOCALES = (
    "en_US,ar*,zh*,hr*,cs*,da*,nl*,en*,et*,fi*,fr*,de*,el*,he*,hu*,"
    "id*,it*,ja*,ko*,lv*,lt*,ms*,no*,pl*,pt*,ro*,sk*,sl*,es*,sv*,th*,"
    "tr*,vi*,en_AU,en_CA,en_GB"
)
BAZAARVOICE_ENDPOINT   = "https://api.bazaarvoice.com/data/reviews.json"

# SNeviews model options (used in shared sidebar)
MODEL_OPTIONS      = ["gpt-4o-mini", "gpt-4o", "gpt-4.1", "gpt-5"]
DEFAULT_MODEL      = "gpt-4o-mini"
REASONING_OPTIONS  = ["none", "low", "medium", "high"]
DEFAULT_REASONING  = "low"

NON_VALUES = {"<NA>", "NA", "N/A", "NONE", "-", "", "NAN", "NULL"}

THEME_KEYWORDS: Dict[str, List[str]] = {
    "Cooking performance":      ["crispy","cook","cooking","air fry","bake","broil","reheat","dehydrate","temperature","preheat","evenly","juicy","frozen"],
    "Ease of use":              ["easy","simple","intuitive","buttons","controls","instructions","setup","user friendly","learning curve"],
    "Capacity and footprint":   ["size","capacity","counter","countertop","space","basket","tray","fits","large","small","compact"],
    "Cleaning and maintenance": ["clean","cleanup","dishwasher","wash","mess","grease","sticky","scrub"],
    "Build quality":            ["broke","broken","durable","quality","plastic","flimsy","stopped working","defect","replacement","warranty"],
    "Noise, odor, and heat":    ["noise","noisy","loud","odor","smell","hot","heat","steam","fan"],
    "Design and aesthetics":    ["design","looks","sleek","beautiful","style","appearance","color"],
    "Value and price":          ["price","worth","value","expensive","cost","money","deal"],
    "Service and shipping":     ["shipping","delivery","customer service","support","return","replacement","arrived","damaged","missing"],
}

STOPWORDS = {
    "a","about","after","again","all","also","am","an","and","any","are","as","at",
    "be","because","been","before","being","best","better","but","by","can","could",
    "did","do","does","don","down","even","every","for","from","get","got","great",
    "had","has","have","he","her","here","hers","him","his","how","i","if","in",
    "into","is","it","its","just","like","love","made","make","many","me","more",
    "most","much","my","new","no","not","now","of","on","one","only","or","other",
    "our","out","over","product","really","so","some","than","that","the","their",
    "them","then","there","these","they","this","to","too","use","used","using",
    "very","was","we","well","were","what","when","which","while","with","would",
    "you","your",
}

PERSONAS: Dict[str, Dict[str, Any]] = {
    "Product Development": {
        "blurb": "Translates reviews into product and feature decisions.",
        "prompt": "Create a report for the product development team. Highlight what customers love, unmet needs, feature gaps, usability friction, and concrete roadmap opportunities. End with the top 5 product actions ranked by impact.",
        "instructions": "You are a senior product strategy analyst. Focus on feature prioritization, user experience, jobs-to-be-done, and roadmap implications. Cite review IDs for important claims.",
    },
    "Quality Engineer": {
        "blurb": "Focuses on failure modes, defects, durability, and root-cause signals.",
        "prompt": "Create a report for a quality engineer. Identify defect patterns, reliability risks, cleaning issues, performance inconsistencies, and probable root-cause hypotheses. Separate confirmed evidence from inference.",
        "instructions": "You are a senior quality and reliability analyst. Prioritize failure modes, defect language, repeat complaints, severity, and probable root causes. Cite review IDs for material claims.",
    },
    "Consumer Insights": {
        "blurb": "Extracts sentiment drivers, purchase motivations, and voice-of-customer insights.",
        "prompt": "Create a report for the consumer insights team. Summarize key sentiment drivers, barriers to adoption, purchase motivations, key use cases, and how tone changes across star ratings and incentivized vs non-incentivized reviews.",
        "instructions": "You are a consumer insights lead. Synthesize sentiment drivers, motivations, barriers, and language. Use concise, executive-ready writing and cite review IDs for important findings.",
    },
}

# Symptomizer column layout
DET_LETTERS  = ["K","L","M","N","O","P","Q","R","S","T"]
DEL_LETTERS  = ["U","V","W","X","Y","Z","AA","AB","AC","AD"]
DET_INDEXES  = [column_index_from_string(c) for c in DET_LETTERS]
DEL_INDEXES  = [column_index_from_string(c) for c in DEL_LETTERS]
META_ORDER   = [("Safety","AE"),("Reliability","AF"),("# of Sessions","AG")]
META_INDEXES = {name: column_index_from_string(col) for name, col in META_ORDER}
AI_DET_HEADERS = [f"AI Symptom Detractor {i}" for i in range(1, 11)]
AI_DEL_HEADERS = [f"AI Symptom Delighter {i}" for i in range(1, 11)]
AI_META_HEADERS = ["AI Safety","AI Reliability","AI # of Sessions"]

SAFETY_ENUM    = ["Not Mentioned","Concern","Positive"]
RELIABILITY_ENUM = ["Not Mentioned","Negative","Neutral","Positive"]
SESSIONS_ENUM  = ["0","1","2–3","4–9","10+","Unknown"]

DEFAULT_PRIORITY_DELIGHTERS = [
    "Overall Satisfaction","Ease Of Use","Effective Results","Visible Improvement",
    "Time Saver","Comfort","Value","Reliability",
]
DEFAULT_PRIORITY_DETRACTORS = [
    "Poor Results","Ease Of Use","Reliability Issue","High Cost","Irritation",
    "Battery Problem","High Noise","Cleaning Difficulty","Setup Issue",
    "Connectivity Issue","Safety Concern",
]


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

class ReviewDownloaderError(Exception):
    pass

@dataclass
class ReviewBatchSummary:
    product_url: str
    product_id: str
    total_reviews: int
    page_size: int
    requests_needed: int
    reviews_downloaded: int


def _safe_text(value: Any, default: str = "") -> str:
    if value is None: return default
    if isinstance(value, (list,tuple,set,dict,pd.Series,pd.DataFrame,pd.Index)): return default
    try:
        missing = pd.isna(value)
    except Exception:
        missing = False
    if isinstance(missing, bool) and missing: return default
    text = str(value).strip()
    return default if text.lower() in {"nan","none","null","<na>"} else text

def _safe_int(value: Any, default: int = 0) -> int:
    try: return int(float(value))
    except Exception: return default

def _safe_bool(value: Any, default: bool = False) -> bool:
    if value is None: return default
    if isinstance(value, bool): return value
    text = _safe_text(value).lower()
    if text in {"true","1","yes","y","t"}: return True
    if text in {"false","0","no","n","f",""}: return False
    return default

def _safe_mean(series: pd.Series) -> Optional[float]:
    if series.empty: return None
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    return float(numeric.mean()) if not numeric.empty else None

def _safe_pct(num: float, den: float) -> float:
    return 0.0 if not den else float(num)/float(den)

def _fmt_money(x: float) -> str:
    try: return f"${x:,.4f}" if x < 1 else f"${x:,.2f}"
    except Exception: return "$0.00"

def _fmt_secs(sec: float) -> str:
    sec = max(0.0, float(sec or 0))
    m = int(sec // 60); s = int(round(sec - m*60))
    return f"{m}:{s:02d}"

def _canon(s: str) -> str:
    return " ".join(str(s).split()).lower().strip()

def _canon_simple(s: str) -> str:
    return "".join(ch for ch in _canon(s) if ch.isalnum())

def _safe_html(s: Any) -> str:
    return html.escape(str(s or ""))

def _chip_html(items: List[Tuple[str,str]]) -> str:
    if not items: return "<span class='chip gray'>No active filters</span>"
    parts = ["<div class='chip-wrap'>"]
    for text, color in items:
        parts.append(f"<span class='chip {color}'>{_safe_html(text)}</span>")
    parts.append("</div>")
    return "".join(parts)

def _is_missing_value(value: Any) -> bool:
    if value is None: return True
    if isinstance(value, (list,tuple,set,dict,pd.Series,pd.DataFrame,pd.Index)): return False
    try: missing = pd.isna(value)
    except Exception: return False
    return bool(missing) if isinstance(missing,(bool,int)) else False

def _format_metric_number(value: Optional[float], digits: int = 2) -> str:
    if value is None or _is_missing_value(value): return "n/a"
    return f"{value:.{digits}f}"

def _format_pct(value: Optional[float], digits: int = 1) -> str:
    if value is None or _is_missing_value(value): return "n/a"
    return f"{100*float(value):.{digits}f}%"

def _truncate_text(text: str, max_chars: int = 420) -> str:
    text = re.sub(r"\s+"," ",_safe_text(text)).strip()
    return text if len(text) <= max_chars else text[:max_chars-3].rstrip()+"..."

def _normalize_text(text: str) -> str:
    return re.sub(r"\s+"," ",str(text).lower()).strip()

def _tokenize(text: str) -> List[str]:
    return [t for t in re.findall(r"[a-z0-9']+",_normalize_text(text))
            if len(t) > 2 and t not in STOPWORDS]

def _slugify(text: str, fallback: str = "custom") -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9]+","_",_safe_text(text).lower())
    cleaned = re.sub(r"_+","_",cleaned).strip("_") or fallback
    return ("prompt_"+cleaned if cleaned[0].isdigit() else cleaned)[:64]

def _first_non_empty(series: pd.Series) -> str:
    for v in series.astype(str):
        v = _safe_text(v)
        if v and v.lower() != "nan": return v
    return ""

def _clean_text(x: Any) -> str:
    if pd.isna(x): return ""
    return str(x).strip()

def _is_filled(val: Any) -> bool:
    if pd.isna(val): return False
    s = str(val).strip()
    return s != "" and s.upper() not in NON_VALUES

def _estimate_tokens(text: str, model_id: str = "gpt-4o-mini") -> int:
    s = str(text or "")
    if not s: return 0
    if _HAS_TIKTOKEN:
        try:
            enc = tiktoken.get_encoding("cl100k_base")
            return int(len(enc.encode(s)))
        except Exception: pass
    return int(max(1, math.ceil(len(s)/4)))


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED OPENAI HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _get_api_key() -> Optional[str]:
    try:
        if "OPENAI_API_KEY" in st.secrets: return str(st.secrets["OPENAI_API_KEY"])
        if "openai" in st.secrets and st.secrets["openai"].get("api_key"):
            return str(st.secrets["openai"]["api_key"])
    except Exception: pass
    return os.getenv("OPENAI_API_KEY")

def _get_openai_client() -> Optional[Any]:
    key = _get_api_key()
    if not (_HAS_OPENAI and key): return None
    try:
        return OpenAI(api_key=key, timeout=60, max_retries=3)
    except TypeError:
        try: return OpenAI(api_key=key)
        except Exception: return None

def _shared_model() -> str:
    return st.session_state.get("shared_model", DEFAULT_MODEL)

def _shared_reasoning() -> str:
    return st.session_state.get("shared_reasoning", DEFAULT_REASONING)

def _show_thinking(message: str):
    ph = st.empty()
    ph.markdown(f"""
    <div class="thinking-overlay">
      <div class="thinking-card">
        <div class="thinking-spinner"></div>
        <div class="thinking-title">OpenAI is working</div>
        <div class="thinking-sub">{_safe_html(message)}</div>
      </div>
    </div>""", unsafe_allow_html=True)
    return ph

def _safe_json_load(s: str) -> Dict[str, Any]:
    s = (s or "").strip()
    if not s: return {}
    try: return json.loads(s)
    except Exception: pass
    try:
        i = s.find("{"); j = s.rfind("}")
        if i >= 0 and j > i: return json.loads(s[i:j+1])
    except Exception: pass
    return {}

def _chat_complete(
    client: Any, *, model: str, messages: List[Dict[str,str]],
    temperature: float = 0.0, response_format: Optional[Dict] = None,
    max_tokens: int = 1200,
) -> str:
    if client is None: return ""
    kwargs: Dict[str,Any] = dict(model=model, temperature=temperature,
                                  messages=messages, max_tokens=max_tokens)
    if response_format: kwargs["response_format"] = response_format
    try:
        resp = client.chat.completions.create(**kwargs)
        return (resp.choices[0].message.content or "").strip()
    except Exception: return ""


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — DATA LAYER
# ═══════════════════════════════════════════════════════════════════════════════

def _get_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent":
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/123.0 Safari/537.36"})
    return s

def _extract_product_id_from_url(url: str) -> Optional[str]:
    path = urlparse(url).path
    m = re.search(r"/([A-Za-z0-9_-]+)\.html(?:$|[?#])", path)
    if m:
        c = m.group(1).strip().upper()
        if re.fullmatch(r"[A-Z0-9_-]{3,}", c): return c
    return None

def _extract_product_id_from_html(h: str) -> Optional[str]:
    for pat in [r'Item\s*No\.?\s*([A-Z0-9_-]{3,})',r'"productId"\s*:\s*"([A-Z0-9_-]{3,})"',
                r'"sku"\s*:\s*"([A-Z0-9_-]{3,})"',r'"model"\s*:\s*"([A-Z0-9_-]{3,})"']:
        m = re.search(pat, h, flags=re.IGNORECASE)
        if m: return m.group(1).strip().upper()
    soup = BeautifulSoup(h, "html.parser")
    text = soup.get_text(" ", strip=True)
    for pat in [r"Item\s*No\.?\s*([A-Z0-9_-]{3,})", r"Model\s*:?\s*([A-Z0-9_-]{3,})"]:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m: return m.group(1).strip().upper()
    return None

def _fetch_reviews_page(session, *, product_id, passkey, displaycode,
                        api_version, page_size, offset, sort, content_locales):
    params = {
        "resource":"reviews","action":"REVIEWS_N_STATS",
        "filter":[f"productid:eq:{product_id}",
                  f"contentlocale:eq:{content_locales}","isratingsonly:eq:false"],
        "filter_reviews":f"contentlocale:eq:{content_locales}",
        "include":"authors,products,comments","filteredstats":"reviews",
        "Stats":"Reviews","limit":int(page_size),"offset":int(offset),
        "limit_comments":3,"sort":sort,
        "passkey":passkey,"apiversion":api_version,"displaycode":displaycode,
    }
    resp = session.get(BAZAARVOICE_ENDPOINT, params=params, timeout=45)
    resp.raise_for_status()
    payload = resp.json()
    if payload.get("HasErrors"):
        raise ReviewDownloaderError(f"Bazaarvoice error: {payload.get('Errors')}")
    return payload

def _is_incentivized(review: Dict) -> bool:
    badges = [str(b).lower() for b in (review.get("BadgesOrder") or [])]
    if any("incentivized" in b for b in badges): return True
    ctx = review.get("ContextDataValues") or {}
    if isinstance(ctx, dict):
        for k, v in ctx.items():
            if "incentivized" in str(k).lower():
                flag = str((v.get("Value","") if isinstance(v,dict) else v)).strip().lower()
                if flag in {"","true","1","yes"}: return True
    return False

def _flatten_review(r: Dict) -> Dict:
    photos = r.get("Photos") or []
    urls = []
    for p in photos:
        sz = p.get("Sizes") or {}
        for sn in ["large","normal","thumbnail"]:
            u = (sz.get(sn) or {}).get("Url")
            if u: urls.append(u); break
    syn = r.get("SyndicationSource") or {}
    return {
        "review_id": r.get("Id"), "product_id": r.get("ProductId"),
        "original_product_name": r.get("OriginalProductName"),
        "title": _safe_text(r.get("Title")), "review_text": _safe_text(r.get("ReviewText")),
        "rating": r.get("Rating"), "is_recommended": r.get("IsRecommended"),
        "user_nickname": r.get("UserNickname"), "author_id": r.get("AuthorId"),
        "user_location": r.get("UserLocation"), "content_locale": r.get("ContentLocale"),
        "submission_time": r.get("SubmissionTime"),
        "moderation_status": r.get("ModerationStatus"),
        "campaign_id": r.get("CampaignId"), "source_client": r.get("SourceClient"),
        "is_featured": r.get("IsFeatured"), "is_syndicated": r.get("IsSyndicated"),
        "syndication_source_name": syn.get("Name"),
        "is_ratings_only": r.get("IsRatingsOnly"),
        "total_positive_feedback_count": r.get("TotalPositiveFeedbackCount"),
        "badges": ", ".join(str(x) for x in (r.get("BadgesOrder") or [])),
        "context_data_json": json.dumps(r.get("ContextDataValues") or {}, ensure_ascii=False),
        "photos_count": len(photos), "photo_urls": " | ".join(urls),
        "incentivized_review": _is_incentivized(r),
        "raw_json": json.dumps(r, ensure_ascii=False),
    }

def _ensure_columns(df: pd.DataFrame, cols: Sequence[str]) -> pd.DataFrame:
    for c in cols:
        if c not in df.columns: df[c] = pd.NA
    return df

def _extract_age_group(val: Any) -> Optional[str]:
    if val is None or (isinstance(val,float) and pd.isna(val)): return None
    payload = val
    if isinstance(payload, str):
        stripped = payload.strip()
        if not stripped: return None
        try: payload = json.loads(stripped)
        except Exception: return None
    if not isinstance(payload, dict): return None
    for k, raw in payload.items():
        if "age" not in str(k).lower(): continue
        candidate = raw.get("Value") or raw.get("Label") if isinstance(raw,dict) else raw
        candidate = _safe_text(candidate)
        if candidate and candidate.lower() not in {"nan","none","null","unknown","prefer not to say"}:
            return candidate
    return None

def _finalize_reviews_df(df: pd.DataFrame) -> pd.DataFrame:
    required = [
        "review_id","product_id","base_sku","sku_item","product_or_sku",
        "original_product_name","title","review_text","rating","is_recommended",
        "content_locale","submission_time","submission_date","submission_month",
        "incentivized_review","is_syndicated","photos_count","photo_urls",
        "title_and_text","retailer","post_link","age_group","user_nickname",
        "user_location","total_positive_feedback_count","source_system","source_file",
    ]
    df = _ensure_columns(df.copy(), required)
    if df.empty:
        for c in ["has_photos","has_media","review_length_chars","review_length_words",
                  "rating_label","year_month_sort"]:
            if c not in df.columns: df[c] = pd.Series(dtype="object")
        return df

    df["review_id"] = df["review_id"].fillna("").astype(str).str.strip()
    missing_ids = df["review_id"].eq("") | df["review_id"].str.lower().isin({"nan","none","null"})
    if missing_ids.any():
        df.loc[missing_ids,"review_id"] = [f"review_{i+1}" for i in range(int(missing_ids.sum()))]

    if "context_data_json" in df.columns:
        df["age_group"] = df["age_group"].fillna(df["context_data_json"].map(_extract_age_group))

    df["rating"]            = pd.to_numeric(df["rating"], errors="coerce")
    df["incentivized_review"]= df["incentivized_review"].fillna(False).astype(bool)
    df["is_syndicated"]      = df["is_syndicated"].fillna(False).astype(bool)
    df["photos_count"]       = pd.to_numeric(df["photos_count"],errors="coerce").fillna(0).astype(int)
    df["title"]              = df["title"].fillna("").astype(str)
    df["review_text"]        = df["review_text"].fillna("").astype(str)
    df["submission_time"]    = pd.to_datetime(df["submission_time"],errors="coerce",utc=True).dt.tz_convert(None)
    df["submission_date"]    = df["submission_time"].dt.date
    df["submission_month"]   = df["submission_time"].dt.to_period("M").astype(str)
    df["content_locale"]     = df["content_locale"].fillna("").astype(str).replace({"":pd.NA})
    df["base_sku"]           = df.get("base_sku",pd.Series(dtype="str")).fillna("").astype(str).str.strip()
    df["sku_item"]           = df.get("sku_item",pd.Series(dtype="str")).fillna("").astype(str).str.strip()
    df["product_id"]         = df["product_id"].fillna("").astype(str).str.strip()
    fallback = df["base_sku"].where(df["base_sku"].ne(""), df["product_id"])
    df["product_or_sku"]     = df["sku_item"].where(df["sku_item"].ne(""), fallback)
    df["product_or_sku"]     = df["product_or_sku"].fillna("").astype(str).str.strip().replace({"":pd.NA})
    df["title_and_text"]     = (df["title"].str.strip()+" "+df["review_text"].str.strip()).str.strip()
    df["has_photos"]         = df["photos_count"] > 0
    df["has_media"]          = df["has_photos"]
    df["review_length_chars"]= df["review_text"].str.len()
    df["review_length_words"]= df["review_text"].str.split().str.len().fillna(0).astype(int)
    df["rating_label"]       = df["rating"].map(lambda x: f"{int(x)} star" if pd.notna(x) else "Unknown")
    df["year_month_sort"]    = pd.to_datetime(df["submission_month"],format="%Y-%m",errors="coerce")
    sort_cols = [c for c in ["submission_time","review_id"] if c in df.columns]
    if sort_cols:
        df = df.sort_values(sort_cols,ascending=[False,False],na_position="last").reset_index(drop=True)
    return df

def _pick_first_col(df: pd.DataFrame, aliases: Sequence[str]) -> Optional[str]:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    for a in aliases:
        c = lookup.get(str(a).strip().lower())
        if c: return c
    return None

def _series_from_aliases(df: pd.DataFrame, aliases: Sequence[str]) -> pd.Series:
    c = _pick_first_col(df, aliases)
    if c is None: return pd.Series([pd.NA]*len(df), index=df.index)
    return df[c]

def _parse_flag(value: Any, *, pos: Sequence[str], neg: Sequence[str]) -> Any:
    text = _safe_text(value).lower()
    if text in {"","nan","none","null","n/a"}: return pd.NA
    if any(text==t.lower() for t in neg): return False
    if any(text==t.lower() for t in pos): return True
    if text.startswith(("not ","non ")): return False
    return True

def _normalize_uploaded_df(raw_df: pd.DataFrame, *, source_name: str = "") -> pd.DataFrame:
    working = raw_df.copy()
    working.columns = [str(c).strip() for c in working.columns]
    n = pd.DataFrame(index=working.index)
    n["review_id"]              = _series_from_aliases(working,["Event Id","Event ID","Review ID","Review Id","Id"])
    n["product_id"]             = _series_from_aliases(working,["Base SKU","Product ID","Product Id","ProductId","BaseSKU"])
    n["base_sku"]               = _series_from_aliases(working,["Base SKU","BaseSKU"])
    n["sku_item"]               = _series_from_aliases(working,["SKU Item","SKU","Child SKU","Variant SKU","Item Number","Item No"])
    n["original_product_name"]  = _series_from_aliases(working,["Product Name","Product","Name"])
    n["review_text"]            = _series_from_aliases(working,["Review Text","Review","Body","Content"])
    n["title"]                  = _series_from_aliases(working,["Title","Review Title","Headline"])
    n["post_link"]              = _series_from_aliases(working,["Post Link","URL","Review URL","Product URL"])
    n["rating"]                 = _series_from_aliases(working,["Rating (num)","Rating","Stars","Star Rating"])
    n["submission_time"]        = _series_from_aliases(working,["Opened date","Opened Date","Submission Time","Review Date","Date"])
    n["content_locale"]         = _series_from_aliases(working,["Content Locale","Locale","Location","Country"])
    n["retailer"]               = _series_from_aliases(working,["Retailer","Merchant","Channel"])
    n["age_group"]              = _series_from_aliases(working,["Age Group","Age","Age Range"])
    n["user_location"]          = _series_from_aliases(working,["Location","Country"])
    n["user_nickname"]          = pd.NA
    n["total_positive_feedback_count"] = pd.NA
    n["is_recommended"]         = pd.NA
    n["photos_count"]           = 0
    n["photo_urls"]             = pd.NA
    n["source_file"]            = source_name or pd.NA
    n["source_system"]          = "Uploaded file"
    seeded = _series_from_aliases(working,["Seeded Flag","Seeded","Incentivized"])
    n["incentivized_review"]    = seeded.map(lambda v: _parse_flag(v,
        pos=["seeded","incentivized","yes","true","1"],
        neg=["not seeded","not incentivized","no","false","0"]))
    syndicated = _series_from_aliases(working,["Syndicated Flag","Syndicated"])
    n["is_syndicated"]          = syndicated.map(lambda v: _parse_flag(v,
        pos=["syndicated","yes","true","1"],neg=["not syndicated","no","false","0"]))
    return _finalize_reviews_df(n)

def _read_uploaded_file(uploaded_file: Any) -> pd.DataFrame:
    fname = getattr(uploaded_file,"name","uploaded_file")
    raw = uploaded_file.getvalue()
    suffix = fname.lower().rsplit(".",1)[-1] if "." in fname else "csv"
    if suffix == "csv":
        try:    raw_df = pd.read_csv(io.BytesIO(raw))
        except UnicodeDecodeError: raw_df = pd.read_csv(io.BytesIO(raw),encoding="latin-1")
    elif suffix in {"xlsx","xls","xlsm"}:
        raw_df = pd.read_excel(io.BytesIO(raw))
    else:
        raise ReviewDownloaderError(f"Unsupported file type: {fname}")
    if raw_df.empty: raise ReviewDownloaderError(f"{fname} is empty.")
    return _normalize_uploaded_df(raw_df, source_name=fname)

def _load_uploaded_files(uploaded_files: Sequence[Any]) -> Dict[str,Any]:
    if not uploaded_files:
        raise ReviewDownloaderError("Upload at least one file.")
    with st.spinner("Reading uploaded files…"):
        frames = [_read_uploaded_file(f) for f in uploaded_files]
    combined = pd.concat(frames, ignore_index=True)
    combined["review_id"] = combined["review_id"].astype(str)
    combined = combined.drop_duplicates(subset=["review_id"],keep="first").reset_index(drop=True)
    combined = _finalize_reviews_df(combined)
    inferred_id = _first_non_empty(combined["base_sku"].fillna("")) or \
                  _first_non_empty(combined["product_id"].fillna("")) or "UPLOADED_REVIEWS"
    file_names = [getattr(f,"name","uploaded_file") for f in uploaded_files]
    src_label = file_names[0] if len(file_names)==1 else f"{len(file_names)} uploaded files"
    summary = ReviewBatchSummary(
        product_url="", product_id=inferred_id,
        total_reviews=len(combined), page_size=max(len(combined),1),
        requests_needed=0, reviews_downloaded=len(combined),
    )
    return {"summary":summary,"reviews_df":combined,"source_type":"uploaded","source_label":src_label}

def _load_product_reviews(product_url: str) -> Dict[str,Any]:
    product_url = product_url.strip()
    if not re.match(r"^https?://",product_url,flags=re.IGNORECASE):
        product_url = "https://"+product_url
    session = _get_session()
    with st.spinner("Loading product page…"):
        resp = session.get(product_url, timeout=30); resp.raise_for_status()
        product_html = resp.text
    product_id = _extract_product_id_from_url(product_url) or _extract_product_id_from_html(product_html)
    if not product_id:
        raise ReviewDownloaderError("Could not find product ID on page.")
    with st.spinner("Checking review volume…"):
        payload = _fetch_reviews_page(session, product_id=product_id,
            passkey=DEFAULT_PASSKEY, displaycode=DEFAULT_DISPLAYCODE,
            api_version=DEFAULT_API_VERSION, page_size=1, offset=0,
            sort=DEFAULT_SORT, content_locales=DEFAULT_CONTENT_LOCALES)
        total = int(payload.get("TotalResults",0))
    progress = st.progress(0.0, text="Downloading reviews…")
    status   = st.empty()
    offsets  = list(range(0,total,DEFAULT_PAGE_SIZE))
    raw_reviews: List[Dict] = []
    for i, offset in enumerate(offsets,1):
        status.info(f"Pulling page {i}/{len(offsets)}")
        page = _fetch_reviews_page(session, product_id=product_id,
            passkey=DEFAULT_PASSKEY, displaycode=DEFAULT_DISPLAYCODE,
            api_version=DEFAULT_API_VERSION, page_size=DEFAULT_PAGE_SIZE,
            offset=offset, sort=DEFAULT_SORT, content_locales=DEFAULT_CONTENT_LOCALES)
        raw_reviews.extend(page.get("Results") or [])
        progress.progress(i/len(offsets))
    status.success(f"Downloaded {len(raw_reviews)} reviews.")
    rows = [_flatten_review(r) for r in raw_reviews]
    df   = _finalize_reviews_df(pd.DataFrame(rows))
    if not df.empty:
        df["review_id"]      = df["review_id"].astype(str)
        df["product_or_sku"] = df.get("product_or_sku",pd.Series(index=df.index,dtype="object")).fillna(product_id)
        df["base_sku"]       = df.get("base_sku",pd.Series(index=df.index,dtype="object")).fillna(product_id)
        df["product_id"]     = df["product_id"].fillna(product_id)
    summary = ReviewBatchSummary(
        product_url=product_url, product_id=product_id,
        total_reviews=total, page_size=DEFAULT_PAGE_SIZE,
        requests_needed=len(offsets), reviews_downloaded=len(df),
    )
    return {"summary":summary,"reviews_df":df,"source_type":"bazaarvoice","source_label":product_url}


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════

def _compute_metrics(df: pd.DataFrame) -> Dict[str,Any]:
    n = len(df)
    if n == 0:
        return dict(review_count=0,avg_rating=None,avg_rating_non_incentivized=None,
                    pct_low_star=0.0,pct_one_star=0.0,pct_two_star=0.0,pct_five_star=0.0,
                    pct_incentivized=0.0,pct_with_photos=0.0,pct_syndicated=0.0,
                    recommend_rate=None,median_review_words=None,
                    non_incentivized_count=0,low_star_count=0)
    ni = df[~df["incentivized_review"].fillna(False)]
    rec_base = df[df["is_recommended"].notna()]
    rec_rate = _safe_pct(int(rec_base["is_recommended"].astype(bool).sum()),len(rec_base)) if not rec_base.empty else None
    mw = float(df["review_length_words"].median()) if "review_length_words" in df.columns and not df["review_length_words"].dropna().empty else None
    low = df["rating"].isin([1,2])
    return dict(
        review_count=n, avg_rating=_safe_mean(df["rating"]),
        avg_rating_non_incentivized=_safe_mean(ni["rating"]),
        pct_low_star=_safe_pct(int(low.sum()),n),
        pct_one_star=_safe_pct(int((df["rating"]==1).sum()),n),
        pct_two_star=_safe_pct(int((df["rating"]==2).sum()),n),
        pct_five_star=_safe_pct(int((df["rating"]==5).sum()),n),
        pct_incentivized=_safe_pct(int(df["incentivized_review"].fillna(False).sum()),n),
        pct_with_photos=_safe_pct(int(df["has_photos"].fillna(False).sum()),n),
        pct_syndicated=_safe_pct(int(df["is_syndicated"].fillna(False).sum()),n),
        recommend_rate=rec_rate, median_review_words=mw,
        non_incentivized_count=len(ni),
        low_star_count=int(low.sum()),
    )

def _rating_distribution(df: pd.DataFrame) -> pd.DataFrame:
    base = pd.DataFrame({"rating":[1,2,3,4,5]})
    if df.empty:
        base["review_count"]=0; base["share"]=0.0; return base
    grouped = (df.dropna(subset=["rating"]).assign(rating=lambda x:x["rating"].astype(int))
               .groupby("rating",as_index=False).size().rename(columns={"size":"review_count"}))
    merged = base.merge(grouped,how="left",on="rating").fillna({"review_count":0})
    merged["review_count"] = merged["review_count"].astype(int)
    merged["share"] = merged["review_count"] / max(len(df),1)
    return merged

def _monthly_trend(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return pd.DataFrame(columns=["submission_month","review_count","avg_rating","month_start"])
    return (df.dropna(subset=["submission_time"])
            .assign(month_start=lambda x:x["submission_time"].dt.to_period("M").dt.to_timestamp())
            .groupby("month_start",as_index=False)
            .agg(review_count=("review_id","count"),avg_rating=("rating","mean"))
            .assign(submission_month=lambda x:x["month_start"].dt.strftime("%Y-%m"))
            .sort_values("month_start"))

def _compute_themes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["theme","mention_count","mention_rate",
                                     "avg_rating_when_mentioned","low_star_mentions","high_star_mentions"])
    texts = df["title_and_text"].fillna("").astype(str).map(_normalize_text)
    rows = []
    for theme, keywords in THEME_KEYWORDS.items():
        mask = texts.map(lambda t: any(kw in t for kw in keywords))
        sub  = df[mask]
        rows.append(dict(theme=theme, mention_count=int(mask.sum()),
            mention_rate=_safe_pct(int(mask.sum()),len(df)),
            avg_rating_when_mentioned=_safe_mean(sub["rating"]),
            low_star_mentions=int(sub["rating"].isin([1,2]).sum()),
            high_star_mentions=int(sub["rating"].isin([4,5]).sum())))
    return pd.DataFrame(rows).sort_values(["mention_count","low_star_mentions"],ascending=[False,False])

def _top_terms(texts: Iterable[str], *, top_n: int = 12) -> pd.DataFrame:
    counter: Counter = Counter()
    for t in texts: counter.update(_tokenize(t))
    return pd.DataFrame([{"term":t,"count":c} for t,c in counter.most_common(top_n)])

def _apply_filters(df: pd.DataFrame, *, selected_ratings, incentivized_mode,
                   selected_products, selected_locales, recommendation_mode,
                   syndicated_mode, media_mode, date_range, text_query) -> pd.DataFrame:
    if df.empty: return df.copy()
    f = df.copy()
    if selected_ratings: f = f[f["rating"].isin(selected_ratings)]
    if selected_products and "product_or_sku" in f.columns:
        f = f[f["product_or_sku"].fillna("").isin(selected_products)]
    if incentivized_mode == "Non-incentivized only":
        f = f[~f["incentivized_review"].fillna(False)]
    elif incentivized_mode == "Incentivized only":
        f = f[f["incentivized_review"].fillna(False)]
    if selected_locales:
        f = f[f["content_locale"].fillna("Unknown").isin(selected_locales)]
    if recommendation_mode == "Recommended only":
        f = f[f["is_recommended"].fillna(False)]
    elif recommendation_mode == "Not recommended only":
        f = f[f["is_recommended"].notna() & ~f["is_recommended"].fillna(False)]
    if syndicated_mode == "Syndicated only": f = f[f["is_syndicated"].fillna(False)]
    elif syndicated_mode == "Non-syndicated only": f = f[~f["is_syndicated"].fillna(False)]
    if media_mode == "With photos only": f = f[f["has_photos"].fillna(False)]
    elif media_mode == "No photos only": f = f[~f["has_photos"].fillna(False)]
    if date_range and date_range[0] and date_range[1] and "submission_date" in f.columns:
        f = f[f["submission_date"].notna() &
              (f["submission_date"]>=date_range[0]) &
              (f["submission_date"]<=date_range[1])]
    q = text_query.strip()
    if q: f = f[f["title_and_text"].fillna("").str.contains(re.escape(q),case=False,na=False,regex=True)]
    return f.reset_index(drop=True)

def _build_filter_options(df: pd.DataFrame) -> Dict[str,Any]:
    valid_dates = df["submission_date"].dropna() if "submission_date" in df.columns else pd.Series(dtype="object")
    product_groups = sorted({str(v).strip() for v in df["product_or_sku"].dropna().astype(str)
                             if str(v).strip() and str(v).strip().lower() not in {"nan","none"}}
                            ) if not df.empty else []
    return dict(
        ratings=[1,2,3,4,5],
        product_groups=product_groups,
        locales=sorted(str(l) for l in df["content_locale"].dropna().unique()) if not df.empty else [],
        min_date=valid_dates.min() if not valid_dates.empty else None,
        max_date=valid_dates.max() if not valid_dates.empty else None,
    )

def _describe_filters(*, selected_ratings, selected_products, review_source_mode,
                      selected_locales, recommendation_mode, date_range, text_query) -> str:
    parts: List[str] = []
    if selected_ratings and set(selected_ratings)!={1,2,3,4,5}:
        parts.append("ratings="+",".join(str(r) for r in selected_ratings))
    if selected_products:
        parts.append("sku="+", ".join(selected_products[:4])+("…" if len(selected_products)>4 else ""))
    if review_source_mode != "All reviews": parts.append(f"source={review_source_mode.lower()}")
    if selected_locales: parts.append("locales="+", ".join(selected_locales))
    if recommendation_mode != "All": parts.append(f"recommendation={recommendation_mode.lower()}")
    if date_range and date_range[0] and date_range[1]:
        parts.append(f"dates={date_range[0]} to {date_range[1]}")
    if text_query.strip(): parts.append(f'text="{text_query.strip()}"')
    return "; ".join(parts) if parts else "No active filters"

def _product_display_name(summary: ReviewBatchSummary, df: pd.DataFrame) -> str:
    if not df.empty and "original_product_name" in df.columns:
        name = _first_non_empty(df["original_product_name"].fillna(""))
        if name: return name
    return summary.product_id


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — AI ANALYST
# ═══════════════════════════════════════════════════════════════════════════════

GENERAL_ANALYST_INSTRUCTIONS = textwrap.dedent("""
    You are SharkNinja Review Analyst, an internal voice-of-customer assistant.
    Help product development, quality engineering, and consumer insights teams understand the review base.
    Prioritize the supplied review text over generic product assumptions.
    Ground every material claim in the supplied review dataset.
    Do not invent counts, quotes, or trends that are not supported by the evidence.
    When evidence is mixed or weak, say so clearly.
    Use markdown. Cite supporting review IDs in parentheses, e.g. (review_ids: 12345, 67890).
    Default to a crisp response. Stay under roughly 375 words unless asked for a deep dive.
""").strip()

def _build_persona_instructions(persona_name: Optional[str]) -> str:
    if not persona_name: return GENERAL_ANALYST_INSTRUCTIONS
    p = PERSONAS[persona_name]
    return textwrap.dedent(f"""
        {p['instructions']}
        Ground every important finding in the supplied review dataset.
        Do not invent facts, counts, or quotes.
        If evidence is mixed or weak, say so explicitly.
        Use markdown. Cite review IDs in parentheses, e.g. (review_ids: 12345, 67890).
        Keep the report compact and decision-ready. Stay under ~375 words unless asked.
        End with a short action list tailored to the audience.
    """).strip()

def _select_relevant_reviews(df: pd.DataFrame, question: str, max_reviews: int = 18) -> pd.DataFrame:
    if df.empty: return df.copy()
    working = df.copy()
    working["search_blob"] = working["title_and_text"].fillna("").astype(str).map(_normalize_text)
    qtokens = _tokenize(question)
    def score(row):
        s = 0.0; text = row["search_blob"]
        for t in qtokens:
            if t in text: s += 3+text.count(t)
        rating = row.get("rating")
        if any(t in {"defect","broken","issue","problem","negative","bad"} for t in qtokens):
            if pd.notna(rating): s += max(0,6-float(rating))
        if not _safe_bool(row.get("incentivized_review"),False): s += 0.5
        if pd.notna(row.get("review_length_words")): s += min(float(row.get("review_length_words",0))/60,2)
        return s
    working["_score"] = working.apply(score, axis=1)
    ranked = working.sort_values(["_score","submission_time"],ascending=[False,False],na_position="last")
    combined = pd.concat([
        ranked.head(max_reviews),
        df[df["rating"].isin([1,2])].head(max_reviews//3 or 1),
        df[df["rating"].isin([4,5])].head(max_reviews//3 or 1),
    ],ignore_index=True).drop_duplicates(subset=["review_id"])
    return combined.head(max_reviews).drop(columns=["search_blob","_score"],errors="ignore")

def _review_snippet_rows(df: pd.DataFrame, *, max_reviews: int = 18) -> List[Dict]:
    rows = []
    for _, row in df.head(max_reviews).iterrows():
        rows.append(dict(
            review_id=_safe_text(row.get("review_id")),
            rating=_safe_int(row.get("rating"),0) if pd.notna(row.get("rating")) else None,
            incentivized_review=_safe_bool(row.get("incentivized_review"),False),
            content_locale=_safe_text(row.get("content_locale")),
            submission_date=_safe_text(row.get("submission_date")),
            title=_truncate_text(row.get("title",""),120),
            snippet=_truncate_text(row.get("review_text",""),520),
        ))
    return rows

def _build_ai_context(*, overall_df, filtered_df, summary, filter_description, question) -> str:
    overall_metrics  = _compute_metrics(overall_df)
    filtered_metrics = _compute_metrics(filtered_df)
    rating_df        = _rating_distribution(filtered_df)
    monthly_df       = _monthly_trend(filtered_df).tail(12)
    relevant = _select_relevant_reviews(filtered_df, question, max_reviews=18)
    recent   = filtered_df.sort_values(["submission_time","review_id"],ascending=[False,False],na_position="last").head(10)
    low_star = filtered_df[filtered_df["rating"].isin([1,2])].head(8)
    hi_star  = filtered_df[filtered_df["rating"].isin([4,5])].head(8)
    evidence = pd.concat([relevant,recent,low_star,hi_star],ignore_index=True).drop_duplicates(subset=["review_id"]).head(28)
    return json.dumps({
        "product": dict(product_id=summary.product_id, product_url=summary.product_url,
                        product_name=_product_display_name(summary,overall_df)),
        "analysis_scope": dict(current_filter_description=filter_description,
                               overall_review_count=len(overall_df),filtered_review_count=len(filtered_df)),
        "metric_snapshot": dict(overall=overall_metrics, filtered=filtered_metrics,
                                rating_distribution_filtered=rating_df.to_dict(orient="records"),
                                monthly_trend_filtered=monthly_df.to_dict(orient="records")),
        "review_text_evidence": _review_snippet_rows(evidence, max_reviews=28),
    }, ensure_ascii=False, indent=2, default=str)

def _call_analyst(*, question, overall_df, filtered_df, summary, filter_description,
                  chat_history, persona_name=None) -> str:
    client = _get_openai_client()
    if client is None: raise ReviewDownloaderError("No OpenAI API key configured.")
    instructions = _build_persona_instructions(persona_name)
    ai_context   = _build_ai_context(overall_df=overall_df,filtered_df=filtered_df,
                                     summary=summary,filter_description=filter_description,question=question)
    msgs: List[Dict[str,str]] = [{"role":m["role"],"content":m["content"]} for m in list(chat_history)[-8:]]
    msgs.append({"role":"user","content":f"User request:\n{question}\n\nReview dataset context (JSON):\n{ai_context}"})
    result = _chat_complete(client, model=_shared_model(), messages=[
        {"role":"system","content":instructions}, *msgs
    ], temperature=0.0, max_tokens=1100)
    if not result: raise ReviewDownloaderError("OpenAI returned an empty answer.")
    return result


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — REVIEW PROMPT (existing tab 4)
# ═══════════════════════════════════════════════════════════════════════════════

REVIEW_PROMPT_STARTER_ROWS = [
    {"column_name":"perceived_loudness",
     "prompt":"How is product loudness described? Use Positive, Negative, Neutral, or Not Mentioned.",
     "labels":"Positive, Negative, Neutral, Not Mentioned"},
    {"column_name":"reliability_risk_signal",
     "prompt":"Does the review mention a product reliability or durability risk? Use Risk Mentioned, Positive Reliability, or Not Mentioned.",
     "labels":"Risk Mentioned, Positive Reliability, Not Mentioned"},
]

def _default_prompt_df() -> pd.DataFrame:
    return pd.DataFrame([REVIEW_PROMPT_STARTER_ROWS[0]])

def _normalize_prompt_definitions(prompt_df: pd.DataFrame, existing_columns: Sequence[str]) -> List[Dict]:
    if prompt_df is None or prompt_df.empty: return []
    normalized = []
    seen: Set[str] = set()
    existing_set = {str(c) for c in existing_columns}
    for _, row in prompt_df.fillna("").iterrows():
        raw_prompt = _safe_text(row.get("prompt"))
        raw_labels = _safe_text(row.get("labels"))
        raw_column = _safe_text(row.get("column_name"))
        if not raw_prompt and not raw_labels and not raw_column: continue
        if not raw_prompt: raise ReviewDownloaderError("Each Review Prompt row needs a prompt.")
        if not raw_labels: raise ReviewDownloaderError("Each Review Prompt row needs labels.")
        labels = [l.strip() for l in raw_labels.split(",") if l.strip()]
        deduped = list(dict.fromkeys(labels))
        if "Not Mentioned" not in deduped and len(deduped) <= 7: deduped.append("Not Mentioned")
        if len(deduped) < 2: raise ReviewDownloaderError("Each prompt needs at least two labels.")
        col = _slugify(raw_column or raw_prompt)
        if col in existing_set and col not in {"review_id"}: col = f"{col}_ai"
        base = col; suffix = 2
        while col in seen: col = f"{base}_{suffix}"; suffix += 1
        seen.add(col)
        normalized.append(dict(column_name=col,
                                display_name=col.replace("_"," ").title(),
                                prompt=raw_prompt, labels=deduped,
                                labels_csv=", ".join(deduped)))
    return normalized

def _build_tagging_schema(prompt_definitions: Sequence[Dict]) -> Dict:
    item_props = {"review_id":{"type":"string"}}
    required   = ["review_id"]
    for p in prompt_definitions:
        item_props[p["column_name"]] = {"type":"string","enum":list(p["labels"])}
        required.append(p["column_name"])
    return {"type":"object","additionalProperties":False,
            "properties":{"results":{"type":"array","items":{"type":"object",
                "additionalProperties":False,"properties":item_props,"required":required}}},
            "required":["results"]}

def _classify_chunk(*, client, chunk_df, prompt_definitions) -> pd.DataFrame:
    prompt_count = max(len(prompt_definitions),1)
    max_out = int(max(1800,min(12000,450+len(chunk_df)*(18+12*prompt_count))))
    reviews_payload = [dict(review_id=_safe_text(row.get("review_id")),
        rating=_safe_int(row.get("rating"),0) if pd.notna(row.get("rating")) else None,
        title=_truncate_text(row.get("title",""),200),
        review_text=_truncate_text(row.get("review_text",""),1000),
        incentivized_review=_safe_bool(row.get("incentivized_review"),False))
        for _, row in chunk_df.iterrows()]
    prompt_payload = [dict(column_name=p["column_name"],prompt=p["prompt"],labels=p["labels"])
                      for p in prompt_definitions]
    instructions = textwrap.dedent("""
        You are a deterministic review-tagging engine.
        For each review and each prompt definition, return exactly one allowed label.
        Base each label only on the supplied review content.
        If the review does not mention the topic, use Not Mentioned when available.
    """).strip()
    result_text = _chat_complete(client, model=_shared_model(),
        messages=[{"role":"system","content":instructions},
                  {"role":"user","content":json.dumps({"prompt_definitions":prompt_payload,"reviews":reviews_payload})}],
        temperature=0.0,
        response_format={"type":"json_schema","name":"review_prompt_tagging",
                          "schema":_build_tagging_schema(prompt_definitions),"strict":True},
        max_tokens=max_out)
    data = _safe_json_load(result_text)
    output_rows = data.get("results") or []
    out_df = pd.DataFrame(output_rows)
    if out_df.empty: raise ReviewDownloaderError("OpenAI returned no row-level prompt results.")
    out_df["review_id"] = out_df["review_id"].astype(str)
    expected = set(chunk_df["review_id"].astype(str))
    returned = set(out_df["review_id"].astype(str))
    if expected != returned:
        missing = sorted(expected-returned); extra = sorted(returned-expected)
        raise ReviewDownloaderError(f"Incomplete batch. Missing: {missing[:5]} Extra: {extra[:5]}")
    for p in prompt_definitions:
        if p["column_name"] not in out_df.columns:
            raise ReviewDownloaderError(f"OpenAI omitted column: {p['column_name']}")
    return out_df

def _run_review_prompt_tagging(*, client, source_df, prompt_definitions, chunk_size) -> pd.DataFrame:
    if source_df.empty: raise ReviewDownloaderError("No reviews in scope.")
    chunks = list(range(0,len(source_df),chunk_size))
    progress = st.progress(0.0,text="Preparing AI review prompt run…")
    status   = st.empty()
    outputs: List[pd.DataFrame] = []
    for i, start in enumerate(chunks,1):
        chunk_df = source_df.iloc[start:start+chunk_size].copy()
        status.info(f"Classifying reviews {start+1}-{min(start+chunk_size,len(source_df))} of {len(source_df)}")
        outputs.append(_classify_chunk(client=client,chunk_df=chunk_df,prompt_definitions=prompt_definitions))
        progress.progress(i/len(chunks))
    status.success(f"Finished tagging {len(source_df):,} reviews.")
    combined = pd.concat(outputs,ignore_index=True).drop_duplicates(subset=["review_id"],keep="last")
    return combined

def _merge_prompt_results(overall_df, prompt_results_df, prompt_definitions) -> pd.DataFrame:
    updated = overall_df.copy()
    rid_series = updated["review_id"].astype(str)
    lookup = prompt_results_df.set_index("review_id")
    for p in prompt_definitions:
        col = p["column_name"]
        if col not in updated.columns: updated[col] = pd.NA
        mapping = lookup[col].to_dict()
        new_vals = rid_series.map(mapping)
        updated[col] = new_vals.where(new_vals.notna(), updated[col])
    return updated

def _summarize_prompt_results(prompt_results_df, prompt_definitions, source_df=None) -> pd.DataFrame:
    merged = prompt_results_df.copy()
    merged["review_id"] = merged["review_id"].astype(str)
    if source_df is not None and not source_df.empty and "review_id" in source_df.columns:
        lookup = source_df[[c for c in ["review_id","rating"] if c in source_df.columns]].copy()
        lookup["review_id"] = lookup["review_id"].astype(str)
        merged = merged.merge(lookup,on="review_id",how="left")
    rows = []
    total = max(len(prompt_results_df),1)
    for p in prompt_definitions:
        col = p["column_name"]
        for label in p["labels"]:
            sub = merged[merged[col]==label]
            rows.append(dict(column_name=col,display_name=p["display_name"],label=str(label),
                review_count=len(sub),share=_safe_pct(len(sub),total),
                avg_rating=_safe_mean(sub["rating"]) if "rating" in sub.columns else None))
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def _autosize_worksheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    for idx, col in enumerate(df.columns, 1):
        series = df[col].head(250).fillna("").astype(str)
        max_len = max([len(str(col))]+[len(v) for v in series.tolist()])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len+2,48)

def _dataframe_for_sql(df: pd.DataFrame) -> pd.DataFrame:
    sql_df = df.copy()
    for col in sql_df.columns:
        if pd.api.types.is_datetime64_any_dtype(sql_df[col]):
            sql_df[col] = sql_df[col].dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        elif pd.api.types.is_bool_dtype(sql_df[col]):
            sql_df[col] = sql_df[col].astype(int)
    return sql_df

def _build_master_excel(summary: ReviewBatchSummary, reviews_df: pd.DataFrame, *,
                        prompt_definitions=None, prompt_summary_df=None,
                        prompt_scope_label="") -> bytes:
    metrics  = _compute_metrics(reviews_df)
    rating_df= _rating_distribution(reviews_df)
    monthly_df=_monthly_trend(reviews_df)
    summary_df = pd.DataFrame([dict(
        product_name=_product_display_name(summary,reviews_df),
        product_id=summary.product_id, product_url=summary.product_url,
        reviews_downloaded=summary.reviews_downloaded,
        avg_rating=metrics.get("avg_rating"),
        avg_rating_non_incentivized=metrics.get("avg_rating_non_incentivized"),
        pct_low_star=metrics.get("pct_low_star"),
        pct_incentivized=metrics.get("pct_incentivized"),
        generated_utc=pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
    )])
    priority_cols = ["review_id","product_id","rating","incentivized_review",
                     "is_recommended","submission_time","content_locale","title","review_text"]
    prompt_cols = [p["column_name"] for p in (prompt_definitions or []) if p["column_name"] in reviews_df.columns]
    ordered = [c for c in priority_cols+prompt_cols if c in reviews_df.columns]
    remaining = [c for c in reviews_df.columns if c not in ordered]
    export_reviews = reviews_df[ordered+remaining]
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        sheets = {"Summary":summary_df,"Reviews":export_reviews,
                  "RatingDistribution":rating_df,"ReviewVolume":monthly_df}
        if prompt_definitions:
            sheets["ReviewPromptDefinitions"] = pd.DataFrame([
                dict(column_name=p["column_name"],display_name=p["display_name"],
                     prompt=p["prompt"],labels=", ".join(p["labels"]),scope=prompt_scope_label)
                for p in prompt_definitions])
        if prompt_summary_df is not None and not prompt_summary_df.empty:
            sheets["ReviewPromptSummary"] = prompt_summary_df
        for sname, df in sheets.items():
            df.to_excel(writer, sheet_name=sname, index=False)
            _autosize_worksheet(writer.sheets[sname], df)
    out.seek(0)
    return out.getvalue()

def _get_master_bundle(summary: ReviewBatchSummary, reviews_df: pd.DataFrame,
                       prompt_artifacts: Optional[Dict]) -> Dict:
    prompt_defs    = (prompt_artifacts or {}).get("definitions") or []
    prompt_sum_df  = (prompt_artifacts or {}).get("summary_df")
    prompt_scope   = (prompt_artifacts or {}).get("scope_label","")
    key = json.dumps(dict(product_id=summary.product_id,review_count=len(reviews_df),
                           columns=sorted(str(c) for c in reviews_df.columns),
                           prompt_sig=(prompt_artifacts or {}).get("definition_signature")),
                     sort_keys=True)
    bundle = st.session_state.get("master_export_bundle")
    if bundle and bundle.get("key")==key: return bundle
    excel_bytes = _build_master_excel(summary,reviews_df,
                                       prompt_definitions=prompt_defs,
                                       prompt_summary_df=prompt_sum_df,
                                       prompt_scope_label=prompt_scope)
    ts = pd.Timestamp.utcnow().strftime("%Y%m%d_%H%M%S")
    bundle = dict(key=key, excel_bytes=excel_bytes,
                  excel_name=f"{summary.product_id}_review_workspace_{ts}.xlsx")
    st.session_state["master_export_bundle"] = bundle
    return bundle


# ═══════════════════════════════════════════════════════════════════════════════
#  SYMPTOMIZER — HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _get_symptom_whitelists_from_file(file_bytes: bytes) -> Tuple[List[str],List[str],Dict[str,List[str]]]:
    bio = io.BytesIO(file_bytes)
    try: df_sym = pd.read_excel(bio, sheet_name="Symptoms")
    except Exception: return [], [], {}
    if df_sym is None or df_sym.empty: return [], [], {}
    df_sym.columns = [str(c).strip() for c in df_sym.columns]
    lowcols = {c.lower():c for c in df_sym.columns}
    alias_col = next((lowcols[c] for c in ["aliases","alias"] if c in lowcols), None)
    label_col = next((lowcols[c] for c in ["symptom","label","name","item"] if c in lowcols), None)
    type_col  = next((lowcols[c] for c in ["type","polarity","category","side"] if c in lowcols), None)
    pos_tags  = {"delighter","delighters","positive","pos","pros"}
    neg_tags  = {"detractor","detractors","negative","neg","cons"}
    def _clean(s: pd.Series) -> List[str]:
        vals = s.dropna().astype(str).str.strip()
        out: List[str] = []; seen: Set[str] = set()
        for v in vals:
            if v and v not in seen: seen.add(v); out.append(v)
        return out
    delighters: List[str] = []; detractors: List[str] = []; alias_map: Dict[str,List[str]] = {}
    if label_col and type_col:
        df_sym[type_col] = df_sym[type_col].astype(str).str.lower().str.strip()
        delighters = _clean(df_sym.loc[df_sym[type_col].isin(pos_tags),label_col])
        detractors = _clean(df_sym.loc[df_sym[type_col].isin(neg_tags),label_col])
        if alias_col:
            for _, row in df_sym.iterrows():
                lbl = str(row.get(label_col,"")).strip()
                als = str(row.get(alias_col,"")).strip()
                if lbl:
                    alias_map[lbl] = [p.strip() for p in als.replace(",","|").split("|") if p.strip()] if als else []
    else:
        for lc, orig in lowcols.items():
            if "delight" in lc or "positive" in lc or lc=="pros": delighters.extend(_clean(df_sym[orig]))
            if "detract" in lc or "negative" in lc or lc=="cons": detractors.extend(_clean(df_sym[orig]))
        delighters = list(dict.fromkeys(delighters)); detractors = list(dict.fromkeys(detractors))
    return delighters, detractors, alias_map

def _ensure_ai_columns(df: pd.DataFrame) -> pd.DataFrame:
    for h in AI_DET_HEADERS+AI_DEL_HEADERS+AI_META_HEADERS:
        if h not in df.columns: df[h] = None
    return df

def _detect_symptom_columns(df: pd.DataFrame) -> Dict[str,List[str]]:
    cols = [str(c).strip() for c in df.columns]
    return dict(
        manual_detractors=[f"Symptom {i}" for i in range(1,11) if f"Symptom {i}" in cols],
        manual_delighters=[f"Symptom {i}" for i in range(11,21) if f"Symptom {i}" in cols],
        ai_detractors=[c for c in cols if c.startswith("AI Symptom Detractor ")],
        ai_delighters=[c for c in cols if c.startswith("AI Symptom Delighter ")],
    )

def _filled_mask(df: pd.DataFrame, cols: List[str]) -> pd.Series:
    if not cols: return pd.Series(False, index=df.index)
    mask = pd.Series(False, index=df.index)
    for c in cols:
        if c not in df.columns: continue
        s = df[c].fillna("").astype(str).str.strip()
        mask |= (s!="") & (~s.str.upper().isin(NON_VALUES))
    return mask

def _detect_missing(df: pd.DataFrame, colmap: Dict[str,List[str]]) -> pd.DataFrame:
    out = df.copy()
    det_cols = colmap["manual_detractors"]+colmap["ai_detractors"]
    del_cols = colmap["manual_delighters"]+colmap["ai_delighters"]
    out["Has_Detractors"]    = _filled_mask(out, det_cols)
    out["Has_Delighters"]    = _filled_mask(out, del_cols)
    out["Needs_Detractors"]  = ~out["Has_Detractors"]
    out["Needs_Delighters"]  = ~out["Has_Delighters"]
    out["Needs_Symptomization"] = out["Needs_Detractors"] & out["Needs_Delighters"]
    return out

def _call_symptomizer_batch(*, client, items, allowed_delighters, allowed_detractors,
                             product_profile="", max_ev_chars=120) -> Dict[int,Dict]:
    out_by_idx: Dict[int,Dict] = {}
    if not items: return out_by_idx
    sys_lines = [
        "You are a high-recall, disciplined review symptomizer for consumer products.",
        "For each review, return relevant delighters (positive themes) and detractors (negative themes).",
        "Only include a label if there is direct textual support in the review.",
        "Return STRICT JSON with schema:",
        '{"items":[{"id":"<id>","detractors":[{"label":"<from allowed list>","evidence":["<exact substring>"]}],'
        '"delighters":[{"label":"<from allowed list>","evidence":["<exact substring>"]}],'
        '"unlisted_detractors":["<SHORT THEME>"],"unlisted_delighters":["<SHORT THEME>"],'
        '"safety":"<enum>","reliability":"<enum>","sessions":"<enum>"}]}',
        "Rules:",
        f"- Evidence must be exact substrings ≤{max_ev_chars} chars. Up to 2 per label.",
        "- Use ONLY the allowed lists. If nothing fits, use unlisted_* for a short reusable theme.",
        "- Cap to maximum 10 detractors and 10 delighters.",
        "SAFETY one of: ['Not Mentioned','Concern','Positive']",
        "RELIABILITY one of: ['Not Mentioned','Negative','Neutral','Positive']",
        "SESSIONS one of: ['0','1','2–3','4–9','10+','Unknown']",
    ]
    if product_profile:
        sys_lines.insert(1, f"Product context: {product_profile[:400]}")
    payload = dict(
        allowed_delighters=allowed_delighters,
        allowed_detractors=allowed_detractors,
        items=[dict(id=str(it["idx"]),review=it["review"],
                    needs_delighters=it.get("needs_del",True),
                    needs_detractors=it.get("needs_det",True))
               for it in items],
    )
    result_text = _chat_complete(client, model=_shared_model(),
        messages=[{"role":"system","content":"\n".join(sys_lines)},
                  {"role":"user","content":json.dumps(payload)}],
        temperature=0.0, response_format={"type":"json_object"},
        max_tokens=4000)
    data = _safe_json_load(result_text)
    items_out = data.get("items") or (data if isinstance(data,list) else [])
    by_id = {str(o.get("id")):o for o in items_out if isinstance(o,dict) and "id" in o}

    def _extract(objs, allowed, side):
        labels = []; ev_map = {}
        for obj in (objs or []):
            if not isinstance(obj,dict): continue
            raw = str(obj.get("label","")).strip()
            # coerce to nearest allowed label
            exact = {_canon_simple(x):x for x in allowed}
            lbl = exact.get(_canon_simple(raw))
            if not lbl:
                m = difflib.get_close_matches(raw, allowed, n=1, cutoff=0.82)
                lbl = m[0] if m else None
            if not lbl: continue
            evs = [str(e)[:max_ev_chars] for e in (obj.get("evidence") or []) if isinstance(e,str) and e.strip()]
            if lbl not in labels: labels.append(lbl); ev_map[lbl]=evs[:2]
            if len(labels)>=10: break
        return labels, ev_map

    for it in items:
        idx  = int(it["idx"])
        obj  = by_id.get(str(idx)) or {}
        dels, ev_del = _extract(obj.get("delighters",[]), allowed_delighters, "del")
        dets, ev_det = _extract(obj.get("detractors",[]), allowed_detractors, "det")
        safety      = str(obj.get("safety","Not Mentioned"))
        reliability = str(obj.get("reliability","Not Mentioned"))
        sessions    = str(obj.get("sessions","Unknown"))
        safety      = safety      if safety      in SAFETY_ENUM      else "Not Mentioned"
        reliability = reliability if reliability in RELIABILITY_ENUM else "Not Mentioned"
        sessions    = sessions    if sessions    in SESSIONS_ENUM    else "Unknown"
        out_by_idx[idx] = dict(dels=dels,dets=dets,ev_del=ev_del,ev_det=ev_det,
                                unl_dels=obj.get("unlisted_delighters",[])[:10],
                                unl_dets=obj.get("unlisted_detractors",[])[:10],
                                safety=safety,reliability=reliability,sessions=sessions)
    return out_by_idx

def _ai_build_symptom_list(*, client, product_description: str, sample_reviews: List[str]) -> Dict:
    sys = textwrap.dedent("""
        You design symptom catalogs for consumer product review analysis.
        Return STRICT JSON:
        {"delighters":[{"label":"<2-4 words Title Case>","rationale":"<short>"}],
         "detractors":[{"label":"<2-4 words Title Case>","rationale":"<short>"}]}
        Rules:
        - Return 8-12 delighters and 8-12 detractors.
        - Labels must be mutually exclusive and reusable across many reviews.
        - Use singular nouns when possible.
        - Cover broad aspects: performance, ease of use, value, reliability, design, safety.
    """).strip()
    payload = dict(product_description=product_description,
                   sample_reviews=sample_reviews[:20])
    result_text = _chat_complete(client, model=_shared_model(),
        messages=[{"role":"system","content":sys},
                  {"role":"user","content":json.dumps(payload)}],
        temperature=0.0, response_format={"type":"json_object"}, max_tokens=1200)
    data = _safe_json_load(result_text)
    return dict(
        delighters=[str(o.get("label","")).strip() for o in (data.get("delighters") or []) if str(o.get("label","")).strip()][:15],
        detractors=[str(o.get("label","")).strip() for o in (data.get("detractors") or []) if str(o.get("label","")).strip()][:15],
    )

def _generate_symptomized_workbook(original_bytes: bytes, updated_df: pd.DataFrame) -> bytes:
    wb = load_workbook(io.BytesIO(original_bytes))
    sheet_name = "Star Walk scrubbed verbatims"
    if sheet_name not in wb.sheetnames: sheet_name = wb.sheetnames[0]
    ws: Worksheet = wb[sheet_name]
    df2 = _ensure_ai_columns(updated_df.copy())
    fill_green = PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
    fill_red   = PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid")
    fill_yel   = PatternFill(start_color="FFF2CC",end_color="FFF2CC",fill_type="solid")
    fill_blu   = PatternFill(start_color="CFE2F3",end_color="CFE2F3",fill_type="solid")
    fill_pur   = PatternFill(start_color="EAD1DC",end_color="EAD1DC",fill_type="solid")
    for i, (rid, row) in enumerate(df2.iterrows(), start=2):
        for j, col_idx in enumerate(DET_INDEXES, 1):
            val = row.get(f"AI Symptom Detractor {j}")
            cv  = None if (pd.isna(val) or str(val).strip()=="") else val
            cell = ws.cell(row=i, column=col_idx, value=cv)
            if cv: cell.fill = fill_red
        for j, col_idx in enumerate(DEL_INDEXES, 1):
            val = row.get(f"AI Symptom Delighter {j}")
            cv  = None if (pd.isna(val) or str(val).strip()=="") else val
            cell = ws.cell(row=i, column=col_idx, value=cv)
            if cv: cell.fill = fill_green
        if _is_filled(row.get("AI Safety")):
            c = ws.cell(row=i, column=META_INDEXES["Safety"], value=str(row["AI Safety"])); c.fill=fill_yel
        if _is_filled(row.get("AI Reliability")):
            c = ws.cell(row=i, column=META_INDEXES["Reliability"], value=str(row["AI Reliability"])); c.fill=fill_blu
        if _is_filled(row.get("AI # of Sessions")):
            c = ws.cell(row=i, column=META_INDEXES["# of Sessions"], value=str(row["AI # of Sessions"])); c.fill=fill_pur
    for c in DET_INDEXES+DEL_INDEXES+list(META_INDEXES.values()):
        try: ws.column_dimensions[get_column_letter(c)].width = 28
        except Exception: pass
    out = io.BytesIO(); wb.save(out); return out.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE INIT
# ═══════════════════════════════════════════════════════════════════════════════

def _init_session_state() -> None:
    st.session_state.setdefault("analysis_dataset", None)
    st.session_state.setdefault("chat_messages", [])
    st.session_state.setdefault("master_export_bundle", None)
    st.session_state.setdefault("prompt_definitions_df", _default_prompt_df())
    st.session_state.setdefault("prompt_builder_suggestion", None)
    st.session_state.setdefault("prompt_run_artifacts", None)
    st.session_state.setdefault("prompt_run_notice", None)
    st.session_state.setdefault("chat_scope_signature", None)
    st.session_state.setdefault("chat_scope_notice", None)
    st.session_state.setdefault("active_main_view", "Dashboard")
    st.session_state.setdefault("workspace_view_selector", "Dashboard")
    st.session_state.setdefault("review_explorer_page", 1)
    st.session_state.setdefault("review_explorer_per_page", 20)
    st.session_state.setdefault("review_explorer_sort", "Newest")
    # shared model
    st.session_state.setdefault("shared_model", DEFAULT_MODEL)
    st.session_state.setdefault("shared_reasoning", DEFAULT_REASONING)
    # symptomizer
    st.session_state.setdefault("sym_delighters", [])
    st.session_state.setdefault("sym_detractors", [])
    st.session_state.setdefault("sym_aliases", {})
    st.session_state.setdefault("sym_symptoms_source", "none")  # none | file | manual | ai
    st.session_state.setdefault("sym_processed_rows", [])
    st.session_state.setdefault("sym_new_candidates", {})
    st.session_state.setdefault("sym_product_profile", "")
    st.session_state.setdefault("sym_scope_choice", "Missing both")
    st.session_state.setdefault("sym_n_to_process", 10)
    st.session_state.setdefault("sym_batch_size", 5)
    st.session_state.setdefault("sym_max_ev_chars", 120)
    st.session_state.setdefault("sym_run_notice", None)

_init_session_state()


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════

def _render_shared_sidebar(df: Optional[pd.DataFrame]) -> Dict[str,Any]:
    api_key = _get_api_key()
    selected_ratings    = [1,2,3,4,5]
    selected_products: List[str] = []
    review_source_mode  = "All reviews"
    selected_locales: List[str] = []
    recommendation_mode = "All"
    date_range: Optional[Tuple[date,date]] = None
    text_query          = ""

    with st.sidebar:
        # ── Shared Model Settings ──────────────────────────────────────────
        st.header("🤖 AI Model Settings")
        st.selectbox("Model", options=MODEL_OPTIONS,
                     index=MODEL_OPTIONS.index(st.session_state.get("shared_model",DEFAULT_MODEL)),
                     key="shared_model",
                     help="Used by the AI Analyst, Review Prompt, and Symptomizer tabs.")
        st.selectbox("Reasoning effort", options=REASONING_OPTIONS,
                     index=REASONING_OPTIONS.index(st.session_state.get("shared_reasoning",DEFAULT_REASONING)),
                     key="shared_reasoning")
        if api_key:
            st.success("OpenAI API key loaded")
        else:
            st.warning("Add OPENAI_API_KEY to Streamlit secrets to enable AI features.")
        st.caption(f"Model: {_shared_model()} · Reasoning: {_shared_reasoning()}")
        st.divider()

        # ── Review Filters ─────────────────────────────────────────────────
        st.header("🔍 Review Filters")
        st.caption("These filters drive the Dashboard, Review Explorer, AI Analyst, Review Prompt, and Symptomizer.")
        if df is None:
            st.info("Build a workspace to unlock filters.")
        else:
            options = _build_filter_options(df)
            RATING_OPTIONS = ["All ratings","1 star","2 stars","3 stars","4 stars","5 stars","1-2 stars","4-5 stars","Custom"]
            rating_mode = st.selectbox("Ratings", options=RATING_OPTIONS, index=0, key="sidebar_rating_mode")
            custom_ratings = None
            if rating_mode == "Custom":
                custom_ratings = st.multiselect("Custom ratings", options=options["ratings"],
                                                default=options["ratings"], key="sidebar_custom_ratings")
            mapping = {"All ratings":[1,2,3,4,5],"1 star":[1],"2 stars":[2],"3 stars":[3],
                       "4 stars":[4],"5 stars":[5],"1-2 stars":[1,2],"4-5 stars":[4,5]}
            selected_ratings = mapping.get(rating_mode, custom_ratings or [1,2,3,4,5])

            review_source_mode = st.selectbox("Review source",
                options=["All reviews","Organic only","Incentivized only"], index=0, key="sidebar_review_source")

            if options["product_groups"] and len(options["product_groups"])>1:
                selected_products = st.multiselect("SKU / product ID", options=options["product_groups"],
                                                   default=[], key="sidebar_product_groups")
            if options["locales"]:
                selected_locales = st.multiselect("Market / locale", options=options["locales"],
                                                  default=[], key="sidebar_locales")
            recommendation_mode = st.selectbox("Recommendation status",
                options=["All","Recommended only","Not recommended only"], index=0, key="sidebar_recommendation")
            if options["min_date"] and options["max_date"]:
                picked = st.date_input("Submission date range",
                    value=(options["min_date"],options["max_date"]),
                    min_value=options["min_date"], max_value=options["max_date"], key="sidebar_date_range")
                if isinstance(picked,tuple) and len(picked)==2: date_range=(picked[0],picked[1])
            text_query = st.text_input("Text contains", value="", key="sidebar_text_query",
                                       placeholder="noise, basket, capacity…")

        st.divider()
        # ── Symptomizer batch settings ─────────────────────────────────────
        st.header("⚡ Symptomizer Settings")
        st.slider("Batch size", 1, 12, key="sym_batch_size")
        st.slider("Max evidence chars", 60, 200, step=10, key="sym_max_ev_chars")

    src_map = {"All reviews":"All reviews","Organic only":"Non-incentivized only","Incentivized only":"Incentivized only"}
    return dict(
        selected_ratings=selected_ratings,
        selected_products=selected_products,
        review_source_mode=review_source_mode,
        incentivized_mode=src_map.get(review_source_mode,"All reviews"),
        selected_locales=selected_locales,
        recommendation_mode=recommendation_mode,
        date_range=date_range,
        text_query=text_query,
        api_key=api_key,
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — RENDER HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _render_metric_card(label: str, value: str, subtext: str) -> None:
    st.markdown(f"""
    <div class="metric-card">
      <div class="metric-label">{label}</div>
      <div class="metric-value">{value}</div>
      <div class="metric-sub">{subtext}</div>
    </div>""", unsafe_allow_html=True)

def _render_workspace_header(summary, overall_df, prompt_artifacts, *, source_type, source_label) -> None:
    bundle = _get_master_bundle(summary, overall_df, prompt_artifacts)
    product_name = _product_display_name(summary, overall_df)
    organic_count = int((~overall_df["incentivized_review"].fillna(False)).sum()) if not overall_df.empty else 0
    if source_type=="uploaded":
        subtitle = f"Source: {source_label} | {summary.reviews_downloaded:,} reviews | {organic_count:,} organic"
    else:
        subtitle = (f"Product ID {summary.product_id} | {summary.reviews_downloaded:,} reviews downloaded | "
                    f"{organic_count:,} organic | {summary.requests_needed} Bazaarvoice requests")
    st.markdown(f"""
    <div class="hero-card">
      <div class="hero-title">{_safe_html(product_name)}</div>
      <div class="hero-sub">{_safe_html(subtitle)}</div>
    </div>""", unsafe_allow_html=True)
    action_cols = st.columns([1.2,1.2,4])
    action_cols[0].download_button("⬇️ Download all reviews", data=bundle["excel_bytes"],
        file_name=bundle["excel_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)
    if action_cols[1].button("🔄 Reset workspace", use_container_width=True):
        for k in ["analysis_dataset","chat_messages","chat_scope_signature","chat_scope_notice",
                  "master_export_bundle","prompt_run_artifacts","prompt_run_notice",
                  "sym_processed_rows","sym_new_candidates","sym_product_profile"]:
            st.session_state[k] = [] if k in {"chat_messages","sym_processed_rows"} else (
                {} if k in {"sym_new_candidates"} else None if k in {"analysis_dataset"} else "")
        st.rerun()
    action_cols[2].caption("The workbook includes Reviews, Rating Distribution, Volume trend, and any AI prompt columns.")

def _render_top_metrics(overall_df, filtered_df) -> None:
    metrics = _compute_metrics(filtered_df)
    cards = [
        ("Reviews in view", f"{metrics['review_count']:,}", f"of {len(overall_df):,} loaded"),
        ("Avg rating", _format_metric_number(metrics["avg_rating"]), "Filtered view"),
        ("Avg rating (organic)", _format_metric_number(metrics["avg_rating_non_incentivized"]),
         f"{metrics['non_incentivized_count']:,} organic reviews"),
        ("% 1-2 star", _format_pct(metrics["pct_low_star"]), f"{metrics['low_star_count']:,} low-star reviews"),
        ("% incentivized", _format_pct(metrics["pct_incentivized"]), "Current view share"),
    ]
    cols = st.columns(len(cards))
    for col, (label,value,subtext) in zip(cols,cards):
        with col: _render_metric_card(label,value,subtext)

def _sort_reviews(df: pd.DataFrame, sort_mode: str) -> pd.DataFrame:
    w = df.copy()
    if sort_mode=="Newest":    return w.sort_values(["submission_time","review_id"],ascending=[False,False],na_position="last")
    if sort_mode=="Oldest":    return w.sort_values(["submission_time","review_id"],ascending=[True,True],na_position="last")
    if sort_mode=="Highest rating": return w.sort_values(["rating","submission_time"],ascending=[False,False],na_position="last")
    if sort_mode=="Lowest rating":  return w.sort_values(["rating","submission_time"],ascending=[True,False],na_position="last")
    if sort_mode=="Longest":   return w.sort_values(["review_length_words","submission_time"],ascending=[False,False],na_position="last")
    return w

def _render_review_card(row: pd.Series) -> None:
    rating_value = _safe_int(row.get("rating"),0) if pd.notna(row.get("rating")) else 0
    filled = "&#9733;" * max(0,min(rating_value,5))
    empty  = "&#9734;" * max(0,5-rating_value)
    title  = _safe_text(row.get("title"),"No title") or "No title"
    review_text = _safe_text(row.get("review_text"),"No written review text.") or "No written review text."
    meta_bits = [b for b in [_safe_text(row.get("submission_date")),_safe_text(row.get("content_locale")),
                              _safe_text(row.get("retailer")),_safe_text(row.get("product_or_sku"))] if b]
    chips = ["Organic" if not _safe_bool(row.get("incentivized_review"),False) else "Incentivized"]
    rec = row.get("is_recommended")
    if not _is_missing_value(rec): chips.append("Recommended" if _safe_bool(rec,False) else "Not recommended")
    if _safe_bool(row.get("has_photos"),False): chips.append(f"Photos: {_safe_int(row.get('photos_count'),0)}")
    # show any AI Symptom columns
    det_tags = [str(row.get(f"AI Symptom Detractor {j}","")) for j in range(1,11)
                if _is_filled(row.get(f"AI Symptom Detractor {j}"))]
    del_tags = [str(row.get(f"AI Symptom Delighter {j}","")) for j in range(1,11)
                if _is_filled(row.get(f"AI Symptom Delighter {j}"))]
    with st.container(border=True):
        top_cols = st.columns([4.6,1.6])
        with top_cols[0]:
            st.markdown(f"<div class='tiny'>{filled}{empty} {rating_value}/5</div>",unsafe_allow_html=True)
            st.markdown(f"**{title}**")
            if meta_bits: st.caption(" | ".join(meta_bits))
        with top_cols[1]: st.caption(" | ".join(chips))
        st.write(review_text)
        if det_tags or del_tags:
            chips_html = "<div class='chip-wrap'>"
            for t in det_tags: chips_html+=f"<span class='chip red'>{_safe_html(t)}</span>"
            for t in del_tags: chips_html+=f"<span class='chip green'>{_safe_html(t)}</span>"
            chips_html+="</div>"
            st.markdown(chips_html, unsafe_allow_html=True)
        footer = [b for b in [f"Review ID: {_safe_text(row.get('review_id'))}",
                               _safe_text(row.get("user_location"))] if b]
        if footer: st.caption(" | ".join(footer))


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — TAB: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

def _render_dashboard(filtered_df: pd.DataFrame) -> None:
    st.subheader("Dashboard")
    st.markdown("<div class='section-sub'>Rating mix, volume, and theme signals for the current filter set.</div>",unsafe_allow_html=True)
    chart_scope = st.radio("Dashboard scope",["All matching reviews","Organic only"],horizontal=True,key="dashboard_chart_scope")
    chart_df = filtered_df.copy()
    if chart_scope=="Organic only":
        chart_df = chart_df[~chart_df["incentivized_review"].fillna(False)].reset_index(drop=True)
    if chart_df.empty:
        st.info("No reviews match the current scope."); return

    rating_df  = _rating_distribution(chart_df)
    monthly_df = _monthly_trend(chart_df)
    theme_df   = _compute_themes(chart_df)

    rating_df["rating_label"]     = rating_df["rating"].map(lambda v: f"{int(v)}★")
    rating_df["count_pct_label"]  = rating_df.apply(lambda r: f"{int(r['review_count']):,} · {_format_pct(r['share'])}", axis=1)

    chart_cols = st.columns([1.05,1.15])
    with chart_cols[0]:
        with st.container(border=True):
            fig = px.bar(rating_df,x="rating_label",y="review_count",text="count_pct_label",
                         title="Rating distribution",
                         category_orders={"rating_label":["1★","2★","3★","4★","5★"]},
                         hover_data={"share":":.1%","review_count":True})
            fig.update_traces(textposition="outside",cliponaxis=False)
            fig.update_layout(margin=dict(l=24,r=24,t=60,b=20),xaxis_title="Star rating",yaxis_title="Review count")
            st.plotly_chart(fig, use_container_width=True)
    with chart_cols[1]:
        with st.container(border=True):
            if monthly_df.empty:
                st.info("No dated reviews for volume chart.")
            else:
                fig2 = make_subplots(specs=[[{"secondary_y":True}]])
                fig2.add_trace(go.Bar(x=monthly_df["month_start"],y=monthly_df["review_count"],name="Review count",opacity=0.62),secondary_y=False)
                fig2.add_trace(go.Scatter(x=monthly_df["month_start"],y=monthly_df["avg_rating"],name="Avg rating",mode="lines+markers"),secondary_y=True)
                fig2.update_layout(title="Review volume over time",margin=dict(l=24,r=24,t=60,b=20),hovermode="x unified")
                fig2.update_xaxes(title_text="Month")
                fig2.update_yaxes(title_text="Review count",secondary_y=False)
                fig2.update_yaxes(title_text="Avg rating",range=[1,5],secondary_y=True)
                st.plotly_chart(fig2, use_container_width=True)

    with st.container(border=True):
        st.markdown("**Theme signals**")
        if not theme_df.empty:
            fig3 = px.bar(theme_df.head(9),x="mention_rate",y="theme",orientation="h",
                          color="avg_rating_when_mentioned",color_continuous_scale="RdYlGn",
                          range_color=[1,5],
                          hover_data={"mention_count":True,"low_star_mentions":True,"high_star_mentions":True},
                          title="Theme mention rate (colored by avg rating when mentioned)")
            fig3.update_layout(margin=dict(l=24,r=24,t=60,b=20),xaxis_tickformat=".0%",yaxis_title="")
            st.plotly_chart(fig3, use_container_width=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — TAB: REVIEW EXPLORER
# ═══════════════════════════════════════════════════════════════════════════════

def _render_review_explorer(*, summary, overall_df, filtered_df, prompt_artifacts) -> None:
    st.subheader("Review Explorer")
    st.markdown(f"<div class='section-sub'>Showing {len(filtered_df):,} reviews out of {len(overall_df):,} loaded.</div>",unsafe_allow_html=True)
    bundle = _get_master_bundle(summary, overall_df, prompt_artifacts)
    top_cols = st.columns([1.3,1.35,1.0,2.05])
    top_cols[0].download_button("⬇️ Download all reviews",data=bundle["excel_bytes"],
        file_name=bundle["excel_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,key="explorer_download")
    sort_mode = top_cols[1].selectbox("Sort reviews",
        options=["Newest","Oldest","Highest rating","Lowest rating","Longest"],
        key="review_explorer_sort")
    per_page  = int(top_cols[2].selectbox("Per page",options=[10,20,30,50],
        key="review_explorer_per_page"))
    top_cols[3].caption("Use sidebar filters to narrow the stream.")
    ordered_df = _sort_reviews(filtered_df, sort_mode).reset_index(drop=True)
    if ordered_df.empty:
        st.info("No reviews match the current filters."); return
    page_count = max(1,math.ceil(len(ordered_df)/max(per_page,1)))
    current_page = max(1,min(int(st.session_state.get("review_explorer_page",1)),page_count))
    pager_cols = st.columns([0.9,0.9,2.3,1.0,0.9,0.9])
    if pager_cols[0].button("⏮",use_container_width=True,disabled=current_page<=1,key="re_first"): current_page=1
    if pager_cols[1].button("←",use_container_width=True,disabled=current_page<=1,key="re_prev"):  current_page=max(1,current_page-1)
    pager_cols[2].markdown(f"<div style='text-align:center;font-weight:700;padding-top:.6rem;'>Page {current_page} of {page_count:,} &nbsp;·&nbsp; {(current_page-1)*per_page+1:,}–{min(current_page*per_page,len(ordered_df)):,} of {len(ordered_df):,}</div>",unsafe_allow_html=True)
    if st.session_state.get("review_explorer_page_input")!=current_page:
        st.session_state["review_explorer_page_input"]=current_page
    current_page=int(pager_cols[3].number_input("Page",min_value=1,max_value=page_count,
        value=current_page,step=1,key="review_explorer_page_input",label_visibility="collapsed"))
    if pager_cols[4].button("→",use_container_width=True,disabled=current_page>=page_count,key="re_next"): current_page=min(page_count,current_page+1)
    if pager_cols[5].button("⏭",use_container_width=True,disabled=current_page>=page_count,key="re_last"): current_page=page_count
    st.session_state["review_explorer_page"]=max(1,min(current_page,page_count))
    start = (st.session_state["review_explorer_page"]-1)*per_page
    for _,row in ordered_df.iloc[start:start+per_page].iterrows():
        _render_review_card(row)


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — TAB: AI ANALYST
# ═══════════════════════════════════════════════════════════════════════════════

def _render_ai_tab(*, settings, overall_df, filtered_df, summary, filter_description) -> None:
    st.subheader("AI — Product & Consumer Insights")
    st.markdown("<div class='section-sub'>Ask anything. The assistant is grounded in the currently filtered review text.</div>",unsafe_allow_html=True)
    if filtered_df.empty:
        st.info("Adjust filters — no reviews in scope."); return

    scope_sig = json.dumps(dict(product_id=summary.product_id,filter_description=filter_description,
                                review_count=len(filtered_df),
                                source_type=(st.session_state.get("analysis_dataset") or {}).get("source_type","bazaarvoice")),sort_keys=True)
    if st.session_state.get("chat_scope_signature")!=scope_sig:
        if st.session_state.get("chat_messages"):
            st.session_state["chat_messages"]=[]
            st.session_state["chat_scope_notice"]="AI chat cleared — scope changed."
        st.session_state["chat_scope_signature"]=scope_sig

    notice = st.session_state.pop("chat_scope_notice",None)
    if notice: st.info(notice)

    with st.container(border=True):
        sc = st.columns([1.2,1.0,2.0])
        sc[0].metric("Reviews in scope",f"{len(filtered_df):,}")
        organic = int((~filtered_df["incentivized_review"].fillna(False)).sum())
        sc[1].metric("Organic reviews",f"{organic:,}")
        sc[2].caption(f"Model: {_shared_model()} · Scope: {filter_description}")

    api_key = settings.get("api_key")
    if not api_key:
        st.warning("Add OPENAI_API_KEY to Streamlit secrets to enable AI features.")
        st.code('OPENAI_API_KEY = "sk-..."', language="toml"); return

    quick_actions = {
        "Executive summary": dict(prompt="Create a concise executive summary. Lead with biggest strengths, biggest risks, key consumer insight, and top 3 actions.",help="Leadership readout.",persona=None),
        "Product Development": dict(prompt=PERSONAS["Product Development"]["prompt"],help=PERSONAS["Product Development"]["blurb"],persona="Product Development"),
        "Quality Engineer":    dict(prompt=PERSONAS["Quality Engineer"]["prompt"],   help=PERSONAS["Quality Engineer"]["blurb"],   persona="Quality Engineer"),
        "Consumer Insights":   dict(prompt=PERSONAS["Consumer Insights"]["prompt"],  help=PERSONAS["Consumer Insights"]["blurb"],  persona="Consumer Insights"),
    }
    quick_trigger: Optional[Tuple] = None
    with st.container(border=True):
        st.markdown("**Quick reports**")
        action_cols = st.columns(4)
        for col,(label,config) in zip(action_cols,quick_actions.items()):
            if col.button(label,use_container_width=True,help=config["help"],key=f"ai_quick_{_slugify(label)}"):
                quick_trigger=(config["persona"],label,config["prompt"])

    chat_container = st.container(border=True)
    with chat_container:
        if not st.session_state["chat_messages"]:
            st.info("Start with a quick report or type a question below.")
        for msg in st.session_state["chat_messages"]:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])

    helper_cols = st.columns([1.8,1.1,1.1])
    helper_cols[0].caption(f"Scope: {filter_description}")
    if helper_cols[1].button("Clear chat",use_container_width=True,key="ai_clear_chat"):
        st.session_state["chat_messages"]=[]; st.rerun()

    user_message = st.chat_input("Ask about complaint drivers, opportunities, quality risks, sentiment…",key="ai_chat_input")
    prompt_to_send = visible_user_message = persona_name = None
    if quick_trigger:
        persona_name, visible_user_message, prompt_to_send = quick_trigger
    elif user_message:
        prompt_to_send = visible_user_message = user_message

    if prompt_to_send and visible_user_message:
        prior = list(st.session_state["chat_messages"])
        st.session_state["chat_messages"].append({"role":"user","content":visible_user_message})
        overlay = _show_thinking("Reviewing the filtered review text and building a grounded answer…")
        try:
            answer = _call_analyst(question=prompt_to_send,overall_df=overall_df,filtered_df=filtered_df,
                                   summary=summary,filter_description=filter_description,
                                   chat_history=prior,persona_name=persona_name)
            if persona_name: answer=f"## {persona_name} report\n\n{answer}"
        except Exception as exc:
            answer = f"OpenAI request failed: {exc}"
        finally:
            overlay.empty()
        st.session_state["chat_messages"].append({"role":"assistant","content":answer})
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
#  SNEVIEWS — TAB: REVIEW PROMPT
# ═══════════════════════════════════════════════════════════════════════════════

def _render_review_prompt_tab(*, settings, overall_df, filtered_df, summary, filter_description) -> None:
    st.subheader("Review Prompt")
    st.markdown("<div class='section-sub'>Create row-level AI tags that become new review columns.</div>",unsafe_allow_html=True)

    with st.container(border=True):
        st.markdown("**Prompt library**")
        starter_cols = st.columns([1.2,1.2,1])
        if starter_cols[0].button("Add starter pack",use_container_width=True,key="prompt_add_starter"):
            new_rows = pd.DataFrame(REVIEW_PROMPT_STARTER_ROWS)
            existing_names = set(st.session_state["prompt_definitions_df"]["column_name"].astype(str))
            to_add = new_rows[~new_rows["column_name"].isin(existing_names)]
            if not to_add.empty:
                st.session_state["prompt_definitions_df"] = pd.concat([st.session_state["prompt_definitions_df"],to_add],ignore_index=True)
            st.rerun()
        if starter_cols[1].button("Reset to starter",use_container_width=True,key="prompt_reset_starter"):
            st.session_state["prompt_definitions_df"]=pd.DataFrame(REVIEW_PROMPT_STARTER_ROWS); st.rerun()
        if starter_cols[2].button("Clear all",use_container_width=True,key="prompt_clear_all"):
            st.session_state["prompt_definitions_df"]=pd.DataFrame(columns=["column_name","prompt","labels"]); st.rerun()

    st.markdown("#### Prompt definitions")
    edited_df = st.data_editor(st.session_state["prompt_definitions_df"],use_container_width=True,
        num_rows="dynamic",hide_index=True,key="prompt_definition_editor",height=360,
        column_config={
            "column_name":st.column_config.TextColumn("Column name",width="medium"),
            "prompt":st.column_config.TextColumn("Review prompt",width="large"),
            "labels":st.column_config.TextColumn("Labels (comma-separated)",width="large"),
        })
    st.session_state["prompt_definitions_df"]=edited_df

    try:
        prompt_definitions = _normalize_prompt_definitions(st.session_state["prompt_definitions_df"],overall_df.columns)
    except ReviewDownloaderError as exc:
        st.error(str(exc)); prompt_definitions=[]

    api_key = settings.get("api_key")
    client  = _get_openai_client()

    with st.container(border=True):
        scope_cols = st.columns([1.25,1.0,1.0,2.45])
        tagging_scope = scope_cols[0].selectbox("Tagging scope",["Current filtered reviews","All loaded reviews"],index=0,key="prompt_tagging_scope")
        scope_df = filtered_df if tagging_scope=="Current filtered reviews" else overall_df
        batch_size= int(st.session_state.get("sym_batch_size",5))
        est_calls = math.ceil(len(scope_df)/max(1,batch_size)) if len(scope_df) else 0
        scope_cols[1].metric("Reviews in scope",f"{len(scope_df):,}")
        scope_cols[2].metric("Planned requests",f"{est_calls:,}")
        scope_cols[3].caption(f"Scope: {tagging_scope.lower()}. Filters: {filter_description}.")
        run_disabled = (not api_key) or (not prompt_definitions) or len(scope_df)==0
        if st.button("▶️ Run Review Prompt",type="primary",use_container_width=True,disabled=run_disabled,key="prompt_run_btn"):
            overlay = _show_thinking("Classifying each review…")
            try:
                prompt_results_df = _run_review_prompt_tagging(client=client,
                    source_df=scope_df.reset_index(drop=True),
                    prompt_definitions=prompt_definitions,chunk_size=batch_size)
                updated = _merge_prompt_results(overall_df,prompt_results_df,prompt_definitions)
                dataset = dict(st.session_state["analysis_dataset"])
                dataset["reviews_df"]=updated; st.session_state["analysis_dataset"]=dataset
                summary_df=_summarize_prompt_results(prompt_results_df,prompt_definitions,source_df=scope_df)
                st.session_state["prompt_run_artifacts"]=dict(
                    definitions=prompt_definitions, summary_df=summary_df,
                    scope_label=tagging_scope, scope_filter_description=filter_description,
                    scope_review_ids=list(prompt_results_df["review_id"].astype(str)),
                    definition_signature=json.dumps([dict(col=p["column_name"],prompt=p["prompt"],labels=p["labels"])
                                                     for p in prompt_definitions],sort_keys=True),
                    review_count=len(prompt_results_df),
                    generated_utc=pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
                )
                st.session_state["master_export_bundle"]=None
                st.session_state["prompt_run_notice"]=f"Finished tagging {len(prompt_results_df):,} reviews."
            except Exception as exc:
                st.error(f"Review Prompt run failed: {exc}")
            finally:
                overlay.empty()
            st.rerun()

    notice = st.session_state.pop("prompt_run_notice",None)
    if notice: st.success(notice)

    prompt_artifacts = st.session_state.get("prompt_run_artifacts")
    if not prompt_artifacts:
        st.info("Run Review Prompt to generate AI columns."); return

    current_sig = json.dumps([dict(col=p["column_name"],prompt=p["prompt"],labels=p["labels"])
                               for p in prompt_definitions],sort_keys=True) if prompt_definitions else ""
    if current_sig!=prompt_artifacts.get("definition_signature"):
        st.info("Prompt definitions changed — re-run to refresh.")

    updated_overall_df = st.session_state["analysis_dataset"]["reviews_df"]
    review_ids = set(str(x) for x in prompt_artifacts.get("scope_review_ids",[]))
    result_scope_df = updated_overall_df[updated_overall_df["review_id"].astype(str).isin(review_ids)].copy()
    bundle = _get_master_bundle(summary, updated_overall_df, prompt_artifacts)
    st.download_button("⬇️ Download tagged review file",data=bundle["excel_bytes"],
        file_name=bundle["excel_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.caption(f"Latest run: {prompt_artifacts.get('generated_utc')} | Scope: {prompt_artifacts.get('scope_label')} | Reviews: {prompt_artifacts.get('review_count'):,}")

    prompt_lookup = {p["display_name"]:p for p in prompt_artifacts["definitions"]}
    prompt_names  = list(prompt_lookup.keys())
    if not prompt_names: st.info("No prompt results yet."); return
    if st.session_state.get("prompt_result_view") not in prompt_names:
        st.session_state["prompt_result_view"]=prompt_names[0]
    selected_prompt_name = st.radio("Prompt result view",options=prompt_names,horizontal=True,
        key="prompt_result_view",label_visibility="collapsed")
    prompt = prompt_lookup[selected_prompt_name]
    prompt_col = prompt["column_name"]
    base_view_df = result_scope_df[result_scope_df[prompt_col].notna()].copy() if prompt_col in result_scope_df.columns else result_scope_df.iloc[0:0]

    label_options = [str(l) for l in prompt_artifacts["summary_df"][prompt_artifacts["summary_df"]["column_name"]==prompt_col]["label"].tolist()]
    selected_labels = st.multiselect("Labels",options=label_options,default=label_options,key=f"prompt_labels_{prompt_col}")
    preview_df = base_view_df[base_view_df[prompt_col].isin(selected_labels)] if selected_labels else base_view_df.iloc[0:0]

    prompt_summary = pd.DataFrame()
    total = max(len(base_view_df),1)
    rows = []
    for label in prompt["labels"]:
        sub = preview_df[preview_df[prompt_col]==label] if prompt_col in preview_df.columns else preview_df.iloc[0:0]
        rows.append(dict(label=label,review_count=len(sub),share=_safe_pct(len(sub),total),
                         avg_rating=_safe_mean(sub["rating"]) if "rating" in sub.columns else None))
    prompt_summary = pd.DataFrame(rows)

    chart_col, table_col = st.columns([1.45,1.05])
    with chart_col:
        with st.container(border=True):
            if prompt_summary.empty: st.info("No tagged reviews match current filters.")
            else:
                fig = px.pie(prompt_summary,names="label",values="review_count",hole=0.42)
                fig.update_layout(margin=dict(l=20,r=20,t=20,b=20))
                st.plotly_chart(fig,use_container_width=True)
    with table_col:
        with st.container(border=True):
            st.markdown(f"**Column** `{prompt_col}`")
            st.write(prompt["prompt"])
            if not prompt_summary.empty:
                ds = prompt_summary.copy()
                ds["avg_rating"] = ds["avg_rating"].map(lambda x: f"{x:.2f}★" if pd.notna(x) else "—")
                ds["share"]      = ds["share"].map(_format_pct)
                st.dataframe(ds[["label","review_count","avg_rating","share"]],
                             use_container_width=True,hide_index=True,height=260)

    preview_cols = [c for c in ["review_id","rating","incentivized_review","submission_time",
                                 "content_locale","title","review_text",prompt_col] if c in preview_df.columns]
    st.markdown("**Tagged review preview**")
    st.dataframe(preview_df[preview_cols].head(50),use_container_width=True,hide_index=True,height=320)


# ═══════════════════════════════════════════════════════════════════════════════
#  SYMPTOMIZER — TAB (5th tab)
# ═══════════════════════════════════════════════════════════════════════════════

def _render_symptomizer_tab(*, settings, overall_df, filtered_df, summary, filter_description) -> None:
    st.subheader("Symptomizer")
    st.markdown(
        "<div class='section-sub'>Row-level AI tagging of delighters and detractors. "
        "Tags are written back into the shared review dataframe so the Dashboard and Explorer reflect them.</div>",
        unsafe_allow_html=True)

    client  = _get_openai_client()
    api_key = settings.get("api_key")

    # ── Step 1: Symptoms catalog ────────────────────────────────────────────
    st.markdown("### Step 1 — Symptoms catalog")

    dataset = st.session_state.get("analysis_dataset") or {}
    source_type = dataset.get("source_type","")

    # Try to load from uploaded file if it's an xlsx
    def _try_load_symptoms_from_file():
        """Attempt to read Symptoms sheet from the uploaded workbook bytes."""
        if source_type not in {"uploaded","bazaarvoice"}: return
        # For uploaded files: if any uploaded file is xlsx, try reading Symptoms sheet
        uploaded_files_bytes = st.session_state.get("_uploaded_raw_bytes")
        if uploaded_files_bytes:
            dels, dets, aliases = _get_symptom_whitelists_from_file(uploaded_files_bytes)
            if dels or dets:
                st.session_state["sym_delighters"] = dels
                st.session_state["sym_detractors"] = dets
                st.session_state["sym_aliases"]    = aliases
                st.session_state["sym_symptoms_source"] = "file"
                return True
        return False

    # Auto-load symptoms from file once
    if st.session_state.get("sym_symptoms_source")=="none":
        _try_load_symptoms_from_file()

    sym_source = st.session_state.get("sym_symptoms_source","none")
    delighters = list(st.session_state.get("sym_delighters") or [])
    detractors = list(st.session_state.get("sym_detractors") or [])

    # ── Symptoms status banner ──────────────────────────────────────────────
    if not delighters and not detractors:
        st.warning("⚠️ No symptoms defined yet. Use the options below to add them — or proceed without a catalog and the AI will still attempt to tag reviews using its built-in knowledge.")
    else:
        st.markdown(
            _chip_html([
                (f"{len(delighters)} delighters","green"),
                (f"{len(detractors)} detractors","red"),
                (f"Source: {sym_source}","blue"),
            ]),
            unsafe_allow_html=True)

    # ── Three options for defining symptoms ────────────────────────────────
    sym_tabs = st.tabs(["📄 Upload workbook", "✏️ Manual entry", "🤖 AI-assisted builder"])

    with sym_tabs[0]:
        st.markdown("**Upload an Excel workbook that contains a 'Symptoms' sheet.**")
        sym_upload = st.file_uploader("Upload workbook with Symptoms sheet", type=["xlsx"],
                                      key="sym_file_uploader")
        if sym_upload:
            raw_bytes = sym_upload.getvalue()
            st.session_state["_uploaded_raw_bytes"] = raw_bytes
            dels, dets, aliases = _get_symptom_whitelists_from_file(raw_bytes)
            if dels or dets:
                st.session_state["sym_delighters"] = dels
                st.session_state["sym_detractors"] = dets
                st.session_state["sym_aliases"]    = aliases
                st.session_state["sym_symptoms_source"] = "file"
                st.success(f"Loaded {len(dels)} delighters and {len(dets)} detractors from the Symptoms sheet.")
                st.rerun()
            else:
                st.error("No Symptoms sheet found, or it was empty. Make sure the sheet is named 'Symptoms' and has columns for Symptom, Type (Delighter/Detractor), and optionally Aliases.")

    with sym_tabs[1]:
        st.markdown("**Enter delighters and detractors as comma- or newline-separated lists.**")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("🟢 **Delighters** (positive themes)")
            del_text = st.text_area("Delighters (one per line or comma-separated)",
                value="\n".join(delighters), height=220, key="sym_del_manual_text")
        with col2:
            st.markdown("🔴 **Detractors** (negative themes)")
            det_text = st.text_area("Detractors (one per line or comma-separated)",
                value="\n".join(detractors), height=220, key="sym_det_manual_text")
        if st.button("💾 Save manual symptoms", use_container_width=True, key="sym_save_manual"):
            def _parse_list(t: str) -> List[str]:
                items = re.split(r"[\n,;|]+", t)
                return [i.strip() for i in items if i.strip()]
            st.session_state["sym_delighters"] = _parse_list(del_text)
            st.session_state["sym_detractors"] = _parse_list(det_text)
            st.session_state["sym_symptoms_source"] = "manual"
            st.success(f"Saved {len(st.session_state['sym_delighters'])} delighters and {len(st.session_state['sym_detractors'])} detractors.")
            st.rerun()

    with sym_tabs[2]:
        st.markdown("**Describe the product and optionally provide a few sample reviews. The AI will generate a first-cut symptom list.**")
        if not api_key:
            st.warning("OpenAI API key required for AI-assisted builder.")
        else:
            product_desc = st.text_area("Product description",
                value=st.session_state.get("sym_product_profile",""),
                placeholder="e.g. 'SharkNinja Ninja Air Fryer XL — a 6-in-1 countertop air fryer with 6 qt basket'",
                height=80, key="sym_product_desc_input")
            sample_reviews_text = ""
            if not overall_df.empty and "review_text" in overall_df.columns:
                sample_reviews_text = "\n---\n".join(
                    overall_df["review_text"].fillna("").astype(str).head(20).tolist())
                st.caption(f"Will use up to 20 sample reviews from the loaded workspace to guide the AI.")
            if st.button("🤖 Generate symptom list", use_container_width=True,
                         disabled=(not api_key), key="sym_ai_build_btn"):
                overlay = _show_thinking("Generating symptom catalog for your product…")
                try:
                    sample_list = [r for r in sample_reviews_text.split("---") if r.strip()][:20]
                    result = _ai_build_symptom_list(client=client,
                                                    product_description=product_desc,
                                                    sample_reviews=sample_list)
                    st.session_state["sym_ai_build_result"] = result
                    st.session_state["sym_product_profile"] = product_desc
                except Exception as exc:
                    st.error(f"AI builder failed: {exc}")
                finally:
                    overlay.empty()
                st.rerun()

            ai_result = st.session_state.get("sym_ai_build_result")
            if ai_result:
                st.markdown("**AI-generated symptoms — review and accept:**")
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("🟢 Delighters")
                    ai_del_text = st.text_area("Edit delighters",
                        value="\n".join(ai_result.get("delighters",[])),
                        height=200, key="sym_ai_del_edit")
                with col2:
                    st.markdown("🔴 Detractors")
                    ai_det_text = st.text_area("Edit detractors",
                        value="\n".join(ai_result.get("detractors",[])),
                        height=200, key="sym_ai_det_edit")
                if st.button("✅ Accept AI-generated symptoms", use_container_width=True, key="sym_accept_ai"):
                    def _parse_list(t: str) -> List[str]:
                        items = re.split(r"[\n,;|]+", t)
                        return [i.strip() for i in items if i.strip()]
                    st.session_state["sym_delighters"] = _parse_list(ai_del_text)
                    st.session_state["sym_detractors"] = _parse_list(ai_det_text)
                    st.session_state["sym_symptoms_source"] = "ai"
                    st.session_state.pop("sym_ai_build_result", None)
                    st.success(f"Accepted {len(st.session_state['sym_delighters'])} delighters and {len(st.session_state['sym_detractors'])} detractors.")
                    st.rerun()

    st.divider()

    # ── Step 2: Run configuration ───────────────────────────────────────────
    st.markdown("### Step 2 — Configure and run")

    delighters = list(st.session_state.get("sym_delighters") or [])
    detractors = list(st.session_state.get("sym_detractors") or [])

    # build working dataframe with missing flags
    colmap = _detect_symptom_columns(overall_df)
    work   = _detect_missing(overall_df, colmap)

    # apply scope filter
    scope_choice = st.selectbox("Symptomization scope",
        ["Missing both","Any missing","All loaded reviews","Current filtered reviews"],
        key="sym_scope_choice")
    if scope_choice=="Missing both":
        target_df = work[(work["Needs_Delighters"]) & (work["Needs_Detractors"])]
    elif scope_choice=="Any missing":
        target_df = work[(work["Needs_Delighters"]) | (work["Needs_Detractors"])]
    elif scope_choice=="Current filtered reviews":
        filtered_ids = set(filtered_df["review_id"].astype(str))
        target_df = work[work["review_id"].astype(str).isin(filtered_ids)]
    else:
        target_df = work

    run_cols = st.columns([1.5,1.0,1.0,1.5])
    n_to_process = run_cols[0].number_input("How many reviews to process",
        min_value=1, max_value=max(1,len(target_df)), step=1, key="sym_n_to_process")
    batch_size = int(st.session_state.get("sym_batch_size",5))
    est_batches = max(1,math.ceil(int(n_to_process)/batch_size)) if n_to_process else 0
    run_cols[1].metric("In scope",f"{len(target_df):,}")
    run_cols[2].metric("Est. batches",f"{est_batches:,}")
    run_cols[3].caption(f"Scope: {scope_choice}")

    # Stats
    need_both = int(work["Needs_Symptomization"].sum())
    need_del  = int(work["Needs_Delighters"].sum())
    need_det  = int(work["Needs_Detractors"].sum())
    st.markdown(
        f"""<div class='hero-grid' style='grid-template-columns:repeat(4,minmax(0,1fr));margin-top:0;'>
          <div class='hero-stat'><div class='label'>Total reviews</div><div class='value'>{len(overall_df):,}</div></div>
          <div class='hero-stat'><div class='label'>Need delighters</div><div class='value'>{need_del:,}</div></div>
          <div class='hero-stat'><div class='label'>Need detractors</div><div class='value'>{need_det:,}</div></div>
          <div class='hero-stat accent'><div class='label'>Missing both</div><div class='value'>{need_both:,}</div></div>
        </div>""",
        unsafe_allow_html=True)

    run_disabled = (not api_key) or (len(target_df)==0)
    run_btn = st.button(f"▶️ Run Symptomizer on {min(int(n_to_process),len(target_df)):,} review(s)",
                        type="primary", use_container_width=True, disabled=run_disabled, key="sym_run_btn")

    if run_disabled and not api_key:
        st.warning("Add OPENAI_API_KEY to Streamlit secrets to enable Symptomizer.")
    elif run_disabled and len(target_df)==0:
        st.info("No reviews match the current scope.")

    notice = st.session_state.pop("sym_run_notice", None)
    if notice: st.success(notice)

    if run_btn:
        rows_to_process = target_df.head(int(n_to_process)).copy()
        prog    = st.progress(0.0, text="Starting Symptomizer…")
        status  = st.empty()
        eta_box = st.empty()
        processed_rows_local: List[Dict] = []
        t0 = time.perf_counter()
        total_n = max(1, len(rows_to_process))
        done = 0

        # Prepare working df
        updated_df = _ensure_ai_columns(overall_df.copy())
        product_profile = st.session_state.get("sym_product_profile","")

        rows_list   = list(rows_to_process.iterrows())
        batches_idx = list(range(0, len(rows_list), batch_size))

        for bi, start in enumerate(batches_idx, 1):
            batch = rows_list[start:start+batch_size]
            items = []
            for idx, row in batch:
                vb = _clean_text(row.get("review_text","") or row.get("title_and_text",""))
                items.append(dict(idx=int(idx), review=vb,
                                  needs_del=bool(row.get("Needs_Delighters",True)),
                                  needs_det=bool(row.get("Needs_Detractors",True))))
            status.info(f"Batch {bi}/{len(batches_idx)} — rows {start+1}–{min(start+batch_size,len(rows_list))}")
            outs: Dict[int,Dict] = {}
            if client is not None:
                try:
                    outs = _call_symptomizer_batch(
                        client=client, items=items,
                        allowed_delighters=delighters or DEFAULT_PRIORITY_DELIGHTERS,
                        allowed_detractors=detractors or DEFAULT_PRIORITY_DETRACTORS,
                        product_profile=product_profile,
                        max_ev_chars=int(st.session_state.get("sym_max_ev_chars",120)))
                except Exception as exc:
                    status.warning(f"Batch {bi} failed: {exc}")

            for it in items:
                idx   = int(it["idx"])
                out   = outs.get(idx, dict(dels=[],dets=[],ev_del={},ev_det={},
                                            unl_dels=[],unl_dets=[],
                                            safety="Not Mentioned",reliability="Not Mentioned",sessions="Unknown"))
                wrote_dets = list(out.get("dets",[]))[:10]
                wrote_dels = list(out.get("dels",[]))[:10]
                for j, lab in enumerate(wrote_dets):
                    updated_df.loc[idx, f"AI Symptom Detractor {j+1}"] = lab
                for j, lab in enumerate(wrote_dels):
                    updated_df.loc[idx, f"AI Symptom Delighter {j+1}"] = lab
                updated_df.loc[idx,"AI Safety"]       = out.get("safety","Not Mentioned")
                updated_df.loc[idx,"AI Reliability"]  = out.get("reliability","Not Mentioned")
                updated_df.loc[idx,"AI # of Sessions"]= out.get("sessions","Unknown")

                for lab in (out.get("unl_dels",[]) or []) + (out.get("unl_dets",[]) or []):
                    lab = lab.strip()
                    if lab:
                        rec = st.session_state["sym_new_candidates"].setdefault(lab,{"count":0,"refs":[]})
                        rec["count"]+=1
                        if len(rec["refs"])<50: rec["refs"].append(idx)

                processed_rows_local.append(dict(
                    idx=idx, wrote_dets=wrote_dets, wrote_dels=wrote_dels,
                    safety=out.get("safety",""), reliability=out.get("reliability",""),
                    sessions=out.get("sessions",""),
                    ev_det=out.get("ev_det",{}), ev_del=out.get("ev_del",{}),
                    unl_dels=out.get("unl_dels",[]), unl_dets=out.get("unl_dets",[]),
                ))
                done += 1

            prog.progress(done/total_n, text=f"{done}/{total_n} reviews processed")
            elapsed = time.perf_counter()-t0
            rate = done/elapsed if elapsed>0 else 0
            rem  = (total_n-done)/rate if rate>0 else 0
            eta_box.markdown(f"**Speed:** {rate*60:.1f} rev/min · **ETA:** ~{_fmt_secs(rem)}")

        # Write back to shared dataset
        dataset = dict(st.session_state["analysis_dataset"])
        dataset["reviews_df"] = updated_df
        st.session_state["analysis_dataset"]   = dataset
        st.session_state["sym_processed_rows"] = processed_rows_local
        st.session_state["master_export_bundle"] = None  # invalidate cache
        status.success(f"✅ Symptomized {done:,} reviews.")
        st.session_state["sym_run_notice"] = f"Symptomized {done:,} reviews. Tags are now visible in Review Explorer."
        st.rerun()

    st.divider()

    # ── Step 3: Results ─────────────────────────────────────────────────────
    processed = st.session_state.get("sym_processed_rows") or []
    if processed:
        st.markdown("### Step 3 — Results")

        # Summary chips
        total_tags = sum(len(r.get("wrote_dets",[])) + len(r.get("wrote_dels",[])) for r in processed)
        st.markdown(
            _chip_html([
                (f"{len(processed)} reviews tagged","green"),
                (f"{total_tags} total labels written","blue"),
            ]),
            unsafe_allow_html=True)
        st.markdown("")

        # New symptom candidates inbox
        new_cands = {k:v for k,v in (st.session_state.get("sym_new_candidates") or {}).items()
                     if k.strip() and k.strip() not in (delighters+detractors)}
        if new_cands:
            with st.expander(f"🟡 New symptom candidates ({len(new_cands)})", expanded=False):
                st.caption("These labels were suggested by the AI but don't appear in your current catalog. Add them to your symptom list if relevant.")
                cand_df = pd.DataFrame([
                    dict(Add=False, Label=lab, Count=int(rec.get("count",0)))
                    for lab, rec in sorted(new_cands.items(), key=lambda kv:-int(kv[1].get("count",0)))
                ])
                edited_cands = st.data_editor(cand_df, num_rows="fixed", use_container_width=True,
                    hide_index=True, key="sym_cand_editor",
                    column_config={"Add":st.column_config.CheckboxColumn(),
                                   "Label":st.column_config.TextColumn(),
                                   "Count":st.column_config.NumberColumn(format="%d")})
                if st.button("Add selected to Detractors", key="sym_add_cands_det"):
                    to_add = [str(r["Label"]) for _,r in edited_cands.iterrows()
                              if bool(r.get("Add",False)) and str(r.get("Label","")).strip()]
                    if to_add:
                        st.session_state["sym_detractors"] = list(dict.fromkeys(detractors+to_add))
                        st.success(f"Added {len(to_add)} labels to detractors.")
                        st.rerun()

        # Processed review log
        with st.expander(f"📋 Processed review log ({len(processed)} reviews)", expanded=True):
            for rec in processed[-20:]:
                idx = rec.get("idx","?")
                head = (f"Row {idx} — det: {len(rec.get('wrote_dets',[]))} · "
                        f"del: {len(rec.get('wrote_dels',[]))}")
                with st.expander(head):
                    # Try to show verbatim
                    try:
                        verbatim = str(overall_df.loc[int(idx),"review_text"])[:600]
                        st.markdown("**Verbatim**"); st.write(verbatim)
                    except Exception: pass
                    st.markdown(
                        "<div class='chip-wrap'>"
                        +f"<span class='chip yellow'>Safety: {_safe_html(rec.get('safety',''))}</span>"
                        +f"<span class='chip blue'>Reliability: {_safe_html(rec.get('reliability',''))}</span>"
                        +f"<span class='chip purple'>Sessions: {_safe_html(rec.get('sessions',''))}</span>"
                        +"</div>", unsafe_allow_html=True)
                    det_chips = "".join([f"<span class='chip red'>{_safe_html(l)}</span>"
                                         for l in rec.get("wrote_dets",[])])
                    del_chips = "".join([f"<span class='chip green'>{_safe_html(l)}</span>"
                                         for l in rec.get("wrote_dels",[])])
                    if det_chips or del_chips:
                        st.markdown(f"<div class='chip-wrap'>{det_chips}{del_chips}</div>",unsafe_allow_html=True)
                    ev_det = rec.get("ev_det",{}) or {}
                    ev_del = rec.get("ev_del",{}) or {}
                    for lab, evs in {**ev_det,**ev_del}.items():
                        for e in (evs or []):
                            st.caption(f"  · {lab}: {e}")

        # Export with symptomizer output
        st.markdown("**Download tagged workbook**")
        sym_exp_cols = st.columns([1.5,3])
        if sym_exp_cols[0].button("🧾 Prepare symptomized export", use_container_width=True, key="sym_prepare_export"):
            updated_reviews_df = st.session_state["analysis_dataset"]["reviews_df"]
            # Try to get original workbook bytes for template export
            original_bytes = st.session_state.get("_uploaded_raw_bytes")
            if original_bytes:
                sym_export_bytes = _generate_symptomized_workbook(original_bytes, updated_reviews_df)
            else:
                # Fall back to master Excel export
                sym_export_bytes = _build_master_excel(summary, updated_reviews_df)
            st.session_state["sym_export_bytes"] = sym_export_bytes
            st.success("Export prepared.")
        sym_export = st.session_state.get("sym_export_bytes")
        sym_exp_cols[0].download_button("⬇️ Download symptomized file",
            data=sym_export or b"",
            file_name=f"{summary.product_id}_Symptomized.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            disabled=(sym_export is None), key="sym_download_btn")
        sym_exp_cols[1].caption(
            "The export writes AI Symptom Detractor / Delighter columns and Safety / Reliability / # of Sessions "
            "to the exact template column positions (K–AD, AE–AG). If you uploaded the original workbook, "
            "the export preserves all original data.")
    else:
        st.info("Run the Symptomizer above to see results here.")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN APP
# ═══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    st.title(APP_TITLE)
    st.caption(
        "Build a review workspace from a SharkNinja product URL or an uploaded review export, "
        "then filter the voice of customer, explore reviews, chat with an AI analyst, "
        "create row-level AI tags, and run the Symptomizer for structured delighter/detractor tagging.")

    # ── Workspace source selector ───────────────────────────────────────────
    dataset = st.session_state.get("analysis_dataset")

    if dataset:
        cols = st.columns([4.2,1.0])
        cols[0].caption(f"Current workspace · {dataset.get('source_type','').title()} · {dataset.get('source_label','')}")
        if cols[1].button("Clear workspace", use_container_width=True, key="ws_clear"):
            for k in ["analysis_dataset","chat_messages","chat_scope_signature","master_export_bundle",
                      "prompt_run_artifacts","sym_processed_rows","sym_new_candidates",
                      "sym_product_profile","sym_symptoms_source","sym_delighters","sym_detractors",
                      "_uploaded_raw_bytes","sym_export_bytes"]:
                st.session_state[k] = (None if k=="analysis_dataset" else
                                       [] if k in {"chat_messages","sym_processed_rows"} else
                                       {} if k in {"sym_new_candidates","sym_delighters","sym_detractors"} else
                                       "none" if k=="sym_symptoms_source" else
                                       "" if k in {"sym_product_profile"} else None)
            st.rerun()

    source_mode = st.radio("Workspace source",["SharkNinja product URL","Uploaded review file"],
                           horizontal=True, key="workspace_source_mode")

    if source_mode=="SharkNinja product URL":
        product_url = st.text_input("Product URL",
            value="https://www.sharkninja.com/ninja-air-fryer-pro-xl-6-in-1/AF181.html",
            help="Example: https://www.sharkninja.com/ninja-air-fryer-pro-xl-6-in-1/AF181.html")
        if st.button("Build review workspace", type="primary", key="ws_build_url"):
            try:
                new_dataset = _load_product_reviews(product_url)
                st.session_state["analysis_dataset"] = new_dataset
                st.session_state["chat_messages"] = []
                st.session_state["master_export_bundle"] = None
                st.session_state["prompt_run_artifacts"] = None
                st.session_state["sym_processed_rows"] = []
                st.session_state["sym_new_candidates"] = {}
                st.session_state["sym_symptoms_source"] = "none"
                st.rerun()
            except requests.HTTPError as exc: st.error(f"HTTP error: {exc}")
            except ReviewDownloaderError as exc: st.error(str(exc))
            except Exception as exc: st.exception(exc)
    else:
        uploaded_files = st.file_uploader("Upload review export files",
            type=["csv","xlsx","xls"], accept_multiple_files=True,
            help="Supports Axion-style exports and similar CSV/XLSX review files.")
        st.caption("Mapped columns include: Event Id, Base SKU, Review Text, Rating, Opened date, Seeded Flag, Retailer.")
        if st.button("Build review workspace from file", type="primary", key="ws_build_file"):
            try:
                new_dataset = _load_uploaded_files(uploaded_files or [])
                st.session_state["analysis_dataset"] = new_dataset
                st.session_state["chat_messages"] = []
                st.session_state["master_export_bundle"] = None
                st.session_state["prompt_run_artifacts"] = None
                st.session_state["sym_processed_rows"] = []
                st.session_state["sym_new_candidates"] = {}
                st.session_state["sym_symptoms_source"] = "none"
                # store raw bytes for symptomizer template export
                if uploaded_files and len(uploaded_files)==1:
                    fname = getattr(uploaded_files[0],"name","")
                    if fname.lower().endswith(".xlsx"):
                        st.session_state["_uploaded_raw_bytes"] = uploaded_files[0].getvalue()
                st.rerun()
            except ReviewDownloaderError as exc: st.error(str(exc))
            except Exception as exc: st.exception(exc)

    # ── Sidebar + filtered data ──────────────────────────────────────────────
    dataset = st.session_state.get("analysis_dataset")
    settings = _render_shared_sidebar(dataset["reviews_df"] if dataset else None)

    if not dataset:
        st.info("Build a review workspace to unlock the Dashboard, Review Explorer, AI Analyst, Review Prompt, and Symptomizer.")
        return

    summary: ReviewBatchSummary = dataset["summary"]
    overall_df: pd.DataFrame    = dataset["reviews_df"]
    source_type  = dataset.get("source_type","bazaarvoice")
    source_label = dataset.get("source_label","")

    # Map review source mode to incentivized_mode
    src_map = {"All reviews":"All reviews","Organic only":"Non-incentivized only","Incentivized only":"Incentivized only"}
    filtered_df = _apply_filters(overall_df,
        selected_ratings=settings["selected_ratings"],
        incentivized_mode=src_map.get(settings["review_source_mode"],"All reviews"),
        selected_products=settings["selected_products"],
        selected_locales=settings["selected_locales"],
        recommendation_mode=settings["recommendation_mode"],
        syndicated_mode="All", media_mode="All",
        date_range=settings["date_range"],
        text_query=settings["text_query"])
    filter_description = _describe_filters(
        selected_ratings=settings["selected_ratings"],
        selected_products=settings["selected_products"],
        review_source_mode=settings["review_source_mode"],
        selected_locales=settings["selected_locales"],
        recommendation_mode=settings["recommendation_mode"],
        date_range=settings["date_range"],
        text_query=settings["text_query"])

    _render_workspace_header(summary, overall_df, st.session_state.get("prompt_run_artifacts"),
                             source_type=source_type, source_label=source_label)
    _render_top_metrics(overall_df, filtered_df)
    st.caption(f"Filter status: {filter_description}. Showing {len(filtered_df):,} of {len(overall_df):,} reviews.")

    # ── Navigation ───────────────────────────────────────────────────────────
    VIEWS = ["Dashboard","Review Explorer","AI Analyst","Review Prompt","Symptomizer"]
    if st.session_state.get("workspace_view_selector") not in VIEWS:
        st.session_state["workspace_view_selector"] = "Dashboard"
    st.radio("Workspace view", options=VIEWS, horizontal=True, key="workspace_view_selector")
    st.session_state["active_main_view"] = st.session_state["workspace_view_selector"]

    active_view = st.session_state.get("active_main_view","Dashboard")
    common_kwargs = dict(settings=settings, overall_df=overall_df, filtered_df=filtered_df,
                         summary=summary, filter_description=filter_description)

    if active_view=="Dashboard":
        _render_dashboard(filtered_df)
    elif active_view=="Review Explorer":
        _render_review_explorer(summary=summary, overall_df=overall_df, filtered_df=filtered_df,
                                prompt_artifacts=st.session_state.get("prompt_run_artifacts"))
    elif active_view=="AI Analyst":
        _render_ai_tab(**common_kwargs)
    elif active_view=="Review Prompt":
        _render_review_prompt_tab(**common_kwargs)
    elif active_view=="Symptomizer":
        _render_symptomizer_tab(**common_kwargs)


if __name__ == "__main__":
    main()
