# starwalk_ui_v7_5_vector_index.py
# Evidence-Locked Labeling + Overflow Tag + Optional Review Skim (Embeddings)
# Exports: Detractors K–T, Delighters U–AD, Meta AE/AF/AG, Overflow AH
# Requirements: streamlit>=1.28, pandas, openpyxl, numpy, openai>=1.0.0 (optional)

import streamlit as st
import pandas as pd
import numpy as np
import io, os, json, difflib, time, re, html, random
from typing import List, Dict, Tuple, Optional, Set, Any

# -------- Optional: OpenAI --------
try:
    from openai import OpenAI
    _HAS_OPENAI = True
except Exception:
    OpenAI = None  # type: ignore
    _HAS_OPENAI = False

# -------- Excel handling ----------
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

# ------------------- Page Setup -------------------
st.set_page_config(layout="wide", page_title="Review Symptomizer — v7.5 (Evidence • Overflow • Skim)")
st.title("✨ Review Symptomizer — v7.5 (Evidence • Overflow • Skim)")
st.caption("Exact export (K–T dets, U–AD dels) • ETA + presets + overwrite • Undo • New-symptom inbox • Tiles UI • Similarity guard • Evidence-locked labeling • Overflow tag → AH • Optional review skim via embeddings")

# ------------------- Global CSS -------------------
st.markdown(
    """
    <style>
      @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;800&display=swap');
      :root { --brand:#7c3aed; --brand2:#06b6d4; --ok:#16a34a; --bad:#dc2626; --muted:#6b7280; }
      html, body, .stApp { font-family: 'Inter', system-ui, -apple-system, Segoe UI, Roboto, sans-serif; }
      .stApp { background:
        radial-gradient(1200px 500px at 10% -20%, rgba(124,58,237,.18), transparent 60%),
        radial-gradient(1200px 500px at 100% 0%, rgba(6,182,212,.16), transparent 60%);
      }
      .hero { border-radius: 20px; padding: 18px 22px; color: #0b1020;
        background: linear-gradient(180deg, rgba(255,255,255,.95), rgba(255,255,255,.86));
        border: 1px solid rgba(226,232,240,.9);
        box-shadow: 0 10px 30px rgba(16,24,40,.08), 0 2px 6px rgba(16,24,40,.06) inset; }
      .hero-stats { display:flex; gap:14px; flex-wrap:wrap; margin-top: 6px; }
      .stat { background:#fff; border:1px solid #e6eaf0; border-radius:16px; padding:12px 16px; min-width:160px;
        box-shadow: 0 2px 8px rgba(16,24,40,.05); }
      .stat.accent { border-color: rgba(124,58,237,.35); box-shadow: 0 4px 12px rgba(124,58,237,.15); }
      .stat .label{ font-size:11px; letter-spacing:.08em; text-transform:uppercase; color:#64748b; }
      .stat .value{ font-size:28px; font-weight:800; }
      .chip-wrap {display:flex; flex-wrap:wrap; gap:8px;}
      .chip { padding:6px 10px; border-radius:999px; font-size:12.5px; border:1px solid #e6eaf0; background:#fff; box-shadow: 0 1px 2px rgba(16,24,40,.06); }
      .chip.red { background: #fff1f2; border-color:#fecdd3; }
      .chip.green { background: #ecfdf3; border-color:#bbf7d0; }
      .chip.blue { background: #e0f2fe; border-color:#bae6fd; }
      .chip.yellow { background: #fff7ed; border-color:#fed7aa; }
      .chip.purple { background: #f3e8ff; border-color:#e9d5ff; }
      .muted{ color:#64748b; font-size:12px; }
      .chips-block { margin-bottom: 16px; }
      .stButton > button { height: 40px; border-radius: 10px; font-weight: 600;
        background: linear-gradient(180deg, #ffffff, #f7f8fb);
        border: 1px solid #e6eaf0; box-shadow: 0 1px 2px rgba(16,24,40,.04); }
      .stButton > button:hover { border-color: rgba(124,58,237,.35); box-shadow: 0 2px 6px rgba(124,58,237,.15); }
      div.batch-row .stNumberInput input { height: 40px; font-weight: 700; }
      div.batch-row .stButton > button { border-radius: 999px; height: 36px; font-weight: 700; min-width: 72px;
        background: #fff; border: 1px solid rgba(6,182,212,.45); }
      div.batch-row .stButton > button:hover { background: #f0fdff; border-color: var(--brand2); }
      mark.hl { background: #fde68a; padding: 0 .15em; border-radius: .25em; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ------------------- Utilities -------------------
NON_VALUES = {"<NA>", "NA", "N/A", "NONE", "-", "", "NAN", "NULL"}

def clean_text(x):
    if pd.isna(x):
        return ""
    return str(x).strip()

def is_filled(val) -> bool:
    if pd.isna(val):
        return False
    s = str(val).strip()
    return (s != "") and (s.upper() not in NON_VALUES)

@st.cache_data(show_spinner=False)
def get_symptom_whitelists(file_bytes: bytes) -> Tuple[List[str], List[str], Dict[str, List[str]]]:
    bio = io.BytesIO(file_bytes)
    try:
        df_sym = pd.read_excel(bio, sheet_name="Symptoms")
    except Exception:
        return [], [], {}
    if df_sym is None or df_sym.empty:
        return [], [], {}

    df_sym.columns = [str(c).strip() for c in df_sym.columns]
    lowcols = {c.lower(): c for c in df_sym.columns}

    alias_col = next((lowcols.get(c) for c in ["aliases", "alias"] if c in lowcols), None)
    label_col = next((lowcols.get(c) for c in ["symptom", "label", "name", "item"] if c in lowcols), None)
    type_col  = next((lowcols.get(c) for c in ["type", "polarity", "category", "side"] if c in lowcols), None)

    POS_TAGS = {"delighter", "delighters", "positive", "pos", "pros"}
    NEG_TAGS = {"detractor", "detractors", "negative", "neg", "cons"}

    def _clean(series: pd.Series) -> List[str]:
        vals = series.dropna().astype(str).map(str.strip)
        out, seen = [], set()
        for v in vals:
            if v and v not in seen:
                seen.add(v); out.append(v)
        return out

    delighters, detractors, alias_map = [], [], {}

    if label_col and type_col:
        df_sym[type_col] = df_sym[type_col].astype(str).str.lower().str.strip()
        delighters = _clean(df_sym.loc[df_sym[type_col].isin(POS_TAGS), label_col])
        detractors = _clean(df_sym.loc[df_sym[type_col].isin(NEG_TAGS), label_col])
        if alias_col:
            for _, r in df_sym.iterrows():
                lbl = str(r.get(label_col, "")).strip()
                als = str(r.get(alias_col, "")).strip()
                if lbl and als:
                    als_norm = als.replace(",", "|")
                    alias_map[lbl] = [p.strip() for p in als_norm.split("|") if p.strip()]
    else:
        for lc, orig in lowcols.items():
            if ("delight" in lc) or ("positive" in lc) or lc in {"pros"}:
                delighters.extend(_clean(df_sym[orig]))
            if ("detract" in lc) or ("negative" in lc) or lc in {"cons"}:
                detractors.extend(_clean(df_sym[orig]))
        delighters = list(dict.fromkeys(delighters))
        detractors = list(dict.fromkeys(detractors))

    return delighters, detractors, alias_map

# Canonicalization helpers
def _canon(s: str) -> str:
    return " ".join(str(s).split()).lower().strip()
def _canon_simple(s: str) -> str:
    return "".join(ch for ch in _canon(s) if ch.isalnum())

# Evidence highlighting
def highlight_text(text: str, terms: List[str]) -> str:
    safe = html.escape(str(text))
    terms = [t for t in (terms or []) if isinstance(t, str) and len(t.strip()) >= 3]
    if not terms: return safe
    uniq = sorted({t.strip() for t in terms}, key=len, reverse=True)
    try:
        pattern = re.compile("|".join(re.escape(t) for t in uniq), re.IGNORECASE)
    except Exception:
        return safe
    return pattern.sub(lambda m: f"<mark class='hl'>{html.escape(m.group(0))}</mark>", safe)

# Schema detection
def detect_symptom_columns(df: pd.DataFrame) -> Dict[str, List[str]]:
    cols = [str(c).strip() for c in df.columns]
    man_det = [f"Symptom {i}" for i in range(1, 11) if f"Symptom {i}" in cols]
    man_del = [f"Symptom {i}" for i in range(11, 21) if f"Symptom {i}" in cols]
    ai_det  = [c for c in cols if c.startswith("AI Symptom Detractor ")]
    ai_del  = [c for c in cols if c.startswith("AI Symptom Delighter ")]
    return {
        "manual_detractors": man_det,
        "manual_delighters": man_del,
        "ai_detractors": ai_det,
        "ai_delighters": ai_del,
    }

def row_has_any(row: pd.Series, columns: List[str]) -> bool:
    if not columns: return False
    for c in columns:
        if c in row and is_filled(row[c]): return True
    return False

def detect_missing(df: pd.DataFrame, colmap: Dict[str, List[str]]) -> pd.DataFrame:
    det_cols = colmap["manual_detractors"] + colmap["ai_detractors"]
    del_cols = colmap["manual_delighters"] + colmap["ai_delighters"]
    out = df.copy()
    out["Has_Detractors"] = out.apply(lambda r: row_has_any(r, det_cols), axis=1)
    out["Has_Delighters"] = out.apply(lambda r: row_has_any(r, del_cols), axis=1)
    out["Needs_Detractors"] = ~out["Has_Detractors"]
    out["Needs_Delighters"] = ~out["Has_Delighters"]
    out["Needs_Symptomization"] = out["Needs_Detractors"] & out["Needs_Delighters"]
    return out

# ------------------- Fixed template mapping -------------------
DET_LETTERS = ["K","L","M","N","O","P","Q","R","S","T"]
DEL_LETTERS = ["U","V","W","X","Y","Z","AA","AB","AC","AD"]
DET_INDEXES = [column_index_from_string(c) for c in DET_LETTERS]
DEL_INDEXES = [column_index_from_string(c) for c in DEL_LETTERS]

# Meta columns after AD (headers only if blank)
# NEW: add "Label Overflow" at AH (data column in df is "AI Label Overflow")
META_ORDER = [("Safety", "AE"), ("Reliability", "AF"), ("# of Sessions", "AG"), ("Label Overflow", "AH")]
META_INDEXES = {name: column_index_from_string(col) for name, col in META_ORDER}

AI_DET_HEADERS = [f"AI Symptom Detractor {i}" for i in range(1, 11)]
AI_DEL_HEADERS = [f"AI Symptom Delighter {i}" for i in range(1, 11)]
AI_META_HEADERS = ["AI Safety", "AI Reliability", "AI # of Sessions", "AI Label Overflow"]  # <- NEW

def ensure_ai_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    for h in AI_DET_HEADERS + AI_DEL_HEADERS + AI_META_HEADERS:
        if h not in df_in.columns:
            df_in[h] = None
    return df_in

def build_canonical_maps(delighters: List[str], detractors: List[str], alias_map: Dict[str, List[str]]):
    del_map = {_canon(x): x for x in delighters}
    det_map = {_canon(x): x for x in detractors}
    alias_to_label: Dict[str, str] = {}
    for label, aliases in (alias_map or {}).items():
        for a in aliases:
            alias_to_label[_canon(a)] = label
    return del_map, det_map, alias_to_label

# ---------- LLM helpers ----------
SAFETY_ENUM = ["Not Mentioned", "Concern", "Positive"]
RELIABILITY_ENUM = ["Not Mentioned", "Negative", "Neutral", "Positive"]
SESSIONS_ENUM = ["0", "1", "2–3", "4–9", "10+", "Unknown"]

def _symptom_list_version(delighters: List[str], detractors: List[str], aliases: Dict[str, List[str]]) -> str:
    payload = json.dumps({"del": delighters, "det": detractors, "ali": aliases}, sort_keys=True, ensure_ascii=False)
    try:
        import hashlib
        return hashlib.md5(payload.encode("utf-8")).hexdigest()[:10]
    except Exception:
        return f"{len(delighters)}_{len(detractors)}"

def _ensure_label_cache():
    if "_label_cache" not in st.session_state:
        st.session_state["_label_cache"] = {}
    return st.session_state["_label_cache"]

def _openai_labeler(
    verbatim: str,
    client,
    model: str,
    temperature: float,
    delighters: List[str],
    detractors: List[str],
    alias_map: Dict[str, List[str]],
    del_map: Dict[str, str],
    det_map: Dict[str, str],
    alias_to_label: Dict[str, str],
    max_ev_per_label: int = 2,
    max_ev_chars: int = 120,
) -> Tuple[List[str], List[str], List[str], List[str], Dict[str, List[str]], Dict[str, List[str]]]:
    """
    Evidence-locked version with in-session cache.
    Returns:
      dels, dets, unl_dels, unl_dets, ev_del_map, ev_det_map
    """
    if (client is None) or (not verbatim or not verbatim.strip()):
        return [], [], [], [], {}, {}

    v = _symptom_list_version(delighters, detractors, alias_map)
    key = ("lab", _canon(verbatim), model, f"{temperature:.2f}", v, max_ev_per_label, max_ev_chars)
    cache = _ensure_label_cache()
    if key in cache:
        return cache[key]

    sys = "\n".join([
        "You label consumer reviews with predefined symptom lists.",
        "Return STRICT JSON with this schema:",
        '{"detractors":[{"label":"<one from allowed detractors or close alias>","evidence":["<exact substring from review>", "..."]}],',
        ' "delighters":[{"label":"<one from allowed delighters or close alias>","evidence":["<exact substring>", "..."]}],',
        ' "unlisted_detractors":["..."], "unlisted_delighters":["..."]}',
        "",
        "Rules:",
        f"- Evidence MUST be exact substrings from the review. Each ≤ {max_ev_chars} chars. Up to {max_ev_per_label} per label.",
        "- Only include a label if there is clear textual support in the review.",
        "- Use the allowed lists; if close wording appears, pick the closest allowed label or list it under unlisted_*.",
        "- Cap to maximum 10 detractors and 10 delighters.",
    ])
    user_payload = {
        "review": verbatim.strip(),
        "allowed_delighters": delighters,
        "allowed_detractors": detractors
    }

    try:
        resp = client.chat.completions.create(
            model=model,
            temperature=float(temperature),
            messages=[{"role": "system", "content": sys},
                      {"role": "user", "content": json.dumps(user_payload)}],
            response_format={"type": "json_object"}
        )
        content = resp.choices[0].message.content or "{}"
        data = json.loads(content)
    except Exception:
        return [], [], [], [], {}, {}

    # Normalize structure
    raw_dels = data.get("delighters", []) or []
    raw_dets = data.get("detractors", []) or []
    unl_dels = [x for x in (data.get("unlisted_delighters", []) or [])][:10]
    unl_dets = [x for x in (data.get("unlisted_detractors", []) or [])][:10]

    def _canon_map_item(obj: Any, side: str) -> Tuple[Optional[str], List[str]]:
        try:
            lbl_raw = str(obj.get("label", "")).strip()
            evs = [str(e)[:max_ev_chars] for e in (obj.get("evidence", []) or []) if isinstance(e, str) and e.strip()]
        except Exception:
            lbl_raw, evs = "", []
        key2 = _canon(lbl_raw)
        if side == "del":
            label = del_map.get(key2) or alias_to_label.get(key2)
            if label and (label in delighters):
                return label, evs[:max_ev_per_label]
        else:
            label = det_map.get(key2) or alias_to_label.get(key2)
            if label and (label in detractors):
                return label, evs[:max_ev_per_label]
        return None, []

    dels: List[str] = []
    dets: List[str] = []
    ev_del_map: Dict[str, List[str]] = {}
    ev_det_map: Dict[str, List[str]] = {}

    for obj in raw_dels:
        label, evs = _canon_map_item(obj, side="del")
        if label and (label not in dels):
            dels.append(label)
            ev_del_map[label] = evs
        if len(dels) >= 10: break

    for obj in raw_dets:
        label, evs = _canon_map_item(obj, side="det")
        if label and (label not in dets):
            dets.append(label)
            ev_det_map[label] = evs
        if len(dets) >= 10: break

    out = (dels, dets, unl_dels, unl_dets, ev_del_map, ev_det_map)
    cache[key] = out
    return out

def _openai_meta_extractor(verbatim: str, client, model: str, temperature: float) -> Tuple[str, str, str]:
    if (client is None) or (not verbatim or not verbatim.strip()):
        return "Not Mentioned", "Not Mentioned", "Unknown"
    sys = "\n".join([
        "Extract three fields from this consumer review. Use ONLY the allowed values.",
        "SAFETY one of: ['Not Mentioned','Concern','Positive']",
        "RELIABILITY one of: ['Not Mentioned','Negative','Neutral','Positive']",
        "SESSIONS one of: ['0','1','2–3','4–9','10+','Unknown']",
        'Return strict JSON {"safety":"…","reliability":"…","sessions":"…"}',
    ])
    try:
        resp = client.chat.completions.create(
            model=model,
            temperature=float(temperature),
            messages=[{"role": "system", "content": sys},
                      {"role": "user", "content": f'Review:\n"""{verbatim.strip()}"""'}],
            response_format={"type": "json_object"}
        )
        content = resp.choices[0].message.content or "{}"
        data = json.loads(content)
        s = str(data.get("safety", "Not Mentioned")).strip()
        r = str(data.get("reliability", "Not Mentioned")).strip()
        n = str(data.get("sessions", "Unknown")).strip()
        s = s if s in SAFETY_ENUM else "Not Mentioned"
        r = r if r in RELIABILITY_ENUM else "Not Mentioned"
        n = n if n in SESSIONS_ENUM else "Unknown"
        return s, r, n
    except Exception:
        return "Not Mentioned", "Not Mentioned", "Unknown"

# ---------------- Embeddings-based skim (robust) ----------------
EMBED_MODEL = st.secrets.get("OPENAI_EMBED_MODEL",
                             os.getenv("OPENAI_EMBED_MODEL", "text-embedding-3-small"))

def _chunk_text(t: str, max_chars: int = 5000, overlap: int = 200) -> List[str]:
    t = (t or "").strip()
    if not t:
        return []
    chunks = []
    i, n = 0, len(t)
    while i < n:
        j = min(i + max_chars, n)
        chunks.append(t[i:j])
        if j >= n: break
        i = max(0, j - overlap)
    return chunks

@st.cache_data(show_spinner=False)
def build_vector_index(texts: List[str], api_key: Optional[str], batch_size: int = 128) -> Dict[str, Any]:
    """
    Returns: {"vectors": np.ndarray, "chunks": list[str], "index_map": list[tuple[row_idx, chunk_idx]]}
    """
    if not (_HAS_OPENAI and api_key):
        return {"vectors": None, "chunks": [], "index_map": []}

    client = OpenAI(api_key=api_key)

    # Clean & chunk
    cleaned: List[Tuple[int, str]] = []
    for row_idx, t in enumerate(texts or []):
        if isinstance(t, str) and t.strip():
            for c in _chunk_text(t, max_chars=3000, overlap=150):
                cleaned.append((row_idx, c))
    if not cleaned:
        return {"vectors": None, "chunks": [], "index_map": []}

    # Build chunks + index map [(row_idx, local_chunk_idx)]
    index_map: List[Tuple[int, int]] = []
    chunks: List[str] = []
    current_parent = -1
    local_idx = -1
    for row_idx, c in cleaned:
        if row_idx != current_parent:
            current_parent = row_idx
            local_idx = 0
        else:
            local_idx += 1
        index_map.append((row_idx, local_idx))
        chunks.append(c)

    # Embed in batches
    vectors: List[List[float]] = []
    try:
        for i in range(0, len(chunks), batch_size):
            batch = [b for b in chunks[i:i+batch_size] if isinstance(b, str) and b.strip()]
            if not batch: continue
            resp = client.embeddings.create(model=EMBED_MODEL, input=batch)
            vectors.extend([d.embedding for d in resp.data])
    except Exception:
        # Re-raise to put details in logs; Streamlit UI shows guidance below
        raise

    if not vectors:
        return {"vectors": None, "chunks": [], "index_map": []}

    return {
        "vectors": np.array(vectors, dtype=float),
        "chunks": chunks,
        "index_map": index_map,
    }

def _cosine_sim(a: np.ndarray, b: np.ndarray) -> float:
    na = np.linalg.norm(a); nb = np.linalg.norm(b)
    if na == 0 or nb == 0: return 0.0
    return float(np.dot(a, b) / (na * nb))

def embed_query(text: str, api_key: Optional[str]) -> Optional[np.ndarray]:
    if not (_HAS_OPENAI and api_key and text and text.strip()):
        return None
    client = OpenAI(api_key=api_key)
    resp = client.embeddings.create(model=EMBED_MODEL, input=[text.strip()])
    return np.array(resp.data[0].embedding, dtype=float)

def top_k_chunks(query_vec: np.ndarray, index: Dict[str, Any], k: int = 12) -> List[Tuple[int, float, str]]:
    V = index.get("vectors"); chunks = index.get("chunks", [])
    if V is None or V.shape[0] == 0: return []
    sims = [(_cosine_sim(query_vec, V[i]), i) for i in range(V.shape[0])]
    sims.sort(reverse=True)
    out = []
    for s, i in sims[:k]:
        out.append((i, s, chunks[i]))
    return out

def _sample_diverse_chunks(index: Dict[str, Any], n: int = 24, seed: int = 7) -> List[str]:
    chunks = index.get("chunks", [])
    if not chunks: return []
    random.seed(seed)
    if len(chunks) <= n:
        return chunks
    # light “diversity”: sort by length buckets and sample
    chunks_sorted = sorted(chunks, key=len)
    step = max(1, len(chunks_sorted) // n)
    picks = [chunks_sorted[i] for i in range(0, len(chunks_sorted), step)]
    if len(picks) > n:
        picks = random.sample(picks, n)
    return picks

def summarize_product(chunks: List[str], client: Optional[OpenAI], model: str, temperature: float) -> str:
    if not (client and chunks):
        return ""
    # cap total tokens by limiting number and size
    snippets = [c[:400] for c in chunks[:24]]
    prompt = {
        "role": "user",
        "content": json.dumps({
            "task": "Summarize product understanding to aid mass classification.",
            "instructions": [
                "Use only the provided snippets (may be noisy).",
                "Return short bullet points under these headings:",
                "1) Top Detractor Themes",
                "2) Top Delighter Themes",
                "3) Likely Vocabulary / Aliases",
                "4) Ambiguities / Edge Cases to watch",
                "5) Quick Tagging Heuristics (succinct rules)",
            ],
            "snippets": snippets
        })
    }
    try:
        resp = client.chat.completions.create(
            model=model,
            temperature=float(temperature),
            messages=[
                {"role": "system", "content": "You are a rigorous QA analyst for consumer reviews. Be concise and concrete."},
                prompt
            ]
        )
        return (resp.choices[0].message.content or "").strip()
    except Exception:
        return ""

# ------------------- File Upload -------------------
uploaded_file = st.file_uploader("📂 Upload Excel (with 'Star Walk scrubbed verbatims' + 'Symptoms')", type=["xlsx"])
if not uploaded_file: st.stop()

uploaded_bytes = uploaded_file.read(); uploaded_file.seek(0)
try:
    df = pd.read_excel(uploaded_file, sheet_name="Star Walk scrubbed verbatims")
except ValueError:
    uploaded_file.seek(0); df = pd.read_excel(uploaded_file)

if "Verbatim" not in df.columns:
    st.error("Missing 'Verbatim' column."); st.stop()

# Normalize
df.columns = [str(c).strip() for c in df.columns]
df["Verbatim"] = df["Verbatim"].map(clean_text)

# Load Symptoms
DELIGHTERS, DETRACTORS, ALIASES = get_symptom_whitelists(uploaded_bytes)
if not DELIGHTERS and not DETRACTORS:
    st.warning("⚠️ No Symptoms found in 'Symptoms' tab.")
else:
    st.success(f"Loaded {len(DELIGHTERS)} Delighters, {len(DETRACTORS)} Detractors from Symptoms tab.")

# Canonical maps
DEL_MAP, DET_MAP, ALIAS_TO_LABEL = build_canonical_maps(DELIGHTERS, DETRACTORS, ALIASES)

# ------------------- Detection & KPIs -------------------
colmap = detect_symptom_columns(df)
work = detect_missing(df, colmap)

total = len(work)
need_del = int(work["Needs_Delighters"].sum())
need_det = int(work["Needs_Detractors"].sum())
need_both = int(work["Needs_Symptomization"].sum())

st.markdown(f"""
<div class="hero">
  <div class="hero-stats">
    <div class="stat"><div class="label">Total Reviews</div><div class="value">{total:,}</div></div>
    <div class="stat"><div class="label">Need Delighters</div><div class="value">{need_del:,}</div></div>
    <div class="stat"><div class="label">Need Detractors</div><div class="value">{need_det:,}</div></div>
    <div class="stat accent"><div class="label">Missing Both</div><div class="value">{need_both:,}</div></div>
  </div>
</div>
""", unsafe_allow_html=True)

# ------------------- LLM & Similarity Settings -------------------
st.sidebar.header("🤖 LLM Settings")
MODEL_CHOICES = {
    "Fast – GPT-4o-mini": "gpt-4o-mini",
    "Balanced – GPT-4o": "gpt-4o",
    "Advanced – GPT-4.1": "gpt-4.1",
    "Most Advanced – GPT-5": "gpt-5",
}
model_label = st.sidebar.selectbox("Model", list(MODEL_CHOICES.keys()), index=1)
selected_model = MODEL_CHOICES[model_label]
temperature = st.sidebar.slider("Creativity (temperature)", 0.0, 1.0, 0.2, 0.1)

# API Client
api_key = st.secrets.get("OPENAI_API_KEY", os.getenv("OPENAI_API_KEY"))
client = OpenAI(api_key=api_key) if (_HAS_OPENAI and api_key) else None
if client is None:
    st.sidebar.warning("OpenAI not configured — set OPENAI_API_KEY and install 'openai' (>=1.0.0).")

# Similarity guard for new-symptom proposals
sim_threshold = st.sidebar.slider("New-symptom similarity guard", 0.80, 0.99, 0.94, 0.01,
                                  help="Raise to suppress near-duplicates; lower to see more proposals.")

# ----- Evidence settings -----
require_evidence = st.sidebar.checkbox(
    "Require evidence to write labels",
    value=True,
    help="If ON, a label must include ≥1 exact snippet from the review to be written."
)
max_ev_per_label = st.sidebar.slider("Max evidence per label", 1, 3, 2)
max_ev_chars = st.sidebar.slider("Max evidence length (chars)", 40, 200, 120, 10)

# ----- Optional Review Skim (Embeddings) -----
st.subheader("📚 Optional Review Skim (Embeddings)")
col_a, col_b, col_c = st.columns([1.2, 1.2, 2.2])
with col_a:
    do_build = st.button("🔎 Build/Refresh Embedding Index")
with col_b:
    do_skim = st.button("🧭 Generate Product Skim Summary")

if do_build:
    try:
        with st.spinner(f"Embedding {len(df)} reviews with {EMBED_MODEL}..."):
            VINDEX = build_vector_index(df["Verbatim"].astype(str).tolist(), api_key)
            st.session_state["VINDEX"] = VINDEX
        st.success("Embedding index built.")
    except Exception:
        st.error(
            "Embedding request failed. Common causes:\n"
            "• Using a chat model instead of an embedding model\n"
            "• Empty/invalid inputs\n"
            "• Batch too large or text too long\n\n"
            f"Model used: {EMBED_MODEL}"
        )

if do_skim:
    VINDEX = st.session_state.get("VINDEX", {"vectors": None, "chunks": []})
    if VINDEX.get("vectors") is None or not VINDEX.get("chunks"):
        st.info("No embedding index yet. Click **Build/Refresh Embedding Index** first.")
    else:
        with st.spinner("Summarizing..."):
            picks = _sample_diverse_chunks(VINDEX, n=24, seed=13)
            summary = summarize_product(picks, client, selected_model, temperature)
        if summary:
            st.markdown("#### 📌 Product Skim Summary")
            st.write(summary)
        else:
            st.info("No summary produced (check API settings).")

# ------------------- Scope & Preview -------------------
st.subheader("🧪 Symptomize")
scope = st.selectbox(
    "Choose scope",
    ["Missing both", "Any missing", "Missing delighters only", "Missing detractors only"],
    index=0,
)

if scope == "Missing both":
    target = work[(work["Needs_Delighters"]) & (work["Needs_Detractors"])]
elif scope == "Missing delighters only":
    target = work[(work["Needs_Delighters"]) & (~work["Needs_Detractors"])]
elif scope == "Missing detractors only":
    target = work[(~work["Needs_Delighters"]) & (work["Needs_Detractors"])]
else:
    target = work[(work["Needs_Delighters"]) | (work["Needs_Detractors"])]

st.write(f"🔎 **{len(target):,} reviews** match the selected scope.")
with st.expander("Preview in-scope rows", expanded=False):
    preview_cols = ["Verbatim", "Has_Delighters", "Has_Detractors", "Needs_Delighters", "Needs_Detractors"]
    extras = [c for c in ["Star Rating", "Review Date", "Source"] if c in target.columns]
    st.dataframe(target[preview_cols + extras].head(200), use_container_width=True)

# ------------------- Controls -------------------
processed_rows: List[Dict] = []
processed_idx_set: Set[int] = set()
if "undo_stack" not in st.session_state:
    st.session_state["undo_stack"] = []

# Row 1: actions
r1a, r1b, r1c, r1d, r1e = st.columns([1.4, 1.4, 1.8, 1.8, 1.2])
with r1a: run_n_btn = st.button("▶️ Symptomize N", use_container_width=True)
with r1b: run_all_btn = st.button("🚀 Symptomize All", use_container_width=True)
with r1c: overwrite_btn = st.button("♻️ Overwrite & Re-symptomize", use_container_width=True)
with r1d: run_missing_both_btn = st.button("✨ Missing-Both One-Click", use_container_width=True)
with r1e: undo_btn = st.button("↩️ Undo last run", use_container_width=True)

# Small spacer
st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# Row 2: batch size + presets
st.markdown("<div class='batch-row'></div>", unsafe_allow_html=True)
cA, cB, cC, cD, cE = st.columns([1.0, 0.5, 0.5, 0.5, 0.5])

with cA:
    bound_min = 1
    bound_max = max(1, len(target))
    if "n_to_process" not in st.session_state:
        st.session_state["n_to_process"] = min(10, bound_max)
    cur = int(st.session_state.get("n_to_process", 1))
    if cur < bound_min:
        st.session_state["n_to_process"] = bound_min
    elif cur > bound_max:
        st.session_state["n_to_process"] = bound_max

    n_to_process = st.number_input(
        "How many to symptomize (from top of scope)",
        min_value=bound_min,
        max_value=bound_max,
        step=1,
        key="n_to_process",
    )

def _set_n(v: int):
    st.session_state["n_to_process"] = min(max(int(v), 1), max(1, len(target)))
    st.rerun()

with cB: st.button("10",  use_container_width=True, on_click=_set_n, args=(10,))
with cC: st.button("25",  use_container_width=True, on_click=_set_n, args=(25,))
with cD: st.button("50",  use_container_width=True, on_click=_set_n, args=(50,))
with cE: st.button("100", use_container_width=True, on_click=_set_n, args=(100,))

# --- Core runner ---
def _run_symptomize(rows_df: pd.DataFrame, overwrite_mode: bool = False):
    global df
    prog = st.progress(0.0)

    def _fmt_secs(sec: float) -> str:
        m = int(sec // 60); s = int(round(sec - m*60)); return f"{m}:{s:02d}"

    t0 = time.perf_counter(); eta_box = st.empty()

    snapshot: List[Tuple[int, Dict[str, Optional[str]]]] = []

    if overwrite_mode:
        df = ensure_ai_columns(df)
        idxs = rows_df.index.tolist()
        for idx_clear in idxs:
            old_vals = {f"AI Symptom Detractor {j}": df.loc[idx_clear, f"AI Symptom Detractor {j}"] if f"AI Symptom Detractor {j}" in df.columns else None for j in range(1,11)}
            old_vals.update({f"AI Symptom Delighter {j}": df.loc[idx_clear, f"AI Symptom Delighter {j}"] if f"AI Symptom Delighter {j}" in df.columns else None for j in range(1,11)})
            old_vals.update({"AI Safety": df.loc[idx_clear, "AI Safety"] if "AI Safety" in df.columns else None,
                             "AI Reliability": df.loc[idx_clear, "AI Reliability"] if "AI Reliability" in df.columns else None,
                             "AI # of Sessions": df.loc[idx_clear, "AI # of Sessions"] if "AI # of Sessions" in df.columns else None,
                             "AI Label Overflow": df.loc[idx_clear, "AI Label Overflow"] if "AI Label Overflow" in df.columns else None})
            snapshot.append((int(idx_clear), old_vals))
            for j in range(1, 11):
                df.loc[idx_clear, f"AI Symptom Detractor {j}"] = None
                df.loc[idx_clear, f"AI Symptom Delighter {j}"] = None
            df.loc[idx_clear, "AI Safety"] = None
            df.loc[idx_clear, "AI Reliability"] = None
            df.loc[idx_clear, "AI # of Sessions"] = None
            df.loc[idx_clear, "AI Label Overflow"] = None

    total_n = max(1, len(rows_df))
    for k, (idx, row) in enumerate(rows_df.iterrows(), start=1):
        vb = row.get("Verbatim", "")
        needs_deli = bool(row.get("Needs_Delighters", False))
        needs_detr = bool(row.get("Needs_Detractors", False))

        if not overwrite_mode:
            old_vals = {f"AI Symptom Detractor {j}": df.loc[idx, f"AI Symptom Detractor {j}"] if f"AI Symptom Detractor {j}" in df.columns else None for j in range(1,11)}
            old_vals.update({f"AI Symptom Delighter {j}": df.loc[idx, f"AI Symptom Delighter {j}"] if f"AI Symptom Delighter {j}" in df.columns else None for j in range(1,11)})
            old_vals.update({"AI Safety": df.loc[idx, "AI Safety"] if "AI Safety" in df.columns else None,
                             "AI Reliability": df.loc[idx, "AI Reliability"] if "AI Reliability" in df.columns else None,
                             "AI # of Sessions": df.loc[idx, "AI # of Sessions"] if "AI # of Sessions" in df.columns else None,
                             "AI Label Overflow": df.loc[idx, "AI Label Overflow"] if "AI Label Overflow" in df.columns else None})
            snapshot.append((int(idx), old_vals))

        try:
            dels, dets, unl_dels, unl_dets, ev_del_map, ev_det_map = _openai_labeler(
                vb, client, selected_model, temperature,
                DELIGHTERS, DETRACTORS, ALIASES,
                DEL_MAP, DET_MAP, ALIAS_TO_LABEL,
                max_ev_per_label=max_ev_per_label,
                max_ev_chars=max_ev_chars
            ) if client else ([], [], [], [], {}, {})
        except Exception:
            dels, dets, unl_dels, unl_dets, ev_del_map, ev_det_map = [], [], [], [], {}, {}

        try:
            safety, reliability, sessions = _openai_meta_extractor(vb, client, selected_model, temperature) if client else ("Not Mentioned","Not Mentioned","Unknown")
        except Exception:
            safety, reliability, sessions = "Not Mentioned","Not Mentioned","Unknown"

        df = ensure_ai_columns(df)
        wrote_dets, wrote_dels = [], []
        ev_written_det: Dict[str, List[str]] = {}
        ev_written_del: Dict[str, List[str]] = {}

        def _label_allowed(label: str, side: str) -> bool:
            if not require_evidence:
                return True
            evs = (ev_det_map if side == "det" else ev_del_map).get(label, [])
            return len(evs) > 0

        if needs_detr and dets:
            dets_to_write = [lab for lab in dets if _label_allowed(lab, "det")][:10]
            for j, lab in enumerate(dets_to_write):
                col = f"AI Symptom Detractor {j+1}"
                if col not in df.columns: df[col] = None
                df.loc[idx, col] = lab
                ev_written_det[lab] = ev_det_map.get(lab, [])
            wrote_dets = dets_to_write

        if needs_deli and dels:
            dels_to_write = [lab for lab in dels if _label_allowed(lab, "del")][:10]
            for j, lab in enumerate(dels_to_write):
                col = f"AI Symptom Delighter {j+1}"
                if col not in df.columns: df[col] = None
                df.loc[idx, col] = lab
                ev_written_del[lab] = ev_del_map.get(lab, [])
            wrote_dels = dels_to_write

        # Set meta in df
        df.loc[idx, "AI Safety"] = safety
        df.loc[idx, "AI Reliability"] = reliability
        df.loc[idx, "AI # of Sessions"] = sessions

        # Overflow tag: evaluate against DETECTED sets (before trimming)
        more_than_10_dets = len(dets) > 10
        more_than_10_dels = len(dels) > 10
        if more_than_10_dets and more_than_10_dels:
            overflow_tag = "Both>10"
        elif more_than_10_dets:
            overflow_tag = "Detractors>10"
        elif more_than_10_dels:
            overflow_tag = "Delighters>10"
        else:
            overflow_tag = ""
        df.loc[idx, "AI Label Overflow"] = overflow_tag or None

        # Evidence coverage for this row
        total_labels = len(wrote_dets) + len(wrote_dels)
        labels_with_ev = sum(1 for lab in wrote_dets if len(ev_written_det.get(lab, []))>0) + \
                         sum(1 for lab in wrote_dels if len(ev_written_del.get(lab, []))>0)
        row_ev_cov = (labels_with_ev / total_labels) if total_labels else 0.0

        processed_rows.append({
            "Index": int(idx),
            "Verbatim": str(vb),
            "Added_Detractors": wrote_dets,
            "Added_Delighters": wrote_dels,
            "Evidence_Detractors": ev_written_det,
            "Evidence_Delighters": ev_written_del,
            "Unlisted_Detractors": unl_dets,
            "Unlisted_Delighters": unl_dels,
            ">10 Detractors Detected": more_than_10_dets,
            ">10 Delighters Detected": more_than_10_dels,
            "Overflow Tag": overflow_tag,
            "Safety": safety,
            "Reliability": reliability,
            "Sessions": sessions,
            "Evidence_Coverage": row_ev_cov,
        })
        processed_idx_set.add(int(idx))

        prog.progress(k/total_n)
        elapsed = time.perf_counter() - t0
        rate = (k / elapsed) if elapsed > 0 else 0.0
        rem = total_n - k
        eta_sec = (rem / rate) if rate > 0 else 0.0
        eta_box.markdown(f"**Progress:** {k}/{total_n} • **ETA:** ~ {_fmt_secs(eta_sec)} • **Speed:** {rate*60:.1f} rev/min")

    st.session_state["undo_stack"].append({"rows": snapshot})

# Execute by buttons
if client is not None and (run_n_btn or run_all_btn or overwrite_btn or run_missing_both_btn):
    if run_missing_both_btn:
        rows_iter = work[(work["Needs_Delighters"]) & (work["Needs_Detractors"])]
        _run_symptomize(rows_iter, overwrite_mode=False)
    elif overwrite_btn:
        rows_iter = target if run_all_btn else target.head(int(st.session_state.get("n_to_process", 10)))
        _run_symptomize(rows_iter, overwrite_mode=True)
    else:
        rows_iter = target if run_all_btn else target.head(int(st.session_state.get("n_to_process", 10)))
        _run_symptomize(rows_iter, overwrite_mode=False)
    st.success(f"Symptomized {len(processed_rows)} review(s).")

# Undo last run
def _undo_last_run():
    global df
    if not st.session_state["undo_stack"]:
        st.info("Nothing to undo."); return
    snap = st.session_state["undo_stack"].pop()
    for idx, old_vals in snap.get("rows", []):
        for col, val in old_vals.items():
            if col not in df.columns:
                df[col] = None
            df.loc[idx, col] = val
    st.success("Reverted last run.")

if undo_btn:
    _undo_last_run()

# ------------------- Processed Reviews (chips + highlighted evidence) -------------------
if processed_rows:
    st.subheader("🧾 Processed Reviews (this run)")
    ev_cov_vals = [float(r.get("Evidence_Coverage", 0.0)) for r in processed_rows]
    overall_cov = (sum(ev_cov_vals)/len(ev_cov_vals)) if ev_cov_vals else 0.0
    st.caption(f"**Evidence Coverage (this run):** {overall_cov*100:.1f}% of written labels include ≥1 snippet.")

    def _esc(s: str) -> str:
        return (str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;"))

    for rec in processed_rows:
        head = f"Row {rec['Index']} — Dets: {len(rec['Added_Detractors'])} • Dels: {len(rec['Added_Delighters'])}"
        if rec[">10 Detractors Detected"] or rec[">10 Delighters Detected"]:
            head += " • ⚠︎ >10 detected (trimmed to 10)"
        with st.expander(head):
            # Highlight with evidence snippets
            evidence_terms: List[str] = []
            for _, evs in (rec.get("Evidence_Detractors", {}) or {}).items(): evidence_terms.extend(evs or [])
            for _, evs in (rec.get("Evidence_Delighters", {}) or {}).items(): evidence_terms.extend(evs or [])

            st.markdown("**Verbatim (evidence highlighted)**")
            st.markdown(highlight_text(rec["Verbatim"], evidence_terms), unsafe_allow_html=True)

            # Meta + overflow chips
            overflow_chip = ""
            if rec.get("Overflow Tag"):
                overflow_chip = f"<span class='chip'>Overflow: {_esc(rec['Overflow Tag'])}</span>"
            meta_html = (
                "<div class='chips-block chip-wrap'>"
                f"<span class='chip yellow'>Safety: {_esc(rec.get('Safety','Not Mentioned'))}</span>"
                f"<span class='chip blue'>Reliability: {_esc(rec.get('Reliability','Not Mentioned'))}</span>"
                f"<span class='chip purple'># Sessions: {_esc(rec.get('Sessions','Unknown'))}</span>"
                f"{overflow_chip}"
                "</div>"
            )
            st.markdown(meta_html, unsafe_allow_html=True)

            st.markdown("**Detractors added**")
            det_html = "<div class='chip-wrap'>"
            for lab in rec["Added_Detractors"]:
                k = len((rec.get("Evidence_Detractors", {}) or {}).get(lab, []))
                det_html += f"<span class='chip red'>{html.escape(lab)} · Evidence: {k}</span>"
            det_html += "</div>"
            st.markdown(det_html, unsafe_allow_html=True)

            st.markdown("**Delighters added**")
            del_html = "<div class='chip-wrap'>"
            for lab in rec["Added_Delighters"]:
                k = len((rec.get("Evidence_Delighters", {}) or {}).get(lab, []))
                del_html += f"<span class='chip green'>{html.escape(lab)} · Evidence: {k}</span>"
            del_html += "</div>"
            st.markdown(del_html, unsafe_allow_html=True)

            with st.expander("See evidence snippets", expanded=False):
                if rec.get("Evidence_Detractors"):
                    st.markdown("**Detractor evidence**")
                    for lab, evs in rec["Evidence_Detractors"].items():
                        for e in evs: st.write(f"- {e}")
                if rec.get("Evidence_Delighters"):
                    st.markdown("**Delighter evidence**")
                    for lab, evs in rec["Evidence_Delighters"].items():
                        for e in evs: st.write(f"- {e}")

# ------------------- New Symptom Candidates (Approval form) -------------------
cand_del: Dict[str, List[int]] = {}
cand_det: Dict[str, List[int]] = {}
for rec in processed_rows:
    for u in rec.get("Unlisted_Delighters", []) or []:
        cand_del.setdefault(u, []).append(rec["Index"])
    for u in rec.get("Unlisted_Detractors", []) or []:
        cand_det.setdefault(u, []).append(rec["Index"])

whitelist_all = set(DELIGHTERS + DETRACTORS)
alias_all = set([a for lst in ALIASES.values() for a in lst]) if ALIASES else set()
wl_canon = {_canon_simple(x) for x in whitelist_all}
ali_canon = {_canon_simple(x) for x in alias_all}

def _filter_near_dupes(cmap: Dict[str, List[int]], cutoff: float = 0.94) -> Dict[str, List[int]]:
    filtered: Dict[str, List[int]] = {}
    seen_key: Dict[str, str] = {}
    for sym, refs in cmap.items():
        c = _canon_simple(sym)
        if c in wl_canon or c in ali_canon:
            continue
        try:
            m = difflib.get_close_matches(sym, list(whitelist_all), n=1, cutoff=cutoff)
            if m: continue
        except Exception:
            pass
        if c in seen_key:
            filtered[seen_key[c]].extend(refs)
        else:
            filtered[sym] = list(refs); seen_key[c] = sym
    return filtered

cand_del = _filter_near_dupes(cand_del, cutoff=sim_threshold)
cand_det = _filter_near_dupes(cand_det, cutoff=sim_threshold)

if cand_del or cand_det:
    st.subheader("🟡 New Symptom Inbox — Review & Approve")

    def _mk_table(cmap: Dict[str, List[int]], side_label: str) -> pd.DataFrame:
        if not cmap:
            return pd.DataFrame({
                "Add": pd.Series(dtype=bool),
                "Label": pd.Series(dtype=str),
                "Side": pd.Series(dtype=str),
                "Count": pd.Series(dtype=int),
                "Examples": pd.Series(dtype=str),
            })
        rows_tbl = []
        for sym, refs in sorted(cmap.items(), key=lambda kv: (-len(kv[1]), kv[0])):
            examples = []
            for ridx in refs[:3]:
                try:
                    examples.append(str(df.loc[ridx, "Verbatim"]))
                except Exception:
                    pass
            rows_tbl.append({
                "Add": False,
                "Label": sym,
                "Side": side_label,
                "Count": int(len(refs)),
                "Examples": " | ".join(["— "+e[:200] for e in examples])
            })
        tbl = pd.DataFrame(rows_tbl)
        return tbl.astype({"Add": bool, "Label": str, "Side": str, "Count": int, "Examples": str})

    tbl_del = _mk_table(cand_del, "Delighter")
    tbl_det = _mk_table(cand_det, "Detractor")

    with st.form("new_symptom_candidates_form", clear_on_submit=False):
        cA, cB = st.columns(2)
        with cA:
            st.markdown("**Delighter candidates**")
            editor_del = st.data_editor(
                tbl_del,
                num_rows="fixed",
                use_container_width=True,
                column_config={
                    "Add": st.column_config.CheckboxColumn(help="Check to add to the Symptoms sheet"),
                    "Label": st.column_config.TextColumn(),
                    "Side": st.column_config.SelectboxColumn(options=["Delighter","Detractor"]),
                    "Count": st.column_config.NumberColumn(format="%d"),
                    "Examples": st.column_config.TextColumn(width="large"),
                },
                key="cand_del_editor",
            )
        with cB:
            st.markdown("**Detractor candidates**")
            editor_det = st.data_editor(
                tbl_det,
                num_rows="fixed",
                use_container_width=True,
                column_config={
                    "Add": st.column_config.CheckboxColumn(help="Check to add to the Symptoms sheet"),
                    "Label": st.column_config.TextColumn(),
                    "Side": st.column_config.SelectboxColumn(options=["Detractor","Delighter"]),
                    "Count": st.column_config.NumberColumn(format="%d"),
                    "Examples": st.column_config.TextColumn(width="large"),
                },
                key="cand_det_editor",
            )
        add_btn = st.form_submit_button("✅ Add selected to Symptoms & Download updated workbook")

    if add_btn:
        selections: List[Tuple[str, str]] = []
        try:
            if isinstance(editor_del, pd.DataFrame) and not editor_del.empty:
                for _, r_ in editor_del.iterrows():
                    if bool(r_.get("Add", False)) and str(r_.get("Label", "")).strip():
                        side_val = str(r_.get("Side","Delighter")).strip() or "Delighter"
                        selections.append((str(r_["Label"]).strip(), side_val))
        except Exception:
            pass
        try:
            if isinstance(editor_det, pd.DataFrame) and not editor_det.empty:
                for _, r_ in editor_det.iterrows():
                    if bool(r_.get("Add", False)) and str(r_.get("Label", "")).strip():
                        side_val = str(r_.get("Side","Detractor")).strip() or "Detractor"
                        selections.append((str(r_["Label"]).strip(), side_val))
        except Exception:
            pass

        if selections:
            updated_bytes = add_new_symptoms_to_workbook(uploaded_file, selections)
            st.download_button(
                "⬇️ Download 'Symptoms' (updated)",
                data=updated_bytes,
                file_name="Symptoms_Updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success(f"Queued {len(selections)} new label(s) into the Symptoms tab.")
        else:
            st.info("No candidates selected.")

# ------------------- Download Symptomized Workbook -------------------
st.subheader("📦 Download Symptomized Workbook")
try:
    file_base = os.path.splitext(getattr(uploaded_file, 'name', 'Reviews'))[0]
except Exception:
    file_base = 'Reviews'

def generate_template_workbook_bytes(
    original_file,
    updated_df: pd.DataFrame,
    processed_idx: Optional[Set[int]] = None,
    overwrite_processed_slots: bool = False,
) -> bytes:
    """Return workbook bytes with K–T (dets), U–AD (dels), and AE/AF/AG meta (+ AH Label Overflow, headers preserved)."""
    original_file.seek(0)
    wb = load_workbook(original_file)
    sheet_name = "Star Walk scrubbed verbatims"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws: Worksheet = wb[sheet_name]

    df2 = ensure_ai_columns(updated_df.copy())

    for name, col in META_ORDER:
        col_idx = column_index_from_string(col)
        if not ws.cell(row=1, column=col_idx).value:
            ws.cell(row=1, column=col_idx, value=name)

    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_red   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_yel   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_blu   = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    fill_pur   = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")

    pset = set(processed_idx or [])

    def _clear_template_slots(row_i: int):
        for col_idx in DET_INDEXES + DEL_INDEXES + list(META_INDEXES.values()):
            ws.cell(row=row_i, column=col_idx, value=None)

    for i, (rid, r) in enumerate(df2.iterrows(), start=2):
        if overwrite_processed_slots and (int(rid) in pset):
            _clear_template_slots(i)
        for j, col_idx in enumerate(DET_INDEXES, start=1):
            val = r.get(f"AI Symptom Detractor {j}")
            cv = None if (pd.isna(val) or str(val).strip() == "") else val
            cell = ws.cell(row=i, column=col_idx, value=cv)
            if cv is not None: cell.fill = fill_red
        for j, col_idx in enumerate(DEL_INDEXES, start=1):
            val = r.get(f"AI Symptom Delighter {j}")
            cv = None if (pd.isna(val) or str(val).strip() == "") else val
            cell = ws.cell(row=i, column=col_idx, value=cv)
            if cv is not None: cell.fill = fill_green
        safety = r.get("AI Safety"); reliab = r.get("AI Reliability"); sess = r.get("AI # of Sessions")
        overflow = r.get("AI Label Overflow")
        if is_filled(safety):
            c = ws.cell(row=i, column=META_INDEXES["Safety"], value=str(safety)); c.fill = fill_yel
        if is_filled(reliab):
            c = ws.cell(row=i, column=META_INDEXES["Reliability"], value=str(reliab)); c.fill = fill_blu
        if is_filled(sess):
            c = ws.cell(row=i, column=META_INDEXES["# of Sessions"], value=str(sess)); c.fill = fill_pur
        if is_filled(overflow):
            c = ws.cell(row=i, column=META_INDEXES["Label Overflow"], value=str(overflow)); c.fill = fill_yel

    for c in DET_INDEXES + DEL_INDEXES + list(META_INDEXES.values()):
        try: ws.column_dimensions[get_column_letter(c)].width = 28
        except Exception: pass

    out = io.BytesIO(); wb.save(out)
    return out.getvalue()

export_bytes = generate_template_workbook_bytes(
    uploaded_file,
    df,
    processed_idx=processed_idx_set if processed_idx_set else None,
    overwrite_processed_slots=False,
)

st.download_button(
    "⬇️ Download symptomized workbook (XLSX)",
    data=export_bytes,
    file_name=f"{file_base}_Symptomized.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# ------------------- Symptoms Catalog quick export -------------------
st.subheader("🗂️ Download Symptoms Catalog")
sym_df = pd.DataFrame({
    "Symptom": (DELIGHTERS + DETRACTORS),
    "Type":    ["Delighter"]*len(DELIGHTERS) + ["Detractor"]*len(DETRACTORS)
})
if ALIASES:
    alias_rows = [{"Symptom": k, "Aliases": " | ".join(v)} for k, v in ALIASES.items()]
    alias_df = pd.DataFrame(alias_rows)
    sym_df = sym_df.merge(alias_df, how="left", on="Symptom")

sym_bytes = io.BytesIO()
with pd.ExcelWriter(sym_bytes, engine="openpyxl") as xw:
    sym_df.to_excel(xw, index=False, sheet_name="Symptoms")
sym_bytes.seek(0)
st.download_button("⬇️ Download Symptoms Catalog (XLSX)", sym_bytes.getvalue(),
                   file_name=f"{file_base}_Symptoms_Catalog.xlsx",
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ------------------- View Symptoms from Excel Workbook (expander) -------------------
st.subheader("📘 View Symptoms from Excel Workbook")
with st.expander("📘 View Symptoms from Excel Workbook", expanded=False):
    st.markdown("This reflects the **Symptoms** sheet as loaded; use the inbox below to propose additions.")

    tabs = st.tabs(["Delighters", "Detractors", "Aliases", "Meta"])

    def _esc(s: str) -> str:
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _chips(items, color: str):
        items_sorted = sorted({str(x).strip() for x in (items or []) if str(x).strip()})
        if not items_sorted:
            st.write("(none)")
        else:
            htmlchips = "<div class='chip-wrap'>" + "".join([f"<span class='chip {color}'>{_esc(x)}</span>" for x in items_sorted]) + "</div>"
            st.markdown(htmlchips, unsafe_allow_html=True)

    with tabs[0]:
        st.markdown("**Delighter labels from workbook**"); _chips(DELIGHTERS, "green")
    with tabs[1]:
        st.markdown("**Detractor labels from workbook**"); _chips(DETRACTORS, "red")
    with tabs[2]:
        st.markdown("**Aliases (if present)**")
        if ALIASES:
            alias_rows = [{"Label": k, "Aliases": " | ".join(v)} for k, v in sorted(ALIASES.items())]
            st.dataframe(pd.DataFrame(alias_rows), use_container_width=True, hide_index=True)
        else:
            st.write("(no aliases defined)")
    with tabs[3]:
        st.markdown("**Meta fields usage (from this dataset)**")
        df_meta = ensure_ai_columns(df.copy())

        def _count(col: str, order: List[str]) -> pd.DataFrame:
            if col not in df_meta.columns:
                return pd.DataFrame({"Value": order, "Count": [0] * len(order)})
            vc = df_meta[col].fillna("Not Mentioned").astype(str).value_counts().reindex(order, fill_value=0)
            return vc.rename_axis("Value").reset_index(name="Count")

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**Safety**")
            df_s = _count("AI Safety", SAFETY_ENUM)
            st.bar_chart(df_s.set_index("Value")["Count"])
        with c2:
            st.markdown("**Reliability**")
            df_r = _count("AI Reliability", RELIABILITY_ENUM)
            st.bar_chart(df_r.set_index("Value")["Count"])
        with c3:
            st.markdown("**# of Sessions**")
            df_n = _count("AI # of Sessions", SESSIONS_ENUM)
            st.bar_chart(df_n.set_index("Value")["Count"])

# Footer
st.divider()
st.caption("v7.5 — Evidence-locked labeling + Overflow Tag + Embedding Skim. Exports: K–T/U–AD; meta: AE/AF/AG; overflow: AH. No header relabeling.")
