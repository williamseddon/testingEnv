"""
SharkNinja Review Analyst + Symptomizer — Updated and optimized
Updated with:
- Live sidebar filter system with timeframe, stars, core filters, dynamic extra filters, and active filter summary
- Symptom filters shown only when symptom data exists
- Short hoverable Reference tiles for AI Analyst citations only
- More stable model fallbacks for batch / structured operations
- Symptomizer result cards now show detractors and delighters at the bottom like Review Explorer
"""
from __future__ import annotations

import difflib
import gc
import html
import io
import json
import math
import os
import random
import re
import textwrap
import time
from collections import Counter
from dataclasses import dataclass
from datetime import date, timedelta
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple
from urllib.parse import urlparse

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from plotly.subplots import make_subplots

try:
    from openai import OpenAI
    _HAS_OPENAI = True
except ImportError:
    OpenAI = None
    _HAS_OPENAI = False

try:
    import tiktoken
    _TIKTOKEN_ENC = tiktoken.get_encoding("cl100k_base")
    _HAS_TIKTOKEN = True
except Exception:
    tiktoken = None
    _TIKTOKEN_ENC = None
    _HAS_TIKTOKEN = False

st.set_page_config(page_title="SharkNinja Review Analyst", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');
:root {
  --navy:#0f172a; --navy-mid:#1e293b; --navy-soft:#334155;
  --slate-600:#475569; --slate-500:#64748b; --slate-400:#94a3b8;
  --slate-200:#e2e8f0; --slate-100:#f1f5f9; --slate-50:#f8fafc; --white:#ffffff;
  --accent:#6366f1; --accent-bg:rgba(99,102,241,.08);
  --success:#059669; --danger:#dc2626; --warning:#d97706; --info:#2563eb;
  --page-bg:#eef0f4; --surface:#ffffff; --border:#dde1e8; --border-strong:#c8cdd6;
  --shadow-xs:0 1px 2px rgba(15,23,42,.06);
  --shadow-sm:0 1px 4px rgba(15,23,42,.09),0 1px 2px rgba(15,23,42,.05);
  --shadow-md:0 4px 12px rgba(15,23,42,.11),0 2px 4px rgba(15,23,42,.06);
  --shadow-lg:0 8px 28px rgba(15,23,42,.14),0 4px 8px rgba(15,23,42,.07);
  --radius-sm:10px; --radius-md:14px; --radius-lg:18px; --radius-xl:22px;
}
html,body,.stApp{font-family:'Inter',system-ui,-apple-system,sans-serif;color:var(--navy);background:var(--page-bg)!important;}
.main,.block-container,.stMainBlockContainer{background:var(--page-bg)!important;}
.block-container{padding-top:.9rem!important;padding-bottom:2.5rem!important;max-width:1440px!important;}
.hero-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-xl);padding:18px 22px;box-shadow:var(--shadow-sm);margin-bottom:.9rem;}
.metric-card{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:16px 18px 14px;box-shadow:var(--shadow-xs);min-height:108px;display:flex;flex-direction:column;gap:4px;transition:box-shadow .15s,border-color .15s;}
.metric-card:hover{box-shadow:var(--shadow-md);border-color:rgba(99,102,241,.30);}
.metric-card.accent{border-color:rgba(99,102,241,.35);background:linear-gradient(145deg,#eef2ff 0%,var(--surface) 100%);}
.hero-kicker{font-size:10.5px;text-transform:uppercase;letter-spacing:.11em;color:var(--accent);font-weight:700;margin-bottom:3px;}
.hero-title{font-size:22px;font-weight:800;letter-spacing:-.028em;color:var(--navy);line-height:1.15;}
.metric-label{font-size:10.5px;text-transform:uppercase;letter-spacing:.09em;color:var(--slate-500);font-weight:600;}
.metric-value{font-size:clamp(1.6rem,2.1vw,2.1rem);font-weight:800;color:var(--navy);line-height:1;letter-spacing:-.04em;}
.metric-sub{color:var(--slate-500);font-size:12px;line-height:1.35;margin-top:2px;}
.section-title{font-size:18px;font-weight:800;margin:6px 0 8px;color:var(--navy);letter-spacing:-.025em;}
.section-sub{color:var(--slate-500);font-size:13px;margin:0 0 12px;line-height:1.5;}
.badge-row,.chip-wrap{display:flex;gap:6px;flex-wrap:wrap;align-items:center;}
.chip{display:inline-flex;align-items:center;gap:4px;padding:4px 10px;border-radius:999px;font-size:11.5px;font-weight:600;line-height:1;border:1.5px solid transparent;letter-spacing:-.01em;}
.chip.blue{background:#eff6ff;border-color:#bfdbfe;color:#1d4ed8;}
.chip.green{background:#f0fdf4;border-color:#86efac;color:#15803d;}
.chip.red{background:#fff1f2;border-color:#fca5a5;color:#b91c1c;}
.chip.yellow{background:#fefce8;border-color:#fde047;color:#854d0e;}
.chip.indigo{background:#eef2ff;border-color:#c7d2fe;color:#4338ca;}
.chip.gray{background:var(--slate-50);border-color:var(--border);color:var(--slate-600);}
.hero-grid{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:10px;margin-top:12px;}
.hero-stat{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-md);padding:13px 15px;box-shadow:var(--shadow-xs);}
.hero-stat.accent{border-color:rgba(99,102,241,.40);background:linear-gradient(145deg,#eef2ff,var(--surface));}
.hero-stat .label{color:var(--slate-500);font-size:10.5px;text-transform:uppercase;letter-spacing:.08em;font-weight:600;}
.hero-stat .value{font-size:24px;font-weight:800;margin-top:4px;color:var(--navy);letter-spacing:-.035em;}
.stButton>button{border-radius:var(--radius-sm)!important;font-weight:600!important;font-size:13.5px!important;height:38px!important;border:1.5px solid var(--border-strong)!important;background:var(--surface)!important;color:var(--navy-soft)!important;box-shadow:var(--shadow-xs)!important;transition:all .14s ease!important;letter-spacing:-.01em!important;}
.stButton>button:hover{border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(99,102,241,.13)!important;color:var(--accent)!important;}
[data-testid="baseButton-primary"],[data-testid="baseButton-primary"]:hover{background:var(--navy)!important;color:var(--surface)!important;border-color:var(--navy)!important;}
[data-testid="baseButton-primary"]:hover{background:var(--navy-mid)!important;border-color:var(--navy-mid)!important;box-shadow:0 0 0 3px rgba(15,23,42,.14)!important;}
[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea,[data-testid="stNumberInput"] input{border-radius:var(--radius-sm)!important;border-color:var(--border-strong)!important;background:var(--surface)!important;font-family:'Inter',sans-serif!important;font-size:13.5px!important;}
[data-testid="stTextInput"] input:focus,[data-testid="stTextArea"] textarea:focus{border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(99,102,241,.12)!important;}
[data-testid="stSelectbox"]>div>div,[data-testid="stMultiselect"]>div>div{border-radius:var(--radius-sm)!important;border-color:var(--border-strong)!important;background:var(--surface)!important;}
[data-testid="stContainer"][data-border="true"]{border-radius:var(--radius-lg)!important;border-color:var(--border)!important;background:var(--surface)!important;}
[data-testid="stExpander"]{border-radius:var(--radius-md)!important;border-color:var(--border)!important;background:var(--surface)!important;}
[data-testid="stProgressBar"]>div>div{background:var(--accent)!important;border-radius:999px!important;}
[data-testid="stMetric"]{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-md);padding:14px 16px;box-shadow:var(--shadow-xs);}
[data-testid="stDataFrame"]{border-radius:var(--radius-md);overflow:hidden;border:1px solid var(--border);}
[data-testid="stSidebar"]{background:#f5f7fb!important;border-right:1px solid var(--border)!important;}
[data-testid="stSidebar"] .stButton>button{width:100%;}
.ws-status-bar{display:flex;align-items:center;justify-content:space-between;background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:10px 16px;margin-bottom:.5rem;box-shadow:var(--shadow-xs);font-size:13px;gap:12px;flex-wrap:wrap;}
.ws-status-dot{width:8px;height:8px;border-radius:50%;background:var(--success);display:inline-block;margin-right:6px;box-shadow:0 0 0 3px rgba(5,150,105,.18);}
.ws-filter-pill{background:var(--slate-100);border:1px solid var(--border);border-radius:999px;padding:3px 10px;font-size:11.5px;font-weight:600;color:var(--slate-600);}
.review-body{font-size:13.5px;line-height:1.6;color:var(--navy);margin:6px 0 4px;white-space:pre-wrap;word-break:break-word;}
.ev-highlight{background:#fef08a;border-radius:3px;padding:0 .15em;cursor:help;position:relative;}
.ev-highlight::after{content:attr(data-tag);position:absolute;left:50%;top:calc(100% + 6px);transform:translateX(-50%);width:min(260px,60vw);background:var(--navy);color:#f8fafc;border-radius:var(--radius-md);padding:.5rem .65rem;font-size:.72rem;line-height:1.35;box-shadow:var(--shadow-lg);white-space:normal;z-index:1000;pointer-events:none;opacity:0;transition:opacity .12s ease;}
.ev-highlight:hover::after{opacity:1;}
.sw-table-wrap{overflow-y:auto;overflow-x:hidden;border-radius:var(--radius-md);border:1px solid var(--border);}
.sw-table{width:100%;border-collapse:collapse;font-size:12.5px;font-family:'Inter',sans-serif;}
.sw-table thead tr{background:var(--slate-50);border-bottom:2px solid var(--border);}
.sw-table thead th{padding:8px 12px;text-align:left;font-size:10.5px;text-transform:uppercase;letter-spacing:.07em;color:var(--slate-500);font-weight:700;white-space:nowrap;}
.sw-table tbody tr{border-bottom:1px solid var(--border);}
.sw-table tbody tr:last-child{border-bottom:none;}
.sw-table tbody tr:hover{background:var(--slate-50);}
.sw-table tbody td{padding:7px 12px;color:var(--navy);max-width:260px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.sw-td-right{text-align:right!important;font-variant-numeric:tabular-nums;}
.sw-star-good{color:var(--success);font-weight:700;}
.sw-star-bad{color:var(--danger);font-weight:700;}
.sw-divider{border:none;border-top:1px solid var(--border);margin:1.4rem 0 1rem;}
.compact-pager-status{display:flex;flex-direction:column;align-items:center;justify-content:center;font-size:13px;font-weight:700;color:var(--navy);height:38px;letter-spacing:-.01em;}
.compact-pager-sub{font-size:11px;font-weight:400;color:var(--slate-400);margin-top:1px;}
.sym-state-banner{background:var(--surface);border:1px dashed var(--border-strong);border-radius:var(--radius-xl);padding:2rem;text-align:center;margin:1rem 0;}
.sym-state-banner .icon{font-size:2.4rem;margin-bottom:.6rem;}
.sym-state-banner .title{font-size:15px;font-weight:800;color:var(--navy);margin-bottom:.4rem;}
.sym-state-banner .sub{font-size:13px;color:var(--slate-500);line-height:1.55;max-width:540px;margin:0 auto;}
.cohort-table{width:100%;border-collapse:collapse;font-size:12.5px;}
.cohort-table th{background:var(--slate-50);padding:7px 12px;font-size:10.5px;text-transform:uppercase;letter-spacing:.07em;color:var(--slate-500);font-weight:700;border-bottom:2px solid var(--border);text-align:left;}
.cohort-table td{padding:6px 12px;border-bottom:1px solid var(--border);color:var(--navy);}
.cohort-table tr:last-child td{border-bottom:none;}
.thinking-overlay{position:fixed;inset:0;background:rgba(15,23,42,.38);display:flex;align-items:center;justify-content:center;z-index:99999;}
.thinking-card{width:min(400px,92vw);background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-xl);box-shadow:var(--shadow-lg);padding:1.6rem;text-align:center;}
.thinking-spinner{width:40px;height:40px;border:3px solid var(--slate-100);border-top-color:var(--navy);border-radius:50%;margin:0 auto 1rem;animation:tw-spin .8s linear infinite;}
.thinking-title{color:var(--navy);font-weight:800;font-size:1.05rem;margin-bottom:.25rem;letter-spacing:-.02em;}
.thinking-sub{color:var(--slate-500);font-size:.92rem;line-height:1.4;}
.nav-tabs-wrap{background:var(--surface);border-radius:var(--radius-xl);padding:8px 10px;border:1px solid var(--border);box-shadow:var(--shadow-sm);margin:1.1rem 0 1.4rem;}
.nav-tabs-label{font-size:11px;color:var(--slate-500);font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px;}
.soft-panel{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-lg);padding:12px 14px;box-shadow:var(--shadow-xs);margin:.55rem 0 .95rem;}
.pill-row{display:flex;flex-wrap:wrap;gap:7px;margin-top:8px;align-items:center;}
.pill{display:inline-flex;align-items:center;gap:6px;padding:5px 10px;border-radius:999px;background:var(--slate-50);border:1px solid var(--border);font-size:11.5px;font-weight:600;color:var(--navy);}
.pill .muted{color:var(--slate-500);font-weight:700;}
.small-muted{font-size:12px;color:var(--slate-500);}
.ref-wrap{display:inline-flex;position:relative;vertical-align:middle;margin-left:4px;margin-right:2px;}
.ref-tile{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:999px;background:#eff6ff;border:1px solid #bfdbfe;color:#1d4ed8;font-size:11.5px;font-weight:700;line-height:1;cursor:help;white-space:nowrap;}
.ref-tip{position:absolute;left:0;top:calc(100% + 8px);width:min(380px,72vw);background:var(--surface);border:1px solid var(--border);border-radius:var(--radius-md);padding:10px 11px;box-shadow:var(--shadow-lg);z-index:1100;opacity:0;pointer-events:none;transition:opacity .12s ease;}
.ref-wrap:hover .ref-tip{opacity:1;}
.ref-item{padding:7px 0;border-bottom:1px solid var(--border);}
.ref-item:last-child{border-bottom:none;padding-bottom:0;}
.ref-item:first-child{padding-top:0;}
.ref-meta{font-size:10.5px;text-transform:uppercase;letter-spacing:.07em;color:var(--slate-500);font-weight:700;margin-bottom:3px;}
.ref-title{font-size:12px;font-weight:700;color:var(--navy);margin-bottom:3px;line-height:1.35;}
.ref-snippet{font-size:11.5px;line-height:1.45;color:var(--slate-600);white-space:normal;}
.ref-empty{font-size:11.5px;color:var(--slate-500);line-height:1.4;}
@keyframes tw-spin{to{transform:rotate(360deg);}}
@media(max-width:1100px){.hero-grid{grid-template-columns:repeat(2,minmax(0,1fr));}}
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════
APP_TITLE           = "SharkNinja Review Analyst"
DEFAULT_PASSKEY     = "caC6wVBHos09eVeBkLIniLUTzrNMMH2XMADEhpHe1ewUw"
DEFAULT_DISPLAYCODE = "15973_3_0-en_us"
DEFAULT_API_VERSION = "5.5"
DEFAULT_PAGE_SIZE   = 100
DEFAULT_SORT        = "SubmissionTime:desc"
DEFAULT_CONTENT_LOCALES = (
    "en_US,ar*,zh*,hr*,cs*,da*,nl*,en*,et*,fi*,fr*,de*,el*,he*,hu*,"
    "id*,it*,ja*,ko*,lv*,lt*,ms*,no*,pl*,pt*,ro*,sk*,sl*,es*,sv*,th*,"
    "tr*,vi*,en_AU,en_CA,en_GB"
)
BAZAARVOICE_ENDPOINT = "https://api.bazaarvoice.com/data/reviews.json"

DEFAULT_PRODUCT_URL = "https://www.sharkninja.com/ninja-air-fryer-pro-xl-6-in-1/AF181.html"
SOURCE_MODE_URL = "SharkNinja product URL"
SOURCE_MODE_FILE = "Uploaded review file"

TAB_DASHBOARD = "📊  Dashboard"
TAB_REVIEW_EXPLORER = "🔍  Review Explorer"
TAB_AI_ANALYST = "🤖  AI Analyst"
TAB_REVIEW_PROMPT = "🏷️  Review Prompt"
TAB_SYMPTOMIZER = "💊  Symptomizer"
WORKSPACE_TABS = [
    TAB_DASHBOARD,
    TAB_REVIEW_EXPLORER,
    TAB_AI_ANALYST,
    TAB_REVIEW_PROMPT,
    TAB_SYMPTOMIZER,
]

MODEL_OPTIONS = [
    "gpt-5.4-mini",
    "gpt-5.4",
    "gpt-5.4-pro",
    "gpt-5.4-nano",
    "gpt-5-chat-latest",
    "gpt-5-mini",
    "gpt-5",
    "gpt-5-nano",
    "gpt-4o-mini",
    "gpt-4o",
    "gpt-4.1",
]
DEFAULT_MODEL = "gpt-5.4-mini"
DEFAULT_REASONING = "none"
STRUCTURED_FALLBACK_MODEL = "gpt-5.4-mini"
AI_VISIBLE_CHAT_MESSAGES = 2
AI_CONTEXT_TOKEN_BUDGET = 10_000
NON_VALUES = {"<NA>","NA","N/A","NONE","-","","NAN","NULL"}
STOPWORDS = {
    "a","about","after","again","all","also","am","an","and","any","are","as","at",
    "be","because","been","before","being","best","better","but","by","can","could",
    "did","do","does","don","down","even","every","for","from","get","got","great",
    "had","has","have","he","her","here","hers","him","his","how","i","if","in",
    "into","is","it","its","just","like","love","made","make","many","me","more",
    "most","much","my","new","no","not","now","of","on","one","only","or","other",
    "our","out","over","product","really","so","some","than","that","the","their",
    "them","then","there","these","they","this","to","too","use","used","using",
    "very","was","we","well","were","what","when","which","while","with","would",
    "you","your",
}
PERSONAS: Dict[str, Dict[str, Any]] = {
    "Product Development": {
        "blurb": "Translates reviews into product and feature decisions.",
        "prompt": "Create a report for the product development team. Highlight what customers love, unmet needs, feature gaps, usability friction, and concrete roadmap opportunities. End with the top 5 product actions ranked by impact.",
        "instructions": (
            "You are a senior product strategy analyst specialising in consumer appliances.\n"
            "Structure your response with these exact sections:\n"
            "## What Customers Love\n## Unmet Needs & Feature Gaps\n"
            "## Usability Friction\n## Roadmap Opportunities\n## Top 5 Actions (ranked)\n"
            "Cite review IDs inline as (review_ids: 12345, 67890) for every material claim.\n"
            "Be specific — name exact features, not vague categories.\n"
            "Keep each section to 3-5 bullet points. Total response ≤ 500 words."
        ),
    },
    "Quality Engineer": {
        "blurb": "Focuses on failure modes, defects, durability, and root-cause signals.",
        "prompt": "Create a report for a quality engineer. Identify defect patterns, reliability risks, cleaning issues, performance inconsistencies, and probable root-cause hypotheses. Separate confirmed evidence from inference.",
        "instructions": (
            "You are a senior quality and reliability analyst for consumer appliances.\n"
            "Sections:\n"
            "## Confirmed Defect Patterns\n## Reliability & Durability Risks\n"
            "## Root-Cause Hypotheses\n## Cleaning & Maintenance Issues\n"
            "## Risk Severity Matrix (High/Med/Low)\n"
            "Mark speculative claims as [INFERRED]. Cite review IDs for every confirmed finding.\n"
            "Prioritise by frequency × severity. Total ≤ 500 words."
        ),
    },
    "Consumer Insights": {
        "blurb": "Extracts sentiment drivers, purchase motivations, and voice-of-customer insights.",
        "prompt": "Create a report for the consumer insights team. Summarize key sentiment drivers, barriers to adoption, purchase motivations, key use cases, and how tone changes across star ratings and incentivized vs non-incentivized reviews.",
        "instructions": (
            "You are a consumer insights lead specialising in VoC analysis.\n"
            "Sections:\n"
            "## Top Sentiment Drivers (positive)\n## Top Sentiment Drivers (negative)\n"
            "## Purchase Motivations & Jobs-to-be-Done\n## Barriers to Satisfaction\n"
            "## Organic vs Incentivized Tone Differences\n## Key Verbatim Quotes (3-5)\n"
            "Use plain, executive-ready language. Cite review IDs for quotes. Total ≤ 500 words."
        ),
    },
}
DET_LETTERS  = ["K","L","M","N","O","P","Q","R","S","T"]
DEL_LETTERS  = ["U","V","W","X","Y","Z","AA","AB","AC","AD"]
DET_INDEXES  = [column_index_from_string(c) for c in DET_LETTERS]
DEL_INDEXES  = [column_index_from_string(c) for c in DEL_LETTERS]
META_ORDER   = [("Safety","AE"),("Reliability","AF"),("# of Sessions","AG")]
META_INDEXES = {name: column_index_from_string(col) for name,col in META_ORDER}
AI_DET_HEADERS  = [f"AI Symptom Detractor {i}" for i in range(1,11)]
AI_DEL_HEADERS  = [f"AI Symptom Delighter {i}" for i in range(1,11)]
AI_META_HEADERS = ["AI Safety","AI Reliability","AI # of Sessions"]
SAFETY_ENUM      = ["Not Mentioned","Concern","Positive"]
RELIABILITY_ENUM = ["Not Mentioned","Negative","Neutral","Positive"]
SESSIONS_ENUM    = ["0","1","2–3","4–9","10+","Unknown"]
DEFAULT_PRIORITY_DELIGHTERS = ["Overall Satisfaction","Ease Of Use","Effective Results",
    "Visible Improvement","Time Saver","Comfort","Value","Reliability"]
DEFAULT_PRIORITY_DETRACTORS = ["Poor Results","Ease Of Use","Reliability Issue","High Cost",
    "Irritation","Battery Problem","High Noise","Cleaning Difficulty",
    "Setup Issue","Connectivity Issue","Safety Concern"]
REVIEW_PROMPT_STARTER_ROWS = [
    {"column_name":"perceived_loudness",
     "prompt":"How is product loudness described? Positive, Negative, Neutral, or Not Mentioned.",
     "labels":"Positive, Negative, Neutral, Not Mentioned"},
    {"column_name":"reliability_risk_signal",
     "prompt":"Does the review mention a product reliability or durability risk? Risk Mentioned, Positive Reliability, or Not Mentioned.",
     "labels":"Risk Mentioned, Positive Reliability, Not Mentioned"},
]

# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════
class ReviewDownloaderError(Exception):
    pass


@dataclass
class ReviewBatchSummary:
    product_url: str
    product_id: str
    total_reviews: int
    page_size: int
    requests_needed: int
    reviews_downloaded: int


def _safe_text(v, default=""):
    if v is None:
        return default
    if isinstance(v, (list, tuple, set, dict, pd.Series, pd.DataFrame, pd.Index)):
        return default
    try:
        m = pd.isna(v)
    except Exception:
        m = False
    if isinstance(m, bool) and m:
        return default
    t = str(v).strip()
    return default if t.lower() in {"nan", "none", "null", "<na>"} else t


def _safe_int(v, d=0):
    try:
        return int(float(v))
    except Exception:
        return d


def _safe_bool(v, d=False):
    if v is None:
        return d
    if isinstance(v, bool):
        return v
    t = _safe_text(v).lower()
    if t in {"true", "1", "yes", "y", "t"}:
        return True
    if t in {"false", "0", "no", "n", "f", ""}:
        return False
    return d


def _safe_mean(s):
    if s.empty:
        return None
    n = pd.to_numeric(s, errors="coerce").dropna()
    return float(n.mean()) if not n.empty else None


def _safe_pct(num, den):
    return 0.0 if not den else float(num) / float(den)


def _fmt_secs(sec):
    sec = max(0.0, float(sec or 0))
    m = int(sec // 60)
    s = int(round(sec - m * 60))
    return f"{m}:{s:02d}"


def _canon(s):
    return " ".join(str(s).split()).lower().strip()


def _canon_simple(s):
    return "".join(ch for ch in _canon(s) if ch.isalnum())


def _esc(s):
    return html.escape(str(s or ""))


def _chip_html(items):
    if not items:
        return "<span class='chip gray'>No active filters</span>"
    return "<div class='chip-wrap'>" + "".join(f"<span class='chip {c}'>{_esc(t)}</span>" for t, c in items) + "</div>"


def _is_missing(v):
    if v is None:
        return True
    if isinstance(v, (list, tuple, set, dict, pd.Series, pd.DataFrame, pd.Index)):
        return False
    try:
        m = pd.isna(v)
    except Exception:
        return False
    return bool(m) if isinstance(m, (bool, int)) else False


def _fmt_num(v, d=2):
    if v is None or _is_missing(v):
        return "n/a"
    return f"{v:.{d}f}"


def _fmt_pct(v, d=1):
    if v is None or _is_missing(v):
        return "n/a"
    return f"{100 * float(v):.{d}f}%"


def _trunc(text, max_chars=420):
    text = re.sub(r"\s+", " ", _safe_text(text)).strip()
    return text if len(text) <= max_chars else text[:max_chars - 3].rstrip() + "…"


def _norm_text(text):
    return re.sub(r"\s+", " ", str(text).lower()).strip()


def _tokenize(text):
    return [t for t in re.findall(r"[a-z0-9']+", _norm_text(text)) if len(t) > 2 and t not in STOPWORDS]


def _slugify(text, fallback="custom"):
    c = re.sub(r"[^a-zA-Z0-9]+", "_", _safe_text(text).lower())
    c = re.sub(r"_+", "_", c).strip("_") or fallback
    return ("prompt_" + c if c[0].isdigit() else c)[:64]


def _first_non_empty(series):
    for v in series.astype(str):
        v = _safe_text(v)
        if v and v.lower() != "nan":
            return v
    return ""


def _clean_text(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def _is_filled(val):
    if pd.isna(val):
        return False
    s = str(val).strip()
    return s != "" and s.upper() not in NON_VALUES


def _estimate_tokens(text):
    s = str(text or "")
    if not s:
        return 0
    if _HAS_TIKTOKEN and _TIKTOKEN_ENC is not None:
        try:
            return int(len(_TIKTOKEN_ENC.encode(s)))
        except Exception:
            pass
    return int(max(1, math.ceil(len(s) / 4)))

# ═══════════════════════════════════════════════════════════════════════════════
#  OPENAI
# ═══════════════════════════════════════════════════════════════════════════════
def _get_api_key():
    try:
        if "OPENAI_API_KEY" in st.secrets:
            return str(st.secrets["OPENAI_API_KEY"])
        if "openai" in st.secrets and st.secrets["openai"].get("api_key"):
            return str(st.secrets["openai"]["api_key"])
    except Exception:
        pass
    return os.getenv("OPENAI_API_KEY")


@st.cache_resource(show_spinner=False)
def _make_openai_client(api_key: str):
    if not (_HAS_OPENAI and api_key):
        return None
    try:
        return OpenAI(api_key=api_key, timeout=60, max_retries=3)
    except TypeError:
        try:
            return OpenAI(api_key=api_key)
        except Exception:
            return None


def _get_client():
    key = _get_api_key()
    if not (_HAS_OPENAI and key):
        return None
    return _make_openai_client(key)


def _shared_model():
    return st.session_state.get("shared_model", DEFAULT_MODEL)


def _reasoning_options_for_model(model: str) -> List[str]:
    m = _safe_text(model).lower()
    if not m.startswith("gpt-5"):
        return ["none"]
    if m.startswith("gpt-5.4") or m in {"gpt-5-chat-latest", "gpt-5.2", "gpt-5.2-pro"}:
        return ["none", "low", "medium", "high", "xhigh"]
    if m in {"gpt-5", "gpt-5-mini", "gpt-5-nano"}:
        return ["minimal", "low", "medium", "high"]
    return ["none", "low", "medium", "high"]


def _shared_reasoning():
    current_model = _shared_model()
    allowed = _reasoning_options_for_model(current_model)
    cur = _safe_text(st.session_state.get("shared_reasoning", DEFAULT_REASONING)).lower() or DEFAULT_REASONING
    if cur not in allowed:
        cur = "none" if "none" in allowed else allowed[0]
        st.session_state["shared_reasoning"] = cur
    return cur


def _model_supports_reasoning(model: str) -> bool:
    return _safe_text(model).lower().startswith("gpt-5")


def _normalize_reasoning_effort_for_model(model: str, reasoning_effort: Optional[str]) -> Optional[str]:
    if not _model_supports_reasoning(model):
        return None
    allowed = _reasoning_options_for_model(model)
    effort = _safe_text(reasoning_effort).lower()
    if effort in allowed:
        return effort
    if not effort:
        return allowed[0] if allowed else None
    if effort == "none" and "minimal" in allowed:
        return "minimal"
    if effort == "minimal" and "none" in allowed:
        return "none"
    if effort == "xhigh" and "high" in allowed:
        return "high"
    if effort == "high" and "xhigh" in allowed:
        return "high"
    return allowed[0] if allowed else None


def _model_accepts_temperature(model: str, reasoning_effort: Optional[str]) -> bool:
    m = _safe_text(model).lower()
    eff = _safe_text(reasoning_effort).lower()
    if not m.startswith("gpt-5"):
        return True
    if m.startswith("gpt-5.4") or m in {"gpt-5-chat-latest", "gpt-5.2", "gpt-5.2-pro"}:
        return eff in {"", "none"}
    return False


def _split_chat_messages(messages, keep_last=AI_VISIBLE_CHAT_MESSAGES):
    items = list(messages or [])
    keep = max(1, int(keep_last or 1))
    if len(items) <= keep:
        return [], items
    return items[:-keep], items[-keep:]


def _show_thinking(msg):
    ph = st.empty()
    ph.markdown(f"""<div class="thinking-overlay"><div class="thinking-card">
      <div class="thinking-spinner"></div>
      <div class="thinking-title">Working…</div>
      <div class="thinking-sub">{_esc(msg)}</div>
    </div></div>""", unsafe_allow_html=True)
    return ph


def _safe_json_load(s):
    s = (s or "").strip()
    if not s:
        return {}
    try:
        return json.loads(s)
    except Exception:
        pass
    try:
        i = s.find("{")
        j = s.rfind("}")
        if i >= 0 and j > i:
            return json.loads(s[i:j + 1])
    except Exception:
        pass
    return {}


def _prepare_messages_for_model(model: str, messages):
    prepared = []
    use_developer = _safe_text(model).lower().startswith("gpt-5")
    for msg in list(messages or []):
        if not isinstance(msg, dict):
            continue
        item = dict(msg)
        if use_developer and item.get("role") == "system":
            item["role"] = "developer"
        prepared.append(item)
    return prepared


def _build_completion_token_kwargs(max_tokens):
    try:
        limit = int(max_tokens) if max_tokens is not None else None
    except Exception:
        limit = None
    if limit is None or limit <= 0:
        return {}
    return {"max_completion_tokens": limit}


def _chat_complete(client, *, model, messages, temperature=0.0, response_format=None,
                   max_tokens=1200, reasoning_effort=None, _max_retries=3):
    if client is None:
        return ""

    effort = _normalize_reasoning_effort_for_model(model, reasoning_effort)
    kwargs = dict(model=model, messages=_prepare_messages_for_model(model, messages))
    kwargs.update(_build_completion_token_kwargs(max_tokens))
    if response_format:
        kwargs["response_format"] = response_format
    if effort:
        kwargs["reasoning_effort"] = effort
    if temperature is not None and _model_accepts_temperature(model, effort):
        kwargs["temperature"] = temperature

    last_exc = None
    reasoning_enabled = "reasoning_effort" in kwargs
    temperature_enabled = "temperature" in kwargs

    for attempt in range(max(1, _max_retries)):
        try:
            resp = client.chat.completions.create(**kwargs)
            return (resp.choices[0].message.content or "").strip()
        except Exception as exc:
            last_exc = exc
            err = str(exc).lower()

            if "max_completion_tokens" in kwargs and any(k in err for k in (
                "unexpected keyword argument 'max_completion_tokens'",
                'unsupported parameter: "max_completion_tokens"',
                "unsupported parameter: 'max_completion_tokens'",
                "unknown parameter: max_completion_tokens",
                "max_completion_tokens is not supported",
            )):
                token_limit = kwargs.pop("max_completion_tokens", None)
                if token_limit is not None:
                    kwargs["max_tokens"] = token_limit
                continue

            if "max_tokens" in kwargs and any(k in err for k in (
                "unexpected keyword argument 'max_tokens'",
                'unsupported parameter: "max_tokens"',
                "unsupported parameter: 'max_tokens'",
                "use 'max_completion_tokens' instead",
                "deprecated in favor of `max_completion_tokens`",
                "not compatible with o-series models",
                "not compatible with reasoning models",
            )):
                token_limit = kwargs.pop("max_tokens", None)
                if token_limit is not None:
                    kwargs["max_completion_tokens"] = token_limit
                continue

            if reasoning_enabled and any(k in err for k in (
                "reasoning_effort",
                "unknown parameter: reasoning_effort",
                'unsupported parameter: "reasoning_effort"',
                "unsupported parameter: 'reasoning_effort'",
                "invalid reasoning",
                "invalid value for reasoning",
                "does not support reasoning effort",
                "not support reasoning effort",
            )):
                kwargs.pop("reasoning_effort", None)
                reasoning_enabled = False
                continue

            if temperature_enabled and any(k in err for k in (
                "temperature",
                "top_p",
                "only supported when using",
                "not supported when reasoning effort",
                "include these fields will raise",
            )):
                kwargs.pop("temperature", None)
                temperature_enabled = False
                continue

            if any(k in err for k in ("rate_limit", "429", "500", "503", "timeout", "overloaded")):
                time.sleep(min((2 ** attempt) + random.uniform(0, 1), 30))
                continue
            raise

    if last_exc:
        raise last_exc
    return ""


def _model_candidates_for_task(selected_model: str, *, structured: bool = False) -> List[str]:
    preferred = _safe_text(selected_model) or DEFAULT_MODEL
    fallbacks = [preferred]
    if structured:
        fallbacks += [STRUCTURED_FALLBACK_MODEL, DEFAULT_MODEL, "gpt-4.1"]
    else:
        fallbacks += [DEFAULT_MODEL]
    out = []
    seen = set()
    for m in fallbacks:
        if m and m not in seen:
            out.append(m)
            seen.add(m)
    return out


def _chat_complete_with_fallback_models(client, *, model, messages, structured=False, **kwargs):
    last_exc = None
    for candidate in _model_candidates_for_task(model, structured=structured):
        try:
            return _chat_complete(client, model=candidate, messages=messages, **kwargs)
        except Exception as exc:
            last_exc = exc
            continue
    if last_exc:
        raise last_exc
    return ""

# ═══════════════════════════════════════════════════════════════════════════════
#  DATA LAYER
# ═══════════════════════════════════════════════════════════════════════════════
def _get_session():
    s = requests.Session()
    s.headers.update({"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
    return s


def _extract_pid_from_url(url):
    path = urlparse(url).path
    m = re.search(r"/([A-Za-z0-9_-]+)\.html(?:$|[?#])", path)
    if m:
        c = m.group(1).strip().upper()
        if re.fullmatch(r"[A-Z0-9_-]{3,}", c):
            return c
    return None


def _extract_pid_from_html(h):
    for pat in [
        r'Item\s*No\.?\s*([A-Z0-9_-]{3,})',
        r'"productId"\s*:\s*"([A-Z0-9_-]{3,})"',
        r'"sku"\s*:\s*"([A-Z0-9_-]{3,})"',
        r'"model"\s*:\s*"([A-Z0-9_-]{3,})"',
    ]:
        m = re.search(pat, h, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip().upper()
    soup = BeautifulSoup(h, "html.parser")
    text = soup.get_text(" ", strip=True)
    for pat in [r"Item\s*No\.?\s*([A-Z0-9_-]{3,})", r"Model\s*:?\s*([A-Z0-9_-]{3,})"]:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip().upper()
    return None


def _fetch_reviews_page(session, *, product_id, passkey, displaycode, api_version,
                        page_size, offset, sort, content_locales):
    params = dict(
        resource="reviews",
        action="REVIEWS_N_STATS",
        filter=[
            f"productid:eq:{product_id}",
            f"contentlocale:eq:{content_locales}",
            "isratingsonly:eq:false",
        ],
        filter_reviews=f"contentlocale:eq:{content_locales}",
        include="authors,products,comments",
        filteredstats="reviews",
        Stats="Reviews",
        limit=int(page_size),
        offset=int(offset),
        limit_comments=3,
        sort=sort,
        passkey=passkey,
        apiversion=api_version,
        displaycode=displaycode,
    )
    resp = session.get(BAZAARVOICE_ENDPOINT, params=params, timeout=45)
    resp.raise_for_status()
    payload = resp.json()
    if payload.get("HasErrors"):
        raise ReviewDownloaderError(f"BV error: {payload.get('Errors')}")
    return payload


def _is_incentivized(r):
    badges = [str(b).lower() for b in (r.get("BadgesOrder") or [])]
    if any("incentivized" in b for b in badges):
        return True
    ctx = r.get("ContextDataValues") or {}
    if isinstance(ctx, dict):
        for k, v in ctx.items():
            if "incentivized" in str(k).lower():
                flag = str((v.get("Value", "") if isinstance(v, dict) else v)).strip().lower()
                if flag in {"", "true", "1", "yes"}:
                    return True
    return False


def _flatten_review(r):
    photos = r.get("Photos") or []
    urls = []
    for p in photos:
        sz = p.get("Sizes") or {}
        for sn in ["large", "normal", "thumbnail"]:
            u = (sz.get(sn) or {}).get("Url")
            if u:
                urls.append(u)
                break
    syn = r.get("SyndicationSource") or {}
    return dict(
        review_id=r.get("Id"),
        product_id=r.get("ProductId"),
        original_product_name=r.get("OriginalProductName"),
        title=_safe_text(r.get("Title")),
        review_text=_safe_text(r.get("ReviewText")),
        rating=r.get("Rating"),
        is_recommended=r.get("IsRecommended"),
        user_nickname=r.get("UserNickname"),
        author_id=r.get("AuthorId"),
        user_location=r.get("UserLocation"),
        content_locale=r.get("ContentLocale"),
        submission_time=r.get("SubmissionTime"),
        moderation_status=r.get("ModerationStatus"),
        campaign_id=r.get("CampaignId"),
        source_client=r.get("SourceClient"),
        is_featured=r.get("IsFeatured"),
        is_syndicated=r.get("IsSyndicated"),
        syndication_source_name=syn.get("Name"),
        is_ratings_only=r.get("IsRatingsOnly"),
        total_positive_feedback_count=r.get("TotalPositiveFeedbackCount"),
        badges=", ".join(str(x) for x in (r.get("BadgesOrder") or [])),
        context_data_json=json.dumps(r.get("ContextDataValues") or {}, ensure_ascii=False),
        photos_count=len(photos),
        photo_urls=" | ".join(urls),
        incentivized_review=_is_incentivized(r),
        raw_json=json.dumps(r, ensure_ascii=False),
    )


def _ensure_cols(df, cols):
    for c in cols:
        if c not in df.columns:
            df[c] = pd.NA
    return df


def _extract_age_group(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    payload = val
    if isinstance(payload, str):
        stripped = payload.strip()
        if not stripped:
            return None
        try:
            payload = json.loads(stripped)
        except Exception:
            return None
    if not isinstance(payload, dict):
        return None
    for k, raw in payload.items():
        if "age" not in str(k).lower():
            continue
        candidate = raw.get("Value") or raw.get("Label") if isinstance(raw, dict) else raw
        candidate = _safe_text(candidate)
        if candidate and candidate.lower() not in {"nan", "none", "null", "unknown", "prefer not to say"}:
            return candidate
    return None


def _finalize_df(df):
    required = [
        "review_id", "product_id", "base_sku", "sku_item", "product_or_sku",
        "original_product_name", "title", "review_text", "rating", "is_recommended",
        "content_locale", "submission_time", "submission_date", "submission_month",
        "incentivized_review", "is_syndicated", "photos_count", "photo_urls",
        "title_and_text", "retailer", "post_link", "age_group", "user_nickname",
        "user_location", "total_positive_feedback_count", "source_system", "source_file",
    ]
    df = _ensure_cols(df.copy(), required)
    if df.empty:
        for c in ["has_photos", "has_media", "review_length_chars", "review_length_words", "rating_label", "year_month_sort"]:
            if c not in df.columns:
                df[c] = pd.Series(dtype="object")
        return df

    df["review_id"] = df["review_id"].fillna("").astype(str).str.strip()
    missing = df["review_id"].eq("") | df["review_id"].str.lower().isin({"nan", "none", "null"})
    if missing.any():
        df.loc[missing, "review_id"] = [f"review_{i + 1}" for i in range(int(missing.sum()))]
    if "context_data_json" in df.columns:
        df["age_group"] = df["age_group"].fillna(df["context_data_json"].map(_extract_age_group))
    df["rating"] = pd.to_numeric(df["rating"], errors="coerce")
    df["incentivized_review"] = df["incentivized_review"].fillna(False).astype(bool)
    df["is_syndicated"] = df["is_syndicated"].fillna(False).astype(bool)
    df["photos_count"] = pd.to_numeric(df["photos_count"], errors="coerce").fillna(0).astype(int)
    df["title"] = df["title"].fillna("").astype(str)
    df["review_text"] = df["review_text"].fillna("").astype(str)
    df["submission_time"] = pd.to_datetime(df["submission_time"], errors="coerce", utc=True).dt.tz_convert(None)
    df["submission_date"] = df["submission_time"].dt.date
    df["submission_month"] = df["submission_time"].dt.to_period("M").astype(str)
    df["content_locale"] = df["content_locale"].fillna("").astype(str).replace({"": pd.NA})
    df["base_sku"] = df.get("base_sku", pd.Series(dtype="str")).fillna("").astype(str).str.strip()
    df["sku_item"] = df.get("sku_item", pd.Series(dtype="str")).fillna("").astype(str).str.strip()
    df["product_id"] = df["product_id"].fillna("").astype(str).str.strip()
    fallback = df["base_sku"].where(df["base_sku"].ne(""), df["product_id"])
    df["product_or_sku"] = df["sku_item"].where(df["sku_item"].ne(""), fallback)
    df["product_or_sku"] = df["product_or_sku"].fillna("").astype(str).str.strip().replace({"": pd.NA})
    df["title_and_text"] = (df["title"].str.strip() + " " + df["review_text"].str.strip()).str.strip()
    df["has_photos"] = df["photos_count"] > 0
    df["has_media"] = df["has_photos"]
    df["review_length_chars"] = df["review_text"].str.len()
    df["review_length_words"] = df["review_text"].str.split().str.len().fillna(0).astype(int)
    df["rating_label"] = df["rating"].map(lambda x: f"{int(x)} star" if pd.notna(x) else "Unknown")
    df["year_month_sort"] = pd.to_datetime(df["submission_month"], format="%Y-%m", errors="coerce")
    sc = [c for c in ["submission_time", "review_id"] if c in df.columns]
    if sc:
        df = df.sort_values(sc, ascending=[False, False], na_position="last").reset_index(drop=True)
    return df


def _pick_col(df, aliases):
    lk = {str(c).strip().lower(): c for c in df.columns}
    for a in aliases:
        c = lk.get(str(a).strip().lower())
        if c:
            return c
    return None


def _series_alias(df, aliases):
    c = _pick_col(df, aliases)
    if c is None:
        return pd.Series([pd.NA] * len(df), index=df.index)
    return df[c]


def _parse_flag(v, *, pos, neg):
    t = _safe_text(v).lower()
    if t in {"", "nan", "none", "null", "n/a"}:
        return pd.NA
    if any(t == x.lower() for x in neg):
        return False
    if any(t == x.lower() for x in pos):
        return True
    if t.startswith(("not ", "non ")):
        return False
    return True


def _normalize_uploaded_df(raw, *, source_name=""):
    w = raw.copy()
    w.columns = [str(c).strip() for c in w.columns]
    n = pd.DataFrame(index=w.index)
    n["review_id"] = _series_alias(w, ["Event Id", "Event ID", "Review ID", "Review Id", "Id"])
    n["product_id"] = _series_alias(w, ["Base SKU", "Product ID", "Product Id", "ProductId", "BaseSKU"])
    n["base_sku"] = _series_alias(w, ["Base SKU", "BaseSKU"])
    n["sku_item"] = _series_alias(w, ["SKU Item", "SKU", "Child SKU", "Variant SKU", "Item Number", "Item No"])
    n["original_product_name"] = _series_alias(w, ["Product Name", "Product", "Name"])
    n["review_text"] = _series_alias(w, ["Review Text", "Review", "Body", "Content"])
    n["title"] = _series_alias(w, ["Title", "Review Title", "Headline"])
    n["post_link"] = _series_alias(w, ["Post Link", "URL", "Review URL", "Product URL"])
    n["rating"] = _series_alias(w, ["Rating (num)", "Rating", "Stars", "Star Rating"])
    n["submission_time"] = _series_alias(w, ["Opened date", "Opened Date", "Submission Time", "Review Date", "Date"])
    n["content_locale"] = _series_alias(w, ["Content Locale", "Locale", "Location", "Country"])
    n["retailer"] = _series_alias(w, ["Retailer", "Merchant", "Channel"])
    n["age_group"] = _series_alias(w, ["Age Group", "Age", "Age Range"])
    n["user_location"] = _series_alias(w, ["Location", "Country"])
    n["user_nickname"] = pd.NA
    n["total_positive_feedback_count"] = pd.NA
    n["is_recommended"] = pd.NA
    n["photos_count"] = 0
    n["photo_urls"] = pd.NA
    n["source_file"] = source_name or pd.NA
    n["source_system"] = "Uploaded file"
    seeded = _series_alias(w, ["Seeded Flag", "Seeded", "Incentivized"])
    n["incentivized_review"] = seeded.map(lambda v: _parse_flag(v,
        pos=["seeded", "incentivized", "yes", "true", "1"],
        neg=["not seeded", "not incentivized", "no", "false", "0"]))
    syndicated = _series_alias(w, ["Syndicated Flag", "Syndicated"])
    n["is_syndicated"] = syndicated.map(lambda v: _parse_flag(v,
        pos=["syndicated", "yes", "true", "1"],
        neg=["not syndicated", "no", "false", "0"]))
    return _finalize_df(n)


def _read_uploaded_file(f):
    fname = getattr(f, "name", "uploaded_file")
    raw = f.getvalue()
    suffix = fname.lower().rsplit(".", 1)[-1] if "." in fname else "csv"
    if suffix == "csv":
        try:
            raw_df = pd.read_csv(io.BytesIO(raw))
        except UnicodeDecodeError:
            raw_df = pd.read_csv(io.BytesIO(raw), encoding="latin-1")
    elif suffix in {"xlsx", "xls", "xlsm"}:
        raw_df = pd.read_excel(io.BytesIO(raw))
    else:
        raise ReviewDownloaderError(f"Unsupported: {fname}")
    if raw_df.empty:
        raise ReviewDownloaderError(f"{fname} is empty.")
    return _normalize_uploaded_df(raw_df, source_name=fname)


def _load_uploaded_files(files):
    if not files:
        raise ReviewDownloaderError("Upload at least one file.")
    with st.spinner("Reading files…"):
        frames = [_read_uploaded_file(f) for f in files]
    combined = pd.concat(frames, ignore_index=True)
    combined["review_id"] = combined["review_id"].astype(str)
    combined = combined.drop_duplicates(subset=["review_id"], keep="first").reset_index(drop=True)
    combined = _finalize_df(combined)
    pid = (
        _first_non_empty(combined["base_sku"].fillna("")) or
        _first_non_empty(combined["product_id"].fillna("")) or
        "UPLOADED_REVIEWS"
    )
    names = [getattr(f, "name", "file") for f in files]
    src = names[0] if len(names) == 1 else f"{len(names)} uploaded files"
    summary = ReviewBatchSummary(
        product_url="",
        product_id=pid,
        total_reviews=len(combined),
        page_size=max(len(combined), 1),
        requests_needed=0,
        reviews_downloaded=len(combined),
    )
    return dict(summary=summary, reviews_df=combined, source_type="uploaded", source_label=src)


def _load_product_reviews(product_url):
    product_url = product_url.strip()
    if not re.match(r"^https?://", product_url, flags=re.IGNORECASE):
        product_url = "https://" + product_url
    session = _get_session()
    with st.spinner("Loading product page…"):
        resp = session.get(product_url, timeout=30)
        resp.raise_for_status()
        product_html = resp.text
    pid = _extract_pid_from_url(product_url) or _extract_pid_from_html(product_html)
    if not pid:
        raise ReviewDownloaderError("Could not find product ID.")
    with st.spinner("Checking review volume…"):
        payload = _fetch_reviews_page(
            session,
            product_id=pid,
            passkey=DEFAULT_PASSKEY,
            displaycode=DEFAULT_DISPLAYCODE,
            api_version=DEFAULT_API_VERSION,
            page_size=1,
            offset=0,
            sort=DEFAULT_SORT,
            content_locales=DEFAULT_CONTENT_LOCALES,
        )
        total = int(payload.get("TotalResults", 0))
    progress = st.progress(0.0, text="Downloading…")
    status = st.empty()
    offsets = list(range(0, total, DEFAULT_PAGE_SIZE))
    raw_reviews = []
    for i, offset in enumerate(offsets, 1):
        status.info(f"Pulling page {i}/{len(offsets)}")
        page = _fetch_reviews_page(
            session,
            product_id=pid,
            passkey=DEFAULT_PASSKEY,
            displaycode=DEFAULT_DISPLAYCODE,
            api_version=DEFAULT_API_VERSION,
            page_size=DEFAULT_PAGE_SIZE,
            offset=offset,
            sort=DEFAULT_SORT,
            content_locales=DEFAULT_CONTENT_LOCALES,
        )
        raw_reviews.extend(page.get("Results") or [])
        progress.progress(i / len(offsets))
    status.success(f"Downloaded {len(raw_reviews)} reviews.")
    df = _finalize_df(pd.DataFrame([_flatten_review(r) for r in raw_reviews]))
    if not df.empty:
        df["review_id"] = df["review_id"].astype(str)
        df["product_or_sku"] = df.get("product_or_sku", pd.Series(index=df.index, dtype="object")).fillna(pid)
        df["base_sku"] = df.get("base_sku", pd.Series(index=df.index, dtype="object")).fillna(pid)
        df["product_id"] = df["product_id"].fillna(pid)
    summary = ReviewBatchSummary(
        product_url=product_url,
        product_id=pid,
        total_reviews=total,
        page_size=DEFAULT_PAGE_SIZE,
        requests_needed=len(offsets),
        reviews_downloaded=len(df),
    )
    return dict(summary=summary, reviews_df=df, source_type="bazaarvoice", source_label=product_url)

# ═══════════════════════════════════════════════════════════════════════════════
#  ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════
def _df_cache_key(df):
    cols = [c for c in [
        "review_id", "rating", "incentivized_review", "is_recommended",
        "is_syndicated", "photos_count", "has_photos", "submission_time",
        "title_and_text", "review_length_words", "content_locale", "product_or_sku",
    ] if c in df.columns]
    return df[cols].to_json(orient="split", date_format="iso")


@st.cache_data(show_spinner=False, ttl=300)
def _compute_metrics_cached(df_json):
    df = pd.read_json(io.StringIO(df_json), orient="split")
    return _compute_metrics_direct(df)


def _compute_metrics_direct(df):
    n = len(df)
    if n == 0:
        return dict(
            review_count=0,
            avg_rating=None,
            avg_rating_non_incentivized=None,
            pct_low_star=0.0,
            pct_one_star=0.0,
            pct_two_star=0.0,
            pct_five_star=0.0,
            pct_incentivized=0.0,
            pct_with_photos=0.0,
            pct_syndicated=0.0,
            recommend_rate=None,
            median_review_words=None,
            non_incentivized_count=0,
            low_star_count=0,
        )
    ni = df[~df["incentivized_review"].fillna(False)]
    rb = df[df["is_recommended"].notna()]
    rr = _safe_pct(int(rb["is_recommended"].astype(bool).sum()), len(rb)) if not rb.empty else None
    mw = float(df["review_length_words"].median()) if "review_length_words" in df.columns and not df["review_length_words"].dropna().empty else None
    low = df["rating"].isin([1, 2])
    return dict(
        review_count=n,
        avg_rating=_safe_mean(df["rating"]),
        avg_rating_non_incentivized=_safe_mean(ni["rating"]),
        pct_low_star=_safe_pct(int(low.sum()), n),
        pct_one_star=_safe_pct(int((df["rating"] == 1).sum()), n),
        pct_two_star=_safe_pct(int((df["rating"] == 2).sum()), n),
        pct_five_star=_safe_pct(int((df["rating"] == 5).sum()), n),
        pct_incentivized=_safe_pct(int(df["incentivized_review"].fillna(False).sum()), n),
        pct_with_photos=_safe_pct(int(df["has_photos"].fillna(False).sum()), n),
        pct_syndicated=_safe_pct(int(df["is_syndicated"].fillna(False).sum()), n),
        recommend_rate=rr,
        median_review_words=mw,
        non_incentivized_count=len(ni),
        low_star_count=int(low.sum()),
    )


def _get_metrics(df):
    try:
        return _compute_metrics_cached(_df_cache_key(df))
    except Exception:
        return _compute_metrics_direct(df)


@st.cache_data(show_spinner=False, ttl=300)
def _rating_dist_cached(df_json):
    df = pd.read_json(io.StringIO(df_json), orient="split")
    base = pd.DataFrame({"rating": [1, 2, 3, 4, 5]})
    if df.empty:
        base["review_count"] = 0
        base["share"] = 0.0
        return base
    grouped = (
        df.dropna(subset=["rating"])
        .assign(rating=lambda x: x["rating"].astype(int))
        .groupby("rating", as_index=False)
        .size()
        .rename(columns={"size": "review_count"})
    )
    merged = base.merge(grouped, how="left", on="rating").fillna({"review_count": 0})
    merged["review_count"] = merged["review_count"].astype(int)
    merged["share"] = merged["review_count"] / max(len(df), 1)
    return merged


def _rating_dist(df):
    try:
        return _rating_dist_cached(_df_cache_key(df))
    except Exception:
        return pd.DataFrame({"rating": [1, 2, 3, 4, 5], "review_count": [0] * 5, "share": [0.0] * 5})


@st.cache_data(show_spinner=False, ttl=300)
def _monthly_trend_cached(df_json):
    df = pd.read_json(io.StringIO(df_json), orient="split")
    if df.empty:
        return pd.DataFrame(columns=["submission_month", "review_count", "avg_rating", "month_start"])
    df["submission_time"] = pd.to_datetime(df.get("submission_time"), errors="coerce")
    return (
        df.dropna(subset=["submission_time"])
        .assign(month_start=lambda x: x["submission_time"].dt.to_period("M").dt.to_timestamp())
        .groupby("month_start", as_index=False)
        .agg(review_count=("review_id", "count"), avg_rating=("rating", "mean"))
        .assign(submission_month=lambda x: x["month_start"].dt.strftime("%Y-%m"))
        .sort_values("month_start")
    )


def _monthly_trend(df):
    try:
        return _monthly_trend_cached(_df_cache_key(df))
    except Exception:
        return pd.DataFrame(columns=["submission_month", "review_count", "avg_rating", "month_start"])


def _cohort_by_incentivized(df):
    if df.empty:
        return pd.DataFrame()
    w = df.copy()
    w["cohort"] = w["incentivized_review"].fillna(False).map({True: "Incentivized", False: "Organic"})
    w["rating_int"] = pd.to_numeric(w["rating"], errors="coerce")
    w = w.dropna(subset=["rating_int"])
    w["rating_int"] = w["rating_int"].astype(int)
    out = []
    for cohort, grp in w.groupby("cohort"):
        total = max(len(grp), 1)
        for star in [1, 2, 3, 4, 5]:
            cnt = int((grp["rating_int"] == star).sum())
            out.append(dict(cohort=cohort, star=star, count=cnt, pct=cnt / total * 100))
    return pd.DataFrame(out)


def _locale_breakdown(df, top_n=12):
    if df.empty or "content_locale" not in df.columns:
        return pd.DataFrame()
    grp = (
        df.dropna(subset=["content_locale"])
        .groupby("content_locale", as_index=False)
        .agg(count=("review_id", "count"), avg_rating=("rating", "mean"))
        .sort_values("count", ascending=False)
        .head(top_n)
    )
    grp["pct"] = grp["count"] / max(grp["count"].sum(), 1) * 100
    return grp


def _rolling_velocity(df, window=3):
    md = _monthly_trend(df)
    if md.empty:
        return md
    md = md.copy()
    md["rolling_avg"] = md["review_count"].rolling(window, min_periods=1).mean()
    return md


def _review_length_cohort(df):
    if df.empty or "review_length_words" not in df.columns:
        return pd.DataFrame()
    w = df.dropna(subset=["rating", "review_length_words"]).copy()
    w["review_length_words"] = pd.to_numeric(w["review_length_words"], errors="coerce")
    w = w.dropna(subset=["review_length_words"])
    if len(w) < 8:
        return pd.DataFrame()
    try:
        w["length_bin"] = pd.qcut(
            w["review_length_words"],
            q=4,
            labels=["Short (Q1)", "Medium (Q2)", "Long (Q3)", "Very Long (Q4)"],
            duplicates="drop",
        )
    except Exception:
        return pd.DataFrame()
    return (
        w.groupby("length_bin", as_index=False, observed=True)
        .agg(avg_rating=("rating", "mean"), count=("review_id", "count"), median_words=("review_length_words", "median"))
        .rename(columns={"length_bin": "Length Quartile"})
    )


def _top_locations(df, top_n=10):
    if df.empty or "user_location" not in df.columns:
        return pd.DataFrame()
    return (
        df.dropna(subset=["user_location"])
        .groupby("user_location", as_index=False)
        .agg(count=("review_id", "count"), avg_rating=("rating", "mean"))
        .sort_values("count", ascending=False)
        .head(top_n)
    )


def _star_band_trend(df):
    if df.empty:
        return pd.DataFrame()
    md = _monthly_trend(df)
    if md.empty:
        return pd.DataFrame()
    w = df.dropna(subset=["submission_time", "rating"]).copy()
    w["month_start"] = w["submission_time"].dt.to_period("M").dt.to_timestamp()
    w["low"] = w["rating"].isin([1, 2])
    w["high"] = w["rating"].isin([4, 5])
    grp = w.groupby("month_start", as_index=False).agg(total=("review_id", "count"), low_ct=("low", "sum"), high_ct=("high", "sum"))
    grp["pct_low"] = grp["low_ct"] / grp["total"].clip(lower=1) * 100
    grp["pct_high"] = grp["high_ct"] / grp["total"].clip(lower=1) * 100
    return grp.sort_values("month_start")


def _sw_style_fig(fig):
    GRID = "rgba(148,163,184,0.18)"
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter, system-ui, sans-serif"),
        margin=dict(l=44, r=20, t=44, b=36),
    )
    fig.update_xaxes(gridcolor=GRID, zerolinecolor=GRID)
    fig.update_yaxes(gridcolor=GRID, zerolinecolor=GRID)
    return fig


REGION_NAME_MAP = {
    "US": "USA",
    "USA": "USA",
    "GB": "UK",
    "UK": "UK",
    "CA": "Canada",
    "AU": "Australia",
    "DE": "Germany",
    "FR": "France",
    "ES": "Spain",
    "IT": "Italy",
    "JP": "Japan",
    "MX": "Mexico",
    "BR": "Brazil",
    "NL": "Netherlands",
}


def _locale_to_region_label(locale):
    raw = _safe_text(locale).replace("-", "_").strip()
    if not raw:
        return "Unknown"
    parts = [p for p in raw.split("_") if p]
    country = (parts[-1] if parts else raw).upper()
    country = re.sub(r"[^A-Z]", "", country)
    if not country:
        return "Unknown"
    return REGION_NAME_MAP.get(country, country)


def _parse_smoothing_window(label):
    txt = _safe_text(label).lower()
    if txt.startswith("none"):
        return 1
    m = re.search(r"(\d+)", txt)
    return int(m.group(1)) if m else 1


def _cumulative_avg_region_trend(df, *, organic_only=False, top_n=2, smoothing_label="7-day"):
    if df.empty or "submission_time" not in df.columns or "rating" not in df.columns:
        return pd.DataFrame(), []

    w = df.copy()
    w["submission_time"] = pd.to_datetime(w["submission_time"], errors="coerce")
    w["rating"] = pd.to_numeric(w["rating"], errors="coerce")
    w = w.dropna(subset=["submission_time", "rating"]).copy()

    if organic_only and "incentivized_review" in w.columns:
        w = w[~w["incentivized_review"].fillna(False)].copy()

    if w.empty:
        return pd.DataFrame(), []

    w["day"] = w["submission_time"].dt.floor("D")
    w["region"] = w.get("content_locale", pd.Series(index=w.index, dtype="object")).map(_locale_to_region_label).fillna("Unknown")

    full_days = pd.date_range(w["day"].min(), w["day"].max(), freq="D")
    base = pd.DataFrame({"day": full_days})

    overall = w.groupby("day", as_index=False).agg(daily_volume=("review_id", "count"), rating_sum=("rating", "sum"))
    trend = base.merge(overall, on="day", how="left").fillna({"daily_volume": 0, "rating_sum": 0})
    trend["daily_volume"] = trend["daily_volume"].astype(int)
    overall_denom = trend["daily_volume"].cumsum()
    trend["overall_cum_avg"] = np.where(overall_denom > 0, trend["rating_sum"].cumsum() / overall_denom, np.nan)

    region_counts = (
        w[w["region"] != "Unknown"]
        .groupby("region")["review_id"]
        .count()
        .sort_values(ascending=False)
    )
    regions = region_counts.head(top_n).index.tolist()
    if not regions and "Unknown" in set(w["region"]):
        regions = ["Unknown"]

    for region in regions:
        reg = w[w["region"] == region].groupby("day", as_index=False).agg(region_volume=("review_id", "count"), rating_sum=("rating", "sum"))
        reg = base.merge(reg, on="day", how="left").fillna({"region_volume": 0, "rating_sum": 0})
        reg_denom = reg["region_volume"].cumsum()
        trend[f"{region}_cum_avg"] = np.where(reg_denom > 0, reg["rating_sum"].cumsum() / reg_denom, np.nan)

    smoothing_window = _parse_smoothing_window(smoothing_label)
    if smoothing_window > 1:
        for col in [c for c in trend.columns if c.endswith("_cum_avg")]:
            trend[col] = trend[col].rolling(smoothing_window, min_periods=1).mean()

    return trend.sort_values("day").reset_index(drop=True), regions


def _build_volume_bar_series(trend, volume_mode):
    if trend is None or trend.empty:
        return pd.DataFrame(columns=["x", "volume", "width_ms", "label"]), "Reviews/day"

    w = trend[["day", "daily_volume"]].copy()
    w["day"] = pd.to_datetime(w["day"], errors="coerce")
    w["daily_volume"] = pd.to_numeric(w["daily_volume"], errors="coerce").fillna(0)
    w = w.dropna(subset=["day"])
    if w.empty:
        return pd.DataFrame(columns=["x", "volume", "width_ms", "label"]), "Reviews/day"

    mode = _safe_text(volume_mode) or "Reviews/day"
    if mode == "Reviews/week":
        w["bucket_start"] = w["day"].dt.to_period("W-SUN").dt.start_time
        grouped = w.groupby("bucket_start", as_index=False).agg(volume=("daily_volume", "sum"))
        grouped["bucket_days"] = 7.0
        grouped["label"] = grouped["bucket_start"].dt.strftime("Week of %Y-%m-%d")
        axis_title = "Reviews/week"
    elif mode == "Reviews/month":
        w["bucket_start"] = w["day"].dt.to_period("M").dt.start_time
        grouped = w.groupby("bucket_start", as_index=False).agg(volume=("daily_volume", "sum"))
        grouped["bucket_days"] = grouped["bucket_start"].dt.days_in_month.astype(float)
        grouped["label"] = grouped["bucket_start"].dt.strftime("%Y-%m")
        axis_title = "Reviews/month"
    else:
        grouped = w.rename(columns={"day": "bucket_start", "daily_volume": "volume"}).copy()
        grouped["bucket_days"] = 1.0
        grouped["label"] = grouped["bucket_start"].dt.strftime("%Y-%m-%d")
        axis_title = "Reviews/day"

    grouped["x"] = grouped["bucket_start"] + pd.to_timedelta(grouped["bucket_days"] / 2.0, unit="D")
    grouped["width_ms"] = grouped["bucket_days"].map(lambda d: int(pd.Timedelta(days=max(float(d) - 0.15, 0.35)).total_seconds() * 1000))
    grouped["volume"] = pd.to_numeric(grouped["volume"], errors="coerce").fillna(0).astype(int)
    return grouped[["x", "volume", "width_ms", "label"]], axis_title


def _add_axis_break_indicator(fig, *, side="right"):
    if side == "right":
        x0, x1 = 0.988, 0.998
    else:
        x0, x1 = 0.002, 0.012
    y0 = 0.08
    fig.add_shape(type="line", xref="paper", yref="paper", x0=x0, y0=y0, x1=x1, y1=y0 + 0.02, line=dict(color="rgba(71,85,105,0.92)", width=2), layer="above")
    fig.add_shape(type="line", xref="paper", yref="paper", x0=x0, y0=y0 + 0.028, x1=x1, y1=y0 + 0.048, line=dict(color="rgba(71,85,105,0.92)", width=2), layer="above")


def _render_reviews_over_time_chart(df):
    with st.container(border=True):
        st.markdown("<div class='section-title'>📈 Cumulative Avg ★ Over Time by Region (Weighted)</div>", unsafe_allow_html=True)
        st.markdown("<div class='section-sub'>Volume bars stay on the left axis and cumulative average stays on the right axis.</div>", unsafe_allow_html=True)

        c0, c1, c2, c3, c4, c5 = st.columns([1.05, 1, 1, 1, 1.15, 1.3])
        smoothing = c0.selectbox("Smoothing", ["7-day", "14-day", "30-day", "None"], index=0, key="ot_smoothing")
        show_overall = c1.toggle("Show overall", value=True, key="ot_show_overall")
        show_volume = c2.toggle("Show volume bars", value=True, key="ot_show_volume")
        organic_only = c3.toggle("Organic only", value=False, key="ot_organic_only")
        volume_mode = c4.selectbox("Volume bars", ["Reviews/day", "Reviews/week", "Reviews/month"], index=0, key="ot_volume_mode")
        y_view = c5.radio("Y-axis view", ["Zoomed-in", "Full scale"], horizontal=True, key="ot_y_view")

        trend, regions = _cumulative_avg_region_trend(df, organic_only=organic_only, top_n=2, smoothing_label=smoothing)
        if trend.empty:
            st.info("No dated reviews available for the over-time chart.")
            return

        volume_bars, volume_axis_title = _build_volume_bar_series(trend, volume_mode)
        fig = make_subplots(specs=[[{"secondary_y": True}]])

        if show_volume and not volume_bars.empty:
            fig.add_trace(
                go.Bar(
                    x=volume_bars["x"],
                    y=volume_bars["volume"],
                    width=volume_bars["width_ms"],
                    name=volume_axis_title,
                    marker_color="rgba(226,232,240,0.95)",
                    marker_line_color="rgba(226,232,240,0.95)",
                    opacity=0.72,
                    customdata=np.stack([volume_bars["label"]], axis=-1),
                    hovertemplate="%{customdata[0]}<br>Reviews: %{y:,}<extra></extra>",
                ),
                secondary_y=False,
            )

        region_colors = ["#f97316", "#10b981", "#3b82f6", "#ef4444"]
        for idx, region in enumerate(regions):
            col = f"{region}_cum_avg"
            if col not in trend.columns:
                continue
            fig.add_trace(
                go.Scatter(
                    x=trend["day"],
                    y=trend[col],
                    name=region,
                    mode="lines",
                    line=dict(color=region_colors[idx % len(region_colors)], width=2),
                    hovertemplate=f"{region}<br>%{{x|%Y-%m-%d}}<br>Cumulative Avg ★: %{{y:.3f}}<extra></extra>",
                ),
                secondary_y=True,
            )

        if show_overall and "overall_cum_avg" in trend.columns:
            fig.add_trace(
                go.Scatter(
                    x=trend["day"],
                    y=trend["overall_cum_avg"],
                    name="Overall",
                    mode="lines",
                    line=dict(color="#8b5cf6", width=3),
                    hovertemplate="Overall<br>%{x|%Y-%m-%d}<br>Cumulative Avg ★: %{y:.3f}<extra></extra>",
                ),
                secondary_y=True,
            )

        fig.update_layout(
            margin=dict(l=24, r=24, t=16, b=20),
            hovermode="x unified",
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            font_family="Inter",
            legend=dict(orientation="h", y=1.07, x=0),
            bargap=0.06,
        )
        fig.update_xaxes(title_text="", showgrid=False)

        visible_cols = [f"{region}_cum_avg" for region in regions if f"{region}_cum_avg" in trend.columns]
        if show_overall and "overall_cum_avg" in trend.columns:
            visible_cols.append("overall_cum_avg")
        vals = pd.concat([trend[c].dropna() for c in visible_cols], ignore_index=True) if visible_cols else pd.Series(dtype=float)

        if y_view == "Full scale" or vals.empty:
            y_range = [1.0, 5.0]
        else:
            ymin = max(1.0, float(vals.min()) - 0.06)
            ymax = min(5.0, float(vals.max()) + 0.06)
            if ymax - ymin < 0.18:
                mid = (ymin + ymax) / 2
                ymin = max(1.0, mid - 0.09)
                ymax = min(5.0, mid + 0.09)
            y_range = [ymin, ymax]

        fig.update_yaxes(title_text=volume_axis_title if show_volume else "", secondary_y=False, showgrid=False, rangemode="tozero", visible=show_volume)
        fig.update_yaxes(title_text="Cumulative Avg ★", range=y_range, secondary_y=True, showgrid=True, gridcolor="rgba(148,163,184,0.15)")

        if y_view == "Zoomed-in" and y_range[0] > 1.05:
            _add_axis_break_indicator(fig, side="right")

        st.plotly_chart(fig, use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  SYMPTOM ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════
def _get_symptom_col_lists(df):
    ai_det = [c for c in df.columns if c.startswith("AI Symptom Detractor")]
    ai_del = [c for c in df.columns if c.startswith("AI Symptom Delighter")]
    man_det = [f"Symptom {i}" for i in range(1, 11) if f"Symptom {i}" in df.columns]
    man_del = [f"Symptom {i}" for i in range(11, 21) if f"Symptom {i}" in df.columns]
    return (ai_det or man_det), (ai_del or man_del)


def _detect_symptom_state(df):
    det_cols, del_cols = _get_symptom_col_lists(df)
    def _has(cols):
        for c in cols:
            if c not in df.columns:
                continue
            s = df[c].astype(str).str.strip()
            if s.replace({"nan": "", "None": "", "<NA>": ""}).ne("").any():
                return True
        return False
    h_det = _has(det_cols)
    h_del = _has(del_cols)
    if h_det and h_del:
        return "full"
    if h_det or h_del:
        return "partial"
    return "none"


def analyze_symptoms_fast(df_in, symptom_cols):
    _INVALID = {"<NA>", "NA", "N/A", "NULL", "NONE", "NAN", ""}
    cols = [c for c in symptom_cols if c in df_in.columns]
    if not cols:
        return pd.DataFrame(columns=["Item", "Avg Star", "Mentions", "% Total"])
    block = df_in[cols]
    try:
        long = block.stack(dropna=False).reset_index()
    except TypeError:
        long = block.stack().reset_index()
    long.columns = ["__idx", "__col", "symptom"]
    s = long["symptom"].astype("string").str.strip()
    mask = s.map(lambda v: str(v).strip().upper() not in _INVALID and not str(v).startswith("<"), na_action="ignore").fillna(False)
    long = long.loc[mask, ["__idx"]].copy()
    long["symptom"] = s[mask].str.title()
    if long.empty:
        return pd.DataFrame(columns=["Item", "Avg Star", "Mentions", "% Total"])
    counts = long["symptom"].value_counts()
    avg_map = {}
    if "rating" in df_in.columns:
        stars = pd.to_numeric(df_in["rating"], errors="coerce").rename("star")
        tmp = long.drop_duplicates(subset=["__idx", "symptom"]).copy()
        tmp = tmp.join(stars, on="__idx")
        avg_map = tmp.groupby("symptom")["star"].mean().to_dict()
    total = max(1, len(df_in))
    items = counts.index.tolist()
    return pd.DataFrame({
        "Item": [str(x).title() for x in items],
        "Avg Star": [round(float(avg_map[x]), 1) if x in avg_map and not pd.isna(avg_map[x]) else None for x in items],
        "Mentions": counts.values.astype(int),
        "% Total": (counts.values / total * 100).round(1).astype(str) + "%",
    }).sort_values("Mentions", ascending=False, ignore_index=True)


def symptom_table_html(df_in, *, max_height_px=400):
    if df_in is None or df_in.empty:
        return f"<div class='sw-table-wrap' style='max-height:{max_height_px}px;padding:12px;'>No data.</div>"
    cols = [c for c in ["Item", "Mentions", "% Total", "Avg Star", "Net Hit"] if c in df_in.columns]
    th = "".join(f"<th>{html.escape(c)}</th>" for c in cols)
    rows_html = []
    for _, row in df_in[cols].iterrows():
        tds = []
        for c in cols:
            v = row.get(c, "")
            right = "sw-td-right" if c in ("Mentions", "% Total", "Avg Star", "Net Hit") else ""
            if c == "Avg Star":
                try:
                    f = float(v)
                    cls = "sw-star-good" if f >= 4.5 else "sw-star-bad"
                    tds.append(f"<td class='{right} {cls}'>{f:.1f}</td>")
                except Exception:
                    tds.append(f"<td class='{right}'>{html.escape(str(v))}</td>")
            elif c == "Net Hit":
                try:
                    tds.append(f"<td class='{right}'>{float(v):.3f}</td>")
                except Exception:
                    tds.append(f"<td class='{right}'>{html.escape(str(v))}</td>")
            else:
                tds.append(f"<td class='{right}'>{html.escape(str(v))}</td>")
        rows_html.append("<tr>" + "".join(tds) + "</tr>")
    body = "".join(rows_html)
    return (
        f"<div class='sw-table-wrap' style='max-height:{max_height_px}px;'>"
        f"<table class='sw-table'><thead><tr>{th}</tr></thead><tbody>{body}</tbody></table></div>"
    )


def _add_net_hit(tbl, avg_rating):
    if tbl is None or tbl.empty:
        return tbl
    d = tbl.copy()
    d["Mentions"] = pd.to_numeric(d.get("Mentions"), errors="coerce").fillna(0).astype(int)
    gap = max(0.0, 5.0 - float(avg_rating or 0))
    total = float(d["Mentions"].sum()) or 0
    d["Net Hit"] = (gap * (d["Mentions"] / total)).round(3) if total > 0 else 0.0
    return d[[c for c in ["Item", "Mentions", "% Total", "Avg Star", "Net Hit"] if c in d.columns]]


def _opp_scatter(tbl, kind, baseline_avg, *, container_key=""):
    if tbl is None or tbl.empty:
        st.info("No data available.")
        return
    d = tbl.copy()
    d["Mentions"] = pd.to_numeric(d.get("Mentions"), errors="coerce").fillna(0)
    d["Avg Star"] = pd.to_numeric(d.get("Avg Star"), errors="coerce")
    d = d.dropna(subset=["Avg Star"])
    if d.empty:
        st.info("No data available.")
        return
    x = d["Mentions"].astype(float).to_numpy()
    y = d["Avg Star"].astype(float).to_numpy()
    names = d["Item"].astype(str).to_numpy()
    score = (x * np.clip(float(baseline_avg) - y, 0, None) if kind == "detractors" else x * np.clip(y - float(baseline_avg), 0, None))
    show_labels = st.toggle("Show labels", value=False, key=f"opp_lbl_{kind}_{container_key}")
    labels_arr = np.array([""] * len(d), dtype=object)
    if show_labels:
        top_idx = np.argsort(-score)[:10]
        labels_arr[top_idx] = names[top_idx]
    mx = max(float(np.nanmax(x)), 1e-9)
    size = (np.sqrt(x) / np.sqrt(mx)) * 24 + 8
    color = "#ef4444" if kind == "detractors" else "#22c55e"
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=x,
        y=y,
        mode="markers+text" if show_labels else "markers",
        text=labels_arr,
        textposition="top center",
        textfont=dict(size=10, family="Inter"),
        customdata=np.stack([names], axis=1),
        hovertemplate="%{customdata[0]}<br>Mentions=%{x:.0f}<br>Avg ★=%{y:.2f}<extra></extra>",
        marker=dict(size=size, color=color, opacity=0.76, line=dict(width=1, color="rgba(148,163,184,0.38)")),
    ))
    fig.add_hline(y=float(baseline_avg), line_dash="dash", opacity=0.45, annotation_text=f"Avg ★ {baseline_avg:.2f}", annotation_position="right", annotation_font_size=11)
    fig.update_layout(height=420, xaxis_title="Mentions", yaxis_title="Avg ★")
    _sw_style_fig(fig)
    st.plotly_chart(fig, use_container_width=True)
    label = "Fix first — high mentions × below-baseline ★" if kind == "detractors" else "Amplify — high mentions × above-baseline ★"
    top15 = d.copy()
    top15["Score"] = score
    top15 = top15.sort_values("Score", ascending=False).head(15)
    with st.expander(f"📋 {label}", expanded=False):
        ds = top15[["Item", "Mentions", "Avg Star", "Score"]].copy()
        ds["Avg Star"] = ds["Avg Star"].map(lambda v: f"{float(v):.1f}" if pd.notna(v) else "—")
        ds["Score"] = ds["Score"].map(lambda v: f"{float(v):.1f}")
        st.dataframe(ds, use_container_width=True, hide_index=True)


def _render_symptom_bar_chart(tbl, title, color, denom, show_pct):
    if tbl is None or tbl.empty:
        st.info(f"No {title.lower()} data.")
        return
    t = tbl.copy()
    t["Mentions"] = pd.to_numeric(t["Mentions"], errors="coerce").fillna(0)
    t["Pct"] = t["Mentions"] / max(denom, 1) * 100
    x_vals = t["Pct"][::-1] if show_pct else t["Mentions"][::-1]
    x_label = "% of reviews" if show_pct else "Mentions"
    hover = "%{customdata}<br>%: %{x:.1f}%<extra></extra>" if show_pct else "%{customdata}<br>Mentions: %{x:.0f}<extra></extra>"
    fig = go.Figure(go.Bar(x=x_vals, y=t["Item"][::-1], orientation="h", marker_color=color, opacity=0.80, customdata=t["Item"][::-1].astype(str).tolist(), hovertemplate=hover))
    fig.update_layout(title=title, height=max(300, 28 * len(t) + 80), xaxis_title=x_label, yaxis_title="", margin=dict(l=160, r=20, t=46, b=30))
    _sw_style_fig(fig)
    st.plotly_chart(fig, use_container_width=True)


def _render_symptom_dashboard(filtered_df, overall_df=None):
    od = overall_df if overall_df is not None else filtered_df
    sym_state = _detect_symptom_state(od)
    st.markdown("<hr class='sw-divider'>", unsafe_allow_html=True)
    if sym_state == "none":
        st.markdown("""<div class="sym-state-banner">
          <div class="icon">💊</div><div class="title">No symptoms tagged yet</div>
          <div class="sub">Run the <strong>Symptomizer</strong> tab to AI-tag delighters and detractors,
          then return here for the full analytics.<br>
          If your file already contains <em>Symptom 1–20</em> columns they'll appear automatically.</div>
        </div>""", unsafe_allow_html=True)
        return
    if sym_state == "partial":
        det_cols, del_cols = _get_symptom_col_lists(od)
        missing = []
        if not any(od[c].astype(str).str.strip().replace({"nan": "", "<NA>": ""}).ne("").any() for c in det_cols if c in od.columns):
            missing.append("detractors")
        if not any(od[c].astype(str).str.strip().replace({"nan": "", "<NA>": ""}).ne("").any() for c in del_cols if c in od.columns):
            missing.append("delighters")
        if missing:
            st.info(f"Partial tagging — {' and '.join(missing)} not yet labelled.")
    st.markdown("<div class='section-title'>🩺 Detractors & Delighters</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>AI-tagged symptom analysis. Net Hit = each theme's share of the gap-to-5★.</div>", unsafe_allow_html=True)
    det_cols, del_cols = _get_symptom_col_lists(od)
    avg_star = float(_safe_mean(filtered_df["rating"]) or 0)
    det_tbl = _add_net_hit(analyze_symptoms_fast(filtered_df, det_cols), avg_star)
    del_tbl = _add_net_hit(analyze_symptoms_fast(filtered_df, del_cols), avg_star)
    ctrl = st.columns([1.2, 2.8])
    table_limit = int(ctrl[0].selectbox("Rows", [25, 50, 100], index=1, key="sw_tbl_limit"))
    tbl_view = ctrl[1].radio("Table view", ["Split", "Tabs"], horizontal=True, key="sw_tbl_view")
    if tbl_view == "Split":
        sc1, sc2 = st.columns(2)
        with sc1:
            st.markdown("**🔴 Detractors**")
            st.markdown(symptom_table_html(det_tbl.head(table_limit), max_height_px=380), unsafe_allow_html=True)
        with sc2:
            st.markdown("**🟢 Delighters**")
            st.markdown(symptom_table_html(del_tbl.head(table_limit), max_height_px=380), unsafe_allow_html=True)
    else:
        t1, t2 = st.tabs(["🔴 Detractors", "🟢 Delighters"])
        with t1:
            st.markdown(symptom_table_html(det_tbl.head(table_limit), max_height_px=420), unsafe_allow_html=True)
        with t2:
            st.markdown(symptom_table_html(del_tbl.head(table_limit), max_height_px=420), unsafe_allow_html=True)
    try:
        out_xlsx = io.BytesIO()
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            det_tbl.to_excel(writer, sheet_name="Detractors", index=False)
            del_tbl.to_excel(writer, sheet_name="Delighters", index=False)
        ds = st.session_state.get("analysis_dataset") or {}
        pid = (ds.get("summary") and ds["summary"].product_id) or "symptoms"
        st.download_button(
            "⬇️ Download Detractors + Delighters",
            data=out_xlsx.getvalue(),
            file_name=f"{pid}_symptoms.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sw_sym_dl",
        )
    except Exception:
        pass
    st.markdown("<hr class='sw-divider'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>📊 Top Themes</div>", unsafe_allow_html=True)
    bar_ctrl = st.columns([1, 1, 1.2])
    top_n = int(bar_ctrl[0].slider("Top N", 5, 25, 12, 1, key="sw_top_n"))
    org_only = bar_ctrl[1].toggle("Organic only", value=False, key="sw_org_bar")
    show_pct = bar_ctrl[2].toggle("Show %", value=False, key="sw_pct_bar")
    bar_df = filtered_df[~filtered_df["incentivized_review"].fillna(False)] if org_only else filtered_df
    denom = max(1, len(bar_df))
    det_top = analyze_symptoms_fast(bar_df, det_cols).head(top_n)
    del_top = analyze_symptoms_fast(bar_df, del_cols).head(top_n)
    bc1, bc2 = st.columns(2)
    with bc1:
        with st.container(border=True):
            _render_symptom_bar_chart(det_top, "Top Detractors", "#ef4444", denom, show_pct)
    with bc2:
        with st.container(border=True):
            _render_symptom_bar_chart(del_top, "Top Delighters", "#22c55e", denom, show_pct)
    st.markdown("<hr class='sw-divider'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>🎯 Opportunity Matrix</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>Mentions vs Avg ★ · Fix high-mention low-star detractors first · Amplify high-mention high-star delighters.</div>", unsafe_allow_html=True)
    opp_t1, opp_t2 = st.tabs(["🔴 Detractors", "🟢 Delighters"])
    with opp_t1:
        _opp_scatter(det_tbl, "detractors", avg_star, container_key="dash")
    with opp_t2:
        _opp_scatter(del_tbl, "delighters", avg_star, container_key="dash")

# ═══════════════════════════════════════════════════════════════════════════════
#  FILTERS
# ═══════════════════════════════════════════════════════════════════════════════
CORE_REVIEW_FILTER_SPECS = [
    {"key": "product_or_sku", "label": "SKU / Product", "kind": "column", "column": "product_or_sku"},
    {"key": "content_locale", "label": "Market / Locale", "kind": "column", "column": "content_locale"},
    {"key": "retailer", "label": "Retailer", "kind": "column", "column": "retailer"},
    {"key": "source_system", "label": "Source System", "kind": "column", "column": "source_system"},
    {"key": "source_file", "label": "Source File", "kind": "column", "column": "source_file"},
    {"key": "age_group", "label": "Age Group", "kind": "column", "column": "age_group"},
    {"key": "user_location", "label": "Reviewer Location", "kind": "column", "column": "user_location"},
    {"key": "review_type", "label": "Review Type", "kind": "derived"},
    {"key": "recommendation", "label": "Recommendation", "kind": "derived"},
    {"key": "syndication", "label": "Syndication", "kind": "derived"},
    {"key": "media", "label": "Media", "kind": "derived"},
]


def _series_matches_any(series: pd.Series, values: Sequence[str]) -> pd.Series:
    lookup = {str(v).strip().lower() for v in values if str(v).strip()}
    s = series.astype("string").fillna("").str.strip().str.lower()
    return s.isin(lookup)


def _sanitize_multiselect(key: str, options: Sequence[Any], default: Optional[Sequence[Any]] = None):
    opts = list(options or [])
    default_vals = list(default or ["ALL"])
    cur = st.session_state.get(key, default_vals)
    if not isinstance(cur, list):
        cur = [cur]
    cur = [x for x in cur if x in opts]
    if not cur:
        cur = default_vals
    if "ALL" in cur and len(cur) > 1:
        cur = [x for x in cur if x != "ALL"]
    st.session_state[key] = cur


def _sanitize_multiselect_sym(key: str, options: Sequence[str], default: Optional[Sequence[str]] = None):
    opts = list(options or [])
    default_vals = list(default or ["All"])
    cur = st.session_state.get(key, default_vals)
    if not isinstance(cur, list):
        cur = [cur]
    cur = [x for x in cur if x in opts]
    if not cur:
        cur = default_vals
    if "All" in cur and len(cur) > 1:
        cur = [x for x in cur if x != "All"]
    st.session_state[key] = cur


def _reset_review_filters():
    for key in list(st.session_state.keys()):
        if key.startswith("rf_"):
            st.session_state.pop(key, None)
    st.session_state["review_explorer_page"] = 1
    st.session_state["review_filter_signature"] = None


def _filter_series_for_spec(df: pd.DataFrame, spec: Dict[str, Any]) -> pd.Series:
    key = spec["key"]
    if spec["kind"] == "column":
        s = df[spec["column"]].astype("string").fillna("Unknown").str.strip()
        return s.replace("", "Unknown")
    if key == "review_type":
        raw = df.get("incentivized_review", pd.Series(False, index=df.index)).astype("boolean")
        return raw.map({True: "Incentivized", False: "Organic"}).fillna("Unknown")
    if key == "recommendation":
        raw = df.get("is_recommended", pd.Series(pd.NA, index=df.index)).astype("boolean")
        return raw.map({True: "Recommended", False: "Not Recommended"}).fillna("Unknown")
    if key == "syndication":
        raw = df.get("is_syndicated", pd.Series(pd.NA, index=df.index)).astype("boolean")
        return raw.map({True: "Syndicated", False: "Not Syndicated"}).fillna("Unknown")
    if key == "media":
        raw = df.get("has_photos", pd.Series(False, index=df.index)).astype("boolean")
        return raw.map({True: "With Photos", False: "No Photos"}).fillna("Unknown")
    return pd.Series("Unknown", index=df.index)


def _core_filter_specs_for_df(df: pd.DataFrame) -> List[Dict[str, Any]]:
    specs: List[Dict[str, Any]] = []
    for spec in CORE_REVIEW_FILTER_SPECS:
        if spec["kind"] == "column" and spec.get("column") not in df.columns:
            continue
        s = _filter_series_for_spec(df, spec)
        opts = [x for x in sorted({str(v).strip() for v in s.dropna().astype(str) if str(v).strip()}, key=lambda x: x.lower()) if x]
        if not opts:
            continue
        if len(opts) == 1 and opts[0] == "Unknown":
            continue
        specs.append({**spec, "options": ["ALL"] + opts})
    return specs


def _col_options(df: pd.DataFrame, col: str, max_vals: Optional[int] = 250) -> List[str]:
    if col not in df.columns:
        return ["ALL"]
    s = df[col]
    vals = s.astype("string").fillna("Unknown").str.strip().replace("", "Unknown").tolist()
    uniq = list(dict.fromkeys(v for v in vals if str(v).strip()))
    uniq = sorted(uniq, key=lambda x: str(x).lower())
    if max_vals is not None:
        uniq = uniq[: int(max_vals)]
    return ["ALL"] + uniq


def _infer_extra_filter_kind(df: pd.DataFrame, col: str) -> str:
    if col not in df.columns:
        return "categorical"
    s = df[col]
    name = str(col).lower()
    try:
        if pd.api.types.is_datetime64_any_dtype(s):
            return "date"
    except Exception:
        pass
    looks_datey = any(tok in name for tok in ["date", "time", "day", "month", "year"])
    if looks_datey and not pd.api.types.is_numeric_dtype(s):
        try:
            as_dt = pd.to_datetime(s, errors="coerce")
            if as_dt.notna().mean() >= 0.75 and as_dt.nunique(dropna=True) > 2:
                return "date"
        except Exception:
            pass
    try:
        num = pd.to_numeric(s, errors="coerce")
        if num.notna().mean() >= 0.9 and num.nunique(dropna=True) > 6:
            return "numeric"
    except Exception:
        pass
    return "categorical"


def _extra_filter_candidates(df: pd.DataFrame) -> List[str]:
    det_cols, del_cols = _get_symptom_col_lists(df)
    excluded = {
        "review_id", "title", "review_text", "title_and_text", "rating", "rating_label",
        "submission_time", "submission_date", "submission_month", "year_month_sort",
        "incentivized_review", "is_recommended", "is_syndicated", "has_photos", "has_media",
        "photo_urls", "raw_json", "context_data_json", "review_length_chars", "review_length_words",
        "AI Safety", "AI Reliability", "AI # of Sessions",
    }
    excluded.update({spec["column"] for spec in CORE_REVIEW_FILTER_SPECS if spec.get("kind") == "column"})
    excluded.update(set(det_cols + del_cols))
    excluded.update({c for c in df.columns if str(c).startswith("AI Symptom ") or str(c).startswith("Symptom ")})
    return sorted([str(c) for c in df.columns if str(c) not in excluded], key=lambda x: x.lower())


def _symptom_filter_options(df: pd.DataFrame) -> Tuple[List[str], List[str], List[str], List[str]]:
    det_cols, del_cols = _get_symptom_col_lists(df)
    def _values(cols: Sequence[str]) -> List[str]:
        vals: List[str] = []
        for col in cols:
            if col not in df.columns:
                continue
            s = df[col].astype("string").fillna("").str.strip()
            good = s[(s != "") & (~s.str.upper().isin(NON_VALUES))]
            vals.extend(good.tolist())
        return sorted(list(dict.fromkeys(vals)), key=lambda x: x.lower())
    return _values(det_cols), _values(del_cols), list(det_cols), list(del_cols)


def _collect_active_filter_items(df: pd.DataFrame, *, core_specs: Sequence[Dict[str, Any]], extra_cols: Sequence[str], tf: str, start_date, end_date) -> List[Tuple[str, str]]:
    items: List[Tuple[str, str]] = []
    if tf != "All Time":
        if tf == "Custom Range" and start_date and end_date:
            items.append(("Timeframe", f"{start_date} → {end_date}"))
        else:
            items.append(("Timeframe", tf))

    sr_raw = st.session_state.get("rf_sr", ["All"])
    sr_list = sr_raw if isinstance(sr_raw, list) else [sr_raw]
    sr_sel = [x for x in sr_list if str(x).strip() and str(x).lower() != "all"]
    if sr_sel:
        items.append(("Stars", ", ".join(str(x) for x in sr_sel)))

    for spec in core_specs:
        sel = st.session_state.get(f"rf_{spec['key']}", ["ALL"])
        sel_list = sel if isinstance(sel, list) else [sel]
        sel_clean = [x for x in sel_list if str(x).strip() and str(x).upper() != "ALL"]
        if sel_clean:
            items.append((spec["label"], ", ".join(str(x) for x in sel_clean[:4]) + ("" if len(sel_clean) <= 4 else f" +{len(sel_clean) - 4}")))

    for col in extra_cols:
        if col not in df.columns:
            continue
        kind = _infer_extra_filter_kind(df, col)
        rk = f"rf_{col}_range"
        dk = f"rf_{col}_date_range"
        ck = f"rf_{col}_contains"
        if kind == "numeric":
            num = pd.to_numeric(df[col], errors="coerce").dropna()
            if not num.empty and rk in st.session_state and isinstance(st.session_state.get(rk), (tuple, list)) and len(st.session_state.get(rk)) == 2:
                lo, hi = st.session_state[rk]
                base_lo, base_hi = float(num.min()), float(num.max())
                if round(float(lo), 10) != round(base_lo, 10) or round(float(hi), 10) != round(base_hi, 10):
                    items.append((col, f"{float(lo):g} → {float(hi):g}"))
            continue
        if kind == "date":
            dt = pd.to_datetime(df[col], errors="coerce").dropna()
            if not dt.empty and dk in st.session_state and isinstance(st.session_state.get(dk), (tuple, list)) and len(st.session_state.get(dk)) == 2:
                lo, hi = st.session_state[dk]
                base_lo, base_hi = dt.min().date(), dt.max().date()
                if lo and hi and (lo != base_lo or hi != base_hi):
                    items.append((col, f"{lo} → {hi}"))
            continue
        cv = _safe_text(st.session_state.get(ck))
        if cv:
            items.append((col, f"contains: {cv}"))
            continue
        sel = st.session_state.get(f"rf_{col}", ["ALL"])
        sel_list = sel if isinstance(sel, list) else [sel]
        sel_clean = [x for x in sel_list if str(x).strip() and str(x).upper() != "ALL"]
        if sel_clean:
            items.append((col, ", ".join(str(x) for x in sel_clean[:4]) + ("" if len(sel_clean) <= 4 else f" +{len(sel_clean) - 4}")))

    sel_det = [x for x in (st.session_state.get("rf_sym_detract", ["All"]) or []) if str(x).strip() and str(x).lower() != "all"]
    sel_del = [x for x in (st.session_state.get("rf_sym_delight", ["All"]) or []) if str(x).strip() and str(x).lower() != "all"]
    if sel_det:
        items.append(("Detractors", ", ".join(sel_det[:3]) + ("" if len(sel_det) <= 3 else f" +{len(sel_det) - 3}")))
    if sel_del:
        items.append(("Delighters", ", ".join(sel_del[:3]) + ("" if len(sel_del) <= 3 else f" +{len(sel_del) - 3}")))

    kw = _safe_text(st.session_state.get("rf_kw"))
    if kw:
        items.append(("Keyword", kw))
    return items


def _filter_description_from_items(items: Sequence[Tuple[str, str]]) -> str:
    return "; ".join(f"{k}={v}" for k, v in items) if items else "No active filters"


def _apply_live_review_filters(df_base: pd.DataFrame) -> Dict[str, Any]:
    t0 = time.perf_counter()
    if df_base is None or df_base.empty:
        return {"filtered_df": df_base.copy() if isinstance(df_base, pd.DataFrame) else pd.DataFrame(), "active_items": [], "filter_seconds": 0.0, "description": "No active filters"}
    d0 = df_base
    mask = pd.Series(True, index=d0.index)
    today = date.today()
    tf = st.session_state.get("rf_tf", "All Time")
    start_date = end_date = None
    if tf == "Custom Range":
        rng = st.session_state.get("rf_tf_range", (today - timedelta(days=30), today))
        if isinstance(rng, (tuple, list)) and len(rng) == 2:
            start_date, end_date = rng
    elif tf == "Last Week":
        start_date, end_date = today - timedelta(days=7), today
    elif tf == "Last Month":
        start_date, end_date = today - timedelta(days=30), today
    elif tf == "Last Year":
        start_date, end_date = today - timedelta(days=365), today

    date_col = "submission_date" if "submission_date" in d0.columns else ("submission_time" if "submission_time" in d0.columns else None)
    if start_date and end_date and date_col:
        dt = pd.to_datetime(d0[date_col], errors="coerce")
        end_inclusive = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(nanoseconds=1)
        mask &= (dt >= pd.Timestamp(start_date)) & (dt <= end_inclusive)

    sr_raw = st.session_state.get("rf_sr", ["All"])
    sr_list = sr_raw if isinstance(sr_raw, list) else [sr_raw]
    sr_sel = [x for x in sr_list if str(x).strip() and str(x).lower() != "all"]
    if sr_sel and "rating" in d0.columns:
        sr_nums = [int(x) for x in sr_sel if str(x).isdigit()]
        if sr_nums:
            mask &= pd.to_numeric(d0["rating"], errors="coerce").isin(sr_nums)

    core_specs = _core_filter_specs_for_df(d0)
    for spec in core_specs:
        sel = st.session_state.get(f"rf_{spec['key']}", ["ALL"])
        sel_list = sel if isinstance(sel, list) else [sel]
        sel_clean = [x for x in sel_list if str(x).strip() and str(x).upper() != "ALL"]
        if sel_clean:
            mask &= _series_matches_any(_filter_series_for_spec(d0, spec), [str(x) for x in sel_clean])

    extra_cols = [c for c in (st.session_state.get("rf_extra_filter_cols", []) or []) if c in d0.columns]
    for col in extra_cols:
        kind = _infer_extra_filter_kind(d0, col)
        s = d0[col]
        if kind == "numeric":
            rk = f"rf_{col}_range"
            if rk in st.session_state and isinstance(st.session_state.get(rk), (tuple, list)) and len(st.session_state.get(rk)) == 2:
                lo, hi = st.session_state[rk]
                mask &= pd.to_numeric(s, errors="coerce").between(float(lo), float(hi), inclusive="both")
        elif kind == "date":
            dk = f"rf_{col}_date_range"
            if dk in st.session_state and isinstance(st.session_state.get(dk), (tuple, list)) and len(st.session_state.get(dk)) == 2:
                lo, hi = st.session_state[dk]
                if lo and hi:
                    dt = pd.to_datetime(s, errors="coerce")
                    hi_end = pd.Timestamp(hi) + pd.Timedelta(days=1) - pd.Timedelta(nanoseconds=1)
                    mask &= (dt >= pd.Timestamp(lo)) & (dt <= hi_end)
        else:
            ck = f"rf_{col}_contains"
            cv = _safe_text(st.session_state.get(ck)).strip()
            if cv:
                mask &= s.astype("string").fillna("").str.contains(cv, case=False, na=False, regex=False)
            else:
                sel = st.session_state.get(f"rf_{col}", ["ALL"])
                sel_list = sel if isinstance(sel, list) else [sel]
                sel_clean = [x for x in sel_list if str(x).strip() and str(x).upper() != "ALL"]
                if sel_clean:
                    ss = s.astype("string").fillna("")
                    sample = ss.head(200).astype(str)
                    if bool(sample.str.contains(r"\s\|\s", regex=True).any()):
                        toks = [str(x).strip() for x in sel_clean if str(x).strip()]
                        pattern = r"(^|\s*\|\s*)(" + "|".join(re.escape(t) for t in toks) + r")(\s*\|\s*|$)"
                        mask &= ss.str.contains(pattern, case=False, regex=True, na=False)
                    else:
                        mask &= ss.isin([str(x) for x in sel_clean])

    _, _, det_cols, del_cols = _symptom_filter_options(d0)
    sel_det = [x for x in (st.session_state.get("rf_sym_detract", ["All"]) or []) if str(x).strip() and str(x).lower() != "all"]
    sel_del = [x for x in (st.session_state.get("rf_sym_delight", ["All"]) or []) if str(x).strip() and str(x).lower() != "all"]
    if sel_det and det_cols:
        mask &= d0[det_cols].isin(sel_det).any(axis=1)
    if sel_del and del_cols:
        mask &= d0[del_cols].isin(sel_del).any(axis=1)

    kw = _safe_text(st.session_state.get("rf_kw")).strip()
    search_col = "title_and_text" if "title_and_text" in d0.columns else ("review_text" if "review_text" in d0.columns else None)
    if kw and search_col:
        mask &= d0[search_col].astype("string").fillna("").str.contains(kw, case=False, na=False, regex=False)

    filtered = d0.loc[mask].copy()
    active_items = _collect_active_filter_items(d0, core_specs=core_specs, extra_cols=extra_cols, tf=tf, start_date=start_date, end_date=end_date)
    return {"filtered_df": filtered, "active_items": active_items, "filter_seconds": time.perf_counter() - t0, "description": _filter_description_from_items(active_items)}


def _render_active_filter_summary(filter_state: Dict[str, Any], overall_df: pd.DataFrame):
    active_items = filter_state.get("active_items", [])
    pills = []
    for k, v in active_items[:12]:
        pills.append(f"<div class='pill'><span class='muted'>{_esc(k)}:</span> {_esc(v)}</div>")
    st.markdown(
        f"""
<div class="soft-panel">
  <div><b>Active filters</b> • Showing <b>{len(filter_state.get('filtered_df', [])):,}</b> of <b>{len(overall_df):,}</b> reviews
  <span class="small-muted"> (filter time: {float(filter_state.get('filter_seconds', 0.0)):.3f}s)</span>
  </div>
  <div class="pill-row">{''.join(pills) if pills else '<span class="small-muted">None (All data)</span>'}</div>
</div>
""",
        unsafe_allow_html=True,
    )


def _product_name(summary, df):
    if not df.empty and "original_product_name" in df.columns:
        n = _first_non_empty(df["original_product_name"].fillna(""))
        if n:
            return n
    return summary.product_id

# ═══════════════════════════════════════════════════════════════════════════════
#  AI ANALYST
# ═══════════════════════════════════════════════════════════════════════════════
GENERAL_INSTRUCTIONS = textwrap.dedent("""
    You are SharkNinja Review Analyst — an internal voice-of-customer AI assistant.
    ROLE: Synthesise consumer review data into sharp, actionable insights.
    Prioritise evidence from the supplied dataset over generic assumptions.
    ANSWER FORMAT
    • Use clear markdown headings (##, ###).
    • Lead with the most important insight — do not bury the lede.
    • Cite review IDs inline: (review_ids: 12345, 67890).
    • For every quantitative claim state the count or percentage from the data.
    • Mark inferences: [INFERRED].
    • Keep responses ≤ 400 words unless the user explicitly asks for depth.
    • End every response with a "**Next Steps**" section: 2–3 concrete actions.
    GUARDRAILS
    • Do not invent review IDs, quotes, counts, or trends not in the evidence.
    • If the data is insufficient, say so explicitly.
    • Never hallucinate product specs — only cite what reviews mention.
""").strip()


def _persona_instructions(name):
    if not name:
        return GENERAL_INSTRUCTIONS
    return PERSONAS[name]["instructions"]


def _select_relevant(df, question, max_reviews=22):
    if df.empty:
        return df.copy()
    w = df.copy()
    w["blob"] = w["title_and_text"].fillna("").astype(str).map(_norm_text)
    qt = _tokenize(question)

    def score(row):
        s = 0.0
        t = row["blob"]
        for tk in qt:
            if tk in t:
                s += 3 + t.count(tk)
        r = row.get("rating")
        if any(tk in {"defect", "broken", "issue", "problem", "bad", "fail", "broke"} for tk in qt):
            if pd.notna(r):
                s += max(0, 6 - float(r))
        if not _safe_bool(row.get("incentivized_review"), False):
            s += 0.5
        if pd.notna(row.get("review_length_words")):
            s += min(float(row.get("review_length_words", 0)) / 60, 2)
        return s

    w["_sc"] = w.apply(score, axis=1)
    ranked = w.sort_values(["_sc", "submission_time"], ascending=[False, False], na_position="last")
    combined = pd.concat([
        ranked.head(max_reviews),
        df[df["rating"].isin([1, 2])].head(max_reviews // 3 or 1),
        df[df["rating"].isin([4, 5])].head(max_reviews // 3 or 1),
    ], ignore_index=True).drop_duplicates(subset=["review_id"])
    return combined.head(max_reviews).drop(columns=["blob", "_sc"], errors="ignore")


def _snippet_rows(df, *, max_reviews=22):
    rows = []
    for _, row in df.head(max_reviews).iterrows():
        rows.append(dict(
            review_id=_safe_text(row.get("review_id")),
            rating=_safe_int(row.get("rating"), 0) if pd.notna(row.get("rating")) else None,
            incentivized_review=_safe_bool(row.get("incentivized_review"), False),
            content_locale=_safe_text(row.get("content_locale")),
            submission_date=_safe_text(row.get("submission_date")),
            title=_trunc(row.get("title", ""), 120),
            snippet=_trunc(row.get("review_text", ""), 600),
        ))
    return rows


def _build_ai_context(*, overall_df, filtered_df, summary, filter_description, question):
    om = _get_metrics(overall_df)
    fm = _get_metrics(filtered_df)
    try:
        rd = _rating_dist(filtered_df).to_dict(orient="records")
    except Exception:
        rd = []
    try:
        md = _monthly_trend(filtered_df).tail(12).to_dict(orient="records")
    except Exception:
        md = []
    rel = _select_relevant(filtered_df, question, max_reviews=22)
    rec = filtered_df.sort_values(["submission_time", "review_id"], ascending=[False, False], na_position="last").head(10)
    low = filtered_df[filtered_df["rating"].isin([1, 2])].head(8)
    hi = filtered_df[filtered_df["rating"].isin([4, 5])].head(8)
    ev = pd.concat([rel, rec, low, hi], ignore_index=True).drop_duplicates(subset=["review_id"]).head(32)
    payload = dict(
        product=dict(product_id=summary.product_id, product_url=summary.product_url, product_name=_product_name(summary, overall_df)),
        analysis_scope=dict(filter_description=filter_description, overall_review_count=len(overall_df), filtered_review_count=len(filtered_df)),
        metric_snapshot=dict(overall=om, filtered=fm, rating_distribution_filtered=rd, monthly_trend_filtered=md),
        review_text_evidence=_snippet_rows(ev, max_reviews=32),
    )
    full_json = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
    tok = _estimate_tokens(full_json)
    max_ev = 22
    while tok > AI_CONTEXT_TOKEN_BUDGET and max_ev >= 5:
        max_ev -= 4
        payload["review_text_evidence"] = _snippet_rows(ev, max_reviews=max_ev)
        full_json = json.dumps(payload, ensure_ascii=False, indent=2, default=str)
        tok = _estimate_tokens(full_json)
    return full_json


def _call_analyst(*, question, overall_df, filtered_df, summary, filter_description, chat_history, persona_name=None, report_length="Medium"):
    client = _get_client()
    if client is None:
        raise ReviewDownloaderError("No OpenAI API key configured.")
    _length_tokens = {"Short": 600, "Medium": 1200, "Long": 2200}
    _length_note = {
        "Short": "IMPORTANT: Keep your entire response under 150 words. Be direct — one key insight, one action.",
        "Medium": "Keep your response to 300–400 words.",
        "Long": "Provide a thorough 600–800 word response with detailed evidence, examples, and a full Next Steps section.",
    }
    base_instructions = _persona_instructions(persona_name)
    instructions = base_instructions + "\n\n" + _length_note.get(report_length, "")
    max_tok = _length_tokens.get(report_length, 1200)
    ai_ctx = _build_ai_context(overall_df=overall_df, filtered_df=filtered_df, summary=summary, filter_description=filter_description, question=question)
    msgs = [{"role": m["role"], "content": m["content"]} for m in list(chat_history)[-8:]]
    msgs.append({"role": "user", "content": f"User request:\n{question}\n\nReview dataset context (JSON):\n{ai_ctx}"})
    result = _chat_complete_with_fallback_models(
        client,
        model=_shared_model(),
        structured=False,
        messages=[{"role": "system", "content": instructions}, *msgs],
        temperature=0.0,
        max_tokens=max_tok,
        reasoning_effort=_shared_reasoning(),
    )
    if not result:
        raise ReviewDownloaderError("OpenAI returned empty answer.")
    return result

# ═══════════════════════════════════════════════════════════════════════════════
#  REVIEW PROMPT TAGGING
# ═══════════════════════════════════════════════════════════════════════════════
def _default_prompt_df():
    return pd.DataFrame([REVIEW_PROMPT_STARTER_ROWS[0]])


def _normalize_prompt_defs(prompt_df, existing_columns):
    if prompt_df is None or prompt_df.empty:
        return []
    normalized = []
    seen = set()
    existing_set = {str(c) for c in existing_columns}
    for _, row in prompt_df.fillna("").iterrows():
        rp = _safe_text(row.get("prompt"))
        rl = _safe_text(row.get("labels"))
        rc = _safe_text(row.get("column_name"))
        if not rp and not rl and not rc:
            continue
        if not rp:
            raise ReviewDownloaderError("Each prompt row needs a prompt.")
        if not rl:
            raise ReviewDownloaderError("Each prompt row needs labels.")
        labels = [l.strip() for l in rl.split(",") if l.strip()]
        deduped = list(dict.fromkeys(labels))
        if "Not Mentioned" not in deduped and len(deduped) <= 7:
            deduped.append("Not Mentioned")
        if len(deduped) < 2:
            raise ReviewDownloaderError("Each prompt needs at least two labels.")
        col = _slugify(rc or rp)
        if col in existing_set and col not in {"review_id"}:
            col = f"{col}_ai"
        base = col
        suffix = 2
        while col in seen:
            col = f"{base}_{suffix}"
            suffix += 1
        seen.add(col)
        normalized.append(dict(column_name=col, display_name=col.replace("_", " ").title(), prompt=rp, labels=deduped, labels_csv=", ".join(deduped)))
    return normalized


def _build_tagging_schema(prompt_defs):
    item_props = {"review_id": {"type": "string"}}
    required = ["review_id"]
    for p in prompt_defs:
        item_props[p["column_name"]] = {"type": "string", "enum": list(p["labels"])}
        required.append(p["column_name"])
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "results": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": item_props,
                    "required": required,
                },
            },
        },
        "required": ["results"],
    }


def _classify_chunk(*, client, chunk_df, prompt_defs):
    pc = max(len(prompt_defs), 1)
    max_out = int(max(1600, min(9000, 450 + len(chunk_df) * (18 + 10 * pc))))
    reviews_payload = [
        dict(
            review_id=_safe_text(row.get("review_id")),
            rating=_safe_int(row.get("rating"), 0) if pd.notna(row.get("rating")) else None,
            title=_trunc(row.get("title", ""), 200),
            review_text=_trunc(row.get("review_text", ""), 900),
            incentivized_review=_safe_bool(row.get("incentivized_review"), False),
        )
        for _, row in chunk_df.iterrows()
    ]
    prompt_payload = [dict(column_name=p["column_name"], prompt=p["prompt"], labels=p["labels"]) for p in prompt_defs]
    instructions = "You are a deterministic review-tagging engine. For each review and prompt, return exactly one allowed label. If not mentioned, use Not Mentioned."
    user_content = json.dumps({"prompt_definitions": prompt_payload, "reviews": reviews_payload})
    msgs = [{"role": "system", "content": instructions}, {"role": "user", "content": user_content}]
    structured_rf = {"type": "json_schema", "json_schema": {"name": "review_prompt_tagging", "schema": _build_tagging_schema(prompt_defs), "strict": True}}
    result_text = ""
    try:
        result_text = _chat_complete_with_fallback_models(
            client,
            model=_shared_model(),
            structured=True,
            messages=msgs,
            temperature=0.0,
            response_format=structured_rf,
            max_tokens=max_out,
            reasoning_effort=_shared_reasoning(),
        )
    except Exception as exc:
        col_hints = ", ".join(f'{p["column_name"]}: one of [{", ".join(p["labels"])}]' for p in prompt_defs)
        fallback_instructions = (
            "You are a deterministic review-tagging engine. Return ONLY a JSON object with key 'results' containing an array. "
            "Each element must have: review_id (string), " + col_hints + ". Include every review_id from the input. Use 'Not Mentioned' if not applicable."
        )
        result_text = _chat_complete_with_fallback_models(
            client,
            model=_shared_model(),
            structured=True,
            messages=[{"role": "system", "content": fallback_instructions}, {"role": "user", "content": user_content}],
            temperature=0.0,
            response_format={"type": "json_object"},
            max_tokens=max_out,
            reasoning_effort=_shared_reasoning(),
        )
    if not result_text:
        raise ReviewDownloaderError("OpenAI returned an empty response. Check your API key and model selection.")
    data = _safe_json_load(result_text)
    output_rows = data.get("results") or []
    out_df = pd.DataFrame(output_rows)
    if out_df.empty:
        raise ReviewDownloaderError(f"OpenAI returned no tagged rows. Raw response snippet: {result_text[:300]}")
    out_df["review_id"] = out_df["review_id"].astype(str)
    expected = set(chunk_df["review_id"].astype(str))
    returned = set(out_df["review_id"].astype(str))
    if expected != returned:
        miss = sorted(expected - returned)
        if miss:
            import warnings
            warnings.warn(f"Batch partial: missing {miss[:5]}")
        out_df = out_df[out_df["review_id"].isin(expected)]
    return out_df


def _run_review_prompt_tagging(*, client, source_df, prompt_defs, chunk_size):
    if source_df.empty:
        raise ReviewDownloaderError("No reviews in scope.")
    chunks = list(range(0, len(source_df), chunk_size))
    prog = st.progress(0.0, text="Preparing…")
    status = st.empty()
    outputs = []
    errors = []
    for i, start in enumerate(chunks, 1):
        chunk_df = source_df.iloc[start:start + chunk_size].copy()
        status.info(f"Classifying {start + 1}–{min(start + chunk_size, len(source_df))} of {len(source_df)}")
        try:
            outputs.append(_classify_chunk(client=client, chunk_df=chunk_df, prompt_defs=prompt_defs))
        except Exception as exc:
            errors.append(f"Batch {i}: {exc}")
            status.warning(f"Batch {i} failed — {exc}")
        prog.progress(i / len(chunks))
        gc.collect()
    if not outputs:
        err_detail = "; ".join(errors[:3]) if errors else "unknown error"
        raise ReviewDownloaderError(f"All batches failed. First error: {err_detail}")
    if errors:
        status.warning(f"{len(errors)} of {len(chunks)} batch(es) failed — partial results saved.")
    else:
        status.success(f"Finished tagging {len(source_df):,} reviews.")
    return pd.concat(outputs, ignore_index=True).drop_duplicates(subset=["review_id"], keep="last")


def _merge_prompt_results(overall_df, prompt_results_df, prompt_defs):
    updated = overall_df.copy()
    rids = updated["review_id"].astype(str)
    lk = prompt_results_df.set_index("review_id")
    for p in prompt_defs:
        col = p["column_name"]
        if col not in updated.columns:
            updated[col] = pd.NA
        mapping = lk[col].to_dict()
        nv = rids.map(mapping)
        updated[col] = nv.where(nv.notna(), updated[col])
    return updated


def _summarize_prompt_results(prompt_results_df, prompt_defs, source_df=None):
    merged = prompt_results_df.copy()
    merged["review_id"] = merged["review_id"].astype(str)
    if source_df is not None and not source_df.empty and "review_id" in source_df.columns:
        lk = source_df[[c for c in ["review_id", "rating"] if c in source_df.columns]].copy()
        lk["review_id"] = lk["review_id"].astype(str)
        merged = merged.merge(lk, on="review_id", how="left")
    rows = []
    total = max(len(prompt_results_df), 1)
    for p in prompt_defs:
        col = p["column_name"]
        for label in p["labels"]:
            sub = merged[merged[col] == label]
            rows.append(dict(column_name=col, display_name=p["display_name"], label=str(label), review_count=len(sub), share=_safe_pct(len(sub), total), avg_rating=_safe_mean(sub["rating"]) if "rating" in sub.columns else None))
    return pd.DataFrame(rows)

# ═══════════════════════════════════════════════════════════════════════════════
#  EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
def _autosize_ws(ws, df):
    ws.freeze_panes = "A2"
    for idx, col in enumerate(df.columns, 1):
        series = df[col].head(250).fillna("").astype(str)
        max_len = max([len(str(col))] + [len(v) for v in series.tolist()])
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)


def _build_master_excel(summary, reviews_df, *, prompt_defs=None, prompt_summary_df=None, prompt_scope=""):
    metrics = _get_metrics(reviews_df)
    try:
        rd = _rating_dist(reviews_df)
        md = _monthly_trend(reviews_df)
    except Exception:
        rd = pd.DataFrame()
        md = pd.DataFrame()
    summary_df = pd.DataFrame([dict(product_name=_product_name(summary, reviews_df), product_id=summary.product_id, product_url=summary.product_url, reviews_downloaded=summary.reviews_downloaded, avg_rating=metrics.get("avg_rating"), avg_rating_non_incentivized=metrics.get("avg_rating_non_incentivized"), pct_low_star=metrics.get("pct_low_star"), pct_incentivized=metrics.get("pct_incentivized"), generated_utc=pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"))])
    priority_cols = ["review_id", "product_id", "rating", "incentivized_review", "is_recommended", "submission_time", "content_locale", "title", "review_text"]
    pc = [p["column_name"] for p in (prompt_defs or []) if p["column_name"] in reviews_df.columns]
    ordered = [c for c in priority_cols + pc if c in reviews_df.columns]
    remaining = [c for c in reviews_df.columns if c not in ordered]
    exp_reviews = reviews_df[ordered + remaining]
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        sheets = {"Summary": summary_df, "Reviews": exp_reviews, "RatingDistribution": rd, "ReviewVolume": md}
        if prompt_defs:
            sheets["ReviewPromptDefinitions"] = pd.DataFrame([dict(column_name=p["column_name"], display_name=p["display_name"], prompt=p["prompt"], labels=", ".join(p["labels"]), scope=prompt_scope) for p in prompt_defs])
        if prompt_summary_df is not None and not prompt_summary_df.empty:
            sheets["ReviewPromptSummary"] = prompt_summary_df
        for sname, df_ in sheets.items():
            if df_ is None or df_.empty:
                continue
            df_.to_excel(writer, sheet_name=sname, index=False)
            _autosize_ws(writer.sheets[sname], df_)
    out.seek(0)
    return out.getvalue()


def _get_master_bundle(summary, reviews_df, prompt_artifacts):
    pd_ = (prompt_artifacts or {}).get("definitions") or []
    psd = (prompt_artifacts or {}).get("summary_df")
    ps = (prompt_artifacts or {}).get("scope_label", "")
    key = json.dumps(dict(pid=summary.product_id, n=len(reviews_df), cols=sorted(str(c) for c in reviews_df.columns), psig=(prompt_artifacts or {}).get("definition_signature")), sort_keys=True)
    b = st.session_state.get("master_export_bundle")
    if b and b.get("key") == key:
        return b
    xlsx = _build_master_excel(summary, reviews_df, prompt_defs=pd_, prompt_summary_df=psd, prompt_scope=ps)
    ts = pd.Timestamp.utcnow().strftime("%Y%m%d_%H%M%S")
    b = dict(key=key, excel_bytes=xlsx, excel_name=f"{summary.product_id}_review_workspace_{ts}.xlsx")
    st.session_state["master_export_bundle"] = b
    return b

# ═══════════════════════════════════════════════════════════════════════════════
#  SYMPTOMIZER HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _get_symptom_whitelists(file_bytes):
    bio = io.BytesIO(file_bytes)
    try:
        df_sym = pd.read_excel(bio, sheet_name="Symptoms")
    except Exception:
        return [], [], {}
    if df_sym is None or df_sym.empty:
        return [], [], {}
    df_sym.columns = [str(c).strip() for c in df_sym.columns]
    lc = {c.lower(): c for c in df_sym.columns}
    alias_col = next((lc[c] for c in ["aliases", "alias"] if c in lc), None)
    label_col = next((lc[c] for c in ["symptom", "label", "name", "item"] if c in lc), None)
    type_col = next((lc[c] for c in ["type", "polarity", "category", "side"] if c in lc), None)
    pos_tags = {"delighter", "delighters", "positive", "pos", "pros"}
    neg_tags = {"detractor", "detractors", "negative", "neg", "cons"}

    def _clean(s):
        vals = s.dropna().astype(str).str.strip()
        out = []
        seen = set()
        for v in vals:
            if v and v not in seen:
                seen.add(v)
                out.append(v)
        return out

    delighters, detractors, alias_map = [], [], {}
    if label_col and type_col:
        df_sym[type_col] = df_sym[type_col].astype(str).str.lower().str.strip()
        delighters = _clean(df_sym.loc[df_sym[type_col].isin(pos_tags), label_col])
        detractors = _clean(df_sym.loc[df_sym[type_col].isin(neg_tags), label_col])
        if alias_col:
            for _, row in df_sym.iterrows():
                lbl = str(row.get(label_col, "")).strip()
                als = str(row.get(alias_col, "")).strip()
                if lbl:
                    alias_map[lbl] = [p.strip() for p in als.replace(",", "|").split("|") if p.strip()] if als else []
    else:
        for lck, orig in lc.items():
            if "delight" in lck or "positive" in lck or lck == "pros":
                delighters.extend(_clean(df_sym[orig]))
            if "detract" in lck or "negative" in lck or lck == "cons":
                detractors.extend(_clean(df_sym[orig]))
        delighters = list(dict.fromkeys(delighters))
        detractors = list(dict.fromkeys(detractors))
    return delighters, detractors, alias_map


def _ensure_ai_cols(df):
    for h in AI_DET_HEADERS + AI_DEL_HEADERS + AI_META_HEADERS:
        if h not in df.columns:
            df[h] = None
    return df


def _detect_sym_cols(df):
    cols = [str(c).strip() for c in df.columns]
    return dict(
        manual_detractors=[f"Symptom {i}" for i in range(1, 11) if f"Symptom {i}" in cols],
        manual_delighters=[f"Symptom {i}" for i in range(11, 21) if f"Symptom {i}" in cols],
        ai_detractors=[c for c in cols if c.startswith("AI Symptom Detractor ")],
        ai_delighters=[c for c in cols if c.startswith("AI Symptom Delighter ")],
    )


def _filled_mask(df, cols):
    if not cols:
        return pd.Series(False, index=df.index)
    mask = pd.Series(False, index=df.index)
    for c in cols:
        if c not in df.columns:
            continue
        s = df[c].fillna("").astype(str).str.strip()
        mask |= (s != "") & (~s.str.upper().isin(NON_VALUES))
    return mask


def _detect_missing(df, colmap):
    out = df.copy()
    det_cols = colmap["manual_detractors"] + colmap["ai_detractors"]
    del_cols = colmap["manual_delighters"] + colmap["ai_delighters"]
    out["Has_Detractors"] = _filled_mask(out, det_cols)
    out["Has_Delighters"] = _filled_mask(out, del_cols)
    out["Needs_Detractors"] = ~out["Has_Detractors"]
    out["Needs_Delighters"] = ~out["Has_Delighters"]
    out["Needs_Symptomization"] = out["Needs_Detractors"] & out["Needs_Delighters"]
    return out


def _match_label(raw, allowed, aliases=None, cutoff=0.76):
    if not raw or not allowed:
        return None
    raw_s = raw.strip()
    exact = {_canon_simple(x): x for x in allowed}
    lbl = exact.get(_canon_simple(raw_s))
    if lbl:
        return lbl
    if aliases:
        for canonical, als in (aliases or {}).items():
            if canonical not in allowed:
                continue
            for a in (als or []):
                if _canon_simple(raw_s) == _canon_simple(a):
                    return canonical
    m = difflib.get_close_matches(raw_s, allowed, n=1, cutoff=cutoff)
    if m:
        return m[0]
    raw_lower = raw_s.lower()
    for label in allowed:
        if raw_lower in label.lower() or label.lower() in raw_lower:
            return label
    return None


def _validate_evidence(evidence_list, review_text, max_ev_chars=120):
    if not evidence_list or not review_text:
        return []
    rv = review_text.lower()
    out = []
    for e in evidence_list:
        e = str(e).strip()[:max_ev_chars]
        if len(e) >= 4 and e.lower() in rv:
            out.append(e)
    return out[:2]


def _prioritize_for_symptomization(df):
    if df.empty:
        return df
    w = df.copy()
    rating = pd.to_numeric(w.get("rating", pd.Series(dtype=float)), errors="coerce").fillna(3)
    w["_prio"] = (
        (rating <= 2).astype(int) * 4 +
        (rating >= 5).astype(int) * 2 +
        (rating == 4).astype(int) * 1 +
        (w.get("review_length_words", pd.Series(0, index=w.index)).fillna(0).clip(upper=500) / 100)
    )
    kw_det = r"\b(broke|broken|fail|defect|issue|problem|stopped|won't|doesn't|difficult|loud|noise|leak|burned|smoke|stuck|cracked|smells)\b"
    kw_del = r"\b(love|perfect|amazing|excellent|great|easy|simple|quiet|durable|worth|recommend|best)\b"
    text = w.get("title_and_text", w.get("review_text", pd.Series("", index=w.index))).fillna("")
    w["_prio"] += text.str.lower().str.count(kw_det, flags=re.IGNORECASE).clip(upper=3)
    w["_prio"] += text.str.lower().str.count(kw_del, flags=re.IGNORECASE).clip(upper=2) * 0.5
    return w.sort_values("_prio", ascending=False).drop(columns=["_prio"])

def _call_symptomizer_batch(*, client, items, allowed_delighters, allowed_detractors,
                             product_profile="", max_ev_chars=120, aliases=None):
    out_by_idx = {}
    if not items:
        return out_by_idx
    det_list = "\n".join(f"  - {l}" for l in allowed_detractors) or "  (none defined)"
    del_list = "\n".join(f"  - {l}" for l in allowed_delighters) or "  (none defined)"
    system_prompt = f"""You are an expert consumer product review analyst for SharkNinja.
Your job: for each review, identify every symptom from the catalog that is EXPLICITLY mentioned.

{f"Product context: {product_profile[:500]}" if product_profile else ""}

═══ ALLOWED DETRACTORS (problems / complaints) ═══
{det_list}

═══ ALLOWED DELIGHTERS (positives / strengths) ═══
{del_list}

═══ CLASSIFICATION ENUMS ═══
safety      → {SAFETY_ENUM}
reliability → {RELIABILITY_ENUM}
sessions    → {SESSIONS_ENUM}

═══ STRICT RULES ═══
1. HIGH RECALL: Find EVERY applicable symptom. A 1-star review may have 5+ detractors — tag them all.
2. EXACT LABELS: Use label text EXACTLY as it appears in the catalog above. No paraphrasing.
3. EVIDENCE: Each evidence string must be verbatim text from the review (4–{max_ev_chars} chars). Max 2 per label.
4. NO INFERENCE: Only tag what is explicitly stated or clearly described. Never assume.
5. UNLISTED: If an important theme is NOT in the catalog, add it to unlisted_detractors or unlisted_delighters as a 2-5 word noun phrase.
6. ALL IDs: Return a result for EVERY review id in the input, even if no symptoms apply.

═══ OUTPUT SCHEMA (strict JSON) ═══
{{"items":[{{
  "id":"<review_id_string>",
  "detractors":[{{"label":"<exact catalog label>","evidence":["<verbatim text>"]}}],
  "delighters":[{{"label":"<exact catalog label>","evidence":["<verbatim text>"]}}],
  "unlisted_detractors":["<2-5 word theme>"],
  "unlisted_delighters":["<2-5 word theme>"],
  "safety":"<enum value>",
  "reliability":"<enum value>",
  "sessions":"<enum value>"
}}]}}"""
    payload = dict(items=[dict(id=str(it["idx"]), review=it["review"], needs_delighters=it.get("needs_del", True), needs_detractors=it.get("needs_det", True)) for it in items])
    max_out = min(7000, max(1800, 230 * len(items) + 400))
    result_text = _chat_complete_with_fallback_models(
        client,
        model=_shared_model(),
        structured=True,
        messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": json.dumps(payload)}],
        temperature=0.0,
        response_format={"type": "json_object"},
        max_tokens=max_out,
        reasoning_effort=_shared_reasoning(),
    )
    data = _safe_json_load(result_text)
    items_out = data.get("items") or (data if isinstance(data, list) else [])
    by_id = {str(o.get("id")): o for o in items_out if isinstance(o, dict) and "id" in o}
    for it in items:
        idx = int(it["idx"])
        review_text = it.get("review", "")
        obj = by_id.get(str(idx)) or {}

        def _extract_side(objs, allowed):
            labels = []
            ev_map = {}
            for obj2 in (objs or []):
                if not isinstance(obj2, dict):
                    continue
                raw = str(obj2.get("label", "")).strip()
                lbl = _match_label(raw, allowed, aliases=aliases)
                if not lbl:
                    continue
                raw_evs = [str(e) for e in (obj2.get("evidence") or []) if isinstance(e, str)]
                validated = _validate_evidence(raw_evs, review_text, max_ev_chars)
                if not validated:
                    validated = [str(e).strip()[:max_ev_chars] for e in raw_evs if str(e).strip()][:1]
                if lbl not in labels:
                    labels.append(lbl)
                    ev_map[lbl] = validated[:2]
                if len(labels) >= 10:
                    break
            return labels, ev_map

        dels, ev_del = _extract_side(obj.get("delighters", []), allowed_delighters)
        dets, ev_det = _extract_side(obj.get("detractors", []), allowed_detractors)
        safety = str(obj.get("safety", "Not Mentioned")).strip()
        reliability = str(obj.get("reliability", "Not Mentioned")).strip()
        sessions = str(obj.get("sessions", "Unknown")).strip()
        safety = safety if safety in SAFETY_ENUM else "Not Mentioned"
        reliability = reliability if reliability in RELIABILITY_ENUM else "Not Mentioned"
        sessions = sessions if sessions in SESSIONS_ENUM else "Unknown"
        out_by_idx[idx] = dict(
            dels=dels,
            dets=dets,
            ev_del=ev_del,
            ev_det=ev_det,
            unl_dels=[str(x).strip() for x in (obj.get("unlisted_delighters") or []) if str(x).strip()][:10],
            unl_dets=[str(x).strip() for x in (obj.get("unlisted_detractors") or []) if str(x).strip()][:10],
            safety=safety,
            reliability=reliability,
            sessions=sessions,
        )
    return out_by_idx


def _ai_build_symptom_list(*, client, product_description, sample_reviews):
    sys = textwrap.dedent("""
        You are a consumer insights expert building a symptom catalog for product review analysis.
        Your catalog will be used to tag thousands of reviews with recurring themes.
        Every label must be: specific, general enough to recur, mutually exclusive, 2-5 words Title Case.
        DETRACTORS = Problems, failures, complaints, frustrations
        DELIGHTERS = Praised features, positive surprises
        REQUIRED coverage: Performance, Ease Of Use, Cleaning, Noise Level, Build Quality,
        Size/Capacity, Value, Setup, Results Quality, Safety, Temperature Control,
        Timer/Controls, Accessories, Customer Support, Connectivity, Battery Life
        OUTPUT — strict JSON only:
        {"delighters":[{"label":"<Title Case 2-5 words>","rationale":"<why>"}],
         "detractors":[{"label":"<Title Case 2-5 words>","rationale":"<why>"}]}
        Aim for 15-30 labels per side based on actual review patterns.
    """).strip()
    payload = dict(product_description=product_description or "SharkNinja consumer appliance", sample_reviews=sample_reviews[:30])
    result_text = _chat_complete_with_fallback_models(
        client,
        model=_shared_model(),
        structured=True,
        messages=[{"role": "system", "content": sys}, {"role": "user", "content": json.dumps(payload)}],
        temperature=0.0,
        response_format={"type": "json_object"},
        max_tokens=3200,
        reasoning_effort=_shared_reasoning(),
    )
    data = _safe_json_load(result_text)
    return dict(
        delighters=[str(o.get("label", "")).strip() for o in (data.get("delighters") or []) if str(o.get("label", "")).strip()],
        detractors=[str(o.get("label", "")).strip() for o in (data.get("detractors") or []) if str(o.get("label", "")).strip()],
    )


def _gen_symptomized_workbook(original_bytes, updated_df):
    wb = load_workbook(io.BytesIO(original_bytes))
    sheet_name = "Star Walk scrubbed verbatims"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    df2 = _ensure_ai_cols(updated_df.copy())
    fg = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fr = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fy = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fb = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
    fp = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")
    for i, (_, row) in enumerate(df2.iterrows(), start=2):
        for j, ci in enumerate(DET_INDEXES, 1):
            val = row.get(f"AI Symptom Detractor {j}")
            cv = None if (pd.isna(val) or str(val).strip() == "") else val
            cell = ws.cell(row=i, column=ci, value=cv)
            if cv:
                cell.fill = fr
        for j, ci in enumerate(DEL_INDEXES, 1):
            val = row.get(f"AI Symptom Delighter {j}")
            cv = None if (pd.isna(val) or str(val).strip() == "") else val
            cell = ws.cell(row=i, column=ci, value=cv)
            if cv:
                cell.fill = fg
        if _is_filled(row.get("AI Safety")):
            c = ws.cell(row=i, column=META_INDEXES["Safety"], value=str(row["AI Safety"]))
            c.fill = fy
        if _is_filled(row.get("AI Reliability")):
            c = ws.cell(row=i, column=META_INDEXES["Reliability"], value=str(row["AI Reliability"]))
            c.fill = fb
        if _is_filled(row.get("AI # of Sessions")):
            c = ws.cell(row=i, column=META_INDEXES["# of Sessions"], value=str(row["AI # of Sessions"]))
            c.fill = fp
    for c in DET_INDEXES + DEL_INDEXES + list(META_INDEXES.values()):
        try:
            ws.column_dimensions[get_column_letter(c)].width = 28
        except Exception:
            pass
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _dedup_candidates(raw):
    def _norm(s):
        s = s.strip().lower()
        s = re.sub(r"^(not\s+too\s+|not\s+very\s+|not\s+overly\s+|not\s+)", "", s)
        s = re.sub(r"[^a-z0-9 ]", " ", s)
        return re.sub(r"\s+", " ", s).strip()

    labels = sorted(raw.keys(), key=lambda l: -int(raw[l].get("count", 0)))
    merged = {}
    used = set()
    for a in labels:
        if a in used:
            continue
        merged[a] = dict(raw[a])
        used.add(a)
        na = _norm(a)
        for b in labels:
            if b in used or b == a:
                continue
            nb = _norm(b)
            if difflib.SequenceMatcher(None, na, nb).ratio() >= 0.72 or na in nb or nb in na:
                merged[a]["count"] = int(merged[a].get("count", 0)) + int(raw[b].get("count", 0))
                refs = list(merged[a].get("refs", []))
                for r in raw[b].get("refs", []):
                    if r not in refs and len(refs) < 50:
                        refs.append(r)
                merged[a]["refs"] = refs
                merged[a].setdefault("_merged_from", []).append(b)
                used.add(b)
    return merged


def _try_load_symptoms_from_file():
    raw = st.session_state.get("_uploaded_raw_bytes")
    if not raw:
        return False
    d, t, a = _get_symptom_whitelists(raw)
    if d or t:
        st.session_state.update(sym_delighters=d, sym_detractors=t, sym_aliases=a, sym_symptoms_source="file")
        return True
    return False

# ═══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ═══════════════════════════════════════════════════════════════════════════════
def _init_state():
    defaults = dict(
        analysis_dataset=None,
        chat_messages=[],
        master_export_bundle=None,
        prompt_definitions_df=_default_prompt_df(),
        prompt_builder_suggestion=None,
        prompt_run_artifacts=None,
        prompt_run_notice=None,
        chat_scope_signature=None,
        chat_scope_notice=None,
        review_explorer_page=1,
        review_explorer_per_page=20,
        review_explorer_sort="Newest",
        review_filter_signature=None,
        shared_model=DEFAULT_MODEL,
        shared_reasoning=DEFAULT_REASONING,
        workspace_source_mode=SOURCE_MODE_URL,
        workspace_product_url=DEFAULT_PRODUCT_URL,
        workspace_file_uploader_nonce=0,
        workspace_active_tab=TAB_DASHBOARD,
        workspace_tab_request=None,
        ai_scroll_to_top=False,
        sym_delighters=[],
        sym_detractors=[],
        sym_aliases={},
        sym_symptoms_source="none",
        sym_processed_rows=[],
        sym_new_candidates={},
        sym_product_profile="",
        sym_scope_choice="Missing both",
        sym_n_to_process=10,
        sym_batch_size=5,
        sym_max_ev_chars=120,
        sym_run_notice=None,
        _prompt_defs_cache={},
        _prompt_bundle_ready=False,
    )
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)


def _reset_workspace_state(*, reset_source=True):
    st.session_state["analysis_dataset"] = None
    st.session_state["chat_messages"] = []
    st.session_state["chat_scope_signature"] = None
    st.session_state["chat_scope_notice"] = None
    st.session_state["master_export_bundle"] = None
    st.session_state["prompt_run_artifacts"] = None
    st.session_state["prompt_run_notice"] = None
    st.session_state["review_explorer_page"] = 1
    st.session_state["workspace_active_tab"] = TAB_DASHBOARD
    st.session_state["workspace_tab_request"] = None
    st.session_state["ai_scroll_to_top"] = False
    st.session_state["sym_processed_rows"] = []
    st.session_state["sym_new_candidates"] = {}
    st.session_state["sym_product_profile"] = ""
    st.session_state["sym_run_notice"] = None
    st.session_state["sym_symptoms_source"] = "none"
    st.session_state["sym_delighters"] = []
    st.session_state["sym_detractors"] = []
    st.session_state["sym_aliases"] = {}
    st.session_state["_uploaded_raw_bytes"] = None
    st.session_state["sym_export_bytes"] = None
    st.session_state["_prompt_bundle_ready"] = False
    st.session_state.pop("sym_ai_build_result", None)
    _reset_review_filters()
    if reset_source:
        st.session_state["workspace_source_mode"] = SOURCE_MODE_URL
        st.session_state["workspace_product_url"] = DEFAULT_PRODUCT_URL
        st.session_state["workspace_file_uploader_nonce"] = int(st.session_state.get("workspace_file_uploader_nonce", 0)) + 1


_init_state()
# ═══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
def _render_sidebar(df: Optional[pd.DataFrame]):
    api_key = _get_api_key()
    filter_state = {"filtered_df": df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(), "active_items": [], "filter_seconds": 0.0, "description": "No active filters"}
    with st.sidebar:
        st.markdown("### 🤖 AI Model")
        cur_model = st.session_state.get("shared_model", DEFAULT_MODEL)
        if cur_model not in MODEL_OPTIONS:
            cur_model = DEFAULT_MODEL
            st.session_state["shared_model"] = cur_model
        st.selectbox("Model", options=MODEL_OPTIONS, index=MODEL_OPTIONS.index(cur_model), key="shared_model", help="Used by AI Analyst, Review Prompt, and Symptomizer.")
        effort_options = _reasoning_options_for_model(st.session_state.get("shared_model", DEFAULT_MODEL))
        cur_reasoning = _safe_text(st.session_state.get("shared_reasoning", DEFAULT_REASONING)).lower() or DEFAULT_REASONING
        if cur_reasoning not in effort_options:
            cur_reasoning = "none" if "none" in effort_options else effort_options[0]
            st.session_state["shared_reasoning"] = cur_reasoning
        st.selectbox("Reasoning effort", options=effort_options, index=effort_options.index(cur_reasoning), key="shared_reasoning", help="Applied to GPT-5 family models.")
        if api_key:
            st.markdown("<div class='chip green' style='margin-top:4px'>✓ API key loaded</div>", unsafe_allow_html=True)
        else:
            st.warning("Add OPENAI_API_KEY to Streamlit secrets.")

        st.divider()
        st.markdown("### 🔍 Review Filters")
        st.caption("Applies live to every workspace tab.")
        if st.button("🧹 Clear all filters", use_container_width=True, key="rf_clear_btn"):
            _reset_review_filters()
            st.rerun()
        if df is None:
            st.info("Build a workspace to unlock filters.")
        else:
            with st.expander("🗓️ Timeframe", expanded=False):
                tf_opts = ["All Time", "Last Week", "Last Month", "Last Year", "Custom Range"]
                if st.session_state.get("rf_tf") not in tf_opts:
                    st.session_state["rf_tf"] = "All Time"
                st.selectbox("Select timeframe", options=tf_opts, key="rf_tf")
                if st.session_state.get("rf_tf") == "Custom Range":
                    today = date.today()
                    rng = st.session_state.get("rf_tf_range", (today - timedelta(days=30), today))
                    if not (isinstance(rng, (tuple, list)) and len(rng) == 2):
                        rng = (today - timedelta(days=30), today)
                    st.session_state["rf_tf_range"] = tuple(rng)
                    st.date_input("Start / end", value=st.session_state["rf_tf_range"], key="rf_tf_range")

            with st.expander("⭐ Star rating", expanded=False):
                sr_opts = ["All", 5, 4, 3, 2, 1]
                cur = st.session_state.get("rf_sr", ["All"])
                if not isinstance(cur, list):
                    cur = [cur]
                cur = [v for v in cur if v in sr_opts]
                if not cur:
                    cur = ["All"]
                if "All" in cur and len(cur) > 1:
                    cur = [v for v in cur if v != "All"]
                st.session_state["rf_sr"] = cur
                st.multiselect("Select stars", options=sr_opts, default=st.session_state["rf_sr"], key="rf_sr")

            with st.expander("🧭 Core review filters", expanded=True):
                for spec in _core_filter_specs_for_df(df):
                    key = f"rf_{spec['key']}"
                    _sanitize_multiselect(key, spec["options"], ["ALL"])
                    st.multiselect(spec["label"], options=spec["options"], default=st.session_state[key], key=key)

            det_opts, del_opts, _, _ = _symptom_filter_options(df)
            if det_opts or del_opts:
                with st.expander("🩺 Symptom filters", expanded=False):
                    if det_opts:
                        det_all = ["All"] + det_opts
                        _sanitize_multiselect_sym("rf_sym_detract", det_all, ["All"])
                        st.multiselect("Detractors", options=det_all, default=st.session_state["rf_sym_detract"], key="rf_sym_detract")
                    if del_opts:
                        del_all = ["All"] + del_opts
                        _sanitize_multiselect_sym("rf_sym_delight", del_all, ["All"])
                        st.multiselect("Delighters", options=del_all, default=st.session_state["rf_sym_delight"], key="rf_sym_delight")

            with st.expander("🔎 Keyword", expanded=False):
                st.text_input("Search in title + review text", value=st.session_state.get("rf_kw", ""), key="rf_kw")

            extra_candidates = _extra_filter_candidates(df)
            current_extra = [c for c in (st.session_state.get("rf_extra_filter_cols", []) or []) if c in extra_candidates]
            st.session_state["rf_extra_filter_cols"] = current_extra
            with st.expander("➕ Add Filters (power user)", expanded=False):
                st.caption("Choose additional columns to surface as filters.")
                st.multiselect("Available columns", options=extra_candidates, default=current_extra, key="rf_extra_filter_cols")

            extra_cols = st.session_state.get("rf_extra_filter_cols", []) or []
            if extra_cols:
                with st.expander("🧩 Extra filters", expanded=True):
                    for col in extra_cols:
                        if col not in df.columns:
                            continue
                        kind = _infer_extra_filter_kind(df, col)
                        s = df[col]
                        if kind == "numeric":
                            num = pd.to_numeric(s, errors="coerce").dropna()
                            if num.empty:
                                continue
                            lo, hi = float(num.min()), float(num.max())
                            if lo == hi:
                                st.caption(f"{col}: {lo:g} (constant)")
                                continue
                            key = f"rf_{col}_range"
                            default = st.session_state.get(key, (lo, hi))
                            if not (isinstance(default, (tuple, list)) and len(default) == 2):
                                default = (lo, hi)
                            st.session_state[key] = (float(default[0]), float(default[1]))
                            st.slider(col, min_value=lo, max_value=hi, value=st.session_state[key], key=key)
                        elif kind == "date":
                            dt = pd.to_datetime(s, errors="coerce").dropna()
                            if dt.empty:
                                continue
                            lo, hi = dt.min().date(), dt.max().date()
                            key = f"rf_{col}_date_range"
                            default = st.session_state.get(key, (lo, hi))
                            if not (isinstance(default, (tuple, list)) and len(default) == 2):
                                default = (lo, hi)
                            st.session_state[key] = tuple(default)
                            st.date_input(col, value=st.session_state[key], min_value=lo, max_value=hi, key=key)
                        else:
                            try:
                                nunique = int(s.astype("string").replace({"": pd.NA}).nunique(dropna=True))
                            except Exception:
                                nunique = 0
                            if nunique > 600:
                                st.text_input(f"{col} contains", value=str(st.session_state.get(f"rf_{col}_contains") or ""), key=f"rf_{col}_contains", help="High-cardinality column — using a contains filter for speed.")
                            else:
                                opts = _col_options(df, col, max_vals=None)
                                _sanitize_multiselect(f"rf_{col}", opts, ["ALL"])
                                st.multiselect(col, options=opts, default=st.session_state[f"rf_{col}"], key=f"rf_{col}")

            filter_state = _apply_live_review_filters(df)

        st.divider()
        st.markdown("### ⚡ Symptomizer")
        st.slider("Batch size", 1, 12, key="sym_batch_size")
        st.slider("Max evidence chars", 60, 200, step=10, key="sym_max_ev_chars")
    return {"api_key": api_key, "review_filters": filter_state}

# ═══════════════════════════════════════════════════════════════════════════════
#  RENDER HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def _render_metric_card(label, value, subtext, accent=False):
    cls = "metric-card accent" if accent else "metric-card"
    st.markdown(f"""<div class="{cls}">
      <div class="metric-label">{label}</div>
      <div class="metric-value">{value}</div>
      <div class="metric-sub">{subtext}</div>
    </div>""", unsafe_allow_html=True)


def _render_workspace_header(summary, overall_df, prompt_artifacts, *, source_type, source_label):
    bundle = _get_master_bundle(summary, overall_df, prompt_artifacts)
    product_name = _product_name(summary, overall_df)
    organic = int((~overall_df["incentivized_review"].fillna(False)).sum()) if not overall_df.empty else 0
    n = len(overall_df)
    src_chip = f"Uploaded · {source_label}" if source_type == "uploaded" else f"Bazaarvoice · {summary.product_id}"
    st.markdown(f"""<div class="hero-card">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:16px;flex-wrap:wrap;">
        <div>
          <div class="hero-kicker">Review workspace</div>
          <div class="hero-title">{_esc(product_name)}</div>
        </div>
        <div class="badge-row">
          <span class="chip gray">{_esc(src_chip)}</span>
          <span class="chip indigo">{n:,} reviews</span>
          <span class="chip green">{organic:,} organic</span>
        </div>
      </div>
    </div>""", unsafe_allow_html=True)
    a0, a1, a2 = st.columns([1.2, 1.2, 4])
    a0.download_button("⬇️ Download reviews", data=bundle["excel_bytes"], file_name=bundle["excel_name"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    if a1.button("🔄 Reset workspace", use_container_width=True):
        _reset_workspace_state(reset_source=True)
        st.rerun()
    a2.caption("Export includes Reviews, Rating Distribution, Volume trend, and any AI prompt or Symptomizer columns.")


def _render_top_metrics(overall_df, filtered_df):
    m = _get_metrics(filtered_df)
    cards = [
        ("Reviews in view", f"{m['review_count']:,}", f"of {len(overall_df):,} loaded", False),
        ("Avg rating", _fmt_num(m["avg_rating"]), "Filtered view", False),
        ("Avg rating · organic", _fmt_num(m["avg_rating_non_incentivized"]), f"{m['non_incentivized_count']:,} organic", False),
        ("% 1-2 star", _fmt_pct(m["pct_low_star"]), f"{m['low_star_count']:,} low-star", True),
        ("% incentivized", _fmt_pct(m["pct_incentivized"]), "Current view", False),
    ]
    cols = st.columns(len(cards))
    for col, (label, value, sub, acc) in zip(cols, cards):
        with col:
            _render_metric_card(label, value, sub, accent=acc)


_REVIEW_REF_PATTERN = re.compile(r"\(review_ids?\s*:\s*([^)]+)\)", flags=re.IGNORECASE)


def _reference_preview_rows(review_ids: Sequence[str], df: pd.DataFrame, max_items: int = 4) -> List[Dict[str, str]]:
    if df is None or df.empty:
        return []
    lookup = df.copy()
    lookup["review_id"] = lookup["review_id"].astype(str)
    out = []
    used = set()
    for rid in review_ids:
        if rid in used:
            continue
        used.add(rid)
        hit = lookup[lookup["review_id"] == str(rid)]
        if hit.empty:
            continue
        row = hit.iloc[0]
        title = _safe_text(row.get("title"), "Untitled review") or "Untitled review"
        snippet = _trunc(_safe_text(row.get("review_text")) or _safe_text(row.get("title_and_text")), 220)
        meta = []
        if pd.notna(row.get("rating")):
            meta.append(f"★{_safe_int(row.get('rating'), 0)}")
        if _safe_text(row.get("submission_date")):
            meta.append(_safe_text(row.get("submission_date")))
        if _safe_text(row.get("content_locale")):
            meta.append(_safe_text(row.get("content_locale")))
        out.append({"meta": " · ".join(meta), "title": title, "snippet": snippet})
        if len(out) >= max_items:
            break
    return out


def _reference_tile_html_from_ids(review_ids: Sequence[str], df: pd.DataFrame, *, label: str = "Reference") -> str:
    ids = [str(x).strip() for x in review_ids if str(x).strip()]
    previews = _reference_preview_rows(ids, df)
    if not previews:
        tip = "<div class='ref-empty'>Referenced review preview not available in the loaded dataset.</div>"
    else:
        bits = []
        for item in previews:
            bits.append(
                "<div class='ref-item'>"
                + (f"<div class='ref-meta'>{_esc(item['meta'])}</div>" if item.get("meta") else "")
                + f"<div class='ref-title'>{_esc(item['title'])}</div>"
                + f"<div class='ref-snippet'>{_esc(item['snippet'])}</div>"
                + "</div>"
            )
        extra = max(0, len(ids) - len(previews))
        if extra:
            bits.append(f"<div class='ref-item'><div class='ref-empty'>+{extra} more referenced review(s)</div></div>")
        tip = "".join(bits)
    return f"<span class='ref-wrap'><span class='ref-tile'>{_esc(label)}</span><span class='ref-tip'>{tip}</span></span>"


def _reference_tile_html_for_row(row) -> str:
    rid = _safe_text(row.get("review_id"))
    title = _safe_text(row.get("title"), "Untitled review") or "Untitled review"
    snippet = _trunc(_safe_text(row.get("review_text")) or _safe_text(row.get("title_and_text")), 220)
    meta = []
    if pd.notna(row.get("rating")):
        meta.append(f"★{_safe_int(row.get('rating'), 0)}")
    if _safe_text(row.get("submission_date")):
        meta.append(_safe_text(row.get("submission_date")))
    if _safe_text(row.get("content_locale")):
        meta.append(_safe_text(row.get("content_locale")))
    if rid:
        meta.append("Loaded review")
    tip = (
        "<div class='ref-item'>"
        + (f"<div class='ref-meta'>{_esc(' · '.join(meta))}</div>" if meta else "")
        + f"<div class='ref-title'>{_esc(title)}</div>"
        + f"<div class='ref-snippet'>{_esc(snippet)}</div>"
        + "</div>"
    )
    return f"<span class='ref-wrap'><span class='ref-tile'>Reference</span><span class='ref-tip'>{tip}</span></span>"


def _replace_review_citations_with_reference_tiles(text: str, df: pd.DataFrame) -> str:
    safe = html.escape(str(text or ""), quote=False)
    def repl(match):
        raw = match.group(1)
        ids = [p.strip() for p in re.split(r"[,;]", raw) if p.strip()]
        return _reference_tile_html_from_ids(ids, df, label="Reference")
    return _REVIEW_REF_PATTERN.sub(repl, safe)


def _render_markdown_with_reference_tiles(text: str, df: pd.DataFrame):
    processed = _replace_review_citations_with_reference_tiles(text, df)
    st.markdown(processed, unsafe_allow_html=True)


def _sort_reviews(df, sort_mode):
    w = df.copy()
    if sort_mode == "Newest":
        return w.sort_values(["submission_time", "review_id"], ascending=[False, False], na_position="last")
    if sort_mode == "Oldest":
        return w.sort_values(["submission_time", "review_id"], ascending=[True, True], na_position="last")
    if sort_mode == "Highest rating":
        return w.sort_values(["rating", "submission_time"], ascending=[False, False], na_position="last")
    if sort_mode == "Lowest rating":
        return w.sort_values(["rating", "submission_time"], ascending=[True, False], na_position="last")
    if sort_mode == "Longest":
        return w.sort_values(["review_length_words", "submission_time"], ascending=[False, False], na_position="last")
    return w


def _highlight_evidence(text, evidence_items):
    text_str = str(text)
    if not evidence_items or not text_str.strip():
        return f"<div class='review-body'>{html.escape(text_str)}</div>"
    hits = []
    for ev_text, tag_label in evidence_items:
        if not ev_text.strip():
            continue
        for m in re.compile(re.escape(ev_text.strip()), re.IGNORECASE).finditer(text_str):
            hits.append((m.start(), m.end(), tag_label, m.group()))
    if not hits:
        return f"<div class='review-body'>{html.escape(text_str)}</div>"
    hits.sort(key=lambda h: h[0])
    deduped = []
    cursor = 0
    for h in hits:
        if h[0] >= cursor:
            deduped.append(h)
            cursor = h[1]
    parts = []
    cursor = 0
    for start, end, tag_label, matched in deduped:
        parts.append(html.escape(text_str[cursor:start]))
        tip = html.escape(f"AI tag: {tag_label}")
        parts.append(f'<span class="ev-highlight" data-tag="{tip}">{html.escape(matched)}</span>')
        cursor = end
    parts.append(html.escape(text_str[cursor:]))
    return f"<div class='review-body'>{''.join(parts)}</div>"


def _build_evidence_lookup(processed_rows):
    lookup = {}
    for rec in processed_rows:
        idx = str(rec.get("idx", ""))
        if not idx:
            continue
        entries = []
        for lab, evs in (rec.get("ev_det", {}) or {}).items():
            for e in (evs or []):
                if e and e.strip():
                    entries.append((e.strip(), lab))
        for lab, evs in (rec.get("ev_del", {}) or {}).items():
            for e in (evs or []):
                if e and e.strip():
                    entries.append((e.strip(), lab))
        if entries:
            lookup[idx] = entries
    return lookup


def _symptom_tags_html(det_tags, del_tags):
    if not det_tags and not del_tags:
        return ""
    sym_html = "<div style='margin-top:9px;padding-top:9px;border-top:1px solid var(--border);display:flex;flex-direction:column;gap:6px;'>"
    if det_tags:
        det_chips = "".join(f"<span class='chip red' style='font-size:11px;padding:3px 8px;'>{_esc(t)}</span>" for t in det_tags)
        sym_html += f"<div style='display:flex;align-items:flex-start;gap:7px;flex-wrap:wrap;'><span style='font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:var(--danger);font-weight:700;white-space:nowrap;padding-top:3px;'>Issues</span><div style='display:flex;gap:4px;flex-wrap:wrap;'>{det_chips}</div></div>"
    if del_tags:
        del_chips = "".join(f"<span class='chip green' style='font-size:11px;padding:3px 8px;'>{_esc(t)}</span>" for t in del_tags)
        sym_html += f"<div style='display:flex;align-items:flex-start;gap:7px;flex-wrap:wrap;'><span style='font-size:10px;text-transform:uppercase;letter-spacing:.07em;color:var(--success);font-weight:700;white-space:nowrap;padding-top:3px;'>Strengths</span><div style='display:flex;gap:4px;flex-wrap:wrap;'>{del_chips}</div></div>"
    sym_html += "</div>"
    return sym_html


def _render_review_card(row, evidence_items=None):
    rating_val = _safe_int(row.get("rating"), 0) if pd.notna(row.get("rating")) else 0
    stars = "★" * max(0, min(rating_val, 5)) + "☆" * max(0, 5 - rating_val)
    title = _safe_text(row.get("title"), "No title") or "No title"
    review_text = _safe_text(row.get("review_text"), "—") or "—"
    meta_bits = [b for b in [_safe_text(row.get("submission_date")), _safe_text(row.get("content_locale")), _safe_text(row.get("retailer")), _safe_text(row.get("product_or_sku"))] if b]
    is_organic = not _safe_bool(row.get("incentivized_review"), False)
    status_chips = f"<span class='chip {'gray' if is_organic else 'yellow'}'>{'Organic' if is_organic else 'Incentivized'}</span>"
    rec = row.get("is_recommended")
    if not _is_missing(rec):
        status_chips += f"<span class='chip {'gray' if _safe_bool(rec, False) else 'red'}'>{'Recommended' if _safe_bool(rec, False) else 'Not recommended'}</span>"
    det_tags = [str(row.get(f"AI Symptom Detractor {j}", "")) for j in range(1, 11) if _is_filled(row.get(f"AI Symptom Detractor {j}"))]
    del_tags = [str(row.get(f"AI Symptom Delighter {j}", "")) for j in range(1, 11) if _is_filled(row.get(f"AI Symptom Delighter {j}"))]
    with st.container(border=True):
        top_cols = st.columns([5, 1.5])
        with top_cols[0]:
            st.markdown(f"<span style='color:#f59e0b;letter-spacing:-.01em;'>{stars}</span>&nbsp;<span style='font-size:12px;color:var(--slate-500);font-weight:600;'>{rating_val}/5</span>", unsafe_allow_html=True)
            st.markdown(f"<div style='font-weight:700;font-size:14.5px;color:var(--navy);margin:3px 0 2px;'>{_esc(title)}</div>", unsafe_allow_html=True)
            if meta_bits:
                st.markdown(f"<div style='font-size:12px;color:var(--slate-400);margin-bottom:4px;'>{' · '.join(_esc(b) for b in meta_bits)}</div>", unsafe_allow_html=True)
        with top_cols[1]:
            st.markdown(f"<div class='chip-wrap' style='justify-content:flex-end;gap:4px;flex-wrap:wrap;padding-top:2px;'>{status_chips}</div>", unsafe_allow_html=True)
        if evidence_items:
            st.markdown(_highlight_evidence(review_text, evidence_items), unsafe_allow_html=True)
            st.caption("Yellow highlights = Symptomizer evidence · hover to see the AI tag")
        else:
            st.markdown(f"<div class='review-body'>{html.escape(review_text)}</div>", unsafe_allow_html=True)
        tag_html = _symptom_tags_html(det_tags, del_tags)
        if tag_html:
            st.markdown(tag_html, unsafe_allow_html=True)
        loc = _safe_text(row.get("user_location"))
        if loc:
            st.markdown(
                f"<div style='font-size:11.5px;color:var(--slate-400);margin-top:8px;'>{_esc(loc)}</div>",
                unsafe_allow_html=True,
            )
# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
def _render_dashboard(filtered_df, overall_df=None):
    od = overall_df if overall_df is not None else filtered_df
    st.markdown("<div class='section-title'>Dashboard</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>Rating mix, over-time trend, and cohort analytics for the current filter set.</div>", unsafe_allow_html=True)
    sym_state = _detect_symptom_state(od)
    if sym_state == "none":
        st.markdown("""<div class="sym-state-banner" style="padding:1.5rem 1.8rem;text-align:left;display:flex;align-items:center;gap:16px;margin-bottom:.5rem;">
          <div style="font-size:2.2rem;flex-shrink:0;">💊</div>
          <div style="flex:1;">
            <div class="title" style="margin-bottom:4px;font-size:15px;">No symptoms tagged yet</div>
            <div class="sub" style="max-width:none;font-size:12.5px;">Run the Symptomizer to AI-tag delighters &amp; detractors — they'll surface here once complete.</div>
          </div>
        </div>""", unsafe_allow_html=True)
        if st.button("💊 Go to Symptomizer →", type="primary", key="dash_go_sym", use_container_width=False):
            st.session_state["workspace_tab_request"] = TAB_SYMPTOMIZER
            st.rerun()
        st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)

    scope = st.radio("Scope", ["All matching reviews", "Organic only"], horizontal=True, key="dashboard_scope")
    chart_df = filtered_df.copy()
    if scope == "Organic only":
        chart_df = chart_df[~chart_df["incentivized_review"].fillna(False)].reset_index(drop=True)
    if chart_df.empty:
        st.info("No reviews match the current scope.")
        return

    st.markdown("<div style='height:.5rem'></div>", unsafe_allow_html=True)
    _render_reviews_over_time_chart(chart_df)

    rating_df = _rating_dist(chart_df)
    rating_df["rating_label"] = rating_df["rating"].map(lambda v: f"{int(v)}★")
    rating_df["count_pct_label"] = rating_df.apply(lambda r: f"{int(r['review_count']):,} · {_fmt_pct(r['share'])}", axis=1)

    st.markdown("<div style='height:.75rem'></div>", unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        with st.container(border=True):
            fig = px.bar(rating_df, x="rating_label", y="review_count", text="count_pct_label", title="Rating distribution", category_orders={"rating_label": ["1★", "2★", "3★", "4★", "5★"]}, color="rating", color_discrete_map={"1": "#ef4444", "2": "#f97316", "3": "#eab308", "4": "#84cc16", "5": "#22c55e"}, hover_data={"share": ":.1%", "review_count": True})
            fig.update_traces(textposition="outside", cliponaxis=False, showlegend=False)
            fig.update_layout(margin=dict(l=24, r=24, t=52, b=20), xaxis_title="", yaxis_title="", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter")
            st.plotly_chart(fig, use_container_width=True)
    with c2:
        with st.container(border=True):
            cohort_df = _cohort_by_incentivized(chart_df)
            if cohort_df.empty:
                st.info("No cohort data.")
            else:
                fig_c = px.bar(cohort_df, x="star", y="pct", color="cohort", barmode="group", title="Rating split: Organic vs Incentivized", labels={"star": "Star", "pct": "% of cohort", "cohort": "Cohort"}, color_discrete_map={"Organic": "#6366f1", "Incentivized": "#f59e0b"})
                fig_c.update_layout(xaxis=dict(tickmode="array", tickvals=[1, 2, 3, 4, 5], ticktext=["1★", "2★", "3★", "4★", "5★"]), plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter", margin=dict(l=24, r=24, t=52, b=20), legend=dict(orientation="h", y=1.08, x=0))
                fig_c.update_yaxes(ticksuffix="%")
                st.plotly_chart(fig_c, use_container_width=True)

    st.markdown("<div style='height:.75rem'></div>", unsafe_allow_html=True)
    _render_symptom_dashboard(chart_df, od)

    st.markdown("<div style='height:.75rem'></div>", unsafe_allow_html=True)
    st.markdown("<div class='section-title' style='font-size:15px;'>📊 Sentiment & Market</div>", unsafe_allow_html=True)
    sa1, sa2 = st.columns(2)
    with sa1:
        with st.container(border=True):
            sb_df = _star_band_trend(chart_df)
            if sb_df.empty:
                st.info("Insufficient date data for sentiment trend.")
            else:
                fig_sb = go.Figure()
                fig_sb.add_trace(go.Scatter(x=sb_df["month_start"], y=sb_df["pct_low"], name="% 1-2★", mode="lines+markers", line=dict(color="#ef4444", width=2), marker=dict(size=4), fill="tozeroy", fillcolor="rgba(239,68,68,0.08)"))
                fig_sb.add_trace(go.Scatter(x=sb_df["month_start"], y=sb_df["pct_high"], name="% 4-5★", mode="lines+markers", line=dict(color="#22c55e", width=2), marker=dict(size=4)))
                fig_sb.update_layout(title="Sentiment drift: 1-2★ vs 4-5★ over time", hovermode="x unified", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter", margin=dict(l=24, r=24, t=52, b=20), legend=dict(orientation="h", y=1.08, x=0))
                fig_sb.update_yaxes(ticksuffix="%", title="% of monthly reviews")
                st.plotly_chart(fig_sb, use_container_width=True)
    with sa2:
        with st.container(border=True):
            locale_df = _locale_breakdown(chart_df, top_n=10)
            if locale_df.empty:
                st.info("No locale data.")
            else:
                fig_loc = go.Figure()
                fig_loc.add_trace(go.Bar(x=locale_df["count"], y=locale_df["content_locale"], orientation="h", name="Reviews", marker_color="#6366f1", opacity=0.75, hovertemplate="%{y}<br>%{x:,} reviews<extra></extra>"))
                fig_loc.add_trace(go.Scatter(x=locale_df["avg_rating"] * locale_df["count"].max() / 5, y=locale_df["content_locale"], mode="markers", name="Avg ★ (scaled)", marker=dict(color=locale_df["avg_rating"], colorscale="RdYlGn", cmin=1, cmax=5, size=9, showscale=True, colorbar=dict(title="Avg ★", len=0.6, x=1.02)), hovertemplate="%{y}<br>Avg ★: %{text}<extra></extra>", text=[f"{v:.2f}" for v in locale_df["avg_rating"]]))
                fig_loc.update_layout(title="Top markets by review volume", height=max(260, 26 * len(locale_df) + 80), margin=dict(l=80, r=60, t=52, b=20), barmode="overlay", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter", xaxis_title="Reviews", yaxis_title="", legend=dict(orientation="h", y=1.08, x=0))
                st.plotly_chart(fig_loc, use_container_width=True)

    st.markdown("<div style='height:.75rem'></div>", unsafe_allow_html=True)
    rd1, rd2 = st.columns([1.3, 1])
    with rd1:
        with st.container(border=True):
            len_df = _review_length_cohort(chart_df)
            if len_df.empty:
                st.info("Insufficient data for review-length analysis.")
            else:
                fig_len = go.Figure()
                fig_len.add_trace(go.Bar(x=len_df["Length Quartile"], y=len_df["avg_rating"], text=[f"{v:.2f}★" for v in len_df["avg_rating"]], textposition="outside", marker_color=["#ef4444" if v < 3.5 else "#eab308" if v < 4.2 else "#22c55e" for v in len_df["avg_rating"]], hovertemplate="%{x}<br>Avg ★: %{y:.2f}<br>n=%{customdata}<extra></extra>", customdata=len_df["count"]))
                fig_len.update_layout(title="Review depth vs satisfaction", yaxis_range=[1, 5.2], yaxis_title="Avg ★", xaxis_title="", plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)", font_family="Inter", margin=dict(l=24, r=24, t=52, b=20))
                st.plotly_chart(fig_len, use_container_width=True)
    with rd2:
        with st.container(border=True):
            locs = _top_locations(chart_df, top_n=10)
            if locs.empty:
                st.info("No reviewer location data.")
            else:
                st.markdown("<div style='font-weight:700;font-size:13.5px;color:var(--navy);margin-bottom:8px;'>Top reviewer locations</div>", unsafe_allow_html=True)
                locs_display = locs.copy()
                locs_display["avg_rating"] = locs_display["avg_rating"].map(lambda v: f"{v:.2f}★" if pd.notna(v) else "—")
                locs_display = locs_display.rename(columns={"user_location": "Location", "count": "Reviews", "avg_rating": "Avg ★"})
                st.dataframe(locs_display[["Location", "Reviews", "Avg ★"]], use_container_width=True, hide_index=True, height=280)

# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: REVIEW EXPLORER
# ═══════════════════════════════════════════════════════════════════════════════
def _render_review_explorer(*, summary, overall_df, filtered_df, prompt_artifacts):
    st.markdown("<div class='section-title'>Review Explorer</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='section-sub'>Showing <strong>{len(filtered_df):,}</strong> reviews · Use sidebar filters to narrow the stream.</div>", unsafe_allow_html=True)
    bundle = _get_master_bundle(summary, overall_df, prompt_artifacts)
    tc = st.columns([1.3, 1.35, 0.9, 1.1, 0.85])
    tc[0].download_button("⬇️ Download reviews", data=bundle["excel_bytes"], file_name=bundle["excel_name"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="explorer_dl")
    sort_mode = tc[1].selectbox("Sort", ["Newest", "Oldest", "Highest rating", "Lowest rating", "Longest"], key="review_explorer_sort")
    per_page = int(tc[2].selectbox("Per page", [10, 20, 30, 50], key="review_explorer_per_page"))
    tc[3].markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    show_ev = tc[3].toggle("Evidence highlights", value=True, key="re_show_highlights", help="Highlight Symptomizer evidence in yellow — hover to see the AI tag")
    ordered_df = _sort_reviews(filtered_df, sort_mode).reset_index(drop=True)
    if ordered_df.empty:
        st.info("No reviews match the current filters.")
        return
    page_count = max(1, math.ceil(len(ordered_df) / max(per_page, 1)))
    current_page = max(1, min(int(st.session_state.get("review_explorer_page", 1)), page_count))
    start = (current_page - 1) * per_page
    page_df = ordered_df.iloc[start:start + per_page]
    ev_lookup = _build_evidence_lookup(st.session_state.get("sym_processed_rows") or [])
    for orig_idx, row in page_df.iterrows():
        ev_items = (ev_lookup.get(str(orig_idx)) or ev_lookup.get(str(row.get("review_id", "")))) if show_ev else None
        _render_review_card(row, evidence_items=ev_items)
        st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    with st.container(border=True):
        pc = st.columns([0.8, 0.8, 2.9, 0.85, 0.8, 0.8])
        go_first = pc[0].button("⏮", use_container_width=True, disabled=current_page <= 1, key="re_first")
        go_prev = pc[1].button("‹", use_container_width=True, disabled=current_page <= 1, key="re_prev")
        pc[2].markdown(f"<div class='compact-pager-status'>Page {current_page} of {page_count:,}<span class='compact-pager-sub'>{start + 1:,}–{min(start + per_page, len(ordered_df)):,} of {len(ordered_df):,} reviews</span></div>", unsafe_allow_html=True)
        go_page = int(pc[3].number_input("Go", min_value=1, max_value=page_count, value=current_page, step=1, key="re_page_input", label_visibility="collapsed"))
        go_next = pc[4].button("›", use_container_width=True, disabled=current_page >= page_count, key="re_next")
        go_last = pc[5].button("⏭", use_container_width=True, disabled=current_page >= page_count, key="re_last")
    new_page = current_page
    if go_first:
        new_page = 1
    elif go_prev:
        new_page = max(1, current_page - 1)
    elif go_next:
        new_page = min(page_count, current_page + 1)
    elif go_last:
        new_page = page_count
    elif go_page != current_page:
        new_page = go_page
    if new_page != current_page:
        st.session_state["review_explorer_page"] = new_page
        st.rerun()
    else:
        st.session_state["review_explorer_page"] = current_page

# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: AI ANALYST
# ═══════════════════════════════════════════════════════════════════════════════
def _render_ai_tab(*, settings, overall_df, filtered_df, summary, filter_description):
    st.markdown("<div class='section-title'>AI — Product & Consumer Insights</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>Ask anything. Grounded in the currently filtered review text and evidence.</div>", unsafe_allow_html=True)
    if filtered_df.empty:
        st.info("Adjust filters — no reviews in scope.")
        return
    scope_sig = json.dumps(dict(pid=summary.product_id, fd=filter_description, n=len(filtered_df), st=(st.session_state.get("analysis_dataset") or {}).get("source_type", "bv")), sort_keys=True)
    if st.session_state.get("chat_scope_signature") != scope_sig:
        if st.session_state.get("chat_messages"):
            st.session_state["chat_messages"] = []
            st.session_state["chat_scope_notice"] = "Chat cleared — scope changed."
        st.session_state["chat_scope_signature"] = scope_sig
    notice = st.session_state.pop("chat_scope_notice", None)
    if notice:
        st.info(notice)
    if st.session_state.pop("ai_scroll_to_top", False):
        st.markdown("<script>window.scrollTo({top:0,behavior:'smooth'});window.parent.scrollTo({top:0,behavior:'smooth'});</script>", unsafe_allow_html=True)
    with st.container(border=True):
        sc = st.columns([1, 1, 1, 2])
        sc[0].metric("In scope", f"{len(filtered_df):,}")
        sc[1].metric("Organic", f"{int((~filtered_df['incentivized_review'].fillna(False)).sum()):,}")
        sc[2].metric("Model", _shared_model())
        sc[3].caption(f"Scope: {filter_description}")
    api_key = settings.get("api_key")
    if not api_key:
        st.warning("Add OPENAI_API_KEY to Streamlit secrets.")
        st.code('OPENAI_API_KEY = "sk-..."', language="toml")
        return
    archive_msgs, live_msgs = _split_chat_messages(st.session_state.get("chat_messages") or [], keep_last=AI_VISIBLE_CHAT_MESSAGES)
    with st.container(border=True):
        st.markdown("**Current exchange**")
        if not live_msgs:
            st.info("Start with a quick report below, or type a question.")
        else:
            for msg in live_msgs:
                with st.chat_message(msg["role"]):
                    if msg["role"] == "assistant":
                        _render_markdown_with_reference_tiles(msg["content"], overall_df)
                    else:
                        st.markdown(msg["content"])
    if archive_msgs:
        msg_label = "message" if len(archive_msgs) == 1 else "messages"
        with st.expander(f"🗂️ Chat archive ({len(archive_msgs)} earlier {msg_label})", expanded=False):
            for msg in archive_msgs:
                with st.chat_message(msg["role"]):
                    if msg["role"] == "assistant":
                        _render_markdown_with_reference_tiles(msg["content"], overall_df)
                    else:
                        st.markdown(msg["content"])
    quick_actions = {
        "Executive summary": dict(prompt="Create a concise executive summary. Lead with biggest strengths, biggest risks, key consumer insight, and top 3 actions.", help="Leadership readout.", persona=None),
        "Product Development": dict(prompt=PERSONAS["Product Development"]["prompt"], help=PERSONAS["Product Development"]["blurb"], persona="Product Development"),
        "Quality Engineer": dict(prompt=PERSONAS["Quality Engineer"]["prompt"], help=PERSONAS["Quality Engineer"]["blurb"], persona="Quality Engineer"),
        "Consumer Insights": dict(prompt=PERSONAS["Consumer Insights"]["prompt"], help=PERSONAS["Consumer Insights"]["blurb"], persona="Consumer Insights"),
    }
    quick_trigger = None
    with st.container(border=True):
        st.markdown("**Quick reports**")
        acols = st.columns(4)
        for col, (label, config) in zip(acols, quick_actions.items()):
            if col.button(label, use_container_width=True, help=config["help"], key=f"ai_q_{_slugify(label)}"):
                quick_trigger = (config["persona"], label, config["prompt"])
        lc1, lc2 = st.columns([2, 2])
        lc1.radio("Report length", ["Short", "Medium", "Long"], horizontal=True, index=1, key="ai_report_length", help="Short ≈150 words · Medium ≈350 words · Long ≈700 words")
        lc2.caption("Each report is grounded in the filtered review text and cites review IDs for material claims.")
    helper_cols = st.columns([2, 1, 1])
    helper_cols[0].caption(f"Scope: {filter_description}")
    if helper_cols[1].button("Clear chat", use_container_width=True, key="ai_clear_chat"):
        st.session_state["chat_messages"] = []
        st.session_state["ai_scroll_to_top"] = False
        st.rerun()
    user_message = st.chat_input("Ask about drivers, risks, opportunities, or voice-of-customer themes…", key="ai_chat_input")
    prompt_to_send = visible_user_message = persona_name = None
    if quick_trigger:
        persona_name, visible_user_message, prompt_to_send = quick_trigger
    elif user_message:
        prompt_to_send = visible_user_message = user_message
    if prompt_to_send and visible_user_message:
        prior = list(st.session_state["chat_messages"])
        st.session_state["chat_messages"].append({"role": "user", "content": visible_user_message})
        overlay = _show_thinking("Reviewing the filtered review text…")
        try:
            answer = _call_analyst(question=prompt_to_send, overall_df=overall_df, filtered_df=filtered_df, summary=summary, filter_description=filter_description, chat_history=prior, persona_name=persona_name, report_length=st.session_state.get("ai_report_length", "Medium"))
            if persona_name:
                answer = f"## {persona_name} report\n\n{answer}"
        except Exception as exc:
            answer = f"OpenAI request failed: {exc}"
        finally:
            overlay.empty()
        st.session_state["chat_messages"].append({"role": "assistant", "content": answer})
        st.session_state["ai_scroll_to_top"] = True
        st.rerun()
# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: REVIEW PROMPT
# ═══════════════════════════════════════════════════════════════════════════════
def _render_review_prompt_tab(*, settings, overall_df, filtered_df, summary, filter_description):
    st.markdown("<div class='section-title'>Review Prompt</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>Create row-level AI tags that become new review columns.</div>", unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown("**Prompt library**")
        sc = st.columns([1.2, 1.2, 1])
        if sc[0].button("Add starter pack", use_container_width=True, key="prompt_add"):
            new_rows = pd.DataFrame(REVIEW_PROMPT_STARTER_ROWS)
            existing = set(st.session_state["prompt_definitions_df"]["column_name"].astype(str))
            to_add = new_rows[~new_rows["column_name"].isin(existing)]
            if not to_add.empty:
                st.session_state["prompt_definitions_df"] = pd.concat([st.session_state["prompt_definitions_df"], to_add], ignore_index=True)
            st.rerun()
        if sc[1].button("Reset to starter", use_container_width=True, key="prompt_reset"):
            st.session_state["prompt_definitions_df"] = pd.DataFrame(REVIEW_PROMPT_STARTER_ROWS)
            st.rerun()
        if sc[2].button("Clear all", use_container_width=True, key="prompt_clear"):
            st.session_state["prompt_definitions_df"] = pd.DataFrame(columns=["column_name", "prompt", "labels"])
            st.rerun()
    st.markdown("#### Prompt definitions")
    edited_df = st.data_editor(
        st.session_state["prompt_definitions_df"],
        use_container_width=True,
        num_rows="dynamic",
        hide_index=True,
        key="prompt_def_editor",
        height=320,
        column_config={
            "column_name": st.column_config.TextColumn("Column name", width="medium"),
            "prompt": st.column_config.TextColumn("Prompt", width="large"),
            "labels": st.column_config.TextColumn("Labels (comma-separated)", width="large"),
        },
    )
    st.session_state["prompt_definitions_df"] = edited_df
    _pd_sig = edited_df.to_json() if not edited_df.empty else ""
    _cached_defs = st.session_state.get("_prompt_defs_cache", {})
    if _cached_defs.get("sig") == _pd_sig and _cached_defs.get("cols") == list(overall_df.columns):
        prompt_defs = _cached_defs["defs"]
        prompt_defs_error = _cached_defs.get("error")
    else:
        prompt_defs_error = None
        try:
            prompt_defs = _normalize_prompt_defs(edited_df, overall_df.columns)
        except ReviewDownloaderError as exc:
            prompt_defs = []
            prompt_defs_error = str(exc)
        st.session_state["_prompt_defs_cache"] = dict(sig=_pd_sig, cols=list(overall_df.columns), defs=prompt_defs, error=prompt_defs_error)
    if prompt_defs_error:
        st.error(prompt_defs_error)
    api_key = settings.get("api_key")
    client = _get_client()
    with st.container(border=True):
        sc = st.columns([1.25, 1, 1, 2.45])
        tagging_scope = sc[0].selectbox("Scope", ["Current filtered reviews", "All loaded reviews"], index=0, key="prompt_tagging_scope")
        scope_df = filtered_df if tagging_scope == "Current filtered reviews" else overall_df
        batch_size = int(st.session_state.get("sym_batch_size", 5))
        est = math.ceil(len(scope_df) / max(1, batch_size)) if len(scope_df) else 0
        sc[1].metric("Reviews", f"{len(scope_df):,}")
        sc[2].metric("Requests", f"{est:,}")
        sc[3].caption(f"Scope: {tagging_scope.lower()} · {filter_description}")
        run_disabled = (not api_key) or (not prompt_defs) or len(scope_df) == 0
        if st.button("▶️ Run Review Prompt", type="primary", use_container_width=True, disabled=run_disabled, key="prompt_run_btn"):
            overlay = _show_thinking("Classifying each review…")
            try:
                prd = _run_review_prompt_tagging(client=client, source_df=scope_df.reset_index(drop=True), prompt_defs=prompt_defs, chunk_size=batch_size)
                updated = _merge_prompt_results(overall_df, prd, prompt_defs)
                dataset = dict(st.session_state["analysis_dataset"])
                dataset["reviews_df"] = updated
                st.session_state["analysis_dataset"] = dataset
                summary_df = _summarize_prompt_results(prd, prompt_defs, source_df=scope_df)
                defsig = json.dumps([dict(col=p["column_name"], prompt=p["prompt"], labels=p["labels"]) for p in prompt_defs], sort_keys=True)
                st.session_state["prompt_run_artifacts"] = dict(definitions=prompt_defs, summary_df=summary_df, scope_label=tagging_scope, scope_filter_description=filter_description, scope_review_ids=list(prd["review_id"].astype(str)), definition_signature=defsig, review_count=len(prd), generated_utc=pd.Timestamp.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"))
                st.session_state["master_export_bundle"] = None
                st.session_state["_prompt_bundle_ready"] = False
                st.session_state["prompt_run_notice"] = f"Finished tagging {len(prd):,} reviews."
            except Exception as exc:
                st.error(f"Review Prompt run failed: {exc}")
            finally:
                overlay.empty()
            st.rerun()
    notice = st.session_state.pop("prompt_run_notice", None)
    if notice:
        st.success(notice)
    pa = st.session_state.get("prompt_run_artifacts")
    if not pa:
        st.info("Run Review Prompt to generate AI columns.")
        return
    cur_sig = json.dumps([dict(col=p["column_name"], prompt=p["prompt"], labels=p["labels"]) for p in prompt_defs], sort_keys=True) if prompt_defs else ""
    if cur_sig != pa.get("definition_signature"):
        st.info("Prompt definitions changed — re-run to refresh.")
    updated_overall = st.session_state["analysis_dataset"]["reviews_df"]
    hc = st.columns([1.4, 1.4, 4])
    hc[2].caption(f"Run: {pa.get('generated_utc')} · Scope: {pa.get('scope_label')} · Reviews: {pa.get('review_count'):,}")
    if hc[0].button("🔄 Prepare download", use_container_width=True, key="prompt_prep_dl"):
        with st.spinner("Building export…"):
            _get_master_bundle(summary, updated_overall, pa)
        st.session_state["_prompt_bundle_ready"] = True
        st.rerun()
    bundle = st.session_state.get("master_export_bundle")
    dl_ready = bundle is not None
    hc[1].download_button("⬇️ Download tagged file", data=bundle["excel_bytes"] if dl_ready else b"", file_name=bundle["excel_name"] if dl_ready else "tagged.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", disabled=not dl_ready, key="prompt_dl_btn")
    if not dl_ready:
        st.caption("Click **Prepare download** first to build the export file.")
    plookup = {p["display_name"]: p for p in pa["definitions"]}
    pnames = list(plookup.keys())
    if not pnames:
        st.info("No prompt results yet.")
        return
    if st.session_state.get("prompt_result_view") not in pnames:
        st.session_state["prompt_result_view"] = pnames[0]
    sel = st.radio("Prompt result view", options=pnames, horizontal=True, key="prompt_result_view", label_visibility="collapsed")
    prompt = plookup[sel]
    pc_col = prompt["column_name"]
    rids = set(str(x) for x in pa.get("scope_review_ids", []))
    result_scope = updated_overall.loc[updated_overall["review_id"].astype(str).isin(rids)] if rids else updated_overall.iloc[0:0]
    lopts = [str(l) for l in pa["summary_df"][pa["summary_df"]["column_name"] == pc_col]["label"].tolist()]
    sel_labels = st.multiselect("Labels", options=lopts, default=lopts, key=f"plabels_{pc_col}")
    if pc_col in result_scope.columns and not result_scope.empty:
        _slab = result_scope[pc_col]
        _view = result_scope[_slab.isin(sel_labels)] if sel_labels else result_scope.iloc[0:0]
        _vc = _slab.value_counts() if not result_scope.empty else pd.Series(dtype=int)
        _ar = result_scope.groupby(pc_col)["rating"].mean() if "rating" in result_scope.columns and not result_scope.empty else pd.Series(dtype=float)
    else:
        _view = result_scope.iloc[0:0]
        _vc = pd.Series(dtype=int)
        _ar = pd.Series(dtype=float)
    total = max(len(result_scope), 1)
    ps_rows = [dict(label=l, review_count=int(_vc.get(l, 0)), share=_safe_pct(int(_vc.get(l, 0)), total), avg_rating=float(_ar[l]) if l in _ar.index and pd.notna(_ar[l]) else None) for l in prompt["labels"]]
    ps = pd.DataFrame(ps_rows)
    cc, tc_col = st.columns([1.45, 1.05])
    with cc:
        with st.container(border=True):
            if ps.empty or ps["review_count"].sum() == 0:
                st.info("No tagged reviews match current filters.")
            else:
                fig = px.pie(ps, names="label", values="review_count", hole=0.44, color_discrete_sequence=["#6366f1", "#10b981", "#f59e0b", "#ef4444", "#3b82f6", "#8b5cf6"])
                fig.update_layout(margin=dict(l=20, r=20, t=20, b=20), paper_bgcolor="rgba(0,0,0,0)", font_family="Inter")
                st.plotly_chart(fig, use_container_width=True)
    with tc_col:
        with st.container(border=True):
            st.markdown(f"**Column** `{pc_col}`")
            st.write(prompt["prompt"])
            if not ps.empty:
                ds = ps.copy()
                ds["avg_rating"] = ds["avg_rating"].map(lambda x: f"{x:.2f}★" if pd.notna(x) and x is not None else "—")
                ds["share"] = ds["share"].map(_fmt_pct)
                st.dataframe(ds[["label", "review_count", "avg_rating", "share"]], use_container_width=True, hide_index=True, height=240)
    prevcols = [c for c in ["review_id", "rating", "incentivized_review", "submission_time", "content_locale", "title", "review_text", pc_col] if c in _view.columns]
    st.markdown("**Tagged review preview**")
    st.dataframe(_view[prevcols].head(50), use_container_width=True, hide_index=True, height=300)

# ═══════════════════════════════════════════════════════════════════════════════
#  TAB: SYMPTOMIZER
# ═══════════════════════════════════════════════════════════════════════════════
def _render_symptomizer_tab(*, settings, overall_df, filtered_df, summary, filter_description):
    st.markdown("<div class='section-title'>Symptomizer</div>", unsafe_allow_html=True)
    st.markdown("<div class='section-sub'>Row-level AI tagging of delighters and detractors. Tags write back into the shared dataframe and appear in Review Explorer.</div>", unsafe_allow_html=True)
    client = _get_client()
    api_key = settings.get("api_key")
    sym_source = st.session_state.get("sym_symptoms_source", "none")
    if sym_source == "none":
        _try_load_symptoms_from_file()
    sym_source = st.session_state.get("sym_symptoms_source", "none")
    delighters = list(st.session_state.get("sym_delighters") or [])
    detractors = list(st.session_state.get("sym_detractors") or [])
    st.markdown("### 1 · Symptoms catalog")
    if not delighters and not detractors:
        st.warning("⚠️  No symptoms defined yet. Use the tabs below, or proceed and the AI will use built-in product knowledge.")
    else:
        st.markdown(_chip_html([(f"✓ {len(delighters)} delighters", "green"), (f"✓ {len(detractors)} detractors", "red"), (f"Source: {sym_source}", "indigo")]), unsafe_allow_html=True)
    sym_tabs = st.tabs(["🤖  AI builder", "✏️  Manual entry", "📄  Upload workbook"])
    with sym_tabs[0]:
        if not api_key:
            st.warning("OpenAI API key required.")
        else:
            pdesc = st.text_area("Product description", value=st.session_state.get("sym_product_profile", ""), placeholder="e.g. SharkNinja Ninja Air Fryer XL — 6-in-1 countertop air fryer with 6 qt basket", height=80, key="sym_pdesc")
            if not overall_df.empty and "review_text" in overall_df.columns:
                max_samples = min(200, max(5, len(overall_df)))
                sample_n = st.slider("Sample reviews", min_value=5, max_value=max_samples, value=min(20, max_samples), step=5, key="sym_sample_n")
                st.caption(f"Using {sample_n} of {len(overall_df):,} reviews.")
            else:
                sample_n = 20
            if st.button("🤖 Generate symptom list", type="primary", use_container_width=True, key="sym_ai_build"):
                overlay = _show_thinking("Generating symptom catalog…")
                try:
                    samples = overall_df["review_text"].fillna("").astype(str).head(int(sample_n)).tolist() if not overall_df.empty else []
                    result = _ai_build_symptom_list(client=client, product_description=pdesc, sample_reviews=samples)
                    st.session_state["sym_ai_build_result"] = result
                    st.session_state["sym_product_profile"] = pdesc
                except Exception as exc:
                    st.error(f"AI builder failed: {exc}")
                finally:
                    overlay.empty()
                st.rerun()
            ai_result = st.session_state.get("sym_ai_build_result")
            if ai_result:
                st.markdown("**Review and accept:**")
                r1, r2 = st.columns(2)
                with r1:
                    st.markdown("🟢 Delighters")
                    ai_del = st.text_area("Edit", value="\n".join(ai_result.get("delighters", [])), height=180, key="sym_ai_del_edit")
                with r2:
                    st.markdown("🔴 Detractors")
                    ai_det = st.text_area("Edit", value="\n".join(ai_result.get("detractors", [])), height=180, key="sym_ai_det_edit")
                if st.button("✅ Accept", type="primary", use_container_width=True, key="sym_accept_ai"):
                    def _parse(t):
                        return [i.strip() for i in re.split(r"[\n,;|]+", t) if i.strip()]
                    st.session_state.update(sym_delighters=_parse(ai_del), sym_detractors=_parse(ai_det), sym_symptoms_source="ai")
                    st.session_state.pop("sym_ai_build_result", None)
                    st.success("Accepted.")
                    st.rerun()
    with sym_tabs[1]:
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("🟢 **Delighters**")
            del_text = st.text_area("One per line or comma-separated", value="\n".join(delighters), height=200, key="sym_del_manual")
        with c2:
            st.markdown("🔴 **Detractors**")
            det_text = st.text_area("One per line or comma-separated", value="\n".join(detractors), height=200, key="sym_det_manual")
        if st.button("💾 Save symptoms", use_container_width=True, key="sym_save_manual"):
            def _parse(t):
                return [i.strip() for i in re.split(r"[\n,;|]+", t) if i.strip()]
            st.session_state.update(sym_delighters=_parse(del_text), sym_detractors=_parse(det_text), sym_symptoms_source="manual")
            st.success("Saved.")
            st.rerun()
    with sym_tabs[2]:
        st.markdown("Upload an Excel workbook with a **Symptoms** sheet: columns Symptom, Type (Delighter/Detractor), optional Aliases.")
        sym_upload = st.file_uploader("Upload workbook", type=["xlsx"], key="sym_file_uploader")
        if sym_upload:
            raw = sym_upload.getvalue()
            st.session_state["_uploaded_raw_bytes"] = raw
            d, t, a = _get_symptom_whitelists(raw)
            if d or t:
                st.session_state.update(sym_delighters=d, sym_detractors=t, sym_aliases=a, sym_symptoms_source="file")
                st.success(f"Loaded {len(d)} delighters and {len(t)} detractors.")
                st.rerun()
            else:
                st.error("No 'Symptoms' sheet found or it was empty.")
    st.divider()
    st.markdown("### 2 · Configure and run")
    delighters = list(st.session_state.get("sym_delighters") or [])
    detractors = list(st.session_state.get("sym_detractors") or [])
    colmap = _detect_sym_cols(overall_df)
    work = _detect_missing(overall_df, colmap)
    need_both = int(work["Needs_Symptomization"].sum())
    need_del = int(work["Needs_Delighters"].sum())
    need_det = int(work["Needs_Detractors"].sum())
    st.markdown(f"""<div class="hero-grid" style="grid-template-columns:repeat(4,minmax(0,1fr));margin-top:0;margin-bottom:.8rem;">
      <div class="hero-stat"><div class="label">Total reviews</div><div class="value">{len(overall_df):,}</div></div>
      <div class="hero-stat"><div class="label">Need delighters</div><div class="value">{need_del:,}</div></div>
      <div class="hero-stat"><div class="label">Need detractors</div><div class="value">{need_det:,}</div></div>
      <div class="hero-stat accent"><div class="label">Missing both</div><div class="value">{need_both:,}</div></div>
    </div>""", unsafe_allow_html=True)
    scope_choice = st.selectbox("Scope", ["Missing both", "Any missing", "Current filtered reviews", "All loaded reviews"], key="sym_scope_choice")
    if scope_choice == "Missing both":
        target_df = work[(work["Needs_Delighters"]) & (work["Needs_Detractors"])]
    elif scope_choice == "Any missing":
        target_df = work[(work["Needs_Delighters"]) | (work["Needs_Detractors"])]
    elif scope_choice == "Current filtered reviews":
        fids = set(filtered_df["review_id"].astype(str))
        target_df = work[work["review_id"].astype(str).isin(fids)]
    else:
        target_df = work
    rc = st.columns([1.5, 1, 1, 1, 1])
    n_to_process = rc[0].number_input("Reviews to process", min_value=1, max_value=max(1, len(target_df)), step=1, key="sym_n_to_process")
    batch_size = int(rc[1].number_input("Batch size", min_value=1, max_value=20, value=int(st.session_state.get("sym_batch_size", 5)), step=1, key="sym_batch_size_run"))
    est_batches = max(1, math.ceil(int(n_to_process) / batch_size)) if n_to_process else 0
    rc[2].metric("In scope", f"{len(target_df):,}")
    rc[3].metric("Est. batches", f"{est_batches:,}")
    rc[4].caption(f"Scope: {scope_choice}\nModel: {_shared_model()}")
    run_disabled = (not api_key) or (len(target_df) == 0)
    if run_disabled and not api_key:
        st.warning("Add OPENAI_API_KEY to Streamlit secrets.")
    elif run_disabled:
        st.info("No reviews match the current scope.")
    run_btn = st.button(f"▶️ Symptomize {min(int(n_to_process), len(target_df)):,} review(s)", type="primary", use_container_width=True, disabled=run_disabled, key="sym_run_btn")
    notice = st.session_state.pop("sym_run_notice", None)
    if notice:
        st.success(notice)
    if run_btn:
        prioritized = _prioritize_for_symptomization(target_df.head(int(n_to_process)))
        rows_to_process = prioritized.copy()
        prog = st.progress(0.0, text="Starting…")
        status = st.empty()
        eta_box = st.empty()
        stats_box = st.empty()
        processed_local = []
        t0 = time.perf_counter()
        total_n = max(1, len(rows_to_process))
        done = 0
        failed_count = 0
        total_labels_written = 0
        updated_df = _ensure_ai_cols(overall_df.copy())
        profile = st.session_state.get("sym_product_profile", "")
        aliases = st.session_state.get("sym_aliases", {})
        rows_list = list(rows_to_process.iterrows())
        bidxs = list(range(0, len(rows_list), batch_size))
        empty_out = dict(dels=[], dets=[], ev_del={}, ev_det={}, unl_dels=[], unl_dets=[], safety="Not Mentioned", reliability="Not Mentioned", sessions="Unknown")
        _active_dets = detractors or DEFAULT_PRIORITY_DETRACTORS
        _active_dels = delighters or DEFAULT_PRIORITY_DELIGHTERS
        _ev_chars = int(st.session_state.get("sym_max_ev_chars", 120))
        for bi, start in enumerate(bidxs, 1):
            batch = rows_list[start:start + batch_size]
            items = [dict(idx=int(idx), review=_clean_text(row.get("review_text", "") or row.get("title_and_text", "")), needs_del=bool(row.get("Needs_Delighters", True)), needs_det=bool(row.get("Needs_Detractors", True))) for idx, row in batch]
            status.info(f"Batch {bi}/{len(bidxs)} — reviews {start + 1}–{min(start + batch_size, len(rows_list))}")
            outs = {}
            if client:
                try:
                    outs = _call_symptomizer_batch(client=client, items=items, allowed_delighters=_active_dels, allowed_detractors=_active_dets, product_profile=profile, max_ev_chars=_ev_chars, aliases=aliases)
                except Exception as exc:
                    status.warning(f"Batch {bi} failed ({exc}) — retrying individually…")
                    failed_count += len(items)
                    for it in items:
                        try:
                            single = _call_symptomizer_batch(client=client, items=[it], allowed_delighters=_active_dels, allowed_detractors=_active_dets, product_profile=profile, max_ev_chars=_ev_chars, aliases=aliases)
                            outs.update(single)
                            failed_count -= 1
                        except Exception:
                            pass
            for it in items:
                idx = int(it["idx"])
                out = outs.get(idx, empty_out)
                dets_out = list(out.get("dets", []))[:10]
                dels_out = list(out.get("dels", []))[:10]
                for j, lab in enumerate(dets_out):
                    updated_df.loc[idx, f"AI Symptom Detractor {j + 1}"] = lab
                for j, lab in enumerate(dels_out):
                    updated_df.loc[idx, f"AI Symptom Delighter {j + 1}"] = lab
                updated_df.loc[idx, "AI Safety"] = out.get("safety", "Not Mentioned")
                updated_df.loc[idx, "AI Reliability"] = out.get("reliability", "Not Mentioned")
                updated_df.loc[idx, "AI # of Sessions"] = out.get("sessions", "Unknown")
                total_labels_written += len(dets_out) + len(dels_out)
                for lab in (out.get("unl_dels", []) or []) + (out.get("unl_dets", []) or []):
                    lab = lab.strip()
                    if lab:
                        rec = st.session_state["sym_new_candidates"].setdefault(lab, {"count": 0, "refs": []})
                        rec["count"] += 1
                        if len(rec["refs"]) < 50:
                            rec["refs"].append(idx)
                processed_local.append(dict(idx=idx, wrote_dets=dets_out, wrote_dels=dels_out, safety=out.get("safety", ""), reliability=out.get("reliability", ""), sessions=out.get("sessions", ""), ev_det=out.get("ev_det", {}), ev_del=out.get("ev_del", {}), unl_dels=out.get("unl_dels", []), unl_dets=out.get("unl_dets", [])))
                done += 1
            dataset_ck = dict(st.session_state["analysis_dataset"])
            dataset_ck["reviews_df"] = updated_df.copy()
            st.session_state["analysis_dataset"] = dataset_ck
            st.session_state["sym_processed_rows"] = list(processed_local)
            elapsed = time.perf_counter() - t0
            rate = done / elapsed if elapsed > 0 else 0
            rem = (total_n - done) / rate if rate > 0 else 0
            prog.progress(done / total_n, text=f"{done}/{total_n} processed")
            eta_box.markdown(f"**Speed:** {rate * 60:.1f} rev/min · **ETA:** ~{_fmt_secs(rem)}")
            avg_labels = total_labels_written / max(done, 1)
            stats_box.markdown(f"<span style='font-size:12px;color:var(--slate-500);'>Labels written: **{total_labels_written}** · Avg per review: **{avg_labels:.1f}**" + (f" · ⚠️ {failed_count} failed" if failed_count > 0 else "") + "</span>", unsafe_allow_html=True)
            gc.collect()
        dataset = dict(st.session_state["analysis_dataset"])
        dataset["reviews_df"] = updated_df
        st.session_state.update(analysis_dataset=dataset, sym_processed_rows=processed_local, master_export_bundle=None)
        status.success(f"✅ Symptomized {done:,} reviews — {total_labels_written} labels written ({total_labels_written / max(done, 1):.1f} avg/review).")
        st.session_state["sym_run_notice"] = f"Symptomized {done:,} reviews · {total_labels_written} labels. Tags visible in Review Explorer."
        st.rerun()
    st.divider()
    processed = st.session_state.get("sym_processed_rows") or []
    if not processed:
        st.info("Run the Symptomizer above to see results here.")
        return
    st.markdown("### 3 · Results")
    total_tags = sum(len(r.get("wrote_dets", [])) + len(r.get("wrote_dels", [])) for r in processed)
    st.markdown(_chip_html([(f"{len(processed)} reviews tagged", "green"), (f"{total_tags} labels written", "indigo")]), unsafe_allow_html=True)
    raw_cands = {k: v for k, v in (st.session_state.get("sym_new_candidates") or {}).items() if k.strip() and k.strip() not in (delighters + detractors)}
    new_cands = _dedup_candidates(raw_cands) if raw_cands else {}
    if new_cands:
        with st.expander(f"🟡 New symptom candidates ({len(new_cands)})", expanded=False):
            st.caption("Themes AI suggested not in your catalog. Near-duplicates auto-merged.")
            cand_rows = []
            for lab, rec in sorted(new_cands.items(), key=lambda kv: -int(kv[1].get("count", 0))):
                merged_from = rec.get("_merged_from", [])
                note = f"merged from: {', '.join(merged_from[:3])}" if merged_from else ""
                cand_rows.append(dict(Add=False, Label=lab, Count=int(rec.get("count", 0)), Notes=note))
            cand_df = pd.DataFrame(cand_rows)
            edited_cands = st.data_editor(cand_df, num_rows="fixed", use_container_width=True, hide_index=True, key="sym_cand_editor", column_config={"Add": st.column_config.CheckboxColumn(), "Label": st.column_config.TextColumn(), "Count": st.column_config.NumberColumn(format="%d"), "Notes": st.column_config.TextColumn(disabled=True)})
            cc1, cc2 = st.columns(2)
            if cc1.button("Add selected → Detractors", use_container_width=True, key="sym_add_det"):
                to_add = [str(r["Label"]) for _, r in edited_cands.iterrows() if bool(r.get("Add", False)) and str(r.get("Label", "")).strip()]
                if to_add:
                    st.session_state["sym_detractors"] = list(dict.fromkeys(detractors + to_add))
                    st.success(f"Added {len(to_add)}.")
                    st.rerun()
            if cc2.button("Add selected → Delighters", use_container_width=True, key="sym_add_del"):
                to_add = [str(r["Label"]) for _, r in edited_cands.iterrows() if bool(r.get("Add", False)) and str(r.get("Label", "")).strip()]
                if to_add:
                    st.session_state["sym_delighters"] = list(dict.fromkeys(delighters + to_add))
                    st.success(f"Added {len(to_add)}.")
                    st.rerun()
    updated_reviews = (st.session_state.get("analysis_dataset") or {}).get("reviews_df", overall_df)
    with st.expander(f"📋 Review log — last {min(len(processed), 20)} processed", expanded=True):
        for rec in processed[-20:]:
            idx = rec.get("idx", "?")
            head = f"Row {idx} — {len(rec.get('wrote_dets', []))} issues · {len(rec.get('wrote_dels', []))} strengths"
            with st.expander(head):
                all_ev_items = []
                for lab, evs in {**rec.get("ev_det", {}), **rec.get("ev_del", {})}.items():
                    for e in (evs or []):
                        if e and e.strip():
                            all_ev_items.append((e.strip(), lab))
                try:
                    row = updated_reviews.loc[int(idx)]
                    _render_review_card(row, evidence_items=all_ev_items or None)
                except Exception:
                    try:
                        vb = str(overall_df.loc[int(idx), "review_text"])[:800]
                        st.markdown(f"<div class='review-body'>{html.escape(vb)}</div>", unsafe_allow_html=True)
                    except Exception:
                        pass
                st.markdown("<div class='chip-wrap' style='margin-top:8px;margin-bottom:4px;'>" + f"<span class='chip yellow'>Safety: {_esc(rec.get('safety', ''))}</span>" + f"<span class='chip indigo'>Reliability: {_esc(rec.get('reliability', ''))}</span>" + f"<span class='chip gray'>Sessions: {_esc(rec.get('sessions', ''))}</span>" + "</div>", unsafe_allow_html=True)
    ec1, ec2 = st.columns([1.5, 3])
    if ec1.button("🧾 Prepare export", use_container_width=True, key="sym_prep_export"):
        upd = st.session_state["analysis_dataset"]["reviews_df"]
        orig = st.session_state.get("_uploaded_raw_bytes")
        sym_bytes = _gen_symptomized_workbook(orig, upd) if orig else _build_master_excel(summary, upd)
        st.session_state["sym_export_bytes"] = sym_bytes
        st.success("Export prepared.")
    sym_bytes = st.session_state.get("sym_export_bytes")
    ec1.download_button("⬇️ Download symptomized file", data=sym_bytes or b"", file_name=f"{summary.product_id}_Symptomized.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", disabled=(sym_bytes is None), key="sym_dl")
    ec2.caption("Writes AI Symptom Detractor / Delighter columns and Safety · Reliability · Sessions to columns K–AG.")

# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    st.markdown("""<div style="display:flex;align-items:center;gap:12px;margin-bottom:.2rem;">
      <div style="width:36px;height:36px;background:#0f172a;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:18px;">🦈</div>
      <div>
        <div style="font-size:20px;font-weight:800;letter-spacing:-.03em;color:#0f172a;">SharkNinja Review Analyst</div>
        <div style="font-size:12px;color:#64748b;margin-top:1px;">Voice-of-customer · AI analyst · Symptomizer</div>
      </div>
    </div>""", unsafe_allow_html=True)

    dataset = st.session_state.get("analysis_dataset")
    if dataset:
        bc = st.columns([4.2, 1.0])
        bc[0].caption(f"Active workspace · {dataset.get('source_type', '').title()} · {dataset.get('source_label', '')}")
        if bc[1].button("Clear workspace", use_container_width=True, key="ws_clear"):
            _reset_workspace_state(reset_source=True)
            st.rerun()

    if st.session_state.get("workspace_source_mode") not in {SOURCE_MODE_URL, SOURCE_MODE_FILE}:
        st.session_state["workspace_source_mode"] = SOURCE_MODE_URL

    source_mode = st.radio("Workspace source", [SOURCE_MODE_URL, SOURCE_MODE_FILE], horizontal=True, key="workspace_source_mode")
    if source_mode == SOURCE_MODE_URL:
        st.text_input("Product URL", key="workspace_product_url")
        if st.button("Build review workspace", type="primary", key="ws_build_url"):
            try:
                nd = _load_product_reviews(st.session_state.get("workspace_product_url", DEFAULT_PRODUCT_URL))
                _reset_review_filters()
                st.session_state.update(analysis_dataset=nd, chat_messages=[], master_export_bundle=None, prompt_run_artifacts=None, sym_processed_rows=[], sym_new_candidates={}, sym_symptoms_source="none", workspace_active_tab=TAB_DASHBOARD, workspace_tab_request=None)
                st.rerun()
            except requests.HTTPError as exc:
                st.error(f"HTTP error: {exc}")
            except ReviewDownloaderError as exc:
                st.error(str(exc))
            except Exception as exc:
                st.exception(exc)
    else:
        uploader_key = f"workspace_files_{int(st.session_state.get('workspace_file_uploader_nonce', 0))}"
        uploaded_files = st.file_uploader("Upload review export files", type=["csv", "xlsx", "xls"], accept_multiple_files=True, help="Supports Axion-style exports and similar CSV/XLSX review files.", key=uploader_key)
        st.caption("Mapped columns: Event Id · Base SKU · Review Text · Rating · Opened date · Seeded Flag · Retailer")
        if st.button("Build review workspace from file", type="primary", key="ws_build_file"):
            try:
                nd = _load_uploaded_files(uploaded_files or [])
                _reset_review_filters()
                st.session_state.update(analysis_dataset=nd, chat_messages=[], master_export_bundle=None, prompt_run_artifacts=None, sym_processed_rows=[], sym_new_candidates={}, sym_symptoms_source="none", workspace_active_tab=TAB_DASHBOARD, workspace_tab_request=None)
                if uploaded_files and len(uploaded_files) == 1:
                    fname = getattr(uploaded_files[0], "name", "")
                    if fname.lower().endswith(".xlsx"):
                        st.session_state["_uploaded_raw_bytes"] = uploaded_files[0].getvalue()
                st.rerun()
            except ReviewDownloaderError as exc:
                st.error(str(exc))
            except Exception as exc:
                st.exception(exc)

    dataset = st.session_state.get("analysis_dataset")
    settings = _render_sidebar(dataset["reviews_df"] if dataset else None)
    if not dataset:
        st.markdown("""<div style="margin-top:2rem;padding:2rem;background:var(--surface,#fff);border:1px solid #dde1e8;border-radius:18px;text-align:center;box-shadow:0 1px 4px rgba(15,23,42,.08);">
          <div style="font-size:2.5rem;margin-bottom:.75rem;">📊</div>
          <div style="font-size:16px;font-weight:700;color:#0f172a;margin-bottom:.4rem;">No workspace loaded</div>
          <div style="font-size:13px;color:#64748b;">Enter a SharkNinja product URL or upload a review export above to unlock the Dashboard, Review Explorer, AI Analyst, Review Prompt, and Symptomizer.</div>
        </div>""", unsafe_allow_html=True)
        return

    summary = dataset["summary"]
    overall_df = dataset["reviews_df"]
    source_type = dataset.get("source_type", "bazaarvoice")
    source_label = dataset.get("source_label", "")
    filter_state = settings["review_filters"]
    filtered_df = filter_state["filtered_df"]
    filter_description = filter_state["description"]
    new_filter_sig = json.dumps(filter_state["active_items"], default=str)
    if st.session_state.get("review_filter_signature") != new_filter_sig:
        st.session_state["review_filter_signature"] = new_filter_sig
        st.session_state["review_explorer_page"] = 1

    _render_workspace_header(summary, overall_df, st.session_state.get("prompt_run_artifacts"), source_type=source_type, source_label=source_label)
    _render_top_metrics(overall_df, filtered_df)
    _render_active_filter_summary(filter_state, overall_df)

    pending_tab = st.session_state.pop("workspace_tab_request", None)
    if pending_tab in WORKSPACE_TABS:
        st.session_state["workspace_active_tab"] = pending_tab
    elif st.session_state.get("workspace_active_tab") not in WORKSPACE_TABS:
        st.session_state["workspace_active_tab"] = TAB_DASHBOARD
    st.markdown("<div class='nav-tabs-wrap'><div class='nav-tabs-label'>Workspace</div></div>", unsafe_allow_html=True)
    active_tab = st.radio("Workspace tab", WORKSPACE_TABS, horizontal=True, key="workspace_active_tab", label_visibility="collapsed")
    common = dict(settings=settings, overall_df=overall_df, filtered_df=filtered_df, summary=summary, filter_description=filter_description)
    if active_tab == TAB_DASHBOARD:
        _render_dashboard(filtered_df, overall_df)
    elif active_tab == TAB_REVIEW_EXPLORER:
        _render_review_explorer(summary=summary, overall_df=overall_df, filtered_df=filtered_df, prompt_artifacts=st.session_state.get("prompt_run_artifacts"))
    elif active_tab == TAB_AI_ANALYST:
        _render_ai_tab(**common)
    elif active_tab == TAB_REVIEW_PROMPT:
        _render_review_prompt_tab(**common)
    elif active_tab == TAB_SYMPTOMIZER:
        _render_symptomizer_tab(**common)


if __name__ == "__main__":
    main()

