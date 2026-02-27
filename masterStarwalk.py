# starwalk_master_app.py
# Streamlit app: Axion JSON → Star Walk scrubbed verbatims → Dashboard (all-in-one)

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from datetime import datetime, timedelta
from wordcloud import STOPWORDS  # keep for stopword set; not rendering WC
import re
import html as _html
import os
import json
import textwrap
import warnings
import smtplib
from email.message import EmailMessage
from streamlit.components.v1 import html as st_html  # for custom HTML blocks

# New/updated imports
import time
import random
import hashlib
import threading
from typing import List, Optional
from urllib.parse import quote
from io import BytesIO
from collections import Counter

# Timezone
try:
    from zoneinfo import ZoneInfo  # py3.9+
    _NY_TZ = ZoneInfo("America/New_York")
except Exception:
    _NY_TZ = None

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    module="openpyxl",
)

# ---------- Optional text fixer ----------
try:
    from ftfy import fix_text as _ftfy_fix
    _HAS_FTFY = True
except Exception:
    _HAS_FTFY = False
    _ftfy_fix = None

# ---------- OpenAI SDK ----------
try:
    from openai import OpenAI
    _HAS_OPENAI = True
except Exception:
    _HAS_OPENAI = False
    OpenAI = None  # type: ignore

# Optional FAISS for fast vector similarity
try:
    import faiss  # type: ignore
    _HAS_FAISS = True
except Exception:
    _HAS_FAISS = False

# ---------- Local semantic search (fast, offline) ----------
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import linear_kernel
try:
    from rank_bm25 import BM25Okapi
    _HAS_BM25 = True
except Exception:
    _HAS_BM25 = False

try:
    from sentence_transformers import CrossEncoder
    _HAS_RERANKER = True
except Exception:
    _HAS_RERANKER = False

NO_TEMP_MODELS = {"gpt-5", "gpt-5-chat-latest"}


def model_supports_temperature(model_id: str) -> bool:
    if not model_id:
        return True
    if model_id in NO_TEMP_MODELS:
        return False
    return not model_id.startswith("gpt-5")


# ---------- Page config ----------
st.set_page_config(layout="wide", page_title="Star Walk Analysis Dashboard")

# ---------- Force Light Mode ----------
st_html(
    """
<script>
(function () {
  function setLight() {
    try {
      document.documentElement.setAttribute('data-theme','light');
      document.body && document.body.setAttribute('data-theme','light');
      window.localStorage.setItem('theme','light');
    } catch (e) {}
  }
  setLight();
  new MutationObserver(setLight).observe(
    document.documentElement,
    { attributes: true, attributeFilter: ['data-theme'] }
  );
})();
</script>
""",
    height=0,
)

# ---------- Global CSS ----------
GLOBAL_CSS = """
<style>
  :root { scroll-behavior: smooth; scroll-padding-top: 96px; }
  *, ::before, ::after { box-sizing: border-box; }
  @supports (scrollbar-color: transparent transparent){ * { scrollbar-width: thin; scrollbar-color: transparent transparent; } }

  :root{
    --text:#0f172a; --muted:#475569; --muted-2:#64748b;
    --border-strong:#90a7c1; --border:#cbd5e1; --border-soft:#e2e8f0;
    --bg-app:#f6f8fc; --bg-card:#ffffff; --bg-tile:#f8fafc;
    --ring:#3b82f6; --ok:#16a34a; --bad:#dc2626;
    --gap-sm:12px; --gap-md:20px; --gap-lg:32px;
  }
  html[data-theme="dark"], body[data-theme="dark"]{
    --text:rgba(255,255,255,.92); --muted:rgba(255,255,255,.72); --muted-2:rgba(255,255,255,.64);
    --border-strong:rgba(255,255,255,.22); --border:rgba(255,255,255,.16); --border-soft:rgba(255,255,255,.10);
    --bg-app:#0b0e14; --bg-card:rgba(255,255,255,.06); --bg-tile:rgba(255,255,255,.04);
    --ring:#60a5fa; --ok:#34d399; --bad:#f87171;
  }

  html, body, .stApp {
    background: var(--bg-app);
    font-family: "Helvetica Neue", Helvetica, Arial, ui-sans-serif, system-ui, -apple-system, "Segoe UI", Roboto, "Noto Sans", "Liberation Sans", sans-serif;
    color: var(--text);
  }
  .block-container { padding-top:.9rem; padding-bottom:1.2rem; }
  section[data-testid="stSidebar"] .block-container { padding-top:.6rem; }
  mark{ background:#fff2a8; padding:0 .2em; border-radius:3px; }

  .metrics-grid { display:grid; grid-template-columns:repeat(3,minmax(260px,1fr)); gap:17px; }
  @media (max-width:1100px){ .metrics-grid { grid-template-columns:1fr; } }
  .metric-card{ background:var(--bg-card); border-radius:14px; padding:16px; box-shadow:0 0 0 1.5px var(--border-strong), 0 8px 14px rgba(15,23,42,0.06); color:var(--text); }
  .metric-card h4{ margin:.2rem 0 .7rem 0; font-size:1.05rem; color:var(--text); }
  .metric-row{ display:grid; grid-template-columns:repeat(3,1fr); gap:12px; }
  .metric-box{ background:var(--bg-tile); border:1.6px solid var(--border); border-radius:12px; padding:12px; text-align:center; color:var(--text); }
  .metric-label{ color:var(--muted); font-size:.85rem; }
  .metric-kpi{ font-weight:800; font-size:1.8rem; letter-spacing:-0.01em; margin-top:2px; color:var(--text); }

  .review-card{ background:var(--bg-card); border-radius:12px; padding:16px; margin:16px 0 24px; box-shadow:0 0 0 1.5px var(--border-strong), 0 8px 14px rgba(15,23,42,0.06); color:var(--text); }
  .review-card p{ margin:.25rem 0; line-height:1.5; }
  .badges{ display:flex; flex-wrap:wrap; gap:8px; margin-top:6px; }
  .badge{ display:inline-block; padding:4px 10px; border-radius:999px; font-weight:600; font-size:.85rem; border:1.5px solid transparent; }
  .badge.pos{ background:#ecfdf5; border-color:#86efac; color:#065f46; }
  .badge.neg{ background:#fef2f2; border-color:#fca5a5; color:#7f1d1d; }

  [data-testid="stPlotlyChart"]{ margin-top:18px !important; margin-bottom:30px !important; }

  .soft-panel{
    background:var(--bg-card);
    border-radius:14px;
    padding:14px 16px;
    box-shadow:0 0 0 1.2px var(--border-strong), 0 6px 12px rgba(15,23,42,0.05);
    margin:10px 0 14px;
  }
  .small-muted{ color:var(--muted); font-size:.9rem; }
  .kpi-pill{
    display:inline-block; padding:4px 10px; border-radius:999px;
    border:1.3px solid var(--border);
    background:var(--bg-tile);
    font-weight:650; margin-right:8px;
  }
</style>
"""
st.markdown(GLOBAL_CSS, unsafe_allow_html=True)

# ---------- Utilities ----------
_BAD_CHARS_REGEX = re.compile(r"[ÃÂâï€™]")
PII_PAT = re.compile(r"[\w\.-]+@[\w\.-]+|\+?\d[\d\-\s]{6,}\d")


def clean_text(x: str, keep_na: bool = False) -> str:
    if pd.isna(x):
        return pd.NA if keep_na else ""
    s = str(x)
    if s.isascii():
        s = s.strip()
        if s.upper() in {"<NA>", "NA", "N/A", "NULL", "NONE"}:
            return pd.NA if keep_na else ""
        return s
    if _HAS_FTFY:
        try:
            s = _ftfy_fix(s)
        except Exception:
            pass
    if _BAD_CHARS_REGEX.search(s):
        try:
            repaired = s.encode("latin1", errors="ignore").decode("utf-8", errors="ignore")
            if repaired.strip():
                s = repaired
        except Exception:
            pass
    for bad, good in {
        "â€™": "'",
        "â€˜": "‘",
        "â€œ": "“",
        "â€\x9d": "”",
        "â€“": "–",
        "â€”": "—",
        "Â": "",
    }.items():
        s = s.replace(bad, good)
    s = s.strip()
    if s.upper() in {"<NA>", "NA", "N/A", "NULL", "NONE"}:
        return pd.NA if keep_na else ""
    return s


def _mask_pii(s: str) -> str:
    try:
        return PII_PAT.sub("[redacted]", s or "")
    except Exception:
        return s or ""


def esc(x) -> str:
    return _html.escape("" if pd.isna(x) else str(x))


def apply_filter(df: pd.DataFrame, column_name: str, label: str, key: str | None = None):
    options = ["ALL"]
    if column_name in df.columns:
        col = df[column_name].astype("string")
        options += sorted([x for x in col.dropna().unique().tolist() if str(x).strip() != ""])
    selected = st.multiselect(f"Select {label}", options=options, default=["ALL"], key=key)
    if "ALL" not in selected and column_name in df.columns:
        return df[df[column_name].astype("string").isin(selected)], selected
    return df, ["ALL"]


def collect_unique_symptoms(df: pd.DataFrame, cols: list[str]) -> list[str]:
    vals, seen = [], set()
    for c in cols:
        if c in df.columns:
            s = df[c].astype("string").str.strip().dropna()
            for v in pd.unique(s.to_numpy()):
                v = str(v).strip()
                if v and v not in seen:
                    seen.add(v)
                    vals.append(v)
    return vals


def is_valid_symptom_value(x) -> bool:
    if pd.isna(x):
        return False
    s = str(x).strip()
    if not s or s.upper() in {"<NA>", "NA", "N/A", "NULL", "NONE"}:
        return False
    return not bool(re.fullmatch(r"[\W_]+", s))


def analyze_delighters_detractors(filtered_df: pd.DataFrame, symptom_columns: list[str]) -> pd.DataFrame:
    # Vectorized & fast
    cols = [c for c in symptom_columns if c in filtered_df.columns]
    if not cols:
        return pd.DataFrame(columns=["Item", "Avg Star", "Mentions", "% Total"])
    m = filtered_df[cols].melt(value_name="symptom", var_name="col").drop(columns=["col"])
    m["symptom"] = m["symptom"].map(lambda v: clean_text(v, keep_na=True)).astype("string").str.strip()
    m = m[m["symptom"].map(is_valid_symptom_value)]
    if m.empty:
        return pd.DataFrame(columns=["Item", "Avg Star", "Mentions", "% Total"])
    counts = m["symptom"].value_counts()

    if "Star Rating" in filtered_df.columns:
        stars = pd.to_numeric(filtered_df["Star Rating"], errors="coerce")
        avg_map = {}
        for s_val in counts.index:
            rows = filtered_df[cols].isin([s_val]).any(axis=1)
            avg_map[s_val] = round(float(stars[rows].mean()), 1) if rows.any() else None
    else:
        avg_map = {s_val: None for s_val in counts.index}

    total_rows = len(filtered_df) or 1
    out = pd.DataFrame(
        {
            "Item": counts.index.str.title(),
            "Avg Star": [avg_map[s_val] for s_val in counts.index],
            "Mentions": counts.values,
            "% Total": (counts.values / total_rows * 100).round(1).astype(str) + "%",
        }
    )
    return out.sort_values("Mentions", ascending=False, ignore_index=True)


def highlight_html(text: str, keyword: str | None) -> str:
    safe = _html.escape(text or "")
    if keyword:
        try:
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            safe = pattern.sub(lambda m: f"<mark>{m.group(0)}</mark>", safe)
        except re.error:
            pass
    return safe


def _infer_product_label(df_in: pd.DataFrame, fallback_filename: str) -> str:
    base = os.path.splitext(fallback_filename or "Uploaded File")[0]
    if "Model (SKU)" in df_in.columns:
        s = df_in["Model (SKU)"].astype("string").str.strip().replace({"": pd.NA}).dropna()
        if not s.empty:
            top = s.value_counts().head(3).index.tolist()
            if len(top) == 1:
                return str(top[0])
            return " / ".join([str(x) for x in top])
    return base


def _extract_sentence(text: str, keyword: str | None = None, prefer_tail: bool = False) -> str:
    t = (text or "").strip()
    if not t:
        return ""
    # Basic sentence split
    parts = re.split(r"(?<=[.!?\n])\s+", t)
    if keyword:
        k = keyword.lower()
        for p in parts:
            if k in p.lower():
                return p.strip()
    if parts:
        return (parts[-1] if prefer_tail else parts[0]).strip()
    return t[:260]


def _pick_quotes_for_symptom(df_in: pd.DataFrame, symptom: str, cols: list[str], k: int = 2, prefer: str = "low"):
    # prefer = "low" for detractor, "high" for delighter
    if not cols or symptom not in (symptom or ""):
        return []
    d = df_in.copy()
    if "Star Rating" not in d.columns:
        return []
    mask = d[cols].isin([symptom]).any(axis=1)
    sub = d[mask].copy()
    if sub.empty:
        return []
    sub["Star Rating"] = pd.to_numeric(sub["Star Rating"], errors="coerce")
    sub = sub.dropna(subset=["Star Rating"])
    if sub.empty:
        return []
    sub = sub.sort_values("Star Rating", ascending=(prefer == "low"))
    out = []
    for _, row in sub.head(k).iterrows():
        txt = clean_text(row.get("Verbatim", ""))
        txt = _mask_pii(txt)
        sent = _extract_sentence(txt, keyword=symptom, prefer_tail=(prefer == "low"))
        if len(sent) > 280:
            sent = sent[:277] + "…"
        meta = []
        try:
            meta.append(f"{int(row.get('Star Rating'))}★")
        except Exception:
            pass
        for c in ["Source", "Country", "Model (SKU)"]:
            if c in df_in.columns:
                v = row.get(c, pd.NA)
                if pd.notna(v) and str(v).strip():
                    meta.append(str(v).strip())
        if "Review Date" in df_in.columns:
            dv = row.get("Review Date", pd.NaT)
            if pd.notna(dv):
                try:
                    meta.append(pd.to_datetime(dv).strftime("%Y-%m-%d"))
                except Exception:
                    pass
        out.append({"text": sent, "meta": " • ".join(meta) if meta else ""})
    return out


def _detect_trends(df_in: pd.DataFrame, symptom_cols: list[str], min_mentions: int = 3):
    if "Review Date" not in df_in.columns or "Star Rating" not in df_in.columns:
        return []
    d = df_in.copy()
    d["Review Date"] = pd.to_datetime(d["Review Date"], errors="coerce")
    d = d.dropna(subset=["Review Date"])
    if d.empty:
        return []

    # Compare last month vs prior month
    last_date = d["Review Date"].max()
    last_m = pd.Period(last_date, freq="M")
    prev_m = last_m - 1

    last = d[d["Review Date"].dt.to_period("M") == last_m]
    prev = d[d["Review Date"].dt.to_period("M") == prev_m]

    out = []
    if len(prev) > 0:
        pct = ((len(last) - len(prev)) / max(1, len(prev))) * 100
        if (len(last) - len(prev)) >= 5 and pct >= 50:
            out.append(f"Review volume increased from {len(prev)} → {len(last)} in {last_m.strftime('%b %Y')} (↑{pct:.0f}%).")
    elif len(last) >= 5:
        out.append(f"Review volume jumped to {len(last)} in {last_m.strftime('%b %Y')} (from 0 in prior month).")

    # Symptom spikes
    cols = [c for c in symptom_cols if c in d.columns]
    if cols and (not last.empty) and (not prev.empty or not last.empty):
        def _sym_counts(frame: pd.DataFrame):
            vals = []
            for c in cols:
                s = frame[c].astype("string").map(lambda v: clean_text(v, keep_na=True)).astype("string").str.strip()
                s = s[s.map(is_valid_symptom_value)]
                vals.extend([str(x) for x in s.tolist()])
            return Counter(vals)

        c_last = _sym_counts(last)
        c_prev = _sym_counts(prev) if not prev.empty else Counter()

        for sym, cnt in c_last.most_common(12):
            prev_cnt = c_prev.get(sym, 0)
            if prev_cnt == 0 and cnt >= min_mentions:
                out.append(f"New/returning theme: **{str(sym).title()}** with {cnt} mentions in {last_m.strftime('%b %Y')} (0 prior month).")
            elif prev_cnt > 0:
                inc = cnt - prev_cnt
                inc_pct = (inc / prev_cnt) * 100.0
                if inc >= min_mentions and inc_pct >= 50:
                    out.append(f"Spike: **{str(sym).title()}** mentions {prev_cnt} → {cnt} in {last_m.strftime('%b %Y')} (↑{inc_pct:.0f}%).")

    return out[:8]


# ---------- Embedding helpers with backoff, hashing & caching ----------
_EMBED_LOCK = threading.Lock()


def _embed_with_backoff(client, model: str, inputs: List[str], max_attempts: int = 6):
    delay = 1.0
    for attempt in range(1, max_attempts + 1):
        try:
            return client.embeddings.create(model=model, input=inputs)
        except Exception as e:
            status = getattr(e, "status_code", None)
            msg = str(e).lower()
            is_rate = status == 429 or "rate" in msg or "quota" in msg
            if not is_rate or attempt == max_attempts:
                raise
            sleep_s = delay * (1.5 ** (attempt - 1)) + random.uniform(0, 0.5)
            time.sleep(sleep_s)


def _hash_texts(texts: list[str]) -> str:
    h = hashlib.sha256()
    for t in texts:
        h.update(((t or "") + "\x00").encode("utf-8"))
    return h.hexdigest()


def _build_vector_index(texts: list[str], api_key: str, model: str = "text-embedding-3-small"):
    # Builds ONLY on demand (when user asks a question + AI toggle on)
    if not _HAS_OPENAI or not texts:
        return None
    client = OpenAI(api_key=api_key, timeout=60, max_retries=0)
    embs = []
    batch = 128

    with _EMBED_LOCK:
        for i in range(0, len(texts), batch):
            chunk = texts[i : i + batch]
            safe_chunk = [(t or "")[:2000] for t in chunk]
            resp = _embed_with_backoff(client, model, safe_chunk)
            embs.extend([np.array(d.embedding, dtype=np.float32) for d in resp.data])
            time.sleep(0.05 + random.uniform(0, 0.05))

    if not embs:
        return None
    mat = np.vstack(embs).astype(np.float32)
    norms = np.linalg.norm(mat, axis=1, keepdims=True) + 1e-8
    mat_norm = mat / norms
    if _HAS_FAISS:
        index = faiss.IndexFlatIP(mat_norm.shape[1])
        index.add(mat_norm)
        return {"backend": "faiss", "index": index, "texts": texts}
    return (mat, norms, texts)


@st.cache_resource(show_spinner=False)
def _ensure_cache_slot(model: str, content_hash: str, api_key: str):
    # Dummy cache entry to memoize (model, hash) without serializing big data
    return {"ok": True}


def _get_or_build_index(content_hash: str, raw_texts: list[str], api_key: str, model: str) -> Optional[object]:
    memo = st.session_state.setdefault("_vec_idx", {})
    if content_hash in memo:
        return memo[content_hash]
    idx = _build_vector_index(raw_texts, api_key=api_key, model=model)
    memo[content_hash] = idx
    return idx


def vector_search(query: str, index, api_key: str, top_k: int = 8):
    if not _HAS_OPENAI or index is None:
        return []
    client = OpenAI(api_key=api_key)
    qemb = client.embeddings.create(model="text-embedding-3-small", input=[query]).data[0].embedding
    q = np.array(qemb, dtype=np.float32)
    qn = np.linalg.norm(q) + 1e-8
    qn_vec = q / qn
    if isinstance(index, dict) and index.get("backend") == "faiss":
        D, I = index["index"].search(qn_vec.reshape(1, -1), top_k)
        sims = D[0].tolist()
        idxs = I[0].tolist()
        texts = index["texts"]
        return [(texts[i], float(sims[j])) for j, i in enumerate(idxs) if i != -1]
    mat, norms, texts = index
    sims = (mat @ q) / (norms.flatten() * qn)
    idx = np.argsort(-sims)[:top_k]
    return [(texts[i], float(sims[i])) for i in idx]


# ---------- Local retrieval (TF-IDF + optional BM25 + optional reranker) ----------
def _get_or_build_local_text_index(texts: list[str], content_hash: str):
    memo = st.session_state.setdefault("_local_text_idx", {})
    if content_hash in memo:
        return memo[content_hash]

    corpus = [(t or "").strip() for t in texts]
    tfidf = TfidfVectorizer(lowercase=True, strip_accents="unicode", ngram_range=(1, 2), min_df=2, max_df=0.95)
    tfidf_mat = tfidf.fit_transform(corpus)

    bm25 = None
    tokenized = None
    if _HAS_BM25 and st.session_state.get("use_bm25", False):
        tokenized = [t.lower().split() for t in corpus]
        bm25 = BM25Okapi(tokenized)

    memo[content_hash] = {"tfidf": tfidf, "tfidf_mat": tfidf_mat, "bm25": bm25, "tokenized": tokenized, "texts": corpus}
    return memo[content_hash]


def _local_search(query: str, index: dict, top_k: int = 8):
    if not index:
        return []
    q = (query or "").strip()
    if not q:
        return []
    # TF-IDF cosine
    qvec = index["tfidf"].transform([q])
    scores = linear_kernel(qvec, index["tfidf_mat"]).ravel()
    tfidf_top = np.argsort(-scores)[: max(top_k * 3, 20)]

    top = tfidf_top
    # Optional BM25 blend
    if index.get("bm25") is not None:
        bm_scores = index["bm25"].get_scores(q.lower().split())
        # Normalize and blend
        bm = (bm_scores - np.min(bm_scores)) / (np.ptp(bm_scores) + 1e-9)
        tf = (scores - np.min(scores)) / (np.ptp(scores) + 1e-9)
        hybrid = 0.6 * tf + 0.4 * bm
        top = np.argsort(-hybrid)[: top_k * 3]

    top_idx = list(top[:top_k])

    # Optional cross-encoder rerank (local; may download model)
    if _HAS_RERANKER and st.session_state.get("use_reranker", False):
        try:
            ce = st.session_state.get("_cross_encoder")
            if ce is None:
                with st.spinner("Loading local reranker…"):
                    ce = CrossEncoder("cross-encoder/ms-marco-MiniLM-L-6-v2")
                st.session_state["_cross_encoder"] = ce
            cand = [index["texts"][i] for i in top]
            pairs = [[q, c] for c in cand]
            rr = ce.predict(pairs)
            reranked = np.argsort(-rr)[:top_k]
            top_idx = [top[i] for i in reranked]
        except Exception:
            top_idx = list(top[:top_k])

    # Return (text, score) – we’ll return TF-IDF score for transparency
    return [(index["texts"][i], float(scores[i])) for i in top_idx]


# ---------- Local Q&A router ----------
_SYNONYMS = {
    "low_star": ["low star", "1-2", "1–2", "1 and 2", "one and two", "detractor"],
    "avg_star": ["average", "avg", "mean rating"],
    "trend": ["trend", "over time", "by month", "monthly", "week", "weekly"],
    "compare": ["vs", "compare", "difference", "gap"],
    "top": ["top", "biggest", "most common", "leading"],
    "delighters": ["delighter", "delight"],
    "detractors": ["detractor", "complaint", "issue"],
}


def _has_any(q, keys):
    ql = q.lower()
    return any(any(tok in ql for tok in _SYNONYMS.get(k, [])) for k in keys)


def _month_trend(df: pd.DataFrame, col="Star Rating"):
    if "Review Date" not in df.columns or col not in df.columns:
        return pd.DataFrame()
    d = df.copy()
    d["__month"] = pd.to_datetime(d["Review Date"], errors="coerce").dt.to_period("M").astype(str)
    s = pd.to_numeric(d[col], errors="coerce")
    return (
        pd.DataFrame({"__month": d["__month"], col: s})
        .dropna()
        .groupby("__month")[col]
        .agg(["count", "mean"])
        .reset_index()
        .sort_values("__month")
    )


def _cohort_compare(df: pd.DataFrame):
    if "Seeded" not in df.columns or "Star Rating" not in df.columns:
        return None
    d = df.copy()
    d["Seeded"] = d["Seeded"].astype("string").str.upper().fillna("NO")
    s = pd.to_numeric(d["Star Rating"], errors="coerce")
    grp = pd.DataFrame({"Seeded": d["Seeded"], "Star": s}).dropna().groupby("Seeded")["Star"].agg(["count", "mean"])
    org = grp.loc["NO"] if "NO" in grp.index else None
    sed = grp.loc["YES"] if "YES" in grp.index else None
    return grp, org, sed


def _top_symptoms(df: pd.DataFrame, which="detractors", k=5):
    cols = [f"Symptom {i}" for i in (range(1, 11) if which == "detractors" else range(11, 21))]
    data = analyze_delighters_detractors(df, cols)
    return data.head(k)


def _keyword_prevalence(df: pd.DataFrame, term: str):
    if "Verbatim" not in df.columns or not term:
        return 0, 0.0
    ser = df["Verbatim"].astype("string").fillna("")
    cnt = int(ser.str.contains(term, case=False, na=False).sum())
    pct = (cnt / max(1, len(df))) * 100.0
    return cnt, pct


def _route_and_answer_locally(q: str, df: pd.DataFrame, quotes: list[str]) -> str:
    parts = []
    # Trend questions
    if _has_any(q, ["trend"]):
        t = _month_trend(df)
        if not t.empty:
            last = t.tail(2)
            if len(last) == 2:
                delta = last["mean"].iloc[1] - last["mean"].iloc[0]
                parts.append(f"**Monthly trend (avg ★):** last → {last['mean'].iloc[1]:.2f} ({delta:+.2f} vs prev month).")
            parts.append(
                "Top months by avg ★: "
                + ", ".join([f"{r['__month']} ({r['mean']:.2f})" for _, r in t.sort_values("mean", ascending=False).head(3).iterrows()])
            )

    # Compare seeded vs organic
    if _has_any(q, ["compare"]) or "seeded" in q.lower():
        res = _cohort_compare(df)
        if res:
            grp, org, sed = res
            if org is not None and sed is not None:
                gap = sed["mean"] - org["mean"]
                parts.append(
                    f"**Seeded vs Organic (avg ★):** Seeded {sed['mean']:.2f} vs Organic {org['mean']:.2f} (gap {gap:+.2f}). "
                    f"Counts — Seeded {int(sed['count'])}, Organic {int(org['count'])}."
                )

    # Top symptoms
    if _has_any(q, ["top", "detractors"]):
        det = _top_symptoms(df, "detractors", 5)
        if not det.empty:
            parts.append("**Top detractors:** " + "; ".join([f"{r['Item']} (★ {r['Avg Star']}, {int(r['Mentions'])})" for _, r in det.iterrows()]))

    if _has_any(q, ["top", "delighters"]):
        delit = _top_symptoms(df, "delighters", 5)
        if not delit.empty:
            parts.append("**Top delighters:** " + "; ".join([f"{r['Item']} (★ {r['Avg Star']}, {int(r['Mentions'])})" for _, r in delit.iterrows()]))

    # Keyword prevalence
    m = re.search(r'(?:mentions of|frequency of|how many.*mention)\s+"([^"]+)"', q, re.I)
    if not m:
        m = re.search(r"mentions of ([\w\- ]+)", q, re.I)
    if m:
        term = m.group(1).strip()
        cnt, pct = _keyword_prevalence(df, term)
        parts.append(f"**Mentions of “{term}”:** {cnt} reviews ({pct:.1f}%).")

    # Fallback snapshot if nothing matched
    if not parts:
        total = int(len(df))
        avg = float(pd.to_numeric(df.get("Star Rating"), errors="coerce").mean()) if total and "Star Rating" in df.columns else 0.0
        low = float((pd.to_numeric(df.get("Star Rating"), errors="coerce") <= 2).mean() * 100) if total and "Star Rating" in df.columns else 0.0
        parts.append(f"**Snapshot** — {total} reviews; avg ★ {avg:.1f}; % 1–2★ {low:.1f}%.")

    # Add quotes (evidence)
    if quotes:
        parts.append("**Representative quotes:**\n" + "\n".join([f"• “{_mask_pii(qt)}”" for qt in quotes[:5]]))

    return "\n\n".join(parts)


# ---------- Anchors ----------
def anchor(id_: str):
    st.markdown(f"<div id='{id_}'></div>", unsafe_allow_html=True)


def scroll_to(id_: str):
    st.markdown(
        f"""
        <script>
        (function(id){{
          function jump(){{
            const el = document.getElementById(id);
            if(el) {{
              el.scrollIntoView({{behavior:'smooth', block:'start'}});
              window.location.hash = id;
            }}
          }}
          setTimeout(jump, 0); setTimeout(jump, 150); setTimeout(jump, 300); setTimeout(jump, 600);
        }})('{id_}');
        </script>
        """,
        unsafe_allow_html=True,
    )




import io
import json
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

EXCEL_MAX_CHARS = 32767

# =============================
# Helpers
# =============================
def safe_get(d: Dict[str, Any], path: List[str], default: Any = None) -> Any:
    cur: Any = d
    for k in path:
        if isinstance(cur, dict) and k in cur:
            cur = cur[k]
        else:
            return default
    return cur


def as_list(x: Any) -> List[Any]:
    if x is None:
        return []
    if isinstance(x, list):
        return x
    return [x]


def join_list(x: Any, sep: str = " | ") -> Any:
    if x is None:
        return None
    if isinstance(x, list):
        vals = [str(v).strip() for v in x if v is not None and str(v).strip() != ""]
        return sep.join(vals) if vals else None
    return x


def parse_iso_date(x: Any) -> Optional[date]:
    if not x:
        return None
    ts = pd.to_datetime(x, utc=True, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.date()


def title_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    stem = re.sub(r"[_\-]+", " ", stem)
    stem = re.sub(r"([a-z])([A-Z])", r"\1 \2", stem)  # camelCase -> words
    stem = re.sub(r"\s+", " ", stem).strip()
    tokens = stem.split(" ")
    if tokens and any(ch.isdigit() for ch in tokens[0]):
        tokens[0] = tokens[0].upper()
    return " ".join(tokens)


def excel_safe_value(v: Any, list_sep: str = " | ") -> Any:
    """
    Convert arbitrary Python objects into openpyxl-safe scalar types.
    Fixes: ValueError: Cannot convert ['Leshow CN', ...] to Excel
    """
    if v is None:
        return None

    try:
        if pd.isna(v):
            return None
    except Exception:
        pass

    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return None
        return v.to_pydatetime()

    if isinstance(v, (datetime, date, bool, int, float)):
        return v

    if isinstance(v, str):
        return v[: EXCEL_MAX_CHARS - 3] + "..." if len(v) > EXCEL_MAX_CHARS else v

    # numpy scalar
    if hasattr(v, "item") and not isinstance(v, (list, tuple, set, dict)):
        try:
            return excel_safe_value(v.item(), list_sep=list_sep)
        except Exception:
            pass

    if isinstance(v, dict):
        s = json.dumps(v, ensure_ascii=False)
        return s[: EXCEL_MAX_CHARS - 3] + "..." if len(s) > EXCEL_MAX_CHARS else s

    if isinstance(v, (list, tuple, set)):
        parts: List[str] = []
        for x in v:
            if x is None:
                continue
            if isinstance(x, (dict, list, tuple, set)):
                parts.append(json.dumps(x, ensure_ascii=False))
            else:
                parts.append(str(x))
        s = list_sep.join([p.strip() for p in parts if p.strip() != ""])
        if not s:
            return None
        return s[: EXCEL_MAX_CHARS - 3] + "..." if len(s) > EXCEL_MAX_CHARS else s

    s = str(v)
    return s[: EXCEL_MAX_CHARS - 3] + "..." if len(s) > EXCEL_MAX_CHARS else s


# =============================
# JSON -> DataFrames
# =============================
REVIEWS_BASE_COLS: List[Tuple[str, Any]] = [
    ("Record ID", lambda r: r.get("_id")),
    ("Opened Timestamp", lambda r: parse_iso_date(r.get("openedTimestamp"))),
    ("Rating (num)", lambda r: safe_get(r, ["clientAttributes", "Rating (num)"])),
    ("Retailer", lambda r: safe_get(r, ["clientAttributes", "Retailer"])),
    ("Retailer Rating", lambda r: safe_get(r, ["clientAttributes", "Retailer Rating"])),
    ("Model", lambda r: safe_get(r, ["clientAttributes", "Model"])),
    ("Seeded Reviews", lambda r: safe_get(r, ["clientAttributes", "Seeded Reviews"])),
    ("Syndicated/Seeded Reviews", lambda r: safe_get(r, ["clientAttributes", "Syndicated/Seeded Reviews"])),
    ("Location", lambda r: safe_get(r, ["clientAttributes", "Location"])),
    ("Post Link", lambda r: safe_get(r, ["clientAttributes", "Post Link"])),
    ("Title", lambda r: safe_get(r, ["freeText", "Title"])),
    ("Review", lambda r: safe_get(r, ["freeText", "Review"])),
]

REVIEWS_EXTRA_COLS: List[Tuple[str, Any]] = [
    ("Satisfaction Score", lambda r: safe_get(r, ["customAttributes", "Satisfaction Score"])),
    ("Key Review Sentiment_Reviews", lambda r: join_list(safe_get(r, ["customAttributes", "Key Review Sentiment_Reviews"]))),
    ("Key Review Sentiment Type_Reviews", lambda r: join_list(safe_get(r, ["customAttributes", "Key Review Sentiment Type_Reviews"]))),
    ("Dominant Customer Journey Step", lambda r: join_list(safe_get(r, ["customAttributes", "Dominant Customer Journey Step"]))),
    ("Trigger Point_Product", lambda r: join_list(safe_get(r, ["customAttributes", "Trigger Point_Product"]))),
    ("L2 Delighter Component", lambda r: join_list(safe_get(r, ["customAttributes", "L2 Delighter Component"]))),
    ("L2 Delighter Condition", lambda r: join_list(safe_get(r, ["customAttributes", "L2 Delighter Condition"]))),
    ("L2 Delighter Mode", lambda r: join_list(safe_get(r, ["customAttributes", "L2 Delighter Mode"]))),
    ("L3 Non Product Detractors", lambda r: join_list(safe_get(r, ["customAttributes", "L3 Non Product Detractors"]))),
    ("Product_Symptom Component", lambda r: join_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Component"]))),
    ("Product_Symptom Conditions", lambda r: join_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Conditions"]))),
    ("Product_Symptom Mode", lambda r: join_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Mode"]))),
    ("Product Name", lambda r: safe_get(r, ["clientAttributes", "Product Name"])),
    ("Product Category", lambda r: safe_get(r, ["clientAttributes", "Product Category"])),
    ("Base SKU", lambda r: safe_get(r, ["clientAttributes", "Base SKU"])),
    ("Brand", lambda r: safe_get(r, ["clientAttributes", "Brand"])),
    ("Company", lambda r: safe_get(r, ["clientAttributes", "Company"])),
    ("Factory Name", lambda r: safe_get(r, ["clientAttributes", "Factory Name"])),
    ("Translation", lambda r: safe_get(r, ["clientAttributes", "Translation"])),
    ("Event ID", lambda r: r.get("eventId")),
    ("Event Type", lambda r: r.get("eventType")),
    ("Is Linked", lambda r: r.get("isLinked")),
    ("Workspace ID", lambda r: safe_get(r, ["clientAttributes", "Workspace ID"])),
]


def build_reviews_df(records: List[Dict[str, Any]], include_extra: bool = True) -> pd.DataFrame:
    cols = REVIEWS_BASE_COLS + (REVIEWS_EXTRA_COLS if include_extra else [])
    rows = [{name: fn(r) for name, fn in cols} for r in records]
    df = pd.DataFrame(rows)

    if "Rating (num)" in df.columns:
        df["Rating (num)"] = pd.to_numeric(df["Rating (num)"], errors="coerce")

    return df


def build_symptoms_df(records: List[Dict[str, Any]], include_blank_row_when_missing: bool = True) -> pd.DataFrame:
    out_rows: List[Dict[str, Any]] = []

    for r in records:
        rid = r.get("_id")
        opened = parse_iso_date(r.get("openedTimestamp"))
        rating = safe_get(r, ["clientAttributes", "Rating (num)"])
        retailer = safe_get(r, ["clientAttributes", "Retailer"])
        model = safe_get(r, ["clientAttributes", "Model"])

        comps = as_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Component"]))
        conds = as_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Conditions"]))
        modes = as_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Mode"]))

        max_len = max(len(comps), len(conds), len(modes), 0)

        if max_len == 0 and include_blank_row_when_missing:
            out_rows.append(
                {
                    "Record ID": rid,
                    "Opened Timestamp": opened,
                    "Rating": rating,
                    "Retailer": retailer,
                    "Model": model,
                    "Symptom Index": None,
                    "Symptom Component": None,
                    "Symptom Condition": None,
                    "Symptom Mode": None,
                }
            )
            continue

        for i in range(max_len):
            comp = comps[i] if i < len(comps) else None
            cond = conds[i] if i < len(conds) else None
            mode = modes[i] if i < len(modes) else None

            if mode in (None, "", "-", "—"):
                if comp and cond:
                    mode = f"{comp} - {cond}"
                elif cond:
                    mode = f"- {cond}"
                elif comp:
                    mode = f"{comp} -"
                else:
                    mode = "-"

            out_rows.append(
                {
                    "Record ID": rid,
                    "Opened Timestamp": opened,
                    "Rating": rating,
                    "Retailer": retailer,
                    "Model": model,
                    "Symptom Index": i + 1,
                    "Symptom Component": comp,
                    "Symptom Condition": cond,
                    "Symptom Mode": mode,
                }
            )

    df = pd.DataFrame(out_rows)
    if "Rating" in df.columns:
        df["Rating"] = pd.to_numeric(df["Rating"], errors="coerce")
    return df


# =============================
# Summary Data
# =============================
def build_summary_tables(reviews_df: pd.DataFrame, symptoms_df: pd.DataFrame, top_n: int = 10):
    total_reviews = int(len(reviews_df))

    date_min = reviews_df["Opened Timestamp"].min() if "Opened Timestamp" in reviews_df.columns else None
    date_max = reviews_df["Opened Timestamp"].max() if "Opened Timestamp" in reviews_df.columns else None
    date_range_str = ""
    if pd.notna(date_min) and pd.notna(date_max) and date_min and date_max:
        date_range_str = f"{date_min} to {date_max}"

    avg_rating = None
    if "Rating (num)" in reviews_df.columns:
        avg_rating = float(pd.to_numeric(reviews_df["Rating (num)"], errors="coerce").mean())

    rating_counts = (
        reviews_df["Rating (num)"].dropna().astype(int).value_counts().reindex([5, 4, 3, 2, 1], fill_value=0)
    )
    rating_dist = pd.DataFrame(
        {
            "Rating": rating_counts.index.astype(int),
            "Count": rating_counts.values.astype(int),
            "Share": (rating_counts.values / total_reviews) if total_reviews else 0,
        }
    )

    retailer_counts = reviews_df.get("Retailer", pd.Series(dtype=str)).fillna("(blank)").value_counts().head(top_n)
    top_retailers = pd.DataFrame(
        {
            "Retailer": retailer_counts.index,
            "Count": retailer_counts.values.astype(int),
            "Share": (retailer_counts.values / total_reviews) if total_reviews else 0,
        }
    )

    cond_series = symptoms_df.get("Symptom Condition", pd.Series(dtype=str)).fillna("")
    cond_series = cond_series[cond_series.astype(str).str.strip() != ""]
    symptom_rows = int(len(cond_series))
    cond_counts = cond_series.value_counts().head(top_n)
    top_conditions = pd.DataFrame(
        {
            "Condition": cond_counts.index,
            "Count": cond_counts.values.astype(int),
            "Share (of symptom rows)": (cond_counts.values / symptom_rows) if symptom_rows else 0,
        }
    )

    return {
        "total_reviews": total_reviews,
        "date_range_str": date_range_str,
        "avg_rating": avg_rating,
        "rating_dist": rating_dist,
        "top_retailers": top_retailers,
        "top_conditions": top_conditions,
        "symptom_rows": symptom_rows,
    }


# =============================
# Excel Writer (openpyxl)
# =============================
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="1F4E79", bold=True, size=14)
SECTION_FONT = Font(color="1F4E79", bold=False)


def write_df_to_sheet(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([excel_safe_value(v) for v in row])


def add_excel_table(ws, df: pd.DataFrame, table_name: str, style_name: str = "TableStyleMedium9"):
    nrows = len(df) + 1
    ncols = len(df.columns)
    if nrows <= 1 or ncols == 0:
        return

    ref = f"A1:{get_column_letter(ncols)}{nrows}"
    tab = Table(displayName=table_name, ref=ref)

    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    ws.freeze_panes = "A2"


def set_col_widths(ws, df: pd.DataFrame, widths: Dict[str, float]):
    for i, col in enumerate(df.columns, start=1):
        if col in widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[col]


def set_date_format(ws, df: pd.DataFrame, date_cols: List[str]):
    for col_name in date_cols:
        if col_name not in df.columns:
            continue
        col_idx = list(df.columns).index(col_name) + 1
        for r in range(2, len(df) + 2):
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "yyyy-mm-dd"


def apply_hyperlinks(ws, df: pd.DataFrame, url_cols: List[str]):
    for col_name in url_cols:
        if col_name not in df.columns:
            continue
        col_idx = list(df.columns).index(col_name) + 1
        for r in range(2, len(df) + 2):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.style = "Hyperlink"


def wrap_cells(ws, df: pd.DataFrame, wrap_cols: List[str]):
    align = Alignment(wrap_text=True, vertical="top")
    for col_name in wrap_cols:
        if col_name not in df.columns:
            continue
        col_idx = list(df.columns).index(col_name) + 1
        for r in range(1, len(df) + 2):
            ws.cell(row=r, column=col_idx).alignment = align


def build_workbook(
    dataset_title: str,
    reviews_df: pd.DataFrame,
    symptoms_df: pd.DataFrame,
    summary: Dict[str, Any],
    wrap_long_text: bool = False,
) -> bytes:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- Reviews sheet (FIRST TAB) ---
    ws_r = wb.create_sheet("Reviews")
    write_df_to_sheet(ws_r, reviews_df)
    add_excel_table(ws_r, reviews_df, table_name="ReviewsTable")
    apply_hyperlinks(ws_r, reviews_df, url_cols=["Post Link"])
    set_date_format(ws_r, reviews_df, date_cols=["Opened Timestamp"])

    review_widths = {
        "Record ID": 34,
        "Opened Timestamp": 16,
        "Rating (num)": 11,
        "Retailer": 14,
        "Retailer Rating": 14,
        "Model": 12,
        "Seeded Reviews": 15,
        "Syndicated/Seeded Reviews": 22,
        "Location": 10,
        "Post Link": 55,
        "Title": 45,
        "Review": 85,
        "Translation": 55,
        "Factory Name": 22,
    }
    set_col_widths(ws_r, reviews_df, review_widths)
    if wrap_long_text:
        wrap_cells(ws_r, reviews_df, wrap_cols=["Title", "Review", "Translation"])

    # --- Summary sheet ---
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"{dataset_title} — Summary"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Dataset"
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "Total Reviews"
    ws["B4"] = summary["total_reviews"]

    ws["A5"] = "Date Range (Opened)"
    ws["B5"] = summary["date_range_str"]

    ws["A6"] = "Average Rating"
    if summary["avg_rating"] is not None:
        ws["B6"] = round(summary["avg_rating"], 3)

    ws["A8"] = "Rating Distribution"
    ws["A8"].font = SECTION_FONT
    ws["D8"] = "Top Retailers"
    ws["D8"].font = SECTION_FONT

    rd = summary["rating_dist"].copy()
    rd_cols = ["Rating", "Count", "Share"]
    rd = rd[rd_cols]
    start_row, start_col = 9, 1
    for j, col in enumerate(rd_cols):
        cell = ws.cell(row=start_row, column=start_col + j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(rd.itertuples(index=False, name=None), start=1):
        for j, val in enumerate(row):
            c = ws.cell(row=start_row + i, column=start_col + j, value=excel_safe_value(val))
            if j == 2:
                c.number_format = "0.0%"

    tr = summary["top_retailers"].copy()
    tr_cols = ["Retailer", "Count", "Share"]
    tr = tr[tr_cols]
    start_row, start_col = 9, 4
    for j, col in enumerate(tr_cols):
        cell = ws.cell(row=start_row, column=start_col + j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(tr.itertuples(index=False, name=None), start=1):
        for j, val in enumerate(row):
            c = ws.cell(row=start_row + i, column=start_col + j, value=excel_safe_value(val))
            if j == 2:
                c.number_format = "0.0%"

    ws["A22"] = "Top Symptom Conditions"
    ws["A22"].font = SECTION_FONT
    ws["E22"] = "Notes"
    ws["E22"].font = SECTION_FONT

    ws["E23"] = (
        "Symptoms sheet is exploded from\n"
        "customAttributes.taxonomies lists\n"
        "(Component / Condition / Mode)."
    )
    ws["E23"].alignment = Alignment(wrap_text=True, vertical="top")

    tc = summary["top_conditions"].copy()
    tc_cols = ["Condition", "Count", "Share (of symptom rows)"]
    tc = tc[tc_cols]
    start_row, start_col = 23, 1
    for j, col in enumerate(tc_cols):
        cell = ws.cell(row=start_row, column=start_col + j, value=col if j < 2 else "Share")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        if j == 2:
            ws.column_dimensions[get_column_letter(start_col + j)].width = 6
            cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(tc.itertuples(index=False, name=None), start=1):
        for j, val in enumerate(row):
            c = ws.cell(row=start_row + i, column=start_col + j, value=excel_safe_value(val))
            if j == 2:
                c.number_format = "0.0%"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 10

    # --- Symptoms sheet ---
    ws_s = wb.create_sheet("Symptoms")
    write_df_to_sheet(ws_s, symptoms_df)
    add_excel_table(ws_s, symptoms_df, table_name="SymptomsTable")
    set_date_format(ws_s, symptoms_df, date_cols=["Opened Timestamp"])

    symptom_widths = {
        "Record ID": 34,
        "Opened Timestamp": 16,
        "Rating": 8,
        "Retailer": 14,
        "Model": 12,
        "Symptom Index": 12,
        "Symptom Component": 22,
        "Symptom Condition": 38,
        "Symptom Mode": 38,
    }
    set_col_widths(ws_s, symptoms_df, symptom_widths)

    # Make Reviews the active sheet when opening
    wb.active = wb.sheetnames.index("Reviews")

    # Save to bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def _strip_code_fences(s: str) -> str:
    s = s.strip()
    m = re.search(r"```(?:json)?\s*(.*?)```", s, flags=re.DOTALL | re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return s


def _extract_json_substring(s: str) -> str:
    s = s.strip()
    if s.startswith("{") or s.startswith("["):
        return s
    start_candidates = [i for i in [s.find("{"), s.find("[")] if i != -1]
    if not start_candidates:
        return s
    start = min(start_candidates)
    end = max(s.rfind("}"), s.rfind("]"))
    if end != -1 and end > start:
        return s[start : end + 1].strip()
    return s


def _try_parse_json_lines(s: str) -> Optional[List[Dict[str, Any]]]:
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    if len(lines) < 2:
        return None
    objs: List[Dict[str, Any]] = []
    for ln in lines:
        try:
            obj = json.loads(ln)
            if isinstance(obj, dict):
                objs.append(obj)
            else:
                return None
        except Exception:
            return None
    return objs


def loads_flexible_json(text_in: str) -> Tuple[Any, List[str]]:
    warnings: List[str] = []
    s = _strip_code_fences(text_in)
    s = _extract_json_substring(s)

    try:
        return json.loads(s), warnings
    except Exception as e1:
        # Attempt: remove trailing commas
        s2 = re.sub(r",\s*([}\]])", r"\1", s)
        if s2 != s:
            try:
                warnings.append("Removed trailing commas to make JSON valid.")
                return json.loads(s2), warnings
            except Exception:
                warnings.pop()

        # Attempt: JSON Lines
        jl = _try_parse_json_lines(s)
        if jl is not None:
            warnings.append("Detected JSON Lines and parsed each line as a record.")
            return jl, warnings

        raise ValueError("Could not parse input as JSON. Paste valid JSON (or JSON Lines).") from e1


def extract_records(raw: Any) -> List[Dict[str, Any]]:
    if isinstance(raw, dict) and "results" in raw and isinstance(raw["results"], list):
        return raw["results"]
    if isinstance(raw, list):
        return [r for r in raw if isinstance(r, dict)]
    raise ValueError("Unrecognized JSON shape. Expected a dict with `results: []` or a list of record objects.")

# starwalk_converter_app_FIXED.py
# Streamlit app to convert raw website review exports into the
# "Star Walk scrubbed verbatims" format (Symptom 1–10 detractors, 11–20 delighters).

import ast
import io
import re
from typing import Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

APP_VERSION = "2026-02-27-extra-cols-after-symptom20-no-hairtype-v1"

STARWALK_SHEET_NAME = "Star Walk scrubbed verbatims"

# -----------------------------
# Default columns (Hair Type removed)
# -----------------------------
DEFAULT_STARWALK_COLUMNS: List[str] = [
    "Source",
    "Model (SKU)",
    "Seeded",
    "Country",
    "New Review",
    "Review Date",
    "Verbatim Id",
    "Verbatim",
    "Star Rating",
    "Review count per detractor",
] + [f"Symptom {i}" for i in range(1, 21)]

# Default optional extra columns to place after Symptom 20
DEFAULT_EXTRA_AFTER_SYMPTOM20: List[str] = [
    "Key Review Sentiment_Reviews",
    "Key Review Sentiment Type_Reviews",
    "Trigger Point_Product",
    "Dominant Customer Journey Step",
    "L2 Delighter Component",
    "L2 Delighter Mode",
    "L3 Non Product Detractors",
    "Product_Symptom Component",
]

# Used for splitting multi-tag cells like "A; B | C"
DEFAULT_TAG_SPLIT_RE = re.compile(r"[;\|\n,]+")


# -----------------------------
# Helpers
# -----------------------------
def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())


def best_match(target: str, candidates: List[str]) -> Optional[str]:
    """Lightweight matcher: exact normalized match > substring match."""
    if not candidates:
        return None
    t = _norm(target)
    norm_map = {c: _norm(c) for c in candidates}

    for c, nc in norm_map.items():
        if nc == t:
            return c
    for c, nc in norm_map.items():
        if t and (t in nc or nc in t):
            return c
    return None


def resolve_out_col(wanted: str, out_columns: List[str]) -> Optional[str]:
    """
    Resolve an output column name by normalized match so templates with headers like:
      - "Seeded?"
      - "Star Rating (1-5)"
    still map correctly when the app UI key is "Seeded" / "Star Rating".

    Priority:
      1) exact match
      2) normalized exact match
    """
    if wanted in out_columns:
        return wanted
    wn = _norm(wanted)
    for c in out_columns:
        if _norm(c) == wn:
            return c
    return None


@st.cache_data(show_spinner=False)
def read_table(file_bytes: bytes, filename: str, sheet: Optional[str] = None) -> pd.DataFrame:
    if filename.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes))
    if filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet)
    raise ValueError("Unsupported file type. Please upload a CSV or Excel file.")


def dedupe_preserve(items: List[str]) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for x in items:
        s = str(x).strip()
        if not s:
            continue
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out


def parse_tags(value, split_re: re.Pattern) -> List[str]:
    """
    Parse a cell that may contain:
      - NaN/empty
      - a single string
      - delimited strings (e.g., "A; B | C")
      - stringified list: "['A','B']"
      - actual python list/tuple/set (rare but possible)
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []

    if isinstance(value, (list, tuple, set)):
        return dedupe_preserve([str(v).strip() for v in value if str(v).strip()])

    s = str(value).strip()
    if not s:
        return []

    # Try safe parsing first for list-like strings
    if (s.startswith("[") and s.endswith("]")) or (s.startswith("(") and s.endswith(")")):
        try:
            parsed = ast.literal_eval(s)
            if isinstance(parsed, (list, tuple, set)):
                return dedupe_preserve([str(v).strip().strip("'\"") for v in parsed if str(v).strip()])
        except Exception:
            # Fall back to a simple split if literal_eval fails
            inner = s[1:-1].strip()
            if inner:
                parts = [p.strip().strip("'\"") for p in inner.split(",")]
                return dedupe_preserve([p for p in parts if p])
            return []

    parts = [p.strip() for p in split_re.split(s) if p and p.strip()]
    return dedupe_preserve(parts)


# -----------------------------
# Value normalizers
# -----------------------------
def clean_star_rating_value(v):
    """Coerce star rating into a numeric (int/float) and strip words like 'star(s)'."""
    if v is None:
        return pd.NA
    try:
        if pd.isna(v):
            return pd.NA
    except Exception:
        pass

    if isinstance(v, (int, np.integer)):
        return int(v)
    if isinstance(v, (float, np.floating)):
        try:
            if np.isnan(v):
                return pd.NA
        except Exception:
            pass
        fv = float(v)
        return int(fv) if fv.is_integer() else fv

    s = str(v).strip()
    if not s:
        return pd.NA

    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m:
        return pd.NA

    fv = float(m.group(1))
    return int(fv) if fv.is_integer() else fv


def clean_star_rating_series(series: pd.Series) -> pd.Series:
    return series.apply(clean_star_rating_value).astype("object")


_NUMERIC_RE = re.compile(r"^\s*-?\d+(?:\.\d+)?\s*$")


def clean_seeded_value(v):
    """
    Convert Seeded inputs into:
      - "Yes" for seeded-like values (including ['Seeded'])
      - "no" for anything blank/empty/unseeded/false
    """
    if v is None:
        return "no"
    try:
        if pd.isna(v):
            return "no"
    except Exception:
        pass

    if isinstance(v, (bool, np.bool_)):
        return "Yes" if bool(v) else "no"

    if isinstance(v, (int, np.integer, float, np.floating)):
        try:
            return "Yes" if float(v) == 1.0 else "no"
        except Exception:
            return "no"

    s0 = str(v).strip()
    if s0 and _NUMERIC_RE.match(s0):
        try:
            fv = float(s0)
            if fv == 1.0:
                return "Yes"
            if fv == 0.0:
                return "no"
        except Exception:
            pass

    tokens = parse_tags(v, DEFAULT_TAG_SPLIT_RE)

    for t in tokens:
        raw = str(t).strip().lower()
        nt = _norm(t)

        if nt in {"notseeded", "unseeded", "nonseeded"} or raw in {"not seeded", "unseeded", "non seeded"}:
            return "no"
        if nt in {"false", "no", "n", "0"}:
            return "no"

        if nt in {"seeded", "yes", "true", "y", "1"}:
            return "Yes"
        if raw in {"1.0", "1"}:
            return "Yes"
        if "seeded" in nt:
            return "Yes"

    return "no"


def clean_seeded_series(series: pd.Series) -> pd.Series:
    return series.apply(clean_seeded_value).astype("object")


# -----------------------------
# Template helpers
# -----------------------------
def extract_template_header(template_bytes: bytes, sheet_name: str, keep_trailing_blank_cols: int = 5) -> List[str]:
    """
    Extract header row from an Excel sheet using openpyxl so we don't accidentally
    drop trailing blank header columns (common in the Star Walk template).
    """
    wb = load_workbook(io.BytesIO(template_bytes), data_only=False)
    if sheet_name not in wb.sheetnames:
        return DEFAULT_STARWALK_COLUMNS

    ws = wb[sheet_name]
    raw = []
    for c in range(1, ws.max_column + 1):
        raw.append(ws.cell(1, c).value)

    last_named = -1
    for i, v in enumerate(raw):
        if v is not None and str(v).strip() != "":
            last_named = i

    if last_named == -1:
        return DEFAULT_STARWALK_COLUMNS

    max_keep = min(len(raw), last_named + 1 + keep_trailing_blank_cols)
    headers: List[str] = []
    for i in range(max_keep):
        v = raw[i]
        if v is None or str(v).strip() == "":
            headers.append(f"Unnamed: {i}")
        else:
            headers.append(str(v).strip())
    return headers


def load_allowed_l2_from_template(template_bytes: bytes) -> Tuple[Optional[Set[str]], Optional[Set[str]]]:
    """
    If the template has a 'Symptoms' sheet with columns 'Detractors' and 'Delighters',
    use it as an allow-list for L2 tags.
    """
    try:
        df = pd.read_excel(io.BytesIO(template_bytes), sheet_name="Symptoms")
        if "Detractors" not in df.columns or "Delighters" not in df.columns:
            return None, None
        det = set(df["Detractors"].dropna().astype(str).str.strip())
        deli = set(df["Delighters"].dropna().astype(str).str.strip())
        det = {x for x in det if x}
        deli = {x for x in deli if x}
        return det or None, deli or None
    except Exception:
        return None, None


def collect_from_row_tuple(
    row_tup,
    col_idx: Dict[str, int],
    cols: List[str],
    split_re: re.Pattern,
) -> List[str]:
    tags: List[str] = []
    for c in cols:
        i = col_idx.get(c)
        if i is None:
            continue
        tags.extend(parse_tags(row_tup[i], split_re))
    return dedupe_preserve(tags)


def insert_after_symptom20(out_cols: List[str], extra_cols: List[str]) -> List[str]:
    """
    Insert extra columns immediately after Symptom 20.
    If Symptom 20 not found, append at end.
    De-dupes while preserving order.
    """
    base = list(out_cols)
    extras = [c for c in extra_cols if c and str(c).strip()]
    if not extras:
        return base

    # remove extras if already present so insertion doesn't duplicate
    base_no_extras = [c for c in base if c not in extras]

    try:
        idx = base_no_extras.index("Symptom 20")
        return base_no_extras[: idx + 1] + extras + base_no_extras[idx + 1 :]
    except ValueError:
        return base_no_extras + extras


# -----------------------------
# Conversion
# -----------------------------
def convert_to_starwalk(
    src_df: pd.DataFrame,
    out_cols: List[str],
    field_map: Dict[str, Optional[str]],
    l2_det_cols: List[str],
    l2_del_cols: List[str],
    split_regex: str,
    weight_mode: str,
    allowed_det: Optional[Set[str]] = None,
    allowed_del: Optional[Set[str]] = None,
    filter_unknown: bool = False,
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """
    Build Star Walk scrubbed verbatims output:
      - L2 detractors -> Symptom 1-10
      - L2 delighters -> Symptom 11-20
    """
    src_df = src_df.reset_index(drop=True)
    n = len(src_df)

    needed_symptoms = [f"Symptom {i}" for i in range(1, 21)]
    base_cols = list(out_cols)
    for c in needed_symptoms:
        if c not in base_cols:
            base_cols.append(c)
    if "Review count per detractor" not in base_cols:
        base_cols.append("Review count per detractor")

    out = pd.DataFrame(index=range(n), columns=base_cols, dtype="object")

    # Copy mapped fields (with robust output-column resolving for templates)
    out_columns_list = list(out.columns)
    for ui_out_field, in_field in field_map.items():
        out_col = resolve_out_col(ui_out_field, out_columns_list)
        if not out_col:
            continue

        if in_field and in_field in src_df.columns:
            series = src_df[in_field]
            if _norm(out_col) == "seeded":
                out[out_col] = clean_seeded_series(series)
            elif _norm(out_col) == "starrating":
                out[out_col] = clean_star_rating_series(series)
            else:
                out[out_col] = series.values
        else:
            out[out_col] = pd.NA

    # Safety pass (handles template quirks / duplicate header variants)
    for c in list(out.columns):
        if _norm(c) == "seeded":
            out[c] = clean_seeded_series(out[c])
        if _norm(c) == "starrating":
            out[c] = clean_star_rating_series(out[c])

    split_re = re.compile(split_regex)
    src_cols = list(src_df.columns)
    col_idx = {c: i for i, c in enumerate(src_cols)}

    detr_matrix = [[pd.NA] * 10 for _ in range(n)]
    deli_matrix = [[pd.NA] * 10 for _ in range(n)]
    detr_count = np.zeros(n, dtype=int)
    deli_count = np.zeros(n, dtype=int)

    rows_trunc_det = 0
    rows_trunc_del = 0
    unknown_det_total = 0
    unknown_del_total = 0

    for r_i, row in enumerate(src_df.itertuples(index=False, name=None)):
        d_tags = collect_from_row_tuple(row, col_idx, l2_det_cols, split_re)
        l_tags = collect_from_row_tuple(row, col_idx, l2_del_cols, split_re)

        if allowed_det is not None:
            unknown = [t for t in d_tags if t not in allowed_det]
            if unknown:
                unknown_det_total += len(unknown)
                if filter_unknown:
                    d_tags = [t for t in d_tags if t in allowed_det]

        if allowed_del is not None:
            unknown = [t for t in l_tags if t not in allowed_del]
            if unknown:
                unknown_del_total += len(unknown)
                if filter_unknown:
                    l_tags = [t for t in l_tags if t in allowed_del]

        if len(d_tags) > 10:
            rows_trunc_det += 1
        if len(l_tags) > 10:
            rows_trunc_del += 1

        d_tags = d_tags[:10]
        l_tags = l_tags[:10]

        detr_count[r_i] = len(d_tags)
        deli_count[r_i] = len(l_tags)

        for j, t in enumerate(d_tags):
            detr_matrix[r_i][j] = t
        for j, t in enumerate(l_tags):
            deli_matrix[r_i][j] = t

    detr_cols_out = [f"Symptom {i}" for i in range(1, 11)]
    deli_cols_out = [f"Symptom {i}" for i in range(11, 21)]

    out[detr_cols_out] = pd.DataFrame(detr_matrix, columns=detr_cols_out).values
    out[deli_cols_out] = pd.DataFrame(deli_matrix, columns=deli_cols_out).values

    # Weight column behavior (configurable)
    if "Review count per detractor" in out.columns:
        if weight_mode == "Leave blank":
            out["Review count per detractor"] = pd.NA
        elif weight_mode == "Always 1":
            out["Review count per detractor"] = 1.0
        elif weight_mode == "1 / # detractor symptoms (if any)":
            out["Review count per detractor"] = np.where(detr_count > 0, 1.0 / detr_count, pd.NA)
        elif weight_mode == "1 / # delighter symptoms (if any)":
            out["Review count per detractor"] = np.where(deli_count > 0, 1.0 / deli_count, pd.NA)
        else:  # "1 / # total symptoms (detractors + delighters)"
            total = detr_count + deli_count
            out["Review count per detractor"] = np.where(total > 0, 1.0 / total, pd.NA)

    stats = {
        "rows": n,
        "rows_truncated_detractors_gt10": rows_trunc_det,
        "rows_truncated_delighters_gt10": rows_trunc_del,
        "unknown_detractor_tags_total": unknown_det_total,
        "unknown_delighter_tags_total": unknown_del_total,
    }
    return out, stats


def to_excel_one_sheet(df_out: pd.DataFrame, sheet_name: str = STARWALK_SHEET_NAME) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def write_into_template_workbook(template_bytes: bytes, df_out: pd.DataFrame, sheet_name: str = STARWALK_SHEET_NAME) -> bytes:
    """
    Preserve the full template workbook (other sheets, formulas, pivots, etc.) and
    replace the Star Walk scrubbed verbatims sheet data (keeping its header row).

    If df_out contains columns not in the template header, append them to the end
    of the template header so they are included in the export.
    """
    wb = load_workbook(io.BytesIO(template_bytes), data_only=False)

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    # Read template header (row 1)
    header = []
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(1, c).value
        if v is None or str(v).strip() == "":
            header.append(f"Unnamed: {c-1}")
        else:
            header.append(str(v).strip())

    # If sheet is empty, write header from df_out
    if (ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None):
        header = list(df_out.columns)
        for c, name in enumerate(header, start=1):
            ws.cell(1, c).value = name

    # Append missing columns to header (so export always includes them)
    missing = [c for c in df_out.columns if c not in header]
    if missing:
        start = len(header) + 1
        for i, name in enumerate(missing, start=0):
            ws.cell(1, start + i).value = name
        header = header + missing

    # Align df to header columns
    aligned = df_out.copy()
    for col in header:
        if col not in aligned.columns:
            aligned[col] = pd.NA
    aligned = aligned[header]

    # Replace pandas missing with None (so Excel is truly blank)
    aligned = aligned.replace({pd.NA: None})
    aligned = aligned.where(pd.notna(aligned), None)

    # Clear existing rows below header
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Append new data rows
    for r in dataframe_to_rows(aligned, index=False, header=False):
        ws.append(r)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -----------------------------
# UI
# -----------------------------

# =============================
# Master pipeline helpers (JSON -> Star Walk DF)
# =============================

class _NameOnlyUpload:
    """Fallback when user pastes JSON text (no UploadedFile object)."""
    def __init__(self, name: str):
        self.name = name


def _hash_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def canonicalize_starwalk_columns(df_in: pd.DataFrame) -> pd.DataFrame:
    """
    Standardize common Star Walk column header variants so the dashboard works
    even if a template uses headers like 'Seeded?' or 'Star Rating (1-5)'.
    """
    df = df_in.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # Build a normalized -> actual mapping (first occurrence wins)
    existing = list(df.columns)
    norm_to_col = {}
    for c in existing:
        n = _norm(c)
        if n and n not in norm_to_col:
            norm_to_col[n] = c

    desired = {
        "source": "Source",
        "country": "Country",
        "modelsku": "Model (SKU)",
        "seeded": "Seeded",
        "newreview": "New Review",
        "reviewdate": "Review Date",
        "verbatimid": "Verbatim Id",
        "verbatim": "Verbatim",
        "starrating": "Star Rating",
        "reviewcountperdetractor": "Review count per detractor",
    }

    rename = {}
    for n, canon in desired.items():
        if canon in existing:
            continue
        if n in norm_to_col:
            rename[norm_to_col[n]] = canon

    # Symptoms: Symptom1 / symptom 1 / Symptom_01 -> Symptom 1
    for c in existing:
        n = _norm(c)
        m = re.fullmatch(r"symptom(\d+)", n or "")
        if not m:
            continue
        num = int(m.group(1))
        canon = f"Symptom {num}"
        if canon in existing:
            continue
        if c not in rename:
            rename[c] = canon

    if rename:
        # Avoid collisions: if two columns map to the same canon, keep the first.
        seen_target = set()
        safe_rename = {}
        for src, tgt in rename.items():
            if tgt in seen_target or tgt in df.columns:
                continue
            seen_target.add(tgt)
            safe_rename[src] = tgt
        if safe_rename:
            df = df.rename(columns=safe_rename)

    return df


@st.cache_data(show_spinner=False)
def axion_json_to_starwalk(
    raw_text: str,
    *,
    include_extra: bool = True,
    extra_cols_after_symptom20: Optional[List[str]] = None,
    split_regex: str = r"[;\|\n,]+",
    weight_mode: str = "Leave blank",
    template_bytes: Optional[bytes] = None,
    validate_l2_against_template: bool = False,
    filter_unknown_tags: bool = False,
) -> Tuple[pd.DataFrame, Dict[str, int], pd.DataFrame, List[str]]:
    """
    End-to-end:
      Axion JSON export (dict-with-results, list-of-records, or JSON lines)
        -> Reviews DataFrame (aXreviewsConverter logic)
        -> Star Walk scrubbed verbatims DataFrame (axionReviews logic)

    Returns:
      starwalk_df_raw, stats, reviews_df, warnings
    """
    warnings_list: List[str] = []

    raw_obj, parse_warnings = loads_flexible_json(raw_text)
    warnings_list.extend(parse_warnings)

    records = extract_records(raw_obj)
    if not records:
        raise ValueError("Parsed JSON but found 0 record objects.")

    # Build the intermediate Reviews df (same columns your second script expects)
    reviews_df = build_reviews_df(records, include_extra=include_extra)

    # Output columns: default or template header (then insert extra columns after Symptom 20)
    if extra_cols_after_symptom20 is None:
        extra_cols_after_symptom20 = list(DEFAULT_EXTRA_AFTER_SYMPTOM20)

    if template_bytes:
        # Prefer the standard sheet name if present
        try:
            xls = pd.ExcelFile(io.BytesIO(template_bytes))
            sheet = STARWALK_SHEET_NAME if STARWALK_SHEET_NAME in xls.sheet_names else xls.sheet_names[0]
        except Exception:
            sheet = STARWALK_SHEET_NAME

        base_out_cols = extract_template_header(template_bytes, sheet_name=sheet)
        out_cols = insert_after_symptom20(base_out_cols, extra_cols_after_symptom20)
    else:
        out_cols = insert_after_symptom20(list(DEFAULT_STARWALK_COLUMNS), extra_cols_after_symptom20)

    # Fixed mapping from the JSON→Reviews schema into Star Walk columns
    field_map: Dict[str, Optional[str]] = {
        "Source": "Retailer",
        "Model (SKU)": "Model",
        "Seeded": "Seeded Reviews",
        "Country": "Location",
        "New Review": "Syndicated/Seeded Reviews",
        "Review Date": "Opened Timestamp",
        "Verbatim Id": "Record ID",
        "Verbatim": "Review",
        "Star Rating": "Rating (num)",
        # Weight column is computed inside convert_to_starwalk
    }

    # Copy any optional extra columns (if present)
    for c in extra_cols_after_symptom20:
        if c in reviews_df.columns:
            field_map[c] = c

    # Symptoms mapping:
    # Detractors -> Symptom 1-10   (default: Product_Symptom Conditions)
    # Delighters -> Symptom 11-20  (default: L2 Delighter Condition)
    l2_det_cols = [c for c in ["Product_Symptom Conditions"] if c in reviews_df.columns]
    l2_del_cols = [c for c in ["L2 Delighter Condition"] if c in reviews_df.columns]

    # If those aren't present, fall back to any reasonable candidates
    if not l2_det_cols:
        l2_det_cols = [c for c in ["Product_Symptom Mode", "Product_Symptom Component"] if c in reviews_df.columns]
    if not l2_del_cols:
        l2_del_cols = [c for c in ["L2 Delighter Mode", "L2 Delighter Component"] if c in reviews_df.columns]

    allowed_det = allowed_del = None
    if template_bytes and validate_l2_against_template:
        allowed_det, allowed_del = load_allowed_l2_from_template(template_bytes)

    starwalk_df_raw, stats = convert_to_starwalk(
        src_df=reviews_df,
        out_cols=out_cols,
        field_map=field_map,
        l2_det_cols=l2_det_cols,
        l2_del_cols=l2_del_cols,
        split_regex=split_regex,
        weight_mode=weight_mode,
        allowed_det=allowed_det,
        allowed_del=allowed_del,
        filter_unknown=filter_unknown_tags,
    )

    return starwalk_df_raw, stats, reviews_df, warnings_list


@st.cache_data(show_spinner=False)
def clean_starwalk_df_for_dashboard(df_in: pd.DataFrame) -> pd.DataFrame:
    """
    Apply the same cleaning rules as the dashboard's Excel loader,
    but starting from an in-memory DataFrame.
    """
    df_local = canonicalize_starwalk_columns(df_in)

    for col in ["Country", "Source", "Model (SKU)", "Seeded", "New Review"]:
        if col in df_local.columns:
            df_local[col] = df_local[col].astype("string").str.upper()

    if "Star Rating" in df_local.columns:
        df_local["Star Rating"] = pd.to_numeric(df_local["Star Rating"], errors="coerce")

    all_symptom_cols = [c for c in df_local.columns if str(c).startswith("Symptom")]
    for c in all_symptom_cols:
        df_local[c] = df_local[c].apply(lambda v: clean_text(v, keep_na=True)).astype("string")

    if "Verbatim" in df_local.columns:
        df_local["Verbatim"] = df_local["Verbatim"].astype("string").map(clean_text)

    if "Review Date" in df_local.columns:
        df_local["Review Date"] = pd.to_datetime(df_local["Review Date"], errors="coerce")

    return df_local


# ---------- Data Source ----------
st.markdown("### 📁 Data Source")

data_mode = st.radio(
    "Choose input type",
    options=[
        "Axion JSON → Dashboard (recommended)",
        "Star Walk Excel/CSV (already formatted)",
    ],
    horizontal=True,
    key="data_mode",
)

# ---------- Load & clean (cached) ----------
@st.cache_data(show_spinner=False)
def _load_clean_excel(file_bytes: bytes, file_name: str) -> pd.DataFrame:
    # 1) Try named sheet
    # 2) Try to find a sheet containing "Verbatim"
    # 3) Fallback to first sheet
    try:
        if file_name.lower().endswith(".csv"):
            df_local = pd.read_csv(BytesIO(file_bytes))
        else:
            bio = BytesIO(file_bytes)
            try:
                df_local = pd.read_excel(bio, sheet_name="Star Walk scrubbed verbatims")
            except ValueError:
                # Try find a sheet with Verbatim column
                bio2 = BytesIO(file_bytes)
                xls = pd.ExcelFile(bio2)
                candidate = None
                for sh in xls.sheet_names:
                    try:
                        sample = pd.read_excel(xls, sheet_name=sh, nrows=1)
                        cols = [str(c).strip().lower() for c in sample.columns]
                        if any(c == "verbatim" or c.startswith("verbatim") for c in cols):
                            candidate = sh
                            break
                    except Exception:
                        continue
                if candidate:
                    df_local = pd.read_excel(xls, sheet_name=candidate)
                else:
                    bio3 = BytesIO(file_bytes)
                    df_local = pd.read_excel(bio3)
    except Exception as e:
        raise RuntimeError(f"Could not read file: {e}")

    df_local = canonicalize_starwalk_columns(df_local)

    for col in ["Country", "Source", "Model (SKU)", "Seeded", "New Review"]:
        if col in df_local.columns:
            df_local[col] = df_local[col].astype("string").str.upper()

    if "Star Rating" in df_local.columns:
        df_local["Star Rating"] = pd.to_numeric(df_local["Star Rating"], errors="coerce")

    all_symptom_cols = [c for c in df_local.columns if str(c).startswith("Symptom")]
    for c in all_symptom_cols:
        df_local[c] = df_local[c].apply(lambda v: clean_text(v, keep_na=True)).astype("string")

    if "Verbatim" in df_local.columns:
        df_local["Verbatim"] = df_local["Verbatim"].astype("string").map(clean_text)

    if "Review Date" in df_local.columns:
        df_local["Review Date"] = pd.to_datetime(df_local["Review Date"], errors="coerce")

    return df_local


# --- Resolve dataset into `df` + `uploaded_file` (name used downstream) ---
df = None
uploaded_file = None

if data_mode.startswith("Axion JSON"):
    c1, c2 = st.columns([1.25, 1])

    with c1:
        uploaded_json = st.file_uploader("Upload Axion JSON export", type=["json"], key="json_upload")
        with st.expander("…or paste JSON text (optional)", expanded=False):
            pasted_json = st.text_area(
                "Paste JSON (or JSON Lines)",
                height=140,
                key="json_paste",
                placeholder="Paste JSON here (supports dict-with-results, list-of-records, or JSON Lines).",
            )

    with c2:
        st.markdown("**Conversion options**")
        include_extra = st.checkbox("Include extra taxonomy/tag columns", value=True, key="json_include_extra")

        weight_mode = st.selectbox(
            "Weight column behavior (Review count per detractor)",
            options=[
                "Leave blank",
                "Always 1",
                "1 / # detractor symptoms (if any)",
                "1 / # delighter symptoms (if any)",
                "1 / # total symptoms (detractors + delighters)",
            ],
            index=0,
            key="json_weight_mode",
        )

        include_extra_after20 = st.checkbox(
            "Include extra columns after Symptom 20 (for additional filters)",
            value=True,
            key="json_extra_after20",
        )

        template_file = st.file_uploader(
            "Optional: Star Walk template workbook (xlsx)",
            type=["xlsx", "xlsm", "xls"],
            key="json_template",
            help="If provided, the output will mirror the template's header order (useful for downstream pivots).",
        )

        with st.expander("Advanced parsing / validation", expanded=False):
            split_regex = st.text_input(
                "Tag split regex",
                value=r"[;\|\n,]+",
                key="json_split_regex",
                help="Used to split multi-tag cells into individual detractor/delighter tags.",
            )
            validate_l2 = st.checkbox(
                "Validate detractor/delighter tags against template 'Symptoms' sheet",
                value=False,
                key="json_validate_l2",
                help="If the template has a sheet named 'Symptoms' with columns Detractors/Delighters, unknown tags can be counted (or filtered).",
            )
            filter_unknown = st.checkbox(
                "If validating, filter out unknown tags",
                value=False,
                key="json_filter_unknown",
            )
            auto_build = st.checkbox("Auto-build dashboard on change", value=True, key="json_auto_build")
            rebuild = st.button("Rebuild now", type="primary", key="json_rebuild")

    raw_text = None
    raw_bytes = None
    source_name = None

    if uploaded_json is not None and getattr(uploaded_json, "size", 0) > 0:
        raw_bytes = uploaded_json.getvalue()
        raw_text = raw_bytes.decode("utf-8", errors="replace")
        source_name = uploaded_json.name
        uploaded_file = uploaded_json
    elif st.session_state.get("json_paste", "").strip():
        raw_text = st.session_state["json_paste"]
        raw_bytes = raw_text.encode("utf-8", errors="replace")
        source_name = "pasted_json"
        uploaded_file = _NameOnlyUpload(source_name)

    if not raw_text:
        st.info("Upload a JSON file (or paste JSON text) to build the dashboard.")
        st.stop()

    # Determine which extra columns to insert after Symptom 20
    extra_cols_after20 = list(DEFAULT_EXTRA_AFTER_SYMPTOM20) if include_extra_after20 else []

    template_bytes = template_file.getvalue() if template_file is not None else None

    ds_key = "|".join(
        [
            "json",
            _hash_bytes(raw_bytes or b""),
            str(include_extra),
            str(weight_mode),
            str(extra_cols_after20),
            _hash_bytes(template_bytes) if template_bytes else "no_template",
            str(validate_l2),
            str(filter_unknown),
            str(split_regex),
        ]
    )

    # Rebuild if new dataset/options, or if user clicked rebuild
    need_rebuild = st.session_state.get("_master_ds_key") != ds_key
    should_rebuild = rebuild or (st.session_state.get("json_auto_build", True) and need_rebuild)

    if should_rebuild:
        with st.spinner("Converting Axion JSON → Star Walk scrubbed verbatims…"):
            try:
                sw_raw, stats, reviews_df, warnings_list = axion_json_to_starwalk(
                    raw_text,
                    include_extra=include_extra,
                    extra_cols_after_symptom20=extra_cols_after20,
                    split_regex=split_regex,
                    weight_mode=weight_mode,
                    template_bytes=template_bytes,
                    validate_l2_against_template=validate_l2,
                    filter_unknown_tags=filter_unknown,
                )
                sw_clean = clean_starwalk_df_for_dashboard(sw_raw)

                st.session_state["_master_ds_key"] = ds_key
                st.session_state["_master_source_name"] = source_name
                st.session_state["_master_sw_raw"] = sw_raw
                st.session_state["_master_sw_clean"] = sw_clean
                st.session_state["_master_reviews_df"] = reviews_df
                st.session_state["_master_stats"] = stats
                st.session_state["_master_warnings"] = warnings_list

                # Scroll to top on new dataset
                st.session_state["last_uploaded_name"] = source_name
                st.session_state["force_scroll_top_once"] = True
            except Exception as e:
                st.error(f"Conversion failed: {e}")
                st.stop()

    # If we still don't have a built dataset, prompt user
    if st.session_state.get("_master_ds_key") != ds_key:
        st.info("Ready to convert. Click **Rebuild now** (or enable Auto-build).")
        st.stop()

    df = st.session_state.get("_master_sw_clean")
    source_name = st.session_state.get("_master_source_name", source_name)

    # Brief status + downloads
    with st.expander("✅ Conversion status / downloads", expanded=False):
        if st.session_state.get("_master_warnings"):
            for w in st.session_state["_master_warnings"]:
                st.info(w)

        stats = st.session_state.get("_master_stats") or {}
        if stats:
            st.json(stats)

        # Preview
        try:
            st.caption("Preview — Star Walk scrubbed verbatims (first 25 rows)")
            st.dataframe(st.session_state["_master_sw_raw"].head(25), use_container_width=True)
        except Exception:
            pass

        # Downloads
        try:
            starwalk_bytes = to_excel_one_sheet(st.session_state["_master_sw_raw"], sheet_name=STARWALK_SHEET_NAME)
            st.download_button(
                "Download Star Walk Excel (one sheet)",
                data=starwalk_bytes,
                file_name=f"{Path(source_name).stem}_starwalk_scrubbed_verbatims.xlsx" if source_name else "starwalk_scrubbed_verbatims.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.warning(f"Could not generate Star Walk Excel download: {e}")

        if template_bytes:
            try:
                templ_out = write_into_template_workbook(template_bytes, st.session_state["_master_sw_raw"], sheet_name=STARWALK_SHEET_NAME)
                st.download_button(
                    "Download Star Walk Excel (filled template)",
                    data=templ_out,
                    file_name=f"{Path(source_name).stem}_starwalk_filled_template.xlsx" if source_name else "starwalk_filled_template.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.warning(f"Could not write into template workbook: {e}")

        # Optional: intermediate "Clean Reviews" workbook (original JSON→Excel app)
        try:
            st.markdown("---")
            st.caption("Optional: download the intermediate workbook (Reviews / Summary / Symptoms).")
            if st.checkbox("Enable intermediate workbook download", value=False, key="enable_intermediate"):
                # Re-parse records only when requested (avoids extra work for dashboard-only use)
                raw_obj, _warn = loads_flexible_json(raw_text)
                records = extract_records(raw_obj)
                dataset_title = title_from_filename(source_name) if source_name else "Axion Export"
                reviews_df = st.session_state.get("_master_reviews_df") or build_reviews_df(records, include_extra=include_extra)
                symptoms_df = build_symptoms_df(records, include_blank_row_when_missing=True)
                summary = build_summary_tables(reviews_df, symptoms_df, top_n=10)
                wb_bytes = build_workbook(
                    dataset_title=f"{dataset_title} Reviews",
                    reviews_df=reviews_df,
                    symptoms_df=symptoms_df,
                    summary=summary,
                    wrap_long_text=False,
                )
                st.download_button(
                    "Download Clean Reviews workbook (3 sheets)",
                    data=wb_bytes,
                    file_name=f"{Path(source_name).stem}_clean_reviews.xlsx" if source_name else "clean_reviews.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.warning(f"Intermediate workbook download unavailable: {e}")

else:
    uploaded_file = st.file_uploader("Upload your Star Walk Excel/CSV file", type=["xlsx", "csv"], key="excel_upload")

    if uploaded_file and st.session_state.get("last_uploaded_name") != uploaded_file.name:
        st.session_state["last_uploaded_name"] = uploaded_file.name
        st.session_state["force_scroll_top_once"] = True

    if not uploaded_file:
        st.info("Please upload a Star Walk Excel/CSV file to get started.")
        st.stop()

    try:
        df = _load_clean_excel(uploaded_file.getvalue(), uploaded_file.name)
    except Exception as e:
        st.error(str(e))
        st.stop()


# ---------- Sidebar filters ----------
st.sidebar.header("🔍 Filters")

with st.sidebar.expander("🗓️ Timeframe", expanded=False):
    tz_today = datetime.now(_NY_TZ).date() if _NY_TZ else datetime.today().date()
    timeframe = st.selectbox(
        "Select Timeframe",
        options=["All Time", "Last Week", "Last Month", "Last Year", "Custom Range"],
        key="tf",
    )
    today = tz_today
    start_date, end_date = None, None
    if timeframe == "Custom Range":
        sel = st.date_input(
            label="Date Range",
            value=(today - timedelta(days=30), today),
            min_value=datetime(2000, 1, 1).date(),
            max_value=today,
            label_visibility="collapsed",
        )
        if isinstance(sel, tuple) and len(sel) == 2:
            start_date, end_date = sel
        else:
            start_date = end_date = sel
    elif timeframe == "Last Week":
        start_date, end_date = today - timedelta(days=7), today
    elif timeframe == "Last Month":
        start_date, end_date = today - timedelta(days=30), today
    elif timeframe == "Last Year":
        start_date, end_date = today - timedelta(days=365), today

filtered = df.copy()
if start_date and end_date and "Review Date" in filtered.columns:
    dt = pd.to_datetime(filtered["Review Date"], errors="coerce")
    end_inclusive = pd.Timestamp(end_date) + pd.Timedelta(days=1) - pd.Timedelta(nanoseconds=1)
    filtered = filtered[(dt >= pd.Timestamp(start_date)) & (dt <= end_inclusive)]

with st.sidebar.expander("🌟 Star Rating", expanded=False):
    selected_ratings = st.multiselect("Select Star Ratings", options=["All"] + [1, 2, 3, 4, 5], default=["All"], key="sr")
if "All" not in selected_ratings and "Star Rating" in filtered.columns:
    filtered = filtered[filtered["Star Rating"].isin(selected_ratings)]

with st.sidebar.expander("🌍 Standard Filters", expanded=False):
    filtered, _ = apply_filter(filtered, "Country", "Country", key="f_Country")
    filtered, _ = apply_filter(filtered, "Source", "Source", key="f_Source")
    filtered, _ = apply_filter(filtered, "Model (SKU)", "Model (SKU)", key="f_Model (SKU)")
    filtered, _ = apply_filter(filtered, "Seeded", "Seeded", key="f_Seeded")
    filtered, _ = apply_filter(filtered, "New Review", "New Review", key="f_New Review")

detractor_columns = [f"Symptom {i}" for i in range(1, 11)]
delighter_columns = [f"Symptom {i}" for i in range(11, 21)]
existing_detractor_columns = [c for c in detractor_columns if c in filtered.columns]
existing_delighter_columns = [c for c in delighter_columns if c in filtered.columns]
detractor_symptoms = collect_unique_symptoms(filtered, existing_detractor_columns)
delighter_symptoms = collect_unique_symptoms(filtered, existing_delighter_columns)

with st.sidebar.expander("🩺 Review Symptoms", expanded=False):
    selected_delighter = st.multiselect(
        "Select Delighter Symptoms",
        options=["All"] + sorted(delighter_symptoms),
        default=["All"],
        key="delight",
    )
    selected_detractor = st.multiselect(
        "Select Detractor Symptoms",
        options=["All"] + sorted(detractor_symptoms),
        default=["All"],
        key="detract",
    )
if "All" not in selected_delighter and existing_delighter_columns:
    filtered = filtered[filtered[existing_delighter_columns].isin(selected_delighter).any(axis=1)]
if "All" not in selected_detractor and existing_detractor_columns:
    filtered = filtered[filtered[existing_detractor_columns].isin(selected_detractor).any(axis=1)]

with st.sidebar.expander("🔎 Keyword", expanded=False):
    keyword = st.text_input(
        "Keyword to search (in review text)",
        value="",
        key="kw",
        help="Case-insensitive match in review text. Cleans â€™ → '",
    )
    if keyword and "Verbatim" in filtered.columns:
        mask_kw = filtered["Verbatim"].astype("string").fillna("").str.contains(keyword.strip(), case=False, na=False)
        filtered = filtered[mask_kw]

core_cols = {"Country", "Source", "Model (SKU)", "Seeded", "New Review", "Star Rating", "Review Date", "Verbatim"}
symptom_cols = set([f"Symptom {i}" for i in range(1, 21)])
with st.sidebar.expander("📋 Additional Filters", expanded=False):
    additional_columns = [c for c in df.columns if c not in (core_cols | symptom_cols)]
    if additional_columns:
        for column in additional_columns:
            filtered, _ = apply_filter(filtered, column, column, key=f"f_{column}")
    else:
        st.info("No additional filters available.")

with st.sidebar.expander("📄 Review List", expanded=False):
    rpp_options = [10, 20, 50, 100]
    default_rpp = st.session_state.get("reviews_per_page", 10)
    rpp_index = rpp_options.index(default_rpp) if default_rpp in rpp_options else 0
    rpp = st.selectbox("Reviews per page", options=rpp_options, index=rpp_index, key="rpp")
    if rpp != default_rpp:
        st.session_state["reviews_per_page"] = rpp
        st.session_state["review_page"] = 0

# Clear filters
if st.sidebar.button("🧹 Clear all filters"):
    for k in [
        "tf",
        "sr",
        "kw",
        "delight",
        "detract",
        "rpp",
        "review_page",
        "llm_model",
        "llm_model_label",
        "llm_temp",
        "ai_enabled",
        "ai_cap",
        "use_bm25",
        "use_reranker",
        "ask_q",
        "product_summary_text",
    ] + [k for k in list(st.session_state.keys()) if k.startswith("f_")]:
        if k in st.session_state:
            del st.session_state[k]
    st.rerun()

# ---------- AI / Privacy & Local Intelligence ----------
with st.sidebar.expander("🤖 AI Assistant (LLM)", expanded=False):
    _model_choices = [
        ("Fast & economical – 4o-mini", "gpt-4o-mini"),
        ("Balanced – 4o", "gpt-4o"),
        ("Advanced – 4.1", "gpt-4.1"),
        ("Most advanced – GPT-5", "gpt-5"),
        ("GPT-5 (Chat latest)", "gpt-5-chat-latest"),
    ]
    _default_model = st.session_state.get("llm_model", "gpt-4o-mini")
    _default_idx = next((i for i, (_, mid) in enumerate(_model_choices) if mid == _default_model), 0)
    _label = st.selectbox("Model", options=[l for (l, _) in _model_choices], index=_default_idx, key="llm_model_label")
    st.session_state["llm_model"] = dict(_model_choices)[_label]

    temp_supported = model_supports_temperature(st.session_state["llm_model"])
    st.session_state["llm_temp"] = st.slider(
        "Creativity (temperature)",
        min_value=0.0,
        max_value=1.0,
        value=float(st.session_state.get("llm_temp", 0.2)),
        step=0.1,
        disabled=not temp_supported,
        help=("Lower = more deterministic, higher = more creative. Some models ignore this."),
    )
    if not temp_supported:
        st.caption("ℹ️ This model uses a fixed temperature; the slider is disabled.")

with st.sidebar.expander("🔒 Privacy & AI", expanded=True):
    ai_enabled = st.toggle(
        "Enable AI (send filtered text to OpenAI)",
        value=True,
        key="ai_enabled",
        help="When on and you ask a question, short masked snippets and embeddings are sent to OpenAI.",
    )
    ai_cap = st.number_input(
        "Max reviews to embed per Q",
        min_value=200,
        max_value=5000,
        value=int(st.session_state.get("ai_cap", 1500)),
        step=100,
        key="ai_cap",
    )
    st.caption("Emails/phone numbers are masked before embedding. No remote calls happen until you press Send / Generate.")

with st.sidebar.expander("🧠 Local Intelligence", expanded=False):
    st.toggle("Use BM25 blend (local, fast)", value=st.session_state.get("use_bm25", False), key="use_bm25")
    st.toggle("Use cross-encoder reranker (local, slower)", value=st.session_state.get("use_reranker", False), key="use_reranker")

with st.sidebar.expander("🔑 OpenAI Key (optional override)", expanded=False):
    st.text_input("OPENAI_API_KEY override", value="", type="password", key="api_key_override")
    st.caption("Leave blank to use .streamlit/secrets.toml or environment variable.")

# ---------- Dataset Overview (new; no AI calls) ----------
product_label = _infer_product_label(filtered, uploaded_file.name)
st.markdown("## 🧾 Dataset Overview")
st.markdown(
    f"""
<div class="soft-panel">
  <div><span class="kpi-pill">File</span> <b>{esc(uploaded_file.name)}</b></div>
  <div style="margin-top:6px;"><span class="kpi-pill">Product guess</span> <b>{esc(product_label)}</b></div>
  <div class="small-muted" style="margin-top:8px;">
    Overview is based on the <b>current filtered</b> dataset.
  </div>
</div>
""",
    unsafe_allow_html=True,
)

# Precompute symptom summaries once for reuse (charts + tables + AI context)
detractors_results_full = analyze_delighters_detractors(filtered, existing_detractor_columns)
delighters_results_full = analyze_delighters_detractors(filtered, existing_delighter_columns)

# Trend watchouts (local)
_all_sym_cols = [c for c in [f"Symptom {i}" for i in range(1, 21)] if c in filtered.columns]
trend_watchouts = _detect_trends(filtered, symptom_cols=_all_sym_cols, min_mentions=3)

if trend_watchouts:
    with st.expander("⚠️ Watchouts & Recent Movement (local)", expanded=False):
        st.markdown("\n".join([f"- {t}" for t in trend_watchouts]))

# ---------- Preset Questions (new) ----------
anchor("askdata-anchor")
st.markdown("## 🤖 Ask your data")

if "ask_q" not in st.session_state:
    st.session_state["ask_q"] = ""

st.markdown(
    """
<div class="soft-panel">
  <div style="font-weight:800; font-size:1.05rem;">⚡ Preset Insight Questions</div>
  <div class="small-muted" style="margin-top:4px;">
    Click one to prefill the question box (then click <b>Send</b>). No remote calls until you press Send.
  </div>
</div>
""",
    unsafe_allow_html=True,
)

b1, b2, b3, b4, b5, b6 = st.columns([1, 1, 1, 1, 1, 1])
if b1.button("Executive Summary", use_container_width=True):
    st.session_state["ask_q"] = (
        "Provide an executive summary of this product based on the CURRENT filtered reviews. "
        "Quantify key points (counts/percentages), list top delighters and detractors, and include 4–6 evidence quotes."
    )
    scroll_to("askdata-anchor")
if b2.button("Top Delighters", use_container_width=True):
    st.session_state["ask_q"] = "What are the top delighters? Quantify mentions and include evidence quotes."
    scroll_to("askdata-anchor")
if b3.button("Top Detractors", use_container_width=True):
    st.session_state["ask_q"] = "What are the top detractors? Quantify mentions and include evidence quotes."
    scroll_to("askdata-anchor")
if b4.button("Biggest Improvements", use_container_width=True):
    st.session_state["ask_q"] = (
        "What are the biggest improvement opportunities for this product? "
        "Quantify how common each issue is, and recommend 3–6 actions prioritized by impact."
    )
    scroll_to("askdata-anchor")
if b5.button("Trends / Watchouts", use_container_width=True):
    st.session_state["ask_q"] = "Are there any emerging trends or watchouts? Quantify changes and include evidence."
    scroll_to("askdata-anchor")
if b6.button("Seeded vs Organic", use_container_width=True):
    st.session_state["ask_q"] = "Compare Seeded vs Organic. Quantify the avg rating gap, volumes, and key theme differences."
    scroll_to("askdata-anchor")

st.markdown(
    '<div class="callout warn">⚠️ AI can make mistakes. Numbers are computed from the filtered data via tools/context, '
    "but always double-check important conclusions.</div>",
    unsafe_allow_html=True,
)

api_key_override = (st.session_state.get("api_key_override") or "").strip()
api_key = api_key_override or st.secrets.get("OPENAI_API_KEY", os.getenv("OPENAI_API_KEY"))

if not _HAS_OPENAI:
    st.info("To enable remote LLM, add `openai` to requirements and set `OPENAI_API_KEY`. Local Q&A still works.")
elif not api_key:
    st.info("Set `OPENAI_API_KEY` (env or .streamlit/secrets.toml) for remote LLM. Local Q&A still works.")

# ---------- AI Product Summary (new; button-triggered) ----------
with st.expander("🪄 AI Product Summary (narrative, button-triggered)", expanded=False):
    st.caption("Generates a concise narrative summary with quantified delighters/detractors + evidence quotes. No remote call until you click.")

    def _dataset_fingerprint(df_in: pd.DataFrame) -> str:
        parts = [f"rows={len(df_in)}"]
        if "Review Date" in df_in.columns:
            dmin = pd.to_datetime(df_in["Review Date"], errors="coerce").min()
            dmax = pd.to_datetime(df_in["Review Date"], errors="coerce").max()
            parts.append(f"date_min={dmin}")
            parts.append(f"date_max={dmax}")
        if "Star Rating" in df_in.columns:
            sr = pd.to_numeric(df_in["Star Rating"], errors="coerce")
            parts.append(f"avg={sr.mean():.5f}")
            parts.append(f"sum={sr.sum():.5f}")
        if "Model (SKU)" in df_in.columns:
            top = df_in["Model (SKU)"].astype("string").value_counts().head(3).index.tolist()
            parts.append("sku_top=" + ",".join([str(x) for x in top]))
        return hashlib.sha256("||".join(parts).encode("utf-8")).hexdigest()

    def _metrics_snapshot(df_in: pd.DataFrame) -> dict:
        s = pd.to_numeric(df_in.get("Star Rating"), errors="coerce")
        total = int(len(df_in))
        return {
            "product_guess": product_label,
            "total_reviews": total,
            "avg_star": float(s.mean()) if total and "Star Rating" in df_in.columns else 0.0,
            "median_star": float(s.median()) if total and "Star Rating" in df_in.columns else 0.0,
            "pct_1_2": float((s <= 2).mean() * 100) if total and "Star Rating" in df_in.columns else 0.0,
            "star_counts": s.value_counts().sort_index().to_dict() if "Star Rating" in df_in.columns else {},
        }

    def _top_table(df_tbl: pd.DataFrame, k: int = 6) -> list[dict]:
        if df_tbl is None or df_tbl.empty:
            return []
        out = []
        for _, r in df_tbl.head(k).iterrows():
            out.append(
                {
                    "item": str(r.get("Item", "")),
                    "mentions": int(r.get("Mentions", 0)),
                    "avg_star": None if pd.isna(r.get("Avg Star")) else float(r.get("Avg Star")),
                    "pct_total": str(r.get("% Total", "")),
                }
            )
        return out

    if st.button("Generate / Refresh AI Product Summary", key="gen_product_summary"):
        if not ai_enabled or (not _HAS_OPENAI) or (not api_key):
            st.warning("AI is disabled or no API key available.")
        else:
            # Build context (local, accurate)
            m = _metrics_snapshot(filtered)
            top_del = _top_table(delighters_results_full, 6)
            top_det = _top_table(detractors_results_full, 6)

            # Evidence quotes (local pick from top symptoms)
            quotes = []
            if top_del:
                sym = top_del[0]["item"].title()
                q = _pick_quotes_for_symptom(filtered, sym, existing_delighter_columns, k=2, prefer="high")
                for i, qq in enumerate(q):
                    quotes.append(f"[Q{len(quotes)+1}] “{qq['text']}” — {qq['meta']}")
            if top_det:
                sym = top_det[0]["item"].title()
                q = _pick_quotes_for_symptom(filtered, sym, existing_detractor_columns, k=2, prefer="low")
                for i, qq in enumerate(q):
                    quotes.append(f"[Q{len(quotes)+1}] “{qq['text']}” — {qq['meta']}")

            # Add watchouts lines
            watch = trend_watchouts[:6] if trend_watchouts else []

            ctx = {
                "metrics": m,
                "top_delighters": top_del,
                "top_detractors": top_det,
                "watchouts": watch,
                "evidence_quotes": quotes[:6],
            }

            selected_model = st.session_state.get("llm_model", "gpt-4o-mini")
            llm_temp = float(st.session_state.get("llm_temp", 0.2))

            sys_ctx = (
                "You are a consumer insights expert analyzing ONLY the provided product review dataset.\n"
                "Rules:\n"
                "- Use ONLY the data in DATA_JSON for any numbers.\n"
                "- If you mention a number, it must appear in DATA_JSON.\n"
                "- Include evidence quotes by referencing quote IDs like [Q1], [Q2].\n"
                "- Be concise and executive-friendly.\n"
                "- Provide prioritized recommendations (what to do next) and explain why.\n"
            )

            user_prompt = (
                "Generate an AI product summary for stakeholders.\n\n"
                "Include sections:\n"
                "1) Executive summary (2–4 bullets)\n"
                "2) Top delighters (quantified)\n"
                "3) Top detractors (quantified)\n"
                "4) Watchouts / recent movement (if any)\n"
                "5) Recommended actions (prioritized)\n"
                "6) Evidence (brief) — cite quote IDs in-line\n\n"
                f"DATA_JSON:\n{json.dumps(ctx, ensure_ascii=False)}"
            )

            try:
                client = OpenAI(api_key=api_key)
                req = {
                    "model": selected_model,
                    "messages": [{"role": "system", "content": sys_ctx}, {"role": "user", "content": user_prompt}],
                }
                if model_supports_temperature(selected_model):
                    req["temperature"] = llm_temp

                with st.spinner("Generating summary…"):
                    resp = client.chat.completions.create(**req)
                st.session_state["product_summary_text"] = resp.choices[0].message.content
            except Exception as e:
                st.error(f"Could not generate summary: {e}")

    if st.session_state.get("product_summary_text"):
        st.markdown(st.session_state["product_summary_text"])
        with st.expander("🔎 Summary context (what the AI saw)", expanded=False):
            st.json(
                {
                    "metrics": _metrics_snapshot(filtered),
                    "top_delighters": _top_table(delighters_results_full, 6),
                    "top_detractors": _top_table(detractors_results_full, 6),
                    "watchouts": trend_watchouts[:6],
                }
            )

# ---------- Ask form ----------
with st.form("ask_ai_form", clear_on_submit=True):
    q = st.text_area("Ask a question", key="ask_q", height=90)
    send = st.form_submit_button("Send")

if send and q.strip():
    q = q.strip()

    # Build local search index ON DEMAND (no remote calls)
    verb_series_all = filtered.get("Verbatim", pd.Series(dtype=str)).fillna("").astype(str).map(clean_text)
    local_texts = verb_series_all.tolist()
    content_hash_local = _hash_texts(local_texts)
    local_index = _get_or_build_local_text_index(local_texts, content_hash_local)

    def _get_local_quotes(question: str):
        hits = _local_search(question, local_index, top_k=10)
        quotes = []
        for txt, score in hits[:6]:
            s = (txt or "").strip()
            s = _mask_pii(s)
            if len(s) > 320:
                s = s[:317] + "…"
            if s:
                quotes.append(s)
        return quotes

    final_text = None

    # Safe helpers for pandas.query
    def _safe_query(qs: str) -> bool:
        if not qs or len(qs) > 200:
            return False
        bad = ["__", "@", "import", "exec", "eval", "os.", "pd.", "open(", "read", "write", "globals", "locals", "`", ";", "\n"]
        if any(t in qs.lower() for t in bad):
            return False
        return bool(re.fullmatch(r"[A-Za-z0-9_ .<>=!&|()'\"-]+", qs))

    # tool functions (existing)
    def pandas_count(query: str) -> dict:
        try:
            if not _safe_query(query):
                return {"error": "unsafe query"}
            res = filtered.query(query)
            return {"count": int(len(res))}
        except Exception as e:
            return {"error": str(e)}

    def pandas_mean(column: str, query: str | None = None) -> dict:
        try:
            if column not in filtered.columns:
                return {"error": f"Unknown column {column}"}
            d = filtered if not query else (filtered.query(query) if _safe_query(query) else None)
            if d is None:
                return {"error": "unsafe query"}
            return {"mean": float(pd.to_numeric(d[column], errors="coerce").mean())}
        except Exception as e:
            return {"error": str(e)}

    def symptom_stats(symptom: str) -> dict:
        cols = [c for c in [f"Symptom {i}" for i in range(1, 21)] if c in filtered.columns]
        if not cols:
            return {"count": 0, "avg_star": None}
        mask = filtered[cols].isin([symptom]).any(axis=1)
        d = filtered[mask]
        return {
            "count": int(len(d)),
            "avg_star": float(pd.to_numeric(d["Star Rating"], errors="coerce").mean()) if len(d) and "Star Rating" in d.columns else None,
        }

    def keyword_stats(term: str) -> dict:
        if "Verbatim" not in filtered.columns:
            return {"count": 0, "pct": 0.0}
        ser = filtered["Verbatim"].astype("string").fillna("")
        cnt = int(ser.str.contains(term, case=False, na=False).sum())
        pct = (cnt / max(1, len(filtered))) * 100.0
        return {"count": cnt, "pct": pct}

    def get_metrics_snapshot():
        try:
            s = pd.to_numeric(filtered.get("Star Rating"), errors="coerce")
            return {
                "total_reviews": int(len(filtered)),
                "avg_star": float(s.mean()) if len(filtered) and "Star Rating" in filtered.columns else 0.0,
                "median_star": float(s.median()) if len(filtered) and "Star Rating" in filtered.columns else 0.0,
                "low_star_pct_1_2": float((s <= 2).mean() * 100) if len(filtered) and "Star Rating" in filtered.columns else 0.0,
                "star_counts": s.value_counts().sort_index().to_dict() if "Star Rating" in filtered.columns else {},
            }
        except Exception as e:
            return {"error": str(e)}

    # NEW tools for better quant + evidence
    def top_symptoms(which: str, k: int = 10) -> dict:
        try:
            which_l = (which or "").strip().lower()
            if which_l not in {"detractors", "delighters"}:
                return {"error": "which must be 'detractors' or 'delighters'"}
            cols = existing_detractor_columns if which_l == "detractors" else existing_delighter_columns
            tbl = analyze_delighters_detractors(filtered, cols).head(int(k))
            return {"rows": tbl.to_dict(orient="records")}
        except Exception as e:
            return {"error": str(e)}

    def monthly_volume_and_mean() -> dict:
        try:
            if "Review Date" not in filtered.columns or "Star Rating" not in filtered.columns:
                return {"error": "missing columns"}
            d = filtered.copy()
            d["Review Date"] = pd.to_datetime(d["Review Date"], errors="coerce")
            d["Star Rating"] = pd.to_numeric(d["Star Rating"], errors="coerce")
            d = d.dropna(subset=["Review Date", "Star Rating"])
            if d.empty:
                return {"rows": []}
            d["month"] = d["Review Date"].dt.to_period("M").dt.to_timestamp()
            g = d.groupby("month")["Star Rating"].agg(count="count", mean="mean").reset_index().sort_values("month")
            rows = [{"month": r["month"].strftime("%Y-%m"), "count": int(r["count"]), "mean": float(r["mean"])} for _, r in g.iterrows()]
            return {"rows": rows}
        except Exception as e:
            return {"error": str(e)}

    # Default: high-quality local answer + quotes
    local_quotes = _get_local_quotes(q)
    local_answer = _route_and_answer_locally(q, filtered, local_quotes)

    # If AI disabled or no key/client, just show local
    if not ai_enabled or (not _HAS_OPENAI) or (not api_key):
        final_text = local_answer
        evidence_pack = {"local_quotes": local_quotes[:6], "retrieved_quotes": []}
    else:
        # Remote LLM path (tools-first), ONLY after we have retrieval & numbers ready
        try:
            # Prepare masked texts and cap for embedding (keep your existing approach; add local evidence regardless)
            raw_texts_all = [_mask_pii(t) for t in local_texts]

            # Cap deterministically: sort by MD5 and take first ai_cap
            def _stable_top_k(texts: list[str], k: int) -> list[str]:
                if k >= len(texts):
                    return texts
                keyed = [(hashlib.md5((t or "").encode()).hexdigest(), t) for t in texts]
                keyed.sort(key=lambda x: x[0])
                return [t for _, t in keyed[:k]]

            cap_n = int(ai_cap)
            raw_texts = _stable_top_k(raw_texts_all, cap_n)
            emb_model = "text-embedding-3-small"
            content_hash = _hash_texts(raw_texts)

            _ = _ensure_cache_slot(emb_model, content_hash, api_key)  # dummy cached marker
            index = _get_or_build_index(content_hash, raw_texts, api_key, emb_model)

            retrieved = vector_search(q, index, api_key, top_k=10) if index else []
            retrieved_quotes = []
            for txt, sim in retrieved[:6]:
                s = (txt or "").strip()
                if len(s) > 320:
                    s = s[:317] + "…"
                if s:
                    retrieved_quotes.append(s)

            # Build strong context pack (local, accurate)
            snap = get_metrics_snapshot()
            top_det = detractors_results_full.head(8).to_dict(orient="records")
            top_del = delighters_results_full.head(8).to_dict(orient="records")

            # Add a couple of symptom-specific representative quotes (better evidence)
            symptom_quotes = []
            if top_det:
                sym = str(top_det[0].get("Item", "")).title()
                symptom_quotes.extend(_pick_quotes_for_symptom(filtered, sym, existing_detractor_columns, k=2, prefer="low"))
            if top_del:
                sym = str(top_del[0].get("Item", "")).title()
                symptom_quotes.extend(_pick_quotes_for_symptom(filtered, sym, existing_delighter_columns, k=2, prefer="high"))

            # Build evidence list with IDs
            evidence_lines = []
            # Local search quotes (full corpus coverage)
            for i, s in enumerate(local_quotes[:5], start=1):
                evidence_lines.append(f"[L{i}] “{_mask_pii(s)}”")
            # Vector retrieved quotes (semantic)
            for i, s in enumerate(retrieved_quotes[:5], start=1):
                evidence_lines.append(f"[R{i}] “{_mask_pii(s)}”")
            # Symptom-based picked quotes with attribution
            for i, qq in enumerate(symptom_quotes[:4], start=1):
                evidence_lines.append(f"[S{i}] “{qq.get('text','')}” — {qq.get('meta','')}")

            evidence_pack = {"local_quotes": local_quotes[:6], "retrieved_quotes": retrieved_quotes[:6]}

            selected_model = st.session_state.get("llm_model", "gpt-4o-mini")
            llm_temp = float(st.session_state.get("llm_temp", 0.2))

            tools = [
                {
                    "type": "function",
                    "function": {
                        "name": "pandas_count",
                        "description": "Count rows matching a pandas query over the CURRENT filtered dataset.",
                        "parameters": {"type": "object", "properties": {"query": {"type": "string"}}, "required": ["query"]},
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "pandas_mean",
                        "description": "Mean of a numeric column (optional pandas query).",
                        "parameters": {"type": "object", "properties": {"column": {"type": "string"}, "query": {"type": "string"}}, "required": ["column"]},
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "symptom_stats",
                        "description": "Mentions + avg star for a symptom across Symptom 1–20.",
                        "parameters": {"type": "object", "properties": {"symptom": {"type": "string"}}, "required": ["symptom"]},
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "keyword_stats",
                        "description": "Count + % of reviews with term in Verbatim.",
                        "parameters": {"type": "object", "properties": {"term": {"type": "string"}}, "required": ["term"]},
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "get_metrics_snapshot",
                        "description": "Return current page metrics snapshot.",
                        "parameters": {"type": "object", "properties": {}},
                    },
                },
                # New tools
                {
                    "type": "function",
                    "function": {
                        "name": "top_symptoms",
                        "description": "Return top delighters or detractors with mentions, avg star, and % total.",
                        "parameters": {
                            "type": "object",
                            "properties": {"which": {"type": "string"}, "k": {"type": "integer"}},
                            "required": ["which"],
                        },
                    },
                },
                {
                    "type": "function",
                    "function": {
                        "name": "monthly_volume_and_mean",
                        "description": "Return monthly review volume and avg star for the filtered dataset.",
                        "parameters": {"type": "object", "properties": {}},
                    },
                },
            ]

            # Strong system context (quant + evidence oriented)
            sys_ctx = (
                "You are a consumer insights expert for customer reviews.\n"
                "Hard rules:\n"
                "- Use ONLY numbers from METRICS_JSON or tool outputs. Do NOT invent numbers.\n"
                "- If you mention a number, it must be present in METRICS_JSON or a tool result.\n"
                "- Support claims using evidence quotes. Cite quotes by ID (e.g., [L1], [R2], [S1]).\n"
                "- If evidence is weak, say so.\n\n"
                f"PRODUCT_GUESS={product_label}\n"
                f"ROW_COUNT={len(filtered)}\n\n"
                f"METRICS_JSON:\n{json.dumps(snap, ensure_ascii=False)}\n\n"
                f"TOP_DETRACTORS_JSON:\n{json.dumps(top_det, ensure_ascii=False)}\n\n"
                f"TOP_DELIGHTERS_JSON:\n{json.dumps(top_del, ensure_ascii=False)}\n\n"
                f"WATCHOUTS:\n{json.dumps(trend_watchouts[:6], ensure_ascii=False)}\n\n"
                "EVIDENCE_QUOTES:\n" + ("\n".join(evidence_lines) if evidence_lines else "(none)")
            )

            client = OpenAI(api_key=api_key)
            req = {"model": selected_model, "messages": [{"role": "system", "content": sys_ctx}, {"role": "user", "content": q}], "tools": tools}
            if model_supports_temperature(selected_model):
                req["temperature"] = llm_temp

            with st.spinner("Thinking..."):
                first = client.chat.completions.create(**req)

            msg = first.choices[0].message
            tool_calls = getattr(msg, "tool_calls", []) or []
            tool_msgs = []
            if tool_calls:
                for call in tool_calls:
                    try:
                        args = json.loads(call.function.arguments or "{}")
                        if not isinstance(args, dict):
                            args = {}
                    except Exception:
                        args = {}
                    name = call.function.name
                    out = {"error": "unknown tool"}
                    if name == "pandas_count":
                        out = pandas_count(args.get("query", ""))
                    if name == "pandas_mean":
                        out = pandas_mean(args.get("column", ""), args.get("query"))
                    if name == "symptom_stats":
                        out = symptom_stats(args.get("symptom", ""))
                    if name == "keyword_stats":
                        out = keyword_stats(args.get("term", ""))
                    if name == "get_metrics_snapshot":
                        out = get_metrics_snapshot()
                    if name == "top_symptoms":
                        out = top_symptoms(args.get("which", ""), int(args.get("k", 10) or 10))
                    if name == "monthly_volume_and_mean":
                        out = monthly_volume_and_mean()

                    tool_msgs.append({"tool_call_id": call.id, "role": "tool", "name": name, "content": json.dumps(out)})

            if tool_msgs:
                follow = {
                    "model": selected_model,
                    "messages": [
                        {"role": "system", "content": sys_ctx},
                        {"role": "assistant", "tool_calls": tool_calls, "content": None},
                        *tool_msgs,
                    ],
                }
                if model_supports_temperature(selected_model):
                    follow["temperature"] = llm_temp
                res2 = client.chat.completions.create(**follow)
                final_text = res2.choices[0].message.content
            else:
                final_text = msg.content

        except Exception:
            # Hard fallback: robust local answer
            st.info("AI temporarily unavailable. Showing local analysis with evidence instead.")
            final_text = local_answer
            evidence_pack = {"local_quotes": local_quotes[:6], "retrieved_quotes": []}

    # Render Q/A + evidence
    st.markdown(f"<div class='chat-q'><b>User:</b> {esc(q)}</div>", unsafe_allow_html=True)
    st.markdown(f"<div class='chat-a'><b>Assistant:</b> {final_text}</div>", unsafe_allow_html=True)

    with st.expander("🔎 Evidence used (snippets)", expanded=False):
        st.caption("These are the snippets provided to the model. IDs like [L1], [R2], [S1] should appear in the AI answer if evidence is cited correctly.")
        if evidence_pack.get("local_quotes"):
            st.markdown("**Local retrieval (covers full dataset):**")
            for i, s in enumerate(evidence_pack["local_quotes"], start=1):
                st.write(f"[L{i}] “{_mask_pii(s)}”")
        if evidence_pack.get("retrieved_quotes"):
            st.markdown("**Vector retrieval (semantic):**")
            for i, s in enumerate(evidence_pack["retrieved_quotes"], start=1):
                st.write(f"[R{i}] “{_mask_pii(s)}”")

st.markdown("---")

# ---------- Metrics ----------
st.markdown("## ⭐ Star Rating Metrics")
st.caption("All metrics below reflect the **currently filtered** dataset.")

def pct_12(series: pd.Series) -> float:
    s = pd.to_numeric(series, errors="coerce").dropna()
    return float((s <= 2).mean() * 100) if not s.empty else 0.0

def section_stats(sub: pd.DataFrame) -> tuple[int, float, float]:
    cnt = len(sub)
    if cnt == 0 or "Star Rating" not in sub.columns:
        return 0, 0.0, 0.0
    avg = float(pd.to_numeric(sub["Star Rating"], errors="coerce").mean())
    pct = pct_12(sub["Star Rating"])
    return cnt, avg, pct

# Robust Seeded split
if "Seeded" in filtered.columns:
    seed_mask = filtered["Seeded"].astype("string").str.upper().eq("YES")
else:
    seed_mask = pd.Series(False, index=filtered.index)

all_cnt, all_avg, all_low = section_stats(filtered)
org = filtered[~seed_mask]
seed = filtered[seed_mask]
org_cnt, org_avg, org_low = section_stats(org)
seed_cnt, seed_avg, seed_low = section_stats(seed)

def card_html(title, count, avg, pct):
    return textwrap.dedent(f"""
    <div class="metric-card">
      <h4>{_html.escape(title)}</h4>
      <div class="metric-row">
        <div class="metric-box">
          <div class="metric-label">Count</div>
          <div class="metric-kpi">{count:,}</div>
        </div>
        <div class="metric-box">
          <div class="metric-label">Avg ★</div>
          <div class="metric-kpi">{avg:.1f}</div>
        </div>
        <div class="metric-box">
          <div class="metric-label">% 1–2★</div>
          <div class="metric-kpi">{pct:.1f}%</div>
        </div>
      </div>
    </div>
    """).strip()

st.markdown(
    (
        '<div class="metrics-grid">'
        f'{card_html("All Reviews", all_cnt, all_avg, all_low)}'
        f'{card_html("Organic (non-Seeded)", org_cnt, org_avg, org_low)}'
        f'{card_html("Seeded", seed_cnt, seed_avg, seed_low)}'
        "</div>"
    ),
    unsafe_allow_html=True,
)

# ---------- NEW: Monthly Volume + Avg ★ chart (integrated) ----------
st.markdown("### 📊 Monthly Review Volume + Avg ★")
if "Review Date" not in filtered.columns or "Star Rating" not in filtered.columns:
    st.info("Need 'Review Date' and 'Star Rating' columns to compute this chart.")
else:
    cA, cB, cC = st.columns([1, 1, 1])
    with cA:
        show_seeded_split = st.toggle("Split Seeded vs Organic", value=False, key="month_split_seeded")
    with cB:
        show_volume_month = st.toggle("Show Volume Bars", value=True, key="month_show_volume")
    with cC:
        month_freq = st.selectbox("Bucket", ["Month", "Week"], index=0, key="month_bucket")

    d = filtered.copy()
    d["Review Date"] = pd.to_datetime(d["Review Date"], errors="coerce")
    d["Star Rating"] = pd.to_numeric(d["Star Rating"], errors="coerce")
    d = d.dropna(subset=["Review Date", "Star Rating"])
    if d.empty:
        st.warning("No data after cleaning dates/ratings.")
    else:
        if month_freq == "Week":
            d["bucket"] = d["Review Date"].dt.to_period("W-MON").dt.start_time
        else:
            d["bucket"] = d["Review Date"].dt.to_period("M").dt.to_timestamp()

        overall = d.groupby("bucket")["Star Rating"].agg(count="count", mean="mean").reset_index().sort_values("bucket")

        fig = make_subplots(specs=[[{"secondary_y": True}]])
        if show_volume_month:
            fig.add_trace(
                go.Bar(
                    x=overall["bucket"],
                    y=overall["count"],
                    name="Review volume",
                    opacity=0.30,
                    marker=dict(color="rgba(15, 23, 42, 0.35)"),
                ),
                secondary_y=False,
            )
        fig.add_trace(
            go.Scatter(
                x=overall["bucket"],
                y=overall["mean"],
                name="Avg ★",
                mode="lines+markers",
                line=dict(width=3),
                marker=dict(size=6),
                hovertemplate="Bucket: %{x|%Y-%m-%d}<br>Avg ★: %{y:.2f}<extra></extra>",
            ),
            secondary_y=True,
        )

        if show_seeded_split and "Seeded" in d.columns:
            d2 = d.copy()
            d2["Seeded"] = d2["Seeded"].astype("string").str.upper().fillna("NO")
            for label, mask in [("Organic", d2["Seeded"].ne("YES")), ("Seeded", d2["Seeded"].eq("YES"))]:
                sub = d2[mask].groupby("bucket")["Star Rating"].agg(mean="mean").reset_index().sort_values("bucket")
                if not sub.empty:
                    fig.add_trace(
                        go.Scatter(
                            x=sub["bucket"],
                            y=sub["mean"],
                            name=f"{label} Avg ★",
                            mode="lines",
                            line=dict(width=2, dash="dot"),
                            hovertemplate=f"{label}<br>Bucket: %{{x|%Y-%m-%d}}<br>Avg ★: %{{y:.2f}}<extra></extra>",
                        ),
                        secondary_y=True,
                    )

        fig.update_yaxes(title_text="Review volume", secondary_y=False, rangemode="tozero")
        fig.update_yaxes(title_text="Avg ★", secondary_y=True, range=[1.0, 5.2])
        fig.update_layout(
            template="plotly_white",
            hovermode="x unified",
            margin=dict(l=50, r=50, t=40, b=40),
            legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5),
        )
        st.plotly_chart(fig, use_container_width=True)

# ---------- NEW: Top Delighters/Detractors bar charts (integrated) ----------
st.markdown("### 🧩 Top Delighters & Detractors (Bars)")
c1, c2 = st.columns(2)
with c1:
    st.markdown("**Top Detractors (by mentions)**")
    det = detractors_results_full.head(10)
    if det.empty:
        st.info("No detractor symptoms available.")
    else:
        fig_det = go.Figure(
            go.Bar(
                x=det["Mentions"][::-1],
                y=det["Item"][::-1],
                orientation="h",
                opacity=0.85,
                hovertemplate="%{y}<br>Mentions: %{x}<extra></extra>",
            )
        )
        fig_det.update_layout(template="plotly_white", margin=dict(l=140, r=20, t=20, b=20), height=420)
        st.plotly_chart(fig_det, use_container_width=True)
with c2:
    st.markdown("**Top Delighters (by mentions)**")
    deli = delighters_results_full.head(10)
    if deli.empty:
        st.info("No delighter symptoms available.")
    else:
        fig_del = go.Figure(
            go.Bar(
                x=deli["Mentions"][::-1],
                y=deli["Item"][::-1],
                orientation="h",
                opacity=0.85,
                hovertemplate="%{y}<br>Mentions: %{x}<extra></extra>",
            )
        )
        fig_del.update_layout(template="plotly_white", margin=dict(l=140, r=20, t=20, b=20), height=420)
        st.plotly_chart(fig_del, use_container_width=True)

# ---------- Avg ★ Over Time by Region — Cumulative (Weighted Over Time) ----------
st.markdown("### 📈 Cumulative Avg ★ Over Time by Region (Weighted)")

if "Review Date" not in filtered.columns or "Star Rating" not in filtered.columns:
    st.info("Need 'Review Date' and 'Star Rating' columns to compute this chart.")
else:
    # Controls
    c1, c2, c3, c4, c5 = st.columns([1.1, 1.1, 1.1, 0.9, 0.9])
    with c1:
        bucket_label = st.selectbox("Bucket size", ["Day", "Week", "Month"], index=2, key="region_bucket")
        _freq_map = {"Day": "D", "Week": "W", "Month": "M"}
        freq = _freq_map[bucket_label]
    with c2:
        # Pick the "region" dimension
        _candidates = [c for c in ["Country", "Source", "Model (SKU)"] if c in filtered.columns]
        region_col = st.selectbox("Region field", options=_candidates or ["(none)"], key="region_col")
    with c3:
        top_n = st.number_input("Top regions by volume", 1, 15, value=5, step=1, key="region_topn")
    with c4:
        organic_only = st.toggle("Organic Only", value=False, help="Exclude reviews where Seeded == YES")
    with c5:
        show_volume = st.toggle(
            "Show Volume",
            value=True,
            help="Adds subtle bars + a right axis showing review count per bucket.",
        )

    if region_col not in filtered.columns or region_col == "(none)":
        st.info("No region column found for this chart.")
    else:
        d = filtered.copy()

        # Organic-only filter
        if organic_only and "Seeded" in d.columns:
            d = d[d["Seeded"].astype("string").str.upper().ne("YES")]

        # Clean / coerce
        d["Star Rating"] = pd.to_numeric(d["Star Rating"], errors="coerce")
        d["Review Date"] = pd.to_datetime(d["Review Date"], errors="coerce")
        d = d.dropna(subset=["Review Date", "Star Rating"])

        if d.empty:
            st.warning("No data available for the current selections.")
        else:
            # Default regions = Top-N by count
            counts = d[region_col].astype("string").str.strip().replace({"": pd.NA}).dropna()
            top_regions = counts.value_counts().head(int(top_n)).index.tolist()

            regions_available = sorted(
                r for r in d[region_col].astype("string").dropna().unique().tolist()
                if str(r).strip() != ""
            )
            chosen_regions = st.multiselect(
                "Regions to plot",
                options=regions_available,
                default=[r for r in top_regions if r in regions_available],
                key="region_pick",
            )

            # Optional restrict to chosen regions
            if chosen_regions:
                d = d[d[region_col].astype("string").isin(chosen_regions)]

            if d.empty:
                st.warning("No data after region selection.")
            else:
                # ----------------------------
                # CUMULATIVE WEIGHTED AGGREGATION
                # ----------------------------
                freq_eff = "W-MON" if freq == "W" else freq  # Monday-start weeks

                # Per-region bucket sums/counts
                tmp = (
                    d.assign(_region=d[region_col].astype("string"))
                    .groupby([pd.Grouper(key="Review Date", freq=freq_eff), "_region"])["Star Rating"]
                    .agg(bucket_sum="sum", bucket_count="count")
                    .reset_index()
                    .sort_values(["_region", "Review Date"])
                )
                # Cumulative → weighted average
                tmp["cum_sum"] = tmp.groupby("_region")["bucket_sum"].cumsum()
                tmp["cum_cnt"] = tmp.groupby("_region")["bucket_count"].cumsum()
                tmp["Cumulative Avg ★"] = tmp["cum_sum"] / tmp["cum_cnt"]

                # Overall cumulative line + bucket volume
                overall = (
                    d.groupby(pd.Grouper(key="Review Date", freq=freq_eff))["Star Rating"]
                    .agg(bucket_sum="sum", bucket_count="count")
                    .reset_index()
                    .sort_values("Review Date")
                )
                overall["cum_sum"] = overall["bucket_sum"].cumsum()
                overall["cum_cnt"] = overall["bucket_count"].cumsum()
                overall["Cumulative Avg ★"] = overall["cum_sum"] / overall["cum_cnt"]

                # ---- Plotly chart ----
                fig = go.Figure()

                # Volume bars (RIGHT axis) — add first so bars render behind lines
                if show_volume and not overall.empty:
                    fig.add_trace(
                        go.Bar(
                            x=overall["Review Date"],
                            y=overall["bucket_count"],
                            name="Review volume",
                            yaxis="y2",
                            opacity=0.30,  # easier to see, still subtle
                            marker=dict(color="rgba(15, 23, 42, 0.35)", line=dict(width=0)),
                            hovertemplate="Review volume<br>Bucket end: %{x|%Y-%m-%d}<br>Reviews: %{y}<extra></extra>",
                            showlegend=False,
                        )
                    )

                # Region lines
                plot_regions = chosen_regions or tmp["_region"].unique().tolist()
                for reg in plot_regions:
                    sub = tmp[tmp["_region"] == reg]
                    if sub.empty:
                        continue
                    fig.add_trace(
                        go.Scatter(
                            x=sub["Review Date"],
                            y=sub["Cumulative Avg ★"],
                            mode="lines+markers",
                            name=str(reg),
                            line=dict(width=2),
                            marker=dict(size=5),
                            hovertemplate=(
                                f"{region_col}: {reg}<br>"
                                "Bucket end: %{x|%Y-%m-%d}<br>"
                                "Cumulative Avg ★: %{y:.3f}<br>"
                                "Cum. Reviews: %{customdata}<extra></extra>"
                            ),
                            customdata=sub["cum_cnt"],
                        )
                    )

                # Overall dashed line
                if not overall.empty:
                    fig.add_trace(
                        go.Scatter(
                            x=overall["Review Date"],
                            y=overall["Cumulative Avg ★"],
                            mode="lines",
                            name="Overall",
                            line=dict(width=3, dash="dash"),
                            hovertemplate=(
                                "Overall<br>"
                                "Bucket end: %{x|%Y-%m-%d}<br>"
                                "Cumulative Avg ★: %{y:.3f}<br>"
                                "Cum. Reviews: %{customdata}<extra></extra>"
                            ),
                            customdata=overall["cum_cnt"],
                        )
                    )

                # Axis formatting based on bucket
                _tickformat = {"D": "%b %d, %Y", "W": "%b %d, %Y", "M": "%b %Y"}[freq]
                fig.update_xaxes(tickformat=_tickformat, automargin=True)
                fig.update_yaxes(automargin=True)

                # Title/legend/layout (legend below x-axis)
                title_bucket = {"D": "Daily", "W": "Weekly", "M": "Monthly"}[freq]
                title_org = " • Organic Only" if organic_only else ""
                fig.update_layout(
                    title=f"<b>{title_bucket} Cumulative (Weighted) Avg ★ by {region_col}{title_org}</b>",
                    xaxis=dict(title="Date", showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                    yaxis=dict(title="Cumulative Avg ★", showgrid=True, gridcolor="rgba(0,0,0,0.06)"),
                    yaxis2=dict(
                        title="Review volume",
                        overlaying="y",
                        side="right",
                        showgrid=False,
                        rangemode="tozero",
                        autorange=True,
                        visible=bool(show_volume),
                    ),
                    barmode="overlay",
                    hovermode="x unified",
                    plot_bgcolor="white",
                    template="plotly_white",
                    margin=dict(l=60, r=(60 if show_volume else 40), t=70, b=100),
                    legend=dict(orientation="h", yanchor="top", y=-0.28, xanchor="center", x=0.5, bgcolor="rgba(255,255,255,0)"),
                )

                # Extra spacing / no clipping (ONLY set LEFT axis range; do NOT touch yaxis2)
                ys = []
                for tr in fig.data:
                    if getattr(tr, "yaxis", "y") not in (None, "y"):
                        continue
                    if getattr(tr, "y", None) is not None:
                        try:
                            ys.extend([float(v) for v in tr.y if v is not None])
                        except Exception:
                            pass

                if ys:
                    y_min, y_max = min(ys), max(ys)
                    pad = max(0.1, (y_max - y_min) * 0.08)
                    lo = max(1.0, y_min - pad)
                    hi = min(5.2, y_max + pad)
                    fig.update_layout(yaxis_range=[lo, hi])

                fig.update_traces(cliponaxis=False, selector=dict(type="scatter"))
                st.plotly_chart(fig, use_container_width=True)

# ---------- NEW: Country × Source breakdown (integrated) ----------
st.markdown("### 🧭 Country × Source Breakdown")
if "Country" in filtered.columns and "Source" in filtered.columns and "Star Rating" in filtered.columns:
    cA, cB, cC = st.columns([1, 1, 1])
    with cA:
        top_countries = st.number_input("Top Countries", 3, 30, value=10, step=1, key="cx_top_countries")
    with cB:
        top_sources = st.number_input("Top Sources", 3, 30, value=10, step=1, key="cx_top_sources")
    with cC:
        show_heatmap = st.toggle("Show Avg ★ Heatmap", value=False, key="cx_show_heatmap")

    d = filtered.copy()
    d["Star Rating"] = pd.to_numeric(d["Star Rating"], errors="coerce")
    d = d.dropna(subset=["Star Rating"])
    if not d.empty:
        # Limit to top-N for readability
        topC = d["Country"].astype("string").value_counts().head(int(top_countries)).index.tolist()
        topS = d["Source"].astype("string").value_counts().head(int(top_sources)).index.tolist()
        d = d[d["Country"].astype("string").isin(topC) & d["Source"].astype("string").isin(topS)]

        pivot = d.groupby(["Country", "Source"])["Star Rating"].agg(count="count", mean="mean").reset_index()
        pivot["mean"] = pivot["mean"].round(2)

        # Create a wide table: show count and mean side-by-side (stacked columns)
        wide_count = pivot.pivot(index="Country", columns="Source", values="count").fillna(0).astype(int)
        wide_mean = pivot.pivot(index="Country", columns="Source", values="mean").round(2)

        # Display combined table in a readable way
        with st.expander("📋 Table (Count + Avg ★)", expanded=True):
            st.caption("Count table shown first, Avg ★ table shown second.")
            st.markdown("**Counts**")
            st.dataframe(wide_count, use_container_width=True)
            st.markdown("**Avg ★**")
            st.dataframe(wide_mean, use_container_width=True)

        if show_heatmap:
            hm = wide_mean.copy()
            fig_hm = go.Figure(
                data=go.Heatmap(
                    z=hm.values,
                    x=list(hm.columns),
                    y=list(hm.index),
                    hovertemplate="Country=%{y}<br>Source=%{x}<br>Avg ★=%{z}<extra></extra>",
                )
            )
            fig_hm.update_layout(template="plotly_white", margin=dict(l=80, r=20, t=30, b=40), height=520)
            st.plotly_chart(fig_hm, use_container_width=True)
    else:
        st.info("No rating data available after filtering.")
else:
    st.info("Need Country, Source, and Star Rating columns to compute this breakdown.")

# ---------- Symptom Tables ----------
st.markdown("### 🩺 Symptom Tables")
detractors_results = detractors_results_full.head(50)
delighters_results = delighters_results_full.head(50)

view_mode = st.radio("View mode", ["Split", "Tabs"], horizontal=True, index=0)

def _styled_table(df_in: pd.DataFrame):
    if df_in.empty:
        return df_in
    def colstyle(v):
        if pd.isna(v):
            return ""
        try:
            v = float(v)
            if v >= 4.5:
                return "color:#065F46;font-weight:600;"
            if v < 4.5:
                return "color:#7F1D1D;font-weight:600;"
        except Exception:
            pass
        return ""
    return df_in.style.applymap(colstyle, subset=["Avg Star"]).format({"Avg Star": "{:.1f}", "Mentions": "{:.0f}"})

if view_mode == "Split":
    c1, c2 = st.columns([1, 1])
    with c1:
        st.subheader("All Detractors")
        st.dataframe(_styled_table(detractors_results) if not detractors_results.empty else detractors_results, use_container_width=True, hide_index=True)
    with c2:
        st.subheader("All Delighters")
        st.dataframe(_styled_table(delighters_results) if not delighters_results.empty else delighters_results, use_container_width=True, hide_index=True)
else:
    tab1, tab2 = st.tabs(["All Detractors", "All Delighters"])
    with tab1:
        st.dataframe(_styled_table(detractors_results) if not detractors_results.empty else detractors_results, use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(_styled_table(delighters_results) if not delighters_results.empty else delighters_results, use_container_width=True, hide_index=True)

st.markdown("---")

# ---------- Reviews ----------
st.markdown("### 📝 All Reviews")

if not filtered.empty:
    csv_bytes = filtered.to_csv(index=False).encode("utf-8-sig")
    st.download_button("⬇️ Download ALL filtered reviews (CSV)", csv_bytes, file_name="filtered_reviews.csv", mime="text/csv")

if "review_page" not in st.session_state:
    st.session_state["review_page"] = 0
reviews_per_page = st.session_state.get("reviews_per_page", 10)
total_reviews_count = len(filtered)
total_pages = max((total_reviews_count + reviews_per_page - 1) // reviews_per_page, 1)

# Clamp page after any filter changes
st.session_state["review_page"] = min(st.session_state.get("review_page", 0), max(total_pages - 1, 0))

current_page = st.session_state["review_page"]
start_index = current_page * reviews_per_page
end_index = start_index + reviews_per_page
paginated = filtered.iloc[start_index:end_index]

if paginated.empty:
    st.warning("No reviews match the selected criteria.")
else:
    for _, row in paginated.iterrows():
        review_text = row.get("Verbatim", pd.NA)
        review_text = "" if pd.isna(review_text) else clean_text(review_text)
        display_review_html = highlight_html(review_text, st.session_state.get("kw", ""))

        date_val = row.get("Review Date", pd.NaT)
        if pd.isna(date_val):
            date_str = "-"
        else:
            try:
                date_str = pd.to_datetime(date_val).strftime("%Y-%m-%d")
            except Exception:
                date_str = "-"

        def chips(row_i, columns, css_class):
            items = []
            for c in columns:
                val = row_i.get(c, pd.NA)
                if pd.isna(val):
                    continue
                s = str(val).strip()
                if not s or s.upper() in {"<NA>", "NA", "N/A", "-"}:
                    continue
                items.append(f'<span class="badge {css_class}">{_html.escape(s)}</span>')
            return f'<div class="badges">{"".join(items)}</div>' if items else "<i>None</i>"

        delighter_message = chips(row, existing_delighter_columns, "pos")
        detractor_message = chips(row, existing_detractor_columns, "neg")

        star_val = row.get("Star Rating", 0)
        try:
            star_int = int(star_val) if pd.notna(star_val) else 0
        except Exception:
            star_int = 0

        st.markdown(
            f"""
            <div class="review-card">
                <p><strong>Source:</strong> {esc(row.get('Source'))} | <strong>Model:</strong> {esc(row.get('Model (SKU)'))}</p>
                <p><strong>Country:</strong> {esc(row.get('Country'))} | <strong>Date:</strong> {esc(date_str)}</p>
                <p><strong>Rating:</strong> {'⭐' * star_int} ({esc(row.get('Star Rating'))}/5)</p>
                <p><strong>Review:</strong> {display_review_html}</p>
                <div><strong>Delighter Symptoms:</strong> {delighter_message}</div>
                <div><strong>Detractor Symptoms:</strong> {detractor_message}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

# Pager
p1, p2, p3, p4, p5 = st.columns([1, 1, 2, 1, 1])
with p1:
    if st.button("⏮ First", disabled=current_page == 0):
        st.session_state["review_page"] = 0
        st.rerun()
with p2:
    if st.button("⬅ Prev", disabled=current_page == 0):
        st.session_state["review_page"] = max(current_page - 1, 0)
        st.rerun()
with p3:
    showing_from = 0 if total_reviews_count == 0 else start_index + 1
    showing_to = min(end_index, total_reviews_count)
    st.markdown(
        f"<div style='text-align:center;font-weight:700;'>Page {current_page + 1} of {total_pages} • Showing {showing_from}–{showing_to} of {total_reviews_count}</div>",
        unsafe_allow_html=True,
    )
with p4:
    if st.button("Next ➡", disabled=current_page >= total_pages - 1):
        st.session_state["review_page"] = min(current_page + 1, total_pages - 1)
        st.rerun()
with p5:
    if st.button("Last ⏭", disabled=current_page >= total_pages - 1):
        st.session_state["review_page"] = total_pages - 1
        st.rerun()

st.markdown("---")

# ---------- Feedback ----------
anchor("feedback-anchor")
st.markdown("## 💬 Submit Feedback")
st.caption("Tell us what to improve. We care about making this tool user-centric.")

def send_feedback_via_email(subject: str, body: str) -> bool:
    try:
        host = st.secrets.get("SMTP_HOST")
        port = int(st.secrets.get("SMTP_PORT", 587))
        user = st.secrets.get("SMTP_USER")
        pwd = st.secrets.get("SMTP_PASS")
        sender = st.secrets.get("SMTP_FROM", user or "")
        to = st.secrets.get("SMTP_TO", "wseddon@sharkninja.com")
        if not (host and port and sender and to):
            return False
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = sender
        msg["To"] = to
        msg.set_content(body)
        with smtplib.SMTP(host, port, timeout=15) as s:
            s.starttls()
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)
        return True
    except Exception as e:
        st.info(f"Could not send via SMTP: {e}")
        return False

with st.form("feedback_form", clear_on_submit=True):
    name = st.text_input("Your name (optional)")
    email = st.text_input("Your email (optional)")
    message = st.text_area("Feedback / feature request", placeholder="Type your feedback here…", height=140)
    submitted = st.form_submit_button("Submit feedback")
if submitted:
    if not message.strip():
        st.warning("Please enter some feedback before submitting.")
    else:
        body = f"Name: {name or '-'}\nEmail: {email or '-'}\n\nFeedback:\n{message}"
        ok = send_feedback_via_email("Star Walk — Feedback", body)
        if ok:
            st.success("Thanks! Your feedback was sent.")
        else:
            quoted = quote(message or "", safe="")
            st.link_button(
                "Open email to wseddon@sharkninja.com",
                url=f"mailto:wseddon@sharkninja.com?subject=Star%20Walk%20Feedback&body={quoted}",
            )

# one-time scroll on fresh upload
if st.session_state.get("force_scroll_top_once"):
    st.session_state["force_scroll_top_once"] = False
    st.markdown("<script>window.scrollTo({top:0,behavior:'auto'});</script>", unsafe_allow_html=True)
