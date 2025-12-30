# starwalk_converter_app.py
# Streamlit app to convert raw website review exports into the
# "Star Walk scrubbed verbatims" format (Symptom 1–10 detractors, 11–20 delighters).
#
# FIXED (Seeded): Guaranteed conversion of values like "['Seeded']" -> "yes"
# - Cleaning is applied at mapping time (cannot be skipped)
# - Cleaning is idempotent (re-cleaning "yes" stays "yes")
# - Final safety pass re-enforces the rule before writing
#
# FIXED (Star Rating): Guaranteed numeric-only output (e.g., "4 star" -> 4)

import io
import re
import ast
from typing import Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


APP_VERSION = "2025-12-30-seeded-yes-FIXED"
STARWALK_SHEET_NAME = "Star Walk scrubbed verbatims"

# Default header (mirrors typical Star Walk workbook; includes two trailing blank header columns)
DEFAULT_STARWALK_COLUMNS: List[str] = [
    "Source", "Model (SKU)", "Seeded", "Country", "New Review", "Review Date",
    "Verbatim Id", "Verbatim", "Star Rating", "Review count per detractor",
    "Symptom 1", "Symptom 2", "Symptom 3", "Symptom 4", "Symptom 5",
    "Symptom 6", "Symptom 7", "Symptom 8", "Symptom 9", "Symptom 10",
    "Symptom 11", "Symptom 12", "Symptom 13", "Symptom 14", "Symptom 15",
    "Symptom 16", "Symptom 17", "Symptom 18", "Symptom 19", "Symptom 20",
    "Hair Type", "Unnamed: 31", "Unnamed: 32",
]

DEFAULT_TAG_SPLIT_RE = re.compile(r"[;\|\n,]+")

SEED_TRUE = {"seeded", "yes", "true", "y", "1"}
SEED_FALSE = {"notseeded", "no", "false", "n", "0", "unseeded", "nonseeded"}


# -----------------------------
# Helpers
# -----------------------------
def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())


def best_match(target: str, candidates: List[str]) -> Optional[str]:
    """Lightweight matcher: exact normalized match > substring match."""
    if not candidates:
        return None
    t = _norm(target)
    norm_map = {c: _norm(c) for c in candidates}

    for c, nc in norm_map.items():
        if nc == t:
            return c
    for c, nc in norm_map.items():
        if t and (t in nc or nc in t):
            return c
    return None


def _safe_isna(v) -> bool:
    try:
        return pd.isna(v)
    except Exception:
        return False


@st.cache_data(show_spinner=False)
def read_table(file_bytes: bytes, filename: str, sheet: Optional[str] = None) -> pd.DataFrame:
    if filename.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes))
    if filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet)
    raise ValueError("Unsupported file type. Please upload a CSV or Excel file.")


def dedupe_preserve(items: List[str]) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for x in items:
        s = str(x).strip()
        if not s:
            continue
        if s not in seen:
            seen.add(s)
            out.append(s)
    return out


def parse_tags(value, split_re: re.Pattern) -> List[str]:
    """
    Parse a cell that may contain:
      - NaN/empty
      - a single string
      - delimited strings (e.g., "A; B | C")
      - stringified list: "['A','B']"
      - actual list/tuple/set (rare but possible)
    """
    if value is None or _safe_isna(value):
        return []

    if isinstance(value, (list, tuple, set)):
        return dedupe_preserve([str(v).strip() for v in value if str(v).strip()])

    s = str(value).strip()
    if not s:
        return []

    # Try literal_eval for list-like strings when possible (more robust than manual split)
    if (s.startswith("[") and s.endswith("]")) or (s.startswith("(") and s.endswith(")")):
        try:
            obj = ast.literal_eval(s)
            if isinstance(obj, (list, tuple, set)):
                return dedupe_preserve([str(v).strip() for v in obj if str(v).strip()])
        except Exception:
            # Fallback to manual parsing if literal_eval fails
            inner = s[1:-1].strip()
            if inner:
                parts = [p.strip().strip("'\"") for p in inner.split(",")]
                return dedupe_preserve([p for p in parts if p])
            return []

    parts = [p.strip() for p in split_re.split(s) if p and p.strip()]
    return dedupe_preserve(parts)


# -----------------------------
# Cleaning: Star Rating (numeric only)
# -----------------------------
def clean_star_rating_value(v):
    """Coerce star rating into a numeric (int/float) and strip words like 'star(s)'."""
    if v is None or _safe_isna(v):
        return pd.NA

    # Already numeric
    if isinstance(v, (int, np.integer)):
        return int(v)
    if isinstance(v, (float, np.floating)):
        fv = float(v)
        if np.isnan(fv):
            return pd.NA
        return int(fv) if fv.is_integer() else fv

    s = str(v).strip()
    if not s:
        return pd.NA

    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m:
        return pd.NA

    fv = float(m.group(1))
    return int(fv) if fv.is_integer() else fv


def clean_star_rating_series(series: pd.Series) -> pd.Series:
    return series.apply(clean_star_rating_value).astype("object")


# -----------------------------
# Cleaning: Seeded -> "yes" only
# -----------------------------
def clean_seeded_value(v):
    """
    Convert values like:
      - ['Seeded'] / ["Seeded"] / 'Seeded' / 'yes' / True -> 'yes'
    Everything else -> blank (pd.NA)

    IMPORTANT:
    - Token-based equality (no substring matching) to avoid false positives.
    - Idempotent: if it's already 'yes', it stays 'yes'.
    """
    if v is None or _safe_isna(v):
        return pd.NA

    # Boolean / numeric handling
    if isinstance(v, (bool, np.bool_)):
        return "yes" if bool(v) else pd.NA
    if isinstance(v, (int, np.integer)):
        return "yes" if int(v) == 1 else pd.NA
    if isinstance(v, (float, np.floating)):
        fv = float(v)
        if np.isnan(fv):
            return pd.NA
        return "yes" if fv == 1.0 else pd.NA

    # Parse into tokens (handles "['Seeded']" safely)
    tokens = parse_tags(v, DEFAULT_TAG_SPLIT_RE)
    for t in tokens:
        nt = _norm(t)
        if nt in SEED_TRUE:
            return "yes"
        if nt in SEED_FALSE:
            return pd.NA

    return pd.NA


def clean_seeded_series(series: pd.Series) -> pd.Series:
    return series.apply(clean_seeded_value).astype("object")


def apply_final_cleaning(df: pd.DataFrame) -> pd.DataFrame:
    """Final safety pass: enforce Seeded + Star Rating rules no matter what."""
    for c in list(df.columns):
        nc = _norm(c)
        if nc == "starrating":
            df[c] = clean_star_rating_series(df[c])
        elif nc == "seeded":
            df[c] = clean_seeded_series(df[c])
    return df


def find_col_by_norm(cols: List[str], target_norm: str) -> Optional[str]:
    for c in cols:
        if _norm(c) == target_norm:
            return c
    return None


# -----------------------------
# Template handling
# -----------------------------
def extract_template_header(template_bytes: bytes, sheet_name: str, keep_trailing_blank_cols: int = 5) -> List[str]:
    """
    Extract header row from an Excel sheet using openpyxl so we don't accidentally
    drop trailing blank header columns (common in the Star Walk template).
    """
    wb = load_workbook(io.BytesIO(template_bytes), data_only=False)
    if sheet_name not in wb.sheetnames:
        return DEFAULT_STARWALK_COLUMNS

    ws = wb[sheet_name]
    raw = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        raw.append(v)

    # Find last named header cell
    last_named = -1
    for i, v in enumerate(raw):
        if v is not None and str(v).strip() != "":
            last_named = i

    if last_named == -1:
        return DEFAULT_STARWALK_COLUMNS

    max_keep = min(len(raw), last_named + 1 + keep_trailing_blank_cols)
    headers: List[str] = []
    for i in range(max_keep):
        v = raw[i]
        if v is None or str(v).strip() == "":
            headers.append(f"Unnamed: {i}")
        else:
            headers.append(str(v).strip())
    return headers


def load_allowed_l2_from_template(template_bytes: bytes) -> Tuple[Optional[Set[str]], Optional[Set[str]]]:
    """
    If the template has a 'Symptoms' sheet with columns 'Detractors' and 'Delighters',
    use it as an allow-list for L2 tags.
    """
    try:
        df = pd.read_excel(io.BytesIO(template_bytes), sheet_name="Symptoms")
        if "Detractors" not in df.columns or "Delighters" not in df.columns:
            return None, None
        det = set(df["Detractors"].dropna().astype(str).str.strip())
        deli = set(df["Delighters"].dropna().astype(str).str.strip())
        det = {x for x in det if x}
        deli = {x for x in deli if x}
        return det or None, deli or None
    except Exception:
        return None, None


# -----------------------------
# Conversion
# -----------------------------
def collect_from_row_tuple(
    row_tup,
    col_idx: Dict[str, int],
    cols: List[str],
    split_re: re.Pattern,
) -> List[str]:
    tags: List[str] = []
    for c in cols:
        i = col_idx.get(c)
        if i is None:
            continue
        tags.extend(parse_tags(row_tup[i], split_re))
    return dedupe_preserve(tags)


def convert_to_starwalk(
    src_df: pd.DataFrame,
    out_cols: List[str],
    field_map: Dict[str, Optional[str]],
    l2_det_cols: List[str],
    l2_del_cols: List[str],
    split_regex: str,
    weight_mode: str,
    allowed_det: Optional[Set[str]] = None,
    allowed_del: Optional[Set[str]] = None,
    filter_unknown: bool = False,
) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """
    Build Star Walk scrubbed verbatims output:
      - L2 detractors -> Symptom 1-10
      - L2 delighters -> Symptom 11-20
    """
    src_df = src_df.reset_index(drop=True)
    n = len(src_df)

    needed_symptoms = [f"Symptom {i}" for i in range(1, 21)]
    base_cols = list(out_cols)
    for c in needed_symptoms:
        if c not in base_cols:
            base_cols.append(c)
    if "Review count per detractor" not in base_cols:
        base_cols.append("Review count per detractor")

    out = pd.DataFrame(index=range(n), columns=base_cols, dtype="object")

    # Copy mapped fields (WITH guaranteed cleaning for Seeded and Star Rating)
    for out_field, in_field in field_map.items():
        if out_field not in out.columns:
            continue

        if in_field and in_field in src_df.columns:
            series = src_df[in_field]

            if _norm(out_field) == "seeded":
                out[out_field] = clean_seeded_series(series)
            elif _norm(out_field) == "starrating":
                out[out_field] = clean_star_rating_series(series)
            else:
                out[out_field] = series.values
        else:
            out[out_field] = pd.NA

    split_re = re.compile(split_regex)
    src_cols = list(src_df.columns)
    col_idx = {c: i for i, c in enumerate(src_cols)}

    detr_matrix = [[pd.NA] * 10 for _ in range(n)]
    deli_matrix = [[pd.NA] * 10 for _ in range(n)]
    detr_count = np.zeros(n, dtype=int)
    deli_count = np.zeros(n, dtype=int)

    rows_trunc_det = 0
    rows_trunc_del = 0
    unknown_det_total = 0
    unknown_del_total = 0

    for r_i, row in enumerate(src_df.itertuples(index=False, name=None)):
        d_tags = collect_from_row_tuple(row, col_idx, l2_det_cols, split_re)
        l_tags = collect_from_row_tuple(row, col_idx, l2_del_cols, split_re)

        if allowed_det is not None:
            unknown = [t for t in d_tags if t not in allowed_det]
            if unknown:
                unknown_det_total += len(unknown)
                if filter_unknown:
                    d_tags = [t for t in d_tags if t in allowed_det]

        if allowed_del is not None:
            unknown = [t for t in l_tags if t not in allowed_del]
            if unknown:
                unknown_del_total += len(unknown)
                if filter_unknown:
                    l_tags = [t for t in l_tags if t in allowed_del]

        if len(d_tags) > 10:
            rows_trunc_det += 1
        if len(l_tags) > 10:
            rows_trunc_del += 1

        d_tags = d_tags[:10]
        l_tags = l_tags[:10]

        detr_count[r_i] = len(d_tags)
        deli_count[r_i] = len(l_tags)

        for j, t in enumerate(d_tags):
            detr_matrix[r_i][j] = t
        for j, t in enumerate(l_tags):
            deli_matrix[r_i][j] = t

    detr_cols_out = [f"Symptom {i}" for i in range(1, 11)]
    deli_cols_out = [f"Symptom {i}" for i in range(11, 21)]

    out[detr_cols_out] = pd.DataFrame(detr_matrix, columns=detr_cols_out).values
    out[deli_cols_out] = pd.DataFrame(deli_matrix, columns=deli_cols_out).values

    # Weight column behavior (configurable)
    if "Review count per detractor" in out.columns:
        if weight_mode == "Leave blank":
            out["Review count per detractor"] = pd.NA
        elif weight_mode == "Always 1":
            out["Review count per detractor"] = 1.0
        elif weight_mode == "1 / # detractor symptoms (if any)":
            out["Review count per detractor"] = np.where(detr_count > 0, 1.0 / detr_count, pd.NA)
        elif weight_mode == "1 / # delighter symptoms (if any)":
            out["Review count per detractor"] = np.where(deli_count > 0, 1.0 / deli_count, pd.NA)
        else:  # "1 / # total symptoms (detractors + delighters)"
            total = detr_count + deli_count
            out["Review count per detractor"] = np.where(total > 0, 1.0 / total, pd.NA)

    # FINAL SAFETY: enforce Seeded + Star Rating rules no matter what
    out = apply_final_cleaning(out)

    stats = {
        "rows": n,
        "rows_truncated_detractors_gt10": rows_trunc_det,
        "rows_truncated_delighters_gt10": rows_trunc_del,
        "unknown_detractor_tags_total": unknown_det_total,
        "unknown_delighter_tags_total": unknown_del_total,
    }
    return out, stats


def to_excel_one_sheet(df_out: pd.DataFrame, sheet_name: str = STARWALK_SHEET_NAME) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def write_into_template_workbook(template_bytes: bytes, df_out: pd.DataFrame, sheet_name: str = STARWALK_SHEET_NAME) -> bytes:
    """
    Preserve the full template workbook (other sheets, formulas, pivots, etc.) and
    replace the Star Walk scrubbed verbatims sheet data (keeping its header row).
    """
    df_out = apply_final_cleaning(df_out.copy())

    wb = load_workbook(io.BytesIO(template_bytes), data_only=False)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Read template header (row 1) to determine column order we should write
    header = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None or str(v).strip() == "":
            header.append(f"Unnamed: {c-1}")
        else:
            header.append(str(v).strip())

    # If the sheet is empty, use df_out's columns and write header
    if (ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None):
        header = list(df_out.columns)
        for c, name in enumerate(header, start=1):
            ws.cell(1, c).value = name

    # Align df to header columns (extra df cols are dropped; missing are added as NA)
    aligned = df_out.copy()
    for col in header:
        if col not in aligned.columns:
            aligned[col] = pd.NA
    aligned = aligned[header]

    # Re-apply cleaning post-align (in case header name differs)
    aligned = apply_final_cleaning(aligned)

    # Clear existing rows below header
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Append new data rows
    for r in dataframe_to_rows(aligned, index=False, header=False):
        ws.append(r)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# -----------------------------
# UI
# -----------------------------
st.set_page_config(page_title="Star Walk Formatter", layout="wide")
st.title("Star Walk Scrubbed Verbatims Converter")
st.caption(f"App version: `{APP_VERSION}`")

st.markdown(
    """
This app converts a raw website review export into the **Star Walk scrubbed verbatims** format.

It will populate:
- **Symptom 1–10** with **Level 2 Detractors**
- **Symptom 11–20** with **Level 2 Delighters**
"""
)

colA, colB = st.columns(2)
with colA:
    src_file = st.file_uploader("Upload raw website reviews (CSV/XLSX)", type=["csv", "xlsx", "xlsm", "xls"])
with colB:
    template_file = st.file_uploader("Optional: Upload Star Walk template workbook (XLSX)", type=["xlsx", "xlsm", "xls"])

if not src_file:
    st.stop()

src_bytes = src_file.getvalue()
src_sheet = None
if src_file.name.lower().endswith((".xlsx", ".xlsm", ".xls")):
    try:
        xl = pd.ExcelFile(io.BytesIO(src_bytes))
        src_sheet = st.selectbox("Select source sheet", xl.sheet_names, index=0)
    except Exception:
        src_sheet = None

src_df = read_table(src_bytes, src_file.name, sheet=src_sheet)

st.subheader("Source Preview")
st.dataframe(src_df.head(50), use_container_width=True)

# Determine output columns
out_cols = DEFAULT_STARWALK_COLUMNS
allowed_det = allowed_del = None

output_mode_options = ["One-sheet output (fast)"]
if template_file:
    output_mode_options.append("Use template workbook (preserve other tabs/formulas)")

output_mode = st.radio("Output mode", output_mode_options, index=(len(output_mode_options) - 1))

if template_file:
    tbytes = template_file.getvalue()
    try:
        txl = pd.ExcelFile(io.BytesIO(tbytes))
        default_idx = txl.sheet_names.index(STARWALK_SHEET_NAME) if STARWALK_SHEET_NAME in txl.sheet_names else 0
        template_sheet = st.selectbox("Template sheet to mirror columns", txl.sheet_names, index=default_idx)
        out_cols = extract_template_header(tbytes, template_sheet)
    except Exception:
        st.warning("Could not read template; using default Star Walk columns.")
        out_cols = DEFAULT_STARWALK_COLUMNS

    allowed_det, allowed_del = load_allowed_l2_from_template(tbytes)

st.subheader("Field Mapping (optional but recommended)")
all_src_cols = list(src_df.columns)

# Preferred defaults for the most common "raw website reviews" export schema
PREFERRED_CORE_DEFAULTS: Dict[str, List[str]] = {
    "Source": ["Retailer"],
    "Model (SKU)": ["SKU Item"],
    "Seeded": ["Seeded Reviews [AX]"],
    "Country": ["Location"],
    "New Review": ["Syndicated Reviews [AX]"],
    "Review Date": ["Opened date", "Opened Date"],
    "Verbatim": ["Review"],
    "Star Rating": ["Rating"],
}

def pick_preferred(field: str) -> Optional[str]:
    prefs = PREFERRED_CORE_DEFAULTS.get(field, [])
    for p in prefs:
        if p in all_src_cols:
            return p
    for p in prefs:
        pn = _norm(p)
        for c in all_src_cols:
            if _norm(c) == pn:
                return c
    return None

suggest = {
    "Source": pick_preferred("Source") or best_match("Source", all_src_cols),
    "Model (SKU)": pick_preferred("Model (SKU)") or best_match("Model (SKU)", all_src_cols) or best_match("SKU", all_src_cols),
    "Seeded": pick_preferred("Seeded") or best_match("Seeded", all_src_cols),
    "Country": pick_preferred("Country") or best_match("Country", all_src_cols),
    "New Review": pick_preferred("New Review") or best_match("New Review", all_src_cols),
    "Review Date": pick_preferred("Review Date") or best_match("Review Date", all_src_cols) or best_match("Date", all_src_cols),
    "Verbatim Id": None,
    "Verbatim": pick_preferred("Verbatim") or best_match("Verbatim", all_src_cols) or best_match("Review", all_src_cols),
    "Star Rating": pick_preferred("Star Rating") or best_match("Star Rating", all_src_cols) or best_match("Rating", all_src_cols),
    "Hair Type": None,
}

field_map: Dict[str, Optional[str]] = {}
core_fields = ["Source", "Model (SKU)", "Seeded", "Country", "New Review", "Review Date", "Verbatim Id", "Verbatim", "Star Rating", "Hair Type"]
left, right = st.columns(2)
for i, f in enumerate(core_fields):
    with (left if i < (len(core_fields) + 1) // 2 else right):
        options = [None] + all_src_cols
        default_val = suggest.get(f)
        default_idx = options.index(default_val) if default_val in all_src_cols else 0
        field_map[f] = st.selectbox(
            f"Output '{f}' comes from source column:",
            options=options,
            index=default_idx,
            key=f"map_{f}",
        )

st.subheader("L2 Tag Columns (these populate Symptom 1–20)")
with st.expander("Select Level 2 detractor/delighter columns", expanded=True):
    # Heuristics that work for Axion exports (+ general)
    l2_det_guess = [c for c in all_src_cols if "l2" in _norm(c) and "detr" in _norm(c)]
    l2_del_guess = [c for c in all_src_cols if "l2" in _norm(c) and ("delight" in _norm(c) or "promot" in _norm(c) or "delighter" in _norm(c))]

    l2_detractor_cols = st.multiselect(
        "Source column(s) that contain **Level 2 Detractors**",
        options=all_src_cols,
        default=l2_det_guess,
    )
    l2_delighter_cols = st.multiselect(
        "Source column(s) that contain **Level 2 Delighters**",
        options=all_src_cols,
        default=l2_del_guess,
    )

    split_choice = st.selectbox(
        "How are multiple tags separated inside a cell?",
        options=[
            "Semicolon / Comma / Pipe / Newline (recommended)",
            "Semicolon only ;",
            "Comma only ,",
            "Pipe only |",
            "Newline only",
        ],
        index=0,
    )

    if split_choice == "Semicolon / Comma / Pipe / Newline (recommended)":
        split_regex = r"[;\|\n,]+"
    elif split_choice == "Semicolon only ;":
        split_regex = r"[;]+"
    elif split_choice == "Comma only ,":
        split_regex = r"[,]+"
    elif split_choice == "Pipe only |":
        split_regex = r"[\|]+"
    else:
        split_regex = r"[\n]+"

st.subheader("Weight / Count Column")
weight_mode = st.selectbox(
    "How should 'Review count per detractor' be filled?",
    options=[
        "Leave blank",
        "1 / # detractor symptoms (if any)",
        "1 / # delighter symptoms (if any)",
        "1 / # total symptoms (detractors + delighters)",
        "Always 1",
    ],
    index=0,
)

validate_l2 = False
filter_unknown = False
if template_file and (allowed_det or allowed_del):
    with st.expander("Optional: Validate L2 tags against template 'Symptoms' list", expanded=False):
        validate_l2 = st.checkbox("Enable validation (report unknown tags)", value=True)
        filter_unknown = st.checkbox("Drop tags not in the template list", value=False)

st.subheader("Convert & Download")
build = st.button("Convert to Star Walk scrubbed verbatims")

if build:
    if not l2_detractor_cols and not l2_delighter_cols:
        st.error("Please select at least one L2 detractor and/or L2 delighter column.")
        st.stop()

    used_allowed_det = allowed_det if validate_l2 else None
    used_allowed_del = allowed_del if validate_l2 else None

    out_df, stats = convert_to_starwalk(
        src_df=src_df,
        out_cols=out_cols,
        field_map=field_map,
        l2_det_cols=l2_detractor_cols,
        l2_del_cols=l2_delighter_cols,
        split_regex=split_regex,
        weight_mode=weight_mode,
        allowed_det=used_allowed_det,
        allowed_del=used_allowed_del,
        filter_unknown=filter_unknown,
    )

    st.success("Conversion complete.")

    # Diagnostics: Seeded normalization must be only "yes" or blank
    seeded_col = find_col_by_norm(list(out_df.columns), "seeded")
    if seeded_col:
        seeded_vals = out_df[seeded_col]
        seeded_yes = int((seeded_vals == "yes").sum())
        nonempty = seeded_vals.dropna()
        bad = nonempty[(nonempty != "yes") & (nonempty.astype(str).str.strip() != "")]
        st.caption(f"Seeded normalization: **{seeded_yes}** yes out of **{len(out_df)}** rows")
        if len(bad) > 0:
            st.error(f"Seeded still contains non-'yes' values (showing first 10): {bad.head(10).tolist()}")

    c1, c2, c3, c4 = st.columns(4)
    detr_counts = out_df[[f"Symptom {i}" for i in range(1, 11)]].notna().sum(axis=1)
    deli_counts = out_df[[f"Symptom {i}" for i in range(11, 21)]].notna().sum(axis=1)

    with c1:
        st.metric("Rows", stats["rows"])
    with c2:
        st.metric("Avg detractors / row", round(float(detr_counts.mean()), 2))
    with c3:
        st.metric("Avg delighters / row", round(float(deli_counts.mean()), 2))
    with c4:
        st.metric("Rows truncated (>10 tags)", stats["rows_truncated_detractors_gt10"] + stats["rows_truncated_delighters_gt10"])

    if validate_l2:
        st.info(
            f"Unknown tags found — detractors: {stats['unknown_detractor_tags_total']}, "
            f"delighters: {stats['unknown_delighter_tags_total']}"
        )

    st.dataframe(out_df.head(100), use_container_width=True)

    if template_file and output_mode.startswith("Use template"):
        xbytes = write_into_template_workbook(template_file.getvalue(), out_df, sheet_name=STARWALK_SHEET_NAME)
        fname = "starwalk_converted_TEMPLATE.xlsx"
    else:
        xbytes = to_excel_one_sheet(out_df, sheet_name=STARWALK_SHEET_NAME)
        fname = "starwalk_scrubbed_verbatims_converted.xlsx"

    st.download_button(
        "Download Excel",
        data=xbytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )




