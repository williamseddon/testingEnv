import io
import json
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

EXCEL_MAX_CHARS = 32767

# =============================
# Helpers
# =============================
def safe_get(d: Dict[str, Any], path: List[str], default: Any = None) -> Any:
    cur: Any = d
    for k in path:
        if isinstance(cur, dict) and k in cur:
            cur = cur[k]
        else:
            return default
    return cur


def as_list(x: Any) -> List[Any]:
    if x is None:
        return []
    if isinstance(x, list):
        return x
    return [x]


def join_list(x: Any, sep: str = " | ") -> Any:
    if x is None:
        return None
    if isinstance(x, list):
        vals = [str(v).strip() for v in x if v is not None and str(v).strip() != ""]
        return sep.join(vals) if vals else None
    return x


def parse_iso_date(x: Any) -> Optional[date]:
    if not x:
        return None
    ts = pd.to_datetime(x, utc=True, errors="coerce")
    if pd.isna(ts):
        return None
    return ts.date()


def title_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    stem = re.sub(r"[_\-]+", " ", stem)
    stem = re.sub(r"([a-z])([A-Z])", r"\1 \2", stem)  # camelCase -> words
    stem = re.sub(r"\s+", " ", stem).strip()
    tokens = stem.split(" ")
    if tokens and any(ch.isdigit() for ch in tokens[0]):
        tokens[0] = tokens[0].upper()
    return " ".join(tokens)


def excel_safe_value(v: Any, list_sep: str = " | ") -> Any:
    """
    Convert arbitrary Python objects into openpyxl-safe scalar types.
    Fixes: ValueError: Cannot convert ['Leshow CN', ...] to Excel
    """
    if v is None:
        return None

    try:
        if pd.isna(v):
            return None
    except Exception:
        pass

    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return None
        return v.to_pydatetime()

    if isinstance(v, (datetime, date, bool, int, float)):
        return v

    if isinstance(v, str):
        return v[: EXCEL_MAX_CHARS - 3] + "..." if len(v) > EXCEL_MAX_CHARS else v

    # numpy scalar
    if hasattr(v, "item") and not isinstance(v, (list, tuple, set, dict)):
        try:
            return excel_safe_value(v.item(), list_sep=list_sep)
        except Exception:
            pass

    if isinstance(v, dict):
        s = json.dumps(v, ensure_ascii=False)
        return s[: EXCEL_MAX_CHARS - 3] + "..." if len(s) > EXCEL_MAX_CHARS else s

    if isinstance(v, (list, tuple, set)):
        parts: List[str] = []
        for x in v:
            if x is None:
                continue
            if isinstance(x, (dict, list, tuple, set)):
                parts.append(json.dumps(x, ensure_ascii=False))
            else:
                parts.append(str(x))
        s = list_sep.join([p.strip() for p in parts if p.strip() != ""])
        if not s:
            return None
        return s[: EXCEL_MAX_CHARS - 3] + "..." if len(s) > EXCEL_MAX_CHARS else s

    s = str(v)
    return s[: EXCEL_MAX_CHARS - 3] + "..." if len(s) > EXCEL_MAX_CHARS else s


# =============================
# JSON -> DataFrames
# =============================
REVIEWS_BASE_COLS: List[Tuple[str, Any]] = [
    ("Record ID", lambda r: r.get("_id")),
    ("Opened Timestamp", lambda r: parse_iso_date(r.get("openedTimestamp"))),
    ("Rating (num)", lambda r: safe_get(r, ["clientAttributes", "Rating (num)"])),
    ("Retailer", lambda r: safe_get(r, ["clientAttributes", "Retailer"])),
    ("Retailer Rating", lambda r: safe_get(r, ["clientAttributes", "Retailer Rating"])),
    ("Model", lambda r: safe_get(r, ["clientAttributes", "Model"])),
    ("Seeded Reviews", lambda r: safe_get(r, ["clientAttributes", "Seeded Reviews"])),
    ("Syndicated/Seeded Reviews", lambda r: safe_get(r, ["clientAttributes", "Syndicated/Seeded Reviews"])),
    ("Location", lambda r: safe_get(r, ["clientAttributes", "Location"])),
    ("Post Link", lambda r: safe_get(r, ["clientAttributes", "Post Link"])),
    ("Title", lambda r: safe_get(r, ["freeText", "Title"])),
    ("Review", lambda r: safe_get(r, ["freeText", "Review"])),
]

REVIEWS_EXTRA_COLS: List[Tuple[str, Any]] = [
    ("Satisfaction Score", lambda r: safe_get(r, ["customAttributes", "Satisfaction Score"])),
    ("Key Review Sentiment_Reviews", lambda r: join_list(safe_get(r, ["customAttributes", "Key Review Sentiment_Reviews"]))),
    ("Key Review Sentiment Type_Reviews", lambda r: join_list(safe_get(r, ["customAttributes", "Key Review Sentiment Type_Reviews"]))),
    ("Dominant Customer Journey Step", lambda r: join_list(safe_get(r, ["customAttributes", "Dominant Customer Journey Step"]))),
    ("Trigger Point_Product", lambda r: join_list(safe_get(r, ["customAttributes", "Trigger Point_Product"]))),
    ("L2 Delighter Component", lambda r: join_list(safe_get(r, ["customAttributes", "L2 Delighter Component"]))),
    ("L2 Delighter Condition", lambda r: join_list(safe_get(r, ["customAttributes", "L2 Delighter Condition"]))),
    ("L2 Delighter Mode", lambda r: join_list(safe_get(r, ["customAttributes", "L2 Delighter Mode"]))),
    ("L3 Non Product Detractors", lambda r: join_list(safe_get(r, ["customAttributes", "L3 Non Product Detractors"]))),
    ("Product_Symptom Component", lambda r: join_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Component"]))),
    ("Product_Symptom Conditions", lambda r: join_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Conditions"]))),
    ("Product_Symptom Mode", lambda r: join_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Mode"]))),
    ("Product Name", lambda r: safe_get(r, ["clientAttributes", "Product Name"])),
    ("Product Category", lambda r: safe_get(r, ["clientAttributes", "Product Category"])),
    ("Base SKU", lambda r: safe_get(r, ["clientAttributes", "Base SKU"])),
    ("Brand", lambda r: safe_get(r, ["clientAttributes", "Brand"])),
    ("Company", lambda r: safe_get(r, ["clientAttributes", "Company"])),
    ("Factory Name", lambda r: safe_get(r, ["clientAttributes", "Factory Name"])),
    ("Translation", lambda r: safe_get(r, ["clientAttributes", "Translation"])),
    ("Event ID", lambda r: r.get("eventId")),
    ("Event Type", lambda r: r.get("eventType")),
    ("Is Linked", lambda r: r.get("isLinked")),
    ("Workspace ID", lambda r: safe_get(r, ["clientAttributes", "Workspace ID"])),
]


def build_reviews_df(records: List[Dict[str, Any]], include_extra: bool = True) -> pd.DataFrame:
    cols = REVIEWS_BASE_COLS + (REVIEWS_EXTRA_COLS if include_extra else [])
    rows = [{name: fn(r) for name, fn in cols} for r in records]
    df = pd.DataFrame(rows)

    if "Rating (num)" in df.columns:
        df["Rating (num)"] = pd.to_numeric(df["Rating (num)"], errors="coerce")

    return df


def build_symptoms_df(records: List[Dict[str, Any]], include_blank_row_when_missing: bool = True) -> pd.DataFrame:
    out_rows: List[Dict[str, Any]] = []

    for r in records:
        rid = r.get("_id")
        opened = parse_iso_date(r.get("openedTimestamp"))
        rating = safe_get(r, ["clientAttributes", "Rating (num)"])
        retailer = safe_get(r, ["clientAttributes", "Retailer"])
        model = safe_get(r, ["clientAttributes", "Model"])

        comps = as_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Component"]))
        conds = as_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Conditions"]))
        modes = as_list(safe_get(r, ["customAttributes", "taxonomies", "Product_Symptom Mode"]))

        max_len = max(len(comps), len(conds), len(modes), 0)

        if max_len == 0 and include_blank_row_when_missing:
            out_rows.append(
                {
                    "Record ID": rid,
                    "Opened Timestamp": opened,
                    "Rating": rating,
                    "Retailer": retailer,
                    "Model": model,
                    "Symptom Index": None,
                    "Symptom Component": None,
                    "Symptom Condition": None,
                    "Symptom Mode": None,
                }
            )
            continue

        for i in range(max_len):
            comp = comps[i] if i < len(comps) else None
            cond = conds[i] if i < len(conds) else None
            mode = modes[i] if i < len(modes) else None

            if mode in (None, "", "-", "—"):
                if comp and cond:
                    mode = f"{comp} - {cond}"
                elif cond:
                    mode = f"- {cond}"
                elif comp:
                    mode = f"{comp} -"
                else:
                    mode = "-"

            out_rows.append(
                {
                    "Record ID": rid,
                    "Opened Timestamp": opened,
                    "Rating": rating,
                    "Retailer": retailer,
                    "Model": model,
                    "Symptom Index": i + 1,
                    "Symptom Component": comp,
                    "Symptom Condition": cond,
                    "Symptom Mode": mode,
                }
            )

    df = pd.DataFrame(out_rows)
    if "Rating" in df.columns:
        df["Rating"] = pd.to_numeric(df["Rating"], errors="coerce")
    return df


# =============================
# Summary Data
# =============================
def build_summary_tables(reviews_df: pd.DataFrame, symptoms_df: pd.DataFrame, top_n: int = 10):
    total_reviews = int(len(reviews_df))

    date_min = reviews_df["Opened Timestamp"].min() if "Opened Timestamp" in reviews_df.columns else None
    date_max = reviews_df["Opened Timestamp"].max() if "Opened Timestamp" in reviews_df.columns else None
    date_range_str = ""
    if pd.notna(date_min) and pd.notna(date_max) and date_min and date_max:
        date_range_str = f"{date_min} to {date_max}"

    avg_rating = None
    if "Rating (num)" in reviews_df.columns:
        avg_rating = float(pd.to_numeric(reviews_df["Rating (num)"], errors="coerce").mean())

    rating_counts = (
        reviews_df["Rating (num)"].dropna().astype(int).value_counts().reindex([5, 4, 3, 2, 1], fill_value=0)
    )
    rating_dist = pd.DataFrame(
        {
            "Rating": rating_counts.index.astype(int),
            "Count": rating_counts.values.astype(int),
            "Share": (rating_counts.values / total_reviews) if total_reviews else 0,
        }
    )

    retailer_counts = reviews_df.get("Retailer", pd.Series(dtype=str)).fillna("(blank)").value_counts().head(top_n)
    top_retailers = pd.DataFrame(
        {
            "Retailer": retailer_counts.index,
            "Count": retailer_counts.values.astype(int),
            "Share": (retailer_counts.values / total_reviews) if total_reviews else 0,
        }
    )

    cond_series = symptoms_df.get("Symptom Condition", pd.Series(dtype=str)).fillna("")
    cond_series = cond_series[cond_series.astype(str).str.strip() != ""]
    symptom_rows = int(len(cond_series))
    cond_counts = cond_series.value_counts().head(top_n)
    top_conditions = pd.DataFrame(
        {
            "Condition": cond_counts.index,
            "Count": cond_counts.values.astype(int),
            "Share (of symptom rows)": (cond_counts.values / symptom_rows) if symptom_rows else 0,
        }
    )

    return {
        "total_reviews": total_reviews,
        "date_range_str": date_range_str,
        "avg_rating": avg_rating,
        "rating_dist": rating_dist,
        "top_retailers": top_retailers,
        "top_conditions": top_conditions,
        "symptom_rows": symptom_rows,
    }


# =============================
# Excel Writer (openpyxl)
# =============================
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")  # dark blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="1F4E79", bold=True, size=14)
SECTION_FONT = Font(color="1F4E79", bold=False)


def write_df_to_sheet(ws, df: pd.DataFrame):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([excel_safe_value(v) for v in row])


def add_excel_table(ws, df: pd.DataFrame, table_name: str, style_name: str = "TableStyleMedium9"):
    nrows = len(df) + 1
    ncols = len(df.columns)
    if nrows <= 1 or ncols == 0:
        return

    ref = f"A1:{get_column_letter(ncols)}{nrows}"
    tab = Table(displayName=table_name, ref=ref)

    style = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    ws.freeze_panes = "A2"


def set_col_widths(ws, df: pd.DataFrame, widths: Dict[str, float]):
    for i, col in enumerate(df.columns, start=1):
        if col in widths:
            ws.column_dimensions[get_column_letter(i)].width = widths[col]


def set_date_format(ws, df: pd.DataFrame, date_cols: List[str]):
    for col_name in date_cols:
        if col_name not in df.columns:
            continue
        col_idx = list(df.columns).index(col_name) + 1
        for r in range(2, len(df) + 2):
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "yyyy-mm-dd"


def apply_hyperlinks(ws, df: pd.DataFrame, url_cols: List[str]):
    for col_name in url_cols:
        if col_name not in df.columns:
            continue
        col_idx = list(df.columns).index(col_name) + 1
        for r in range(2, len(df) + 2):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.style = "Hyperlink"


def wrap_cells(ws, df: pd.DataFrame, wrap_cols: List[str]):
    align = Alignment(wrap_text=True, vertical="top")
    for col_name in wrap_cols:
        if col_name not in df.columns:
            continue
        col_idx = list(df.columns).index(col_name) + 1
        for r in range(1, len(df) + 2):
            ws.cell(row=r, column=col_idx).alignment = align


def build_workbook(
    dataset_title: str,
    reviews_df: pd.DataFrame,
    symptoms_df: pd.DataFrame,
    summary: Dict[str, Any],
    wrap_long_text: bool = False,
) -> bytes:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- Reviews sheet (FIRST TAB) ---
    ws_r = wb.create_sheet("Reviews")
    write_df_to_sheet(ws_r, reviews_df)
    add_excel_table(ws_r, reviews_df, table_name="ReviewsTable")
    apply_hyperlinks(ws_r, reviews_df, url_cols=["Post Link"])
    set_date_format(ws_r, reviews_df, date_cols=["Opened Timestamp"])

    review_widths = {
        "Record ID": 34,
        "Opened Timestamp": 16,
        "Rating (num)": 11,
        "Retailer": 14,
        "Retailer Rating": 14,
        "Model": 12,
        "Seeded Reviews": 15,
        "Syndicated/Seeded Reviews": 22,
        "Location": 10,
        "Post Link": 55,
        "Title": 45,
        "Review": 85,
        "Translation": 55,
        "Factory Name": 22,
    }
    set_col_widths(ws_r, reviews_df, review_widths)
    if wrap_long_text:
        wrap_cells(ws_r, reviews_df, wrap_cols=["Title", "Review", "Translation"])

    # --- Summary sheet ---
    ws = wb.create_sheet("Summary")
    ws["A1"] = f"{dataset_title} — Summary"
    ws["A1"].font = TITLE_FONT

    ws["A3"] = "Dataset"
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "Total Reviews"
    ws["B4"] = summary["total_reviews"]

    ws["A5"] = "Date Range (Opened)"
    ws["B5"] = summary["date_range_str"]

    ws["A6"] = "Average Rating"
    if summary["avg_rating"] is not None:
        ws["B6"] = round(summary["avg_rating"], 3)

    ws["A8"] = "Rating Distribution"
    ws["A8"].font = SECTION_FONT
    ws["D8"] = "Top Retailers"
    ws["D8"].font = SECTION_FONT

    rd = summary["rating_dist"].copy()
    rd_cols = ["Rating", "Count", "Share"]
    rd = rd[rd_cols]
    start_row, start_col = 9, 1
    for j, col in enumerate(rd_cols):
        cell = ws.cell(row=start_row, column=start_col + j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(rd.itertuples(index=False, name=None), start=1):
        for j, val in enumerate(row):
            c = ws.cell(row=start_row + i, column=start_col + j, value=excel_safe_value(val))
            if j == 2:
                c.number_format = "0.0%"

    tr = summary["top_retailers"].copy()
    tr_cols = ["Retailer", "Count", "Share"]
    tr = tr[tr_cols]
    start_row, start_col = 9, 4
    for j, col in enumerate(tr_cols):
        cell = ws.cell(row=start_row, column=start_col + j, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(tr.itertuples(index=False, name=None), start=1):
        for j, val in enumerate(row):
            c = ws.cell(row=start_row + i, column=start_col + j, value=excel_safe_value(val))
            if j == 2:
                c.number_format = "0.0%"

    ws["A22"] = "Top Symptom Conditions"
    ws["A22"].font = SECTION_FONT
    ws["E22"] = "Notes"
    ws["E22"].font = SECTION_FONT

    ws["E23"] = (
        "Symptoms sheet is exploded from\n"
        "customAttributes.taxonomies lists\n"
        "(Component / Condition / Mode)."
    )
    ws["E23"].alignment = Alignment(wrap_text=True, vertical="top")

    tc = summary["top_conditions"].copy()
    tc_cols = ["Condition", "Count", "Share (of symptom rows)"]
    tc = tc[tc_cols]
    start_row, start_col = 23, 1
    for j, col in enumerate(tc_cols):
        cell = ws.cell(row=start_row, column=start_col + j, value=col if j < 2 else "Share")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        if j == 2:
            ws.column_dimensions[get_column_letter(start_col + j)].width = 6
            cell.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(tc.itertuples(index=False, name=None), start=1):
        for j, val in enumerate(row):
            c = ws.cell(row=start_row + i, column=start_col + j, value=excel_safe_value(val))
            if j == 2:
                c.number_format = "0.0%"

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 10

    # --- Symptoms sheet ---
    ws_s = wb.create_sheet("Symptoms")
    write_df_to_sheet(ws_s, symptoms_df)
    add_excel_table(ws_s, symptoms_df, table_name="SymptomsTable")
    set_date_format(ws_s, symptoms_df, date_cols=["Opened Timestamp"])

    symptom_widths = {
        "Record ID": 34,
        "Opened Timestamp": 16,
        "Rating": 8,
        "Retailer": 14,
        "Model": 12,
        "Symptom Index": 12,
        "Symptom Component": 22,
        "Symptom Condition": 38,
        "Symptom Mode": 38,
    }
    set_col_widths(ws_s, symptoms_df, symptom_widths)

    # Make Reviews the active sheet when opening
    wb.active = wb.sheetnames.index("Reviews")

    # Save to bytes
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# =============================
# Streamlit UI
# =============================
st.set_page_config(page_title="JSON → Clean Excel (Reviews)", layout="wide")
st.title("JSON → Clean Excel Converter (Reviews format)")
st.caption(
    "Upload a JSON file **or** paste JSON text. The app will parse it, normalize list fields, "
    "and export a formatted Excel workbook (Reviews / Summary / Symptoms)."
)


# -----------------------------
# Flexible JSON parsing
# -----------------------------
def _strip_code_fences(s: str) -> str:
    s = s.strip()
    m = re.search(r"```(?:json)?\s*(.*?)```", s, flags=re.DOTALL | re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return s


def _extract_json_substring(s: str) -> str:
    s = s.strip()
    if s.startswith("{") or s.startswith("["):
        return s
    start_candidates = [i for i in [s.find("{"), s.find("[")] if i != -1]
    if not start_candidates:
        return s
    start = min(start_candidates)
    end = max(s.rfind("}"), s.rfind("]"))
    if end != -1 and end > start:
        return s[start : end + 1].strip()
    return s


def _try_parse_json_lines(s: str) -> Optional[List[Dict[str, Any]]]:
    lines = [ln.strip() for ln in s.splitlines() if ln.strip()]
    if len(lines) < 2:
        return None
    objs: List[Dict[str, Any]] = []
    for ln in lines:
        try:
            obj = json.loads(ln)
            if isinstance(obj, dict):
                objs.append(obj)
            else:
                return None
        except Exception:
            return None
    return objs


def loads_flexible_json(text_in: str) -> Tuple[Any, List[str]]:
    warnings: List[str] = []
    s = _strip_code_fences(text_in)
    s = _extract_json_substring(s)

    try:
        return json.loads(s), warnings
    except Exception as e1:
        # Attempt: remove trailing commas
        s2 = re.sub(r",\s*([}\]])", r"\1", s)
        if s2 != s:
            try:
                warnings.append("Removed trailing commas to make JSON valid.")
                return json.loads(s2), warnings
            except Exception:
                warnings.pop()

        # Attempt: JSON Lines
        jl = _try_parse_json_lines(s)
        if jl is not None:
            warnings.append("Detected JSON Lines and parsed each line as a record.")
            return jl, warnings

        raise ValueError("Could not parse input as JSON. Paste valid JSON (or JSON Lines).") from e1


def extract_records(raw: Any) -> List[Dict[str, Any]]:
    if isinstance(raw, dict) and "results" in raw and isinstance(raw["results"], list):
        return raw["results"]
    if isinstance(raw, list):
        return [r for r in raw if isinstance(r, dict)]
    raise ValueError("Unrecognized JSON shape. Expected a dict with `results: []` or a list of record objects.")


# =============================
# UI Inputs (Upload OR Paste)
# =============================
st.subheader("1) Provide input JSON")
left, right = st.columns(2)

uploaded = None
pasted = ""

with left:
    uploaded = st.file_uploader("Upload JSON file", type=["json"])
with right:
    pasted = st.text_area(
        "…or paste JSON text",
        height=210,
        placeholder="Paste JSON (or ```json ... ```). Supports dict-with-results, list of records, or JSON Lines.",
    )

prefer_upload = uploaded is not None and getattr(uploaded, "size", 0) > 0
prefer_paste = (not prefer_upload) and pasted.strip() != ""

raw_text: Optional[str] = None
source_name: str = "pasted_json"

if prefer_upload:
    source_name = uploaded.name
    raw_text = uploaded.getvalue().decode("utf-8", errors="replace")
elif prefer_paste:
    source_name = "pasted_json"
    raw_text = pasted

auto_convert = st.checkbox("Auto-convert when input looks complete", value=True)
convert_clicked = st.button("Convert to Excel", type="primary")


def looks_complete_json(s: str) -> bool:
    ss = s.strip()
    if not ss:
        return False
    if ss.endswith(("}", "]")):
        return True
    if ss.endswith("```"):
        return True
    return False


should_run = False
if raw_text is not None:
    if convert_clicked:
        should_run = True
    elif auto_convert:
        if prefer_upload or looks_complete_json(raw_text):
            should_run = True

# =============================
# Options
# =============================
st.subheader("2) Options")
col1, col2, col3, col4 = st.columns(4)
with col1:
    include_extra = st.checkbox("Include extra tag/taxonomy columns", value=True)
with col2:
    wrap_long_text = st.checkbox("Wrap Title/Review text", value=False)
with col3:
    include_blank_symptom_rows = st.checkbox("Keep 1 blank symptom row when no symptoms exist", value=True)
with col4:
    top_n = st.number_input("Top N in Summary", min_value=5, max_value=25, value=10, step=1)

# =============================
# Convert
# =============================
if raw_text is None:
    st.info("Upload a JSON file or paste JSON text to begin.")
elif not should_run:
    if prefer_paste and auto_convert:
        st.info("Paste looks incomplete — finish the JSON (ending in } or ]) or click **Convert to Excel**.")
else:
    try:
        raw_obj, warnings = loads_flexible_json(raw_text)
        for w in warnings:
            st.info(w)

        records = extract_records(raw_obj)
        if not records:
            st.error("Parsed input, but found 0 record objects to convert.")
            st.stop()

        default_title = title_from_filename(source_name) if source_name else "Pasted JSON"
        dataset_title = st.text_input("Dataset title (used in Summary sheet title)", value=default_title)

        reviews_df = build_reviews_df(records, include_extra=include_extra)
        symptoms_df = build_symptoms_df(records, include_blank_row_when_missing=include_blank_symptom_rows)
        summary = build_summary_tables(reviews_df, symptoms_df, top_n=int(top_n))

        st.success(f"Loaded {len(records):,} records.")
        m1, m2, m3 = st.columns(3)
        m1.metric("Total Reviews", summary["total_reviews"])
        if summary["avg_rating"] is not None:
            m2.metric("Average Rating", f"{summary['avg_rating']:.3f}")
        m3.metric("Symptom rows (non-blank conditions)", summary["symptom_rows"])

        with st.expander("Show parsed JSON structure (first record)", expanded=False):
            st.json(records[0])

        st.subheader("Preview — Reviews (first 50 rows)")
        st.dataframe(reviews_df.head(50), use_container_width=True)

        st.subheader("Preview — Symptoms (first 50 rows)")
        st.dataframe(symptoms_df.head(50), use_container_width=True)

        excel_bytes = build_workbook(
            dataset_title=f"{dataset_title} Reviews" if "review" not in dataset_title.lower() else dataset_title,
            reviews_df=reviews_df,
            symptoms_df=symptoms_df,
            summary=summary,
            wrap_long_text=wrap_long_text,
        )

        out_name_default = f"{Path(source_name).stem}_clean.xlsx" if source_name else "converted_clean.xlsx"
        out_name = st.text_input("Output filename", value=out_name_default)

        st.download_button(
            "Download Excel",
            data=excel_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.caption(
            "Note: List/dict fields are automatically converted to Excel-safe strings "
            "(joined with ' | ' or JSON). This prevents openpyxl errors like "
            "`ValueError: Cannot convert ['Leshow CN', ...] to Excel`."
        )

    except Exception as e:
        st.exception(e)
